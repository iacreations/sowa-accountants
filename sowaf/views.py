from django.urls import reverse, NoReverseMatch
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from django.db import transaction
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta, time, date
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import now
import openpyxl
import csv
import io
import os
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from collections import OrderedDict

from django.db.models import (
    Sum, Value, F, Q, DecimalField, ExpressionWrapper,
    Count, FloatField, When, Case, OuterRef, Subquery
)
from django.db.models.functions import Coalesce, Cast
from django.core.files import File
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from sales.views import _invoice_analytics
from sales.models import Newinvoice, Payment, PaymentInvoice, SalesReceipt
from accounts.models import Account, JournalEntry, JournalLine
from accounts.utils import deposit_accounts_qs, expense_accounts_qs
from accounts.date_ranges import resolve_date_range, RANGE_LABELS, RANGE_OPTIONS
from expenses.models import Bill, Expense, Cheque
from tenancy.permissions import company_required, module_required
from tenancy.models import Company
from .utils import _supplier_ap_balances_bulk
from .models import Newcustomer, Newsupplier, Newemployee, Newasset


# Constants / helpers
def parse_date_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _dec(val, default="0.00") -> Decimal:
    try:
        return Decimal(str(val)) if val not in (None, "", "None", "null") else Decimal(default)
    except Exception:
        return Decimal(default)


def _safe_date(val, fallback=None):
    """
    Accepts a date or a yyyy-mm-dd string. Returns a date.
    """
    if fallback is None:
        fallback = timezone.localdate()

    if not val:
        return fallback

    if hasattr(val, "year"):
        return val  # already date/datetime

    try:
        return timezone.datetime.fromisoformat(str(val)).date()
    except Exception:
        return fallback


def D(x) -> Decimal:
    try:
        return Decimal(str(x or "0.00"))
    except Exception:
        return Decimal("0.00")


def get_or_create_ar_account(company):
    """
    Tenant-safe A/R control account.
    """
    ar = (
        Account.objects.for_company(company)
        .filter(account_name="Accounts Receivable")
        .first()
    )
    if ar:
        return ar

    return Account.objects.create(
        company=company,
        account_name="Accounts Receivable",
        account_type="CURRENT_ASSET",
        detail_type="Accounts receivable",
        is_active=True,
    )


def post_journal_entry(*, company, date, description, source_type, source_id, lines):
    """
    lines = [
      {"account": Account, "debit": Decimal, "credit": Decimal, "customer": Newcustomer|None, "supplier": Newsupplier|None},
      ...
    ]
    """
    with transaction.atomic():
        je = JournalEntry.objects.create(
            company=company,
            date=date,
            description=description,
            source_type=source_type,
            source_id=source_id,
        )

        for ln in lines:
            JournalLine.objects.create(
                entry=je,
                account=ln["account"],
                debit=D(ln.get("debit")),
                credit=D(ln.get("credit")),
                customer=ln.get("customer"),
                supplier=ln.get("supplier"),
            )
        return je


def customer_ar_balance(customer: Newcustomer, company=None) -> Decimal:
    """
    AR = (DR - CR) on Accounts Receivable lines filtered by this customer.
    This becomes the customer's current open balance.
    """
    company = company or customer.company
    ar = get_or_create_ar_account(company)
    agg = JournalLine.objects.filter(
        entry__company=company,
        account=ar,
        customer=customer
    ).aggregate(
        dr=Sum("debit"),
        cr=Sum("credit"),
    )
    dr = D(agg.get("dr"))
    cr = D(agg.get("cr"))
    return dr - cr


def _get_or_create_opening_equity(company):
    opening_equity = (
        Account.objects.for_company(company)
        .filter(account_name="Opening Balance Equity")
        .first()
    )
    if opening_equity:
        return opening_equity

    return Account.objects.create(
        company=company,
        account_name="Opening Balance Equity",
        account_type="OWNER_EQUITY",
        detail_type="Opening balances",
        is_active=True,
    )


def _get_or_create_ar_control_account(company):
    """
    A/R control must be a CURRENT_ASSET with detail_type 'Accounts Receivable (A/R)'.
    """
    ar = (
        Account.objects.for_company(company)
        .filter(detail_type__iexact="Accounts Receivable (A/R)", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="accounts receivable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="receivable", is_active=True).first()
    )
    if ar:
        return ar

    # Auto-create if missing
    ar = Account.objects.create(
        company=company,
        account_name="Accounts Receivable",
        account_type="CURRENT_ASSET",
        detail_type="Accounts Receivable (A/R)",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return ar


def _get_or_create_ap_control_account(company):
    """
    A/P control must be a CURRENT_LIABILITY with detail_type 'Accounts Payable (A/P)'.
    """
    ap = (
        Account.objects.for_company(company)
        .filter(detail_type__iexact="Accounts Payable (A/P)", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="accounts payable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="payable", is_active=True).first()
    )
    if ap:
        return ap

    # Auto-create if missing
    ap = Account.objects.create(
        company=company,
        account_name="Accounts Payable",
        account_type="CURRENT_LIABILITY",
        detail_type="Accounts Payable (A/P)",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return ap


def _get_or_create_customer_ar_subaccount(customer: Newcustomer) -> Account:
    """
    Creates/returns a child account under A/R control specifically for this customer.
    """
    company = customer.company
    ar_control = _get_or_create_ar_control_account(company)

    # Use a stable unique name
    cust_name = (customer.customer_name or customer.company_name or f"Customer {customer.id}").strip()
    sub_name = f"{cust_name}"

    # try find existing subaccount for this customer (best effort by name+parent)
    acc = Account.objects.for_company(company).filter(
        parent=ar_control,
        is_subaccount=True,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()
    if acc:
        return acc

    # create
    acc = Account.objects.create(
        company=company,
        account_name=sub_name,
        account_type=ar_control.account_type,  # CURRENT_ASSET
        detail_type="Customer ledger (A/R)",
        is_active=True,
        is_subaccount=True,
        parent=ar_control,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return acc


def _get_or_create_supplier_ap_subaccount(supplier: Newsupplier) -> Account:
    """
    Creates/returns a child account under A/P control specifically for this supplier.
    """
    company = supplier.company
    ap_control = _get_or_create_ap_control_account(company)

    sup_name = (supplier.company_name or f"Supplier {supplier.id}").strip()
    sub_name = f"{sup_name}"

    acc = Account.objects.for_company(company).filter(
        parent=ap_control,
        is_subaccount=True,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()
    if acc:
        return acc

    acc = Account.objects.create(
        company=company,
        account_name=sub_name,
        account_type=ap_control.account_type,  # CURRENT_LIABILITY
        detail_type="Supplier ledger (A/P)",
        is_active=True,
        is_subaccount=True,
        parent=ap_control,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return acc


def _upsert_opening_balance_je(
    *,
    company,
    source_type: str,
    source_id: int,
    je_date,
    description: str,
    dr_account: Account,
    cr_account: Account,
    amount: Decimal,
):
    """
    Creates or updates ONE JE for opening balance, replaces its lines.
    Also sets supplier/customer links on the CR side when source_type matches.

    - Supplier opening: source_type="SUPP_OPEN_BALANCE" source_id=supplier.id
      -> CR is supplier AP subaccount, should have supplier_id set.
    - Customer opening: source_type="CUST_OPEN_BALANCE" source_id=customer.id
      -> DR/CR accordingly; usually DR customer AR subaccount, should have customer_id set.
    """
    je_date = _safe_date(je_date, timezone.localdate())

    je = JournalEntry.objects.filter(
        company=company,
        source_type=source_type,
        source_id=source_id
    ).first()

    if amount == 0 or amount == Decimal("0.00"):
        if je:
            je.delete()
        return

    if not je:
        je = JournalEntry.objects.create(
            company=company,
            date=je_date,
            description=description,
            source_type=source_type,
            source_id=source_id,
        )
    else:
        je.date = je_date
        je.description = description
        je.save(update_fields=["date", "description"])

        # replace lines
        JournalLine.objects.filter(entry=je).delete()

    # Decide sub-ledger linkage
    supplier_obj = None
    customer_obj = None

    if source_type == "SUPP_OPEN_BALANCE":
        supplier_obj = Newsupplier.objects.for_company(company).filter(pk=source_id).first()

    if source_type == "CUST_OPEN_BALANCE":
        customer_obj = Newcustomer.objects.for_company(company).filter(pk=source_id).first()

    # DR line (normally Opening Equity)
    JournalLine.objects.create(
        entry=je,
        account=dr_account,
        debit=amount,
        credit=Decimal("0.00"),
        supplier=None,
        customer=None,
    )

    # CR line (supplier/customer sub-ledger depending on source_type)
    JournalLine.objects.create(
        entry=je,
        account=cr_account,
        debit=Decimal("0.00"),
        credit=amount,
        supplier=supplier_obj,
        customer=customer_obj,
    )
    return je


def _get_or_create_sales_income_account(company) -> Account:
    acc = (
        Account.objects.for_company(company)
        .filter(account_name__iexact="Sales Income", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(
            account_name__icontains="sales",
            account_type__in=["OPERATING_INCOME", "INVESTING_INCOME"],
            is_active=True
        ).first()
    )
    if acc:
        return acc

    return Account.objects.create(
        company=company,
        account_name="Sales Income",
        account_type="OPERATING_INCOME",
        detail_type="Sales",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )


def _get_or_create_vat_payable_account(company) -> Account:
    acc = (
        Account.objects.for_company(company)
        .filter(account_name__iexact="VAT Payable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(
            account_name__icontains="vat",
            account_type__in=["CURRENT_LIABILITY", "NON_CURRENT_LIABILITY"],
            is_active=True
        ).first()
    )
    if acc:
        return acc

    return Account.objects.create(
        company=company,
        account_name="VAT Payable",
        account_type="CURRENT_LIABILITY",
        detail_type="Taxes payable",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )


def _as_date(d):
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None


def status_for_invoice(inv, total: Decimal, paid: Decimal, balance: Decimal) -> str:
    """
    Simple, consistent status for an invoice.
    total/paid/balance are Decimals precomputed by the caller.
    """
    today = timezone.now().date()
    if balance <= 0:
        return "Paid"
    # overdue if due_date exists and is in the past
    due = getattr(inv, "due_date", None)
    if due and due < today:
        return "Overdue"
    if paid > 0:
        return "Partially paid"
    return "Open"


# HOME CHARTS
DEC = DecimalField(max_digits=18, decimal_places=2)
ZERO = Value(Decimal("0.00"), output_field=DEC)

# COA codes (preferred)
INCOME_TYPES = ["OPERATING_INCOME", "INVESTING_INCOME"]
EXPENSE_TYPES = ["OPERATING_EXPENSE", "INVESTING_EXPENSE", "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE"]

# fallback keywords (because your P&L report uses text matching)
INCOME_KW_RE = r"income|revenue|sales"
EXPENSE_KW_RE = r"expense|expenses|cogs|cost of sales|cost of goods"


def dec(v) -> Decimal:
    """Safe Decimal converter."""
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0.00")


def bankish_q():
    return (
        Q(detail_type__icontains="bank") |
        Q(detail_type__icontains="cash") |
        Q(detail_type__icontains="cash and cash equivalents") |
        Q(detail_type__icontains="cash on hand")
    )


def _income_jl_filter():
    return (
        Q(account__account_type__in=INCOME_TYPES) |
        Q(account__parent__account_type__in=INCOME_TYPES) |
        Q(account__account_name__iregex=INCOME_KW_RE) |
        Q(account__detail_type__iregex=INCOME_KW_RE) |
        Q(account__parent__account_name__iregex=INCOME_KW_RE) |
        Q(account__parent__detail_type__iregex=INCOME_KW_RE)
    )


def _expense_jl_filter():
    return (
        Q(account__account_type__in=EXPENSE_TYPES) |
        Q(account__parent__account_type__in=EXPENSE_TYPES) |
        Q(account__account_name__iregex=EXPENSE_KW_RE) |
        Q(account__detail_type__iregex=EXPENSE_KW_RE) |
        Q(account__parent__account_name__iregex=EXPENSE_KW_RE) |
        Q(account__parent__detail_type__iregex=EXPENSE_KW_RE)
    )


def _to_decimal_number(x) -> Decimal:
    """
    Convert DB numeric (float/int/Decimal/None) safely to Decimal.
    Using str(...) avoids float binary issues more than Decimal(float).
    """
    if x is None:
        return Decimal("0")
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")
from django.urls import reverse, NoReverseMatch
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from django.db import transaction
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta, time, date
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import now
import openpyxl
import csv
import io
import os
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from collections import OrderedDict

from django.db.models import (
    Sum, Value, F, Q, DecimalField, ExpressionWrapper,
    Count, FloatField, When, Case, OuterRef, Subquery
)
from django.db.models.functions import Coalesce, Cast
from django.core.files import File
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from sales.views import _invoice_analytics
from sales.models import Newinvoice, Payment, PaymentInvoice, SalesReceipt
from accounts.models import Account, JournalEntry, JournalLine
from accounts.utils import deposit_accounts_qs, expense_accounts_qs
from accounts.date_ranges import resolve_date_range, RANGE_LABELS, RANGE_OPTIONS
from expenses.models import Bill, Expense, Cheque
from tenancy.permissions import company_required, module_required
from .utils import _supplier_ap_balances_bulk
from .models import Newcustomer, Newsupplier, Newemployee, Newasset


# Constants / helpers
def parse_date_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _dec(val, default="0.00") -> Decimal:
    try:
        return Decimal(str(val)) if val not in (None, "", "None", "null") else Decimal(default)
    except Exception:
        return Decimal(default)


def _safe_date(val, fallback=None):
    """
    Accepts a date or a yyyy-mm-dd string. Returns a date.
    """
    if fallback is None:
        fallback = timezone.localdate()

    if not val:
        return fallback

    if hasattr(val, "year"):
        return val  # already date/datetime

    try:
        return timezone.datetime.fromisoformat(str(val)).date()
    except Exception:
        return fallback


def D(x) -> Decimal:
    try:
        return Decimal(str(x or "0.00"))
    except Exception:
        return Decimal("0.00")


def get_or_create_ar_account(company):
    """
    Tenant-safe A/R control account.
    """
    ar = (
        Account.objects.for_company(company)
        .filter(account_name="Accounts Receivable")
        .first()
    )
    if ar:
        return ar

    return Account.objects.create(
        company=company,
        account_name="Accounts Receivable",
        account_type="CURRENT_ASSET",
        detail_type="Accounts receivable",
        is_active=True,
    )


def post_journal_entry(*, company, date, description, source_type, source_id, lines):
    """
    lines = [
      {"account": Account, "debit": Decimal, "credit": Decimal, "customer": Newcustomer|None, "supplier": Newsupplier|None},
      ...
    ]
    """
    with transaction.atomic():
        je = JournalEntry.objects.create(
            company=company,
            date=date,
            description=description,
            source_type=source_type,
            source_id=source_id,
        )

        for ln in lines:
            JournalLine.objects.create(
                entry=je,
                account=ln["account"],
                debit=D(ln.get("debit")),
                credit=D(ln.get("credit")),
                customer=ln.get("customer"),
                supplier=ln.get("supplier"),
            )
        return je


def customer_ar_balance(customer: Newcustomer, company=None) -> Decimal:
    """
    AR = (DR - CR) on Accounts Receivable lines filtered by this customer.
    This becomes the customer's current open balance.
    """
    company = company or customer.company
    ar = get_or_create_ar_account(company)
    agg = JournalLine.objects.filter(
        entry__company=company,
        account=ar,
        customer=customer
    ).aggregate(
        dr=Sum("debit"),
        cr=Sum("credit"),
    )
    dr = D(agg.get("dr"))
    cr = D(agg.get("cr"))
    return dr - cr


def _get_or_create_opening_equity(company):
    opening_equity = (
        Account.objects.for_company(company)
        .filter(account_name="Opening Balance Equity")
        .first()
    )
    if opening_equity:
        return opening_equity

    return Account.objects.create(
        company=company,
        account_name="Opening Balance Equity",
        account_type="OWNER_EQUITY",
        detail_type="Opening balances",
        is_active=True,
    )


def _get_or_create_ar_control_account(company):
    """
    A/R control must be a CURRENT_ASSET with detail_type 'Accounts Receivable (A/R)'.
    """
    ar = (
        Account.objects.for_company(company)
        .filter(detail_type__iexact="Accounts Receivable (A/R)", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="accounts receivable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="receivable", is_active=True).first()
    )
    if ar:
        return ar

    # Auto-create if missing
    ar = Account.objects.create(
        company=company,
        account_name="Accounts Receivable",
        account_type="CURRENT_ASSET",
        detail_type="Accounts Receivable (A/R)",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return ar


def _get_or_create_ap_control_account(company):
    """
    A/P control must be a CURRENT_LIABILITY with detail_type 'Accounts Payable (A/P)'.
    """
    ap = (
        Account.objects.for_company(company)
        .filter(detail_type__iexact="Accounts Payable (A/P)", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="accounts payable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(account_name__icontains="payable", is_active=True).first()
    )
    if ap:
        return ap

    # Auto-create if missing
    ap = Account.objects.create(
        company=company,
        account_name="Accounts Payable",
        account_type="CURRENT_LIABILITY",
        detail_type="Accounts Payable (A/P)",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return ap


def _get_or_create_customer_ar_subaccount(customer: Newcustomer) -> Account:
    """
    Creates/returns a child account under A/R control specifically for this customer.
    """
    company = customer.company
    ar_control = _get_or_create_ar_control_account(company)

    cust_name = (customer.customer_name or customer.company_name or f"Customer {customer.id}").strip()
    sub_name = f"{cust_name}"

    acc = Account.objects.for_company(company).filter(
        parent=ar_control,
        is_subaccount=True,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()
    if acc:
        return acc

    acc = Account.objects.create(
        company=company,
        account_name=sub_name,
        account_type=ar_control.account_type,
        detail_type="Customer ledger (A/R)",
        is_active=True,
        is_subaccount=True,
        parent=ar_control,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return acc


def _get_or_create_supplier_ap_subaccount(supplier: Newsupplier) -> Account:
    """
    Creates/returns a child account under A/P control specifically for this supplier.
    """
    company = supplier.company
    ap_control = _get_or_create_ap_control_account(company)

    sup_name = (supplier.company_name or f"Supplier {supplier.id}").strip()
    sub_name = f"{sup_name}"

    acc = Account.objects.for_company(company).filter(
        parent=ap_control,
        is_subaccount=True,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()
    if acc:
        return acc

    acc = Account.objects.create(
        company=company,
        account_name=sub_name,
        account_type=ap_control.account_type,
        detail_type="Supplier ledger (A/P)",
        is_active=True,
        is_subaccount=True,
        parent=ap_control,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )
    return acc


def _upsert_opening_balance_je(
    *,
    company,
    source_type: str,
    source_id: int,
    je_date,
    description: str,
    dr_account: Account,
    cr_account: Account,
    amount: Decimal,
):
    """
    Creates or updates ONE JE for opening balance, replaces its lines.
    Also sets supplier/customer links on the CR side when source_type matches.
    """
    je_date = _safe_date(je_date, timezone.localdate())

    je = JournalEntry.objects.filter(
        company=company,
        source_type=source_type,
        source_id=source_id
    ).first()

    if amount == 0 or amount == Decimal("0.00"):
        if je:
            je.delete()
        return

    if not je:
        je = JournalEntry.objects.create(
            company=company,
            date=je_date,
            description=description,
            source_type=source_type,
            source_id=source_id,
        )
    else:
        je.date = je_date
        je.description = description
        je.save(update_fields=["date", "description"])
        JournalLine.objects.filter(entry=je).delete()

    supplier_obj = None
    customer_obj = None

    if source_type == "SUPP_OPEN_BALANCE":
        supplier_obj = Newsupplier.objects.for_company(company).filter(pk=source_id).first()

    if source_type == "CUST_OPEN_BALANCE":
        customer_obj = Newcustomer.objects.for_company(company).filter(pk=source_id).first()

    JournalLine.objects.create(
        entry=je,
        account=dr_account,
        debit=amount,
        credit=Decimal("0.00"),
        supplier=None,
        customer=None,
    )

    JournalLine.objects.create(
        entry=je,
        account=cr_account,
        debit=Decimal("0.00"),
        credit=amount,
        supplier=supplier_obj,
        customer=customer_obj,
    )
    return je


def _get_or_create_sales_income_account(company) -> Account:
    acc = (
        Account.objects.for_company(company)
        .filter(account_name__iexact="Sales Income", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(
            account_name__icontains="sales",
            account_type__in=["OPERATING_INCOME", "INVESTING_INCOME"],
            is_active=True
        ).first()
    )
    if acc:
        return acc

    return Account.objects.create(
        company=company,
        account_name="Sales Income",
        account_type="OPERATING_INCOME",
        detail_type="Sales",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )


def _get_or_create_vat_payable_account(company) -> Account:
    acc = (
        Account.objects.for_company(company)
        .filter(account_name__iexact="VAT Payable", is_active=True).first()
        or Account.objects.for_company(company)
        .filter(
            account_name__icontains="vat",
            account_type__in=["CURRENT_LIABILITY", "NON_CURRENT_LIABILITY"],
            is_active=True
        ).first()
    )
    if acc:
        return acc

    return Account.objects.create(
        company=company,
        account_name="VAT Payable",
        account_type="CURRENT_LIABILITY",
        detail_type="Taxes payable",
        is_active=True,
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )


def _as_date(d):
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None


def status_for_invoice(inv, total: Decimal, paid: Decimal, balance: Decimal) -> str:
    """
    Simple, consistent status for an invoice.
    total/paid/balance are Decimals precomputed by the caller.
    """
    today = timezone.now().date()
    if balance <= 0:
        return "Paid"
    due = getattr(inv, "due_date", None)
    if due and due < today:
        return "Overdue"
    if paid > 0:
        return "Partially paid"
    return "Open"


# HOME CHARTS
DEC = DecimalField(max_digits=18, decimal_places=2)
ZERO = Value(Decimal("0.00"), output_field=DEC)

INCOME_TYPES = ["OPERATING_INCOME", "INVESTING_INCOME"]
EXPENSE_TYPES = ["OPERATING_EXPENSE", "INVESTING_EXPENSE", "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE"]

INCOME_KW_RE = r"income|revenue|sales"
EXPENSE_KW_RE = r"expense|expenses|cogs|cost of sales|cost of goods"


def dec(v) -> Decimal:
    """Safe Decimal converter."""
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0.00")


def bankish_q():
    return (
        Q(detail_type__icontains="bank") |
        Q(detail_type__icontains="cash") |
        Q(detail_type__icontains="cash and cash equivalents") |
        Q(detail_type__icontains="cash on hand")
    )


def _income_jl_filter():
    return (
        Q(account__account_type__in=INCOME_TYPES) |
        Q(account__parent__account_type__in=INCOME_TYPES) |
        Q(account__account_name__iregex=INCOME_KW_RE) |
        Q(account__detail_type__iregex=INCOME_KW_RE) |
        Q(account__parent__account_name__iregex=INCOME_KW_RE) |
        Q(account__parent__detail_type__iregex=INCOME_KW_RE)
    )


def _expense_jl_filter():
    return (
        Q(account__account_type__in=EXPENSE_TYPES) |
        Q(account__parent__account_type__in=EXPENSE_TYPES) |
        Q(account__account_name__iregex=EXPENSE_KW_RE) |
        Q(account__detail_type__iregex=EXPENSE_KW_RE) |
        Q(account__parent__account_name__iregex=EXPENSE_KW_RE) |
        Q(account__parent__detail_type__iregex=EXPENSE_KW_RE)
    )


def _to_decimal_number(x) -> Decimal:
    """
    Convert DB numeric safely to Decimal.
    """
    if x is None:
        return Decimal("0")
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")

@login_required
def home(request):
    # -----------------------------------------
    # SOWA WORKSPACE HOME
    # -----------------------------------------
    if (request.user.is_staff or request.user.is_superuser) and not getattr(request, "company", None):
        context = {
            "pnl_tile": {
                "range_label": "Sowa workspace",
                "net_profit": 0,
                "income": 0,
                "expense": 0,
                "pnl_url": "accounts:report-pnl",
                "pnl_params": "",
            },
            "expenses_tile": {
                "range_label": "Sowa workspace",
                "total": 0,
                "spending_url": "expenses:expenses",
            },
            "bank_rows": [],
            "bs_url": "accounts:report-balance-sheet",
            "invoices_tile": {
                "range_label": "Sowa workspace",
                "overdue_amount": 0,
                "not_due_amount": 0,
                "paid_30_amount": 0,
                "overdue_count": 0,
                "not_due_count": 0,
                "paid_count": 0,
                "invoices_url": "sales:invoices",
            },
            "sales_tile": {
                "range_label": "Sowa workspace",
                "total": 0,
                "labels": [],
                "values": [],
            },
            "exp_labels": [],
            "exp_values": [],
            "range_options": RANGE_OPTIONS,
            "pnl_range_key": request.GET.get("pnl_range", "last_month"),
            "exp_range_key": request.GET.get("exp_range", "last_30_days"),
            "sales_range_key": request.GET.get("sales_range", "this_financial_year_to_date"),
            "inv_range_key": request.GET.get("inv_range", "this_month_to_date"),
        }
        return render(request, "Home.html", context)

    # -----------------------------------------
    # CLIENT WORKSPACE HOME
    # -----------------------------------------
    company = request.company
    if not company:
        messages.error(request, "Select a company first.")
        return redirect("tenancy:choose_company")

    today = timezone.localdate()

    pnl_range_key = request.GET.get("pnl_range", "last_month")
    exp_range_key = request.GET.get("exp_range", "last_30_days")
    sales_range_key = request.GET.get("sales_range", "this_financial_year_to_date")
    inv_range_key = request.GET.get("inv_range", "this_month_to_date")

    pnl_from, pnl_to = resolve_date_range(pnl_range_key)
    exp_from, exp_to = resolve_date_range(exp_range_key)
    sales_from, sales_to = resolve_date_range(sales_range_key)
    inv_from, inv_to = resolve_date_range(inv_range_key)

    pnl_label = RANGE_LABELS.get(pnl_range_key, "Last month")
    exp_label = RANGE_LABELS.get(exp_range_key, "Last 30 days")
    sales_label = RANGE_LABELS.get(sales_range_key, "This financial year to date")
    inv_label = RANGE_LABELS.get(inv_range_key, "This month to date")

    # =========================================================
    # 1) PROFIT & LOSS TILE
    # =========================================================
    jl_pnl = (
        JournalLine.objects
        .select_related("entry", "account", "account__parent")
        .filter(entry__company=company, entry__date__gte=pnl_from, entry__date__lte=pnl_to)
    )

    income_lines = jl_pnl.filter(_income_jl_filter())
    expense_lines = jl_pnl.filter(_expense_jl_filter())

    inc_agg = income_lines.aggregate(
        deb=Coalesce(Sum("debit"), ZERO, output_field=DEC),
        cre=Coalesce(Sum("credit"), ZERO, output_field=DEC),
    )
    exp_agg = expense_lines.aggregate(
        deb=Coalesce(Sum("debit"), ZERO, output_field=DEC),
        cre=Coalesce(Sum("credit"), ZERO, output_field=DEC),
    )

    income_amt = dec(inc_agg["cre"]) - dec(inc_agg["deb"])
    expense_amt = dec(exp_agg["deb"]) - dec(exp_agg["cre"])
    net_profit = income_amt - expense_amt

    pnl_tile = {
        "range_label": pnl_label,
        "net_profit": net_profit,
        "income": income_amt,
        "expense": expense_amt,
        "pnl_url": "accounts:report-pnl",
        "pnl_params": f"?from={pnl_from:%Y-%m-%d}&to={pnl_to:%Y-%m-%d}",
    }

    # =========================================================
    # 2) EXPENSES TILE
    # =========================================================
    jl_exp = (
        JournalLine.objects
        .select_related("entry", "account", "account__parent")
        .filter(entry__company=company, entry__date__gte=exp_from, entry__date__lte=exp_to)
        .filter(_expense_jl_filter())
        .values("account__account_name")
        .annotate(
            deb=Coalesce(Sum("debit"), ZERO, output_field=DEC),
            cre=Coalesce(Sum("credit"), ZERO, output_field=DEC),
        )
        .order_by("account__account_name")
    )

    exp_labels = [r["account__account_name"] or "Expense" for r in jl_exp]
    exp_values_dec = [(dec(r["deb"]) - dec(r["cre"])) for r in jl_exp]
    exp_values = [float(v) for v in exp_values_dec]
    exp_total = sum(exp_values_dec, Decimal("0.00"))

    expenses_tile = {
        "range_label": exp_label,
        "total": exp_total,
        "spending_url": "expenses:expenses",
    }

    # =========================================================
    # 3) BANK ACCOUNTS TILE
    # =========================================================
    bank_accounts = (
        Account.objects.for_company(company)
        .filter(is_active=True)
        .filter(bankish_q())
        .order_by("account_name")
    )

    bank_rows = []
    for acc in bank_accounts:
        agg = JournalLine.objects.filter(
            entry__company=company,
            account=acc,
            entry__date__lte=today
        ).aggregate(
            deb=Coalesce(Sum("debit"), ZERO, output_field=DEC),
            cre=Coalesce(Sum("credit"), ZERO, output_field=DEC),
        )
        bal = dec(agg["deb"]) - dec(agg["cre"])
        bank_rows.append({
            "id": acc.id,
            "name": acc.account_name or "Bank",
            "balance": bal,
            "gl_url": "accounts:general-ledger",
            "gl_params": f"?account_id={acc.id}&to={today:%Y-%m-%d}",
        })

    bs_url = "accounts:report-balance-sheet"

    # =========================================================
    # 4) INVOICES TILE
    # =========================================================
    tz = timezone.get_current_timezone()

    inv_from_dt = timezone.make_aware(datetime.combine(inv_from, time.min), tz)
    inv_to_dt = timezone.make_aware(datetime.combine(inv_to + timedelta(days=1), time.min), tz)

    inv_qs = (
        Newinvoice.objects.for_company(company)
        .filter(date_created__gte=inv_from_dt, date_created__lt=inv_to_dt)
        .annotate(
            total_due_f=Coalesce(Cast("total_due", FloatField()), Value(0.0), output_field=FloatField()),
            total_paid_f=Coalesce(
                Sum(Cast("payments_applied__amount_paid", FloatField())),
                Value(0.0),
                output_field=FloatField(),
            ),
        )
        .order_by("-date_created", "-id")
    )

    overdue_amount = Decimal("0.00")
    unpaid_amount = Decimal("0.00")
    paid_amount = Decimal("0.00")

    overdue_count = 0
    unpaid_count = 0
    paid_count = 0

    for inv in inv_qs:
        total = _to_decimal_number(inv.total_due_f)
        paid = _to_decimal_number(inv.total_paid_f)
        bal = total - paid

        if bal <= Decimal("0.00001"):
            paid_count += 1
            paid_amount += total
        else:
            due = _as_date(getattr(inv, "due_date", None))
            if due and due < today:
                overdue_count += 1
                overdue_amount += bal
            else:
                unpaid_count += 1
                unpaid_amount += bal

    invoices_tile = {
        "range_label": inv_label,
        "overdue_amount": overdue_amount,
        "not_due_amount": unpaid_amount,
        "paid_30_amount": paid_amount,
        "overdue_count": overdue_count,
        "not_due_count": unpaid_count,
        "paid_count": paid_count,
        "invoices_url": "sales:invoices",
    }

    # =========================================================
    # 5) SALES TILE
    # =========================================================
    sales_from_dt = timezone.make_aware(datetime.combine(sales_from, time.min), tz)
    sales_to_dt = timezone.make_aware(datetime.combine(sales_to + timedelta(days=1), time.min), tz)

    inv_rows = (
        Newinvoice.objects.for_company(company)
        .filter(date_created__gte=sales_from_dt, date_created__lt=sales_to_dt)
        .values_list("date_created", "total_due")
        .order_by("date_created")
    )

    months = OrderedDict()
    for created_dt, total_due in inv_rows:
        if not created_dt:
            continue
        key = created_dt.strftime("%b")
        months.setdefault(key, Decimal("0.00"))
        months[key] += dec(total_due)

    sales_tile = {
        "range_label": sales_label,
        "total": sum(months.values(), Decimal("0.00")),
        "labels": list(months.keys()),
        "values": [float(v) for v in months.values()],
    }

    context = {
        "pnl_tile": pnl_tile,
        "expenses_tile": expenses_tile,
        "bank_rows": bank_rows,
        "bs_url": bs_url,
        "invoices_tile": invoices_tile,
        "sales_tile": sales_tile,
        "exp_labels": exp_labels,
        "exp_values": exp_values,
        "range_options": RANGE_OPTIONS,
        "pnl_range_key": pnl_range_key,
        "exp_range_key": exp_range_key,
        "sales_range_key": sales_range_key,
        "inv_range_key": inv_range_key,
    }

    return render(request, "Home.html", context)
# working on the assets
@login_required
@company_required
@module_required("home")
def assets(request):
    company = request.company
    assets = Newasset.objects.for_company(company)
    return render(request, "Assets.html", {"assets": assets})


# adding an asset
@login_required
@company_required
@module_required("home")
@transaction.atomic
def add_assests(request):
    company = request.company

    if request.method == "POST":
        # supplier
        supplier_id = request.POST.get("supplier")
        supplier = None
        if supplier_id:
            supplier = Newsupplier.objects.for_company(company).filter(pk=supplier_id).first()

        # FK: asset account
        asset_account_id = request.POST.get("asset_account")
        asset_account = None
        if asset_account_id:
            asset_account = Account.objects.for_company(company).filter(pk=asset_account_id).first()

        # FK: payment account (cash/bank)
        payment_account_id = request.POST.get("payment_account")
        payment_account = None
        if payment_account_id:
            payment_account = deposit_accounts_qs(company=company).filter(pk=payment_account_id).first()

        # normal fields
        asset_name = request.POST.get("asset_name")
        asset_tag = request.POST.get("asset_tag")
        asset_category = request.POST.get("asset_category")
        asset_description = request.POST.get("asset_description")
        department = request.POST.get("department")
        custodian = request.POST.get("custodian")
        asset_status = request.POST.get("asset_status")
        purchase_price = request.POST.get("purchase_price")

        funding_source = request.POST.get("funding_source")
        life_span = request.POST.get("life_span")
        depreciation_method = request.POST.get("depreciation_method")
        residual_value = request.POST.get("residual_value")

        cost_center = request.POST.get("cost_center")
        asset_condition = request.POST.get("asset_condition")
        maintenance_schedule = request.POST.get("maintenance_schedule")
        insurance_details = request.POST.get("insurance_details")
        notes = request.POST.get("notes")
        asset_attachments = request.FILES.get("asset_attachments")

        # dates: dd/mm/YYYY
        capitalization_date = None
        capitalization_date_str = request.POST.get("capitalization_date")
        if capitalization_date_str:
            try:
                capitalization_date = datetime.strptime(capitalization_date_str, "%d/%m/%Y").date()
            except ValueError:
                capitalization_date = None

        purchase_date = None
        purchase_date_str = request.POST.get("purchase_date")
        if purchase_date_str:
            try:
                purchase_date = datetime.strptime(purchase_date_str, "%d/%m/%Y").date()
            except ValueError:
                purchase_date = None

        warranty = None
        warranty_str = request.POST.get("warranty")
        if warranty_str:
            try:
                warranty = datetime.strptime(warranty_str, "%d/%m/%Y").date()
            except ValueError:
                warranty = None

        asset = Newasset(
            company=company,
            asset_name=asset_name,
            asset_tag=asset_tag,
            asset_category=asset_category,
            asset_description=asset_description,
            department=department,
            custodian=custodian,
            asset_status=asset_status,
            purchase_price=purchase_price,
            purchase_date=purchase_date,
            supplier=supplier,
            warranty=warranty,
            funding_source=funding_source,
            life_span=life_span,
            depreciation_method=depreciation_method,
            residual_value=residual_value,
            capitalization_date=capitalization_date,
            cost_center=cost_center,
            asset_condition=asset_condition,
            maintenance_schedule=maintenance_schedule,
            insurance_details=insurance_details,
            notes=notes,
            asset_attachments=asset_attachments,
            asset_account=asset_account,
            payment_account=payment_account,
        )

        asset.save()

        save_action = request.POST.get("save_action")
        if save_action == "save&new":
            return redirect("sowaf:add-asset")
        return redirect("sowaf:assets")

    suppliers = Newsupplier.objects.for_company(company)
    asset_accounts = Account.objects.for_company(company).filter(
        is_active=True,
        account_type="NON_CURRENT_ASSET"
    ).order_by("account_name", "account_number")

    try:
        payment_accounts = deposit_accounts_qs(company=company)
    except TypeError:
        payment_accounts = deposit_accounts_qs().filter(company=company)

    return render(request, "assets_form.html", {
        "suppliers": suppliers,
        "asset_accounts": asset_accounts,
        "payment_accounts": payment_accounts,
    })


# asset edit
@login_required
@company_required
@module_required("home")
@transaction.atomic
def edit_asset(request, pk):
    company = request.company
    asset = get_object_or_404(Newasset.objects.for_company(company), pk=pk)

    if request.method == "POST":
        asset.asset_name = request.POST.get("asset_name", asset.asset_name)
        asset.asset_tag = request.POST.get("asset_tag", asset.asset_tag)
        asset.asset_category = request.POST.get("asset_category", asset.asset_category)
        asset.asset_description = request.POST.get("asset_description", asset.asset_description)
        asset.department = request.POST.get("department", asset.department)
        asset.custodian = request.POST.get("custodian", asset.custodian)
        asset.asset_status = request.POST.get("asset_status", asset.asset_status)
        asset.purchase_price = request.POST.get("purchase_price", asset.purchase_price)

        asset.funding_source = request.POST.get("funding_source", asset.funding_source)
        asset.life_span = request.POST.get("life_span", asset.life_span)
        asset.depreciation_method = request.POST.get("depreciation_method", asset.depreciation_method)
        asset.residual_value = request.POST.get("residual_value", asset.residual_value)

        asset.cost_center = request.POST.get("cost_center", asset.cost_center)
        asset.asset_condition = request.POST.get("asset_condition", asset.asset_condition)
        asset.maintenance_schedule = request.POST.get("maintenance_schedule", asset.maintenance_schedule)
        asset.insurance_details = request.POST.get("insurance_details", asset.insurance_details)
        asset.notes = request.POST.get("notes", asset.notes)

        # supplier FK
        supplier_id = request.POST.get("supplier")
        asset.supplier = Newsupplier.objects.for_company(company).filter(pk=supplier_id).first() if supplier_id else None

        # asset account FK
        asset_account_id = request.POST.get("asset_account")
        asset.asset_account = Account.objects.for_company(company).filter(pk=asset_account_id).first() if asset_account_id else None

        # payment account FK
        payment_account_id = request.POST.get("payment_account")
        try:
            payment_qs = deposit_accounts_qs(company=company)
        except TypeError:
            payment_qs = deposit_accounts_qs().filter(company=company)
        asset.payment_account = payment_qs.filter(pk=payment_account_id).first() if payment_account_id else None

        # dates
        capitalization_date_str = request.POST.get("capitalization_date")
        if capitalization_date_str:
            try:
                asset.capitalization_date = datetime.strptime(capitalization_date_str, "%d/%m/%Y").date()
            except ValueError:
                pass

        purchase_date_str = request.POST.get("purchase_date")
        if purchase_date_str:
            try:
                asset.purchase_date = datetime.strptime(purchase_date_str, "%d/%m/%Y").date()
            except ValueError:
                pass

        warranty_str = request.POST.get("warranty")
        if warranty_str:
            try:
                asset.warranty = datetime.strptime(warranty_str, "%d/%m/%Y").date()
            except ValueError:
                pass

        # file
        if "asset_attachments" in request.FILES:
            asset.asset_attachments = request.FILES["asset_attachments"]

        asset.save()
        return redirect("sowaf:assets")

    suppliers = Newsupplier.objects.for_company(company)
    asset_accounts = Account.objects.for_company(company).filter(
        is_active=True, account_type="NON_CURRENT_ASSET"
    ).order_by("account_name", "account_number")

    try:
        payment_accounts = deposit_accounts_qs(company=company)
    except TypeError:
        payment_accounts = deposit_accounts_qs().filter(company=company)

    return render(request, "assets_form.html", {
        "asset": asset,
        "suppliers": suppliers,
        "asset_accounts": asset_accounts,
        "payment_accounts": payment_accounts,
    })


# deleting an asset
@login_required
@company_required
@module_required("home")
@transaction.atomic
def delete_asset(request, pk):
    company = request.company
    asset = get_object_or_404(Newasset.objects.for_company(company), pk=pk)
    asset.delete()
    return redirect("sowaf:assets")


# importing assets
@login_required
@company_required
@module_required("home")
def download_assets_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets Template"

    headers = [
        "asset_name", "asset_tag", "asset_category", "asset_description", "department",
        "custodian", "asset_status", "purchase_price", "purchase_date", "supplier",
        "warranty", "funding_source", "life_span", "depreciation_method", "residual_value",
        "accumulated_depreciation", "remaining_value", "asset_account", "capitalization_date",
        "cost_center", "asset_condition", "maintenance_schedule", "insurance_details", "notes",
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="assets_template.xlsx"'
        return response


# functions to handle the date formats
def parse_capitalization_date_safe(capitalization_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(capitalization_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def parse_purchase_date_safe(purchase_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(purchase_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def parse_warranty_safe(warranty):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(warranty), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


@login_required
@company_required
@module_required("home")
@transaction.atomic
def import_assets(request):
    company = request.company

    if request.method != "POST" or "excel_file" not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect("sowaf:assets")

    excel_file = request.FILES["excel_file"]
    file_name = excel_file.name.lower()

    def _resolve_supplier(value):
        if not value:
            return None
        return (
            Newsupplier.objects.for_company(company).filter(company_name__iexact=str(value).strip()).first()
            or Newsupplier.objects.for_company(company).filter(pk=str(value).strip()).first()
        )

    def _resolve_asset_account(value):
        if not value:
            return None
        return (
            Account.objects.for_company(company).filter(account_name__iexact=str(value).strip()).first()
            or Account.objects.for_company(company).filter(account_number__iexact=str(value).strip()).first()
            or Account.objects.for_company(company).filter(pk=str(value).strip()).first()
        )

    try:
        if file_name.endswith(".csv"):
            decoded_file = excel_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader)

            for row in reader:
                (
                    asset_name, asset_tag, asset_category, asset_description, department, custodian,
                    asset_status, purchase_price, purchase_date, supplier, warranty, funding_source,
                    life_span, depreciation_method, residual_value, accumulated_depreciation,
                    remaining_value, asset_account, capitalization_date, cost_center,
                    asset_condition, maintenance_schedule, insurance_details, notes,
                ) = row

                capitalization_date = parse_capitalization_date_safe(capitalization_date)
                purchase_date = parse_purchase_date_safe(purchase_date)
                warranty = parse_warranty_safe(warranty)

                Newasset.objects.create(
                    company=company,
                    asset_name=asset_name,
                    asset_tag=asset_tag,
                    asset_category=asset_category,
                    asset_description=asset_description,
                    department=department,
                    custodian=custodian,
                    asset_status=asset_status,
                    purchase_price=purchase_price,
                    purchase_date=purchase_date,
                    supplier=_resolve_supplier(supplier),
                    warranty=warranty,
                    funding_source=funding_source,
                    life_span=life_span,
                    depreciation_method=depreciation_method,
                    residual_value=residual_value if residual_value not in (None, "") else None,
                    accumulated_depreciation=accumulated_depreciation if accumulated_depreciation not in (None, "") else None,
                    remaining_value=remaining_value if remaining_value not in (None, "") else None,
                    asset_account=_resolve_asset_account(asset_account),
                    capitalization_date=capitalization_date,
                    cost_center=cost_center,
                    asset_condition=asset_condition,
                    maintenance_schedule=maintenance_schedule,
                    insurance_details=insurance_details,
                    notes=notes,
                )

        elif file_name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                    asset_name, asset_tag, asset_category, asset_description, department, custodian,
                    asset_status, purchase_price, purchase_date, supplier, warranty, funding_source,
                    life_span, depreciation_method, residual_value, accumulated_depreciation,
                    remaining_value, asset_account, capitalization_date, cost_center,
                    asset_condition, maintenance_schedule, insurance_details, notes,
                ) = row

                capitalization_date = parse_capitalization_date_safe(capitalization_date)
                purchase_date = parse_purchase_date_safe(purchase_date)
                warranty = parse_warranty_safe(warranty)

                Newasset.objects.create(
                    company=company,
                    asset_name=asset_name,
                    asset_tag=asset_tag,
                    asset_category=asset_category,
                    asset_description=asset_description,
                    department=department,
                    custodian=custodian,
                    asset_status=asset_status,
                    purchase_price=purchase_price,
                    purchase_date=purchase_date,
                    supplier=_resolve_supplier(supplier),
                    warranty=warranty,
                    funding_source=funding_source,
                    life_span=life_span,
                    depreciation_method=depreciation_method,
                    residual_value=residual_value if residual_value not in (None, "") else None,
                    accumulated_depreciation=accumulated_depreciation if accumulated_depreciation not in (None, "") else None,
                    remaining_value=remaining_value if remaining_value not in (None, "") else None,
                    asset_account=_resolve_asset_account(asset_account),
                    capitalization_date=capitalization_date,
                    cost_center=cost_center,
                    asset_condition=asset_condition,
                    maintenance_schedule=maintenance_schedule,
                    insurance_details=insurance_details,
                    notes=notes,
                )
        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect("sowaf:assets")

        messages.success(request, "Asset data imported successfully.")
        return redirect("sowaf:assets")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("sowaf:assets")


# customer view

# -------------------------
# CUSTOMERS LIST (with analytics)
# -------------------------
def must_have_company(request):
    """
    Ensures the user has selected/switched to a company (tenant context).
    CompanyMiddleware sets request.company when session['company_id'] exists.
    """
    if not getattr(request, "company", None):
        messages.error(request, "Please select a company to continue.")
        return False
    return True
@login_required
@company_required
@module_required("home")
def customers(request):
    company = request.company
    q = (request.GET.get("q") or "").strip()

    # ------------------------------
    # 1) INVOICE OUTSTANDING
    # ------------------------------
    paid_per_invoice_sq = (
        PaymentInvoice.objects
        .filter(invoice_id=OuterRef("pk"))
        .values("invoice_id")
        .annotate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))
        .values("total")[:1]
    )

    inv_all = (
        Newinvoice.objects
        .filter(company=company)
        .annotate(
            total_due_dec=Cast(F("total_due"), DecimalField(max_digits=18, decimal_places=2)),
            paid=Coalesce(
                Subquery(
                    paid_per_invoice_sq,
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                ),
                Value(Decimal("0.00"))
            )
        )
        .annotate(
            raw_outstanding=ExpressionWrapper(
                F("total_due_dec") - F("paid"),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            ),
            outstanding=Case(
                When(raw_outstanding__gt=0, then=F("raw_outstanding")),
                default=Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
    )

    per_customer_invoice_open_sq = (
        inv_all.filter(customer_id=OuterRef("pk"))
        .values("customer_id")
        .annotate(sum_outstanding=Coalesce(Sum("outstanding"), Value(Decimal("0.00"))))
        .values("sum_outstanding")[:1]
    )

    # ------------------------------
    # 2) OPENING BALANCE REMAINING (GL-based)
    # ------------------------------
    customer_ar_balance_sq = (
        JournalLine.objects
        .filter(
            entry__company=company,
            account__detail_type__in=["Customer Subledger (A/R)", "Customer ledger (A/R)"],
            account__account_name=OuterRef("customer_name"),
        )
        .values("account__account_name")
        .annotate(
            deb=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
            cred=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
        )
        .annotate(
            bal=ExpressionWrapper(
                F("deb") - F("cred"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
        .values("bal")[:1]
    )

    # ------------------------------
    # 3) CUSTOMERS QUERYSET (combine both) + scope by company
    # ------------------------------
    customers_qs = (
        Newcustomer.objects
        .for_company(company)
        .annotate(
            invoice_remaining=Coalesce(
                Subquery(
                    per_customer_invoice_open_sq,
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                ),
                Value(Decimal("0.00"))
            ),
            opening_remaining=Coalesce(
                Subquery(
                    customer_ar_balance_sq,
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                ),
                Value(Decimal("0.00"))
            ),
        )
        .annotate(
            total_unpaid=ExpressionWrapper(
                F("opening_remaining") + F("invoice_remaining"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
    )

    if q:
        customers_qs = customers_qs.filter(
            Q(customer_name__icontains=q) | Q(company_name__icontains=q)
        )

    # ------------------------------
    # 4) DASHBOARD ANALYTICS
    # ------------------------------
    today = timezone.now().date()
    cutoff = today - timedelta(days=30)

    # UNBILLED = opening balances remaining (GL)
    ar_totals = (
        JournalLine.objects
        .filter(
            entry__company=company,
            account__detail_type__in=["Customer Subledger (A/R)", "Customer ledger (A/R)"]
        )
        .aggregate(
            deb=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
            cred=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
        )
    )
    unbilled_amount = (ar_totals["deb"] or Decimal("0")) - (ar_totals["cred"] or Decimal("0"))

    ar_accounts = (
        Account.objects
        .for_company(company)
        .filter(detail_type__in=["Customer Subledger (A/R)", "Customer ledger (A/R)"])
        .annotate(
            deb=Coalesce(
                Sum(
                    "lines__debit",
                    filter=Q(lines__entry__company=company)
                ),
                Value(Decimal("0.00"))
            ),
            cred=Coalesce(
                Sum(
                    "lines__credit",
                    filter=Q(lines__entry__company=company)
                ),
                Value(Decimal("0.00"))
            ),
        )
        .annotate(
            bal=ExpressionWrapper(
                F("deb") - F("cred"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
    )
    unbilled_count = ar_accounts.filter(bal__gt=0).count()

    open_agg = inv_all.filter(outstanding__gt=0).aggregate(
        amount=Coalesce(Sum("outstanding"), Value(Decimal("0.00"))),
        count=Count("id")
    )

    overdue_agg = inv_all.filter(outstanding__gt=0, due_date__lt=today).aggregate(
        amount=Coalesce(Sum("outstanding"), Value(Decimal("0.00"))),
        count=Count("id")
    )

    recent_paid_from_payments = (
        PaymentInvoice.objects
        .filter(payment__company=company, payment__payment_date__gte=cutoff)
        .aggregate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))
        ["total"]
    ) or Decimal("0.00")

    recent_paid_from_receipts = (
        SalesReceipt.objects
        .filter(company=company, receipt_date__gte=cutoff)
        .aggregate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))
        ["total"]
    ) or Decimal("0.00")

    amount_recent = recent_paid_from_payments + recent_paid_from_receipts
    recent_payments_count = Payment.objects.filter(company=company, payment_date__gte=cutoff).count()
    recent_receipts_count = SalesReceipt.objects.filter(company=company, receipt_date__gte=cutoff).count()
    count_recent = recent_payments_count + recent_receipts_count

    total_for_pct = (
        (unbilled_amount or Decimal("0")) +
        (overdue_agg["amount"] or Decimal("0")) +
        (open_agg["amount"] or Decimal("0")) +
        (amount_recent or Decimal("0"))
    ) or Decimal("1")

    def pct(x):
        return float((x or Decimal("0")) * Decimal("100") / total_for_pct)

    analytics = {
        "amount_unbilled": unbilled_amount,
        "count_unbilled": unbilled_count,

        "amount_overdue": overdue_agg["amount"] or Decimal("0"),
        "count_overdue": int(overdue_agg["count"] or 0),

        "amount_open": open_agg["amount"] or Decimal("0"),
        "count_open": int(open_agg["count"] or 0),

        "amount_recent": amount_recent,
        "count_recent": count_recent,

        "pct_unbilled": pct(unbilled_amount),
        "pct_overdue": pct(overdue_agg["amount"]),
        "pct_open": pct(open_agg["amount"]),
        "pct_recent": pct(amount_recent),
    }

    return render(request, "Customers.html", {
        "customers": customers_qs,
        "q": q,
        "analytics": analytics,
    })


def _cast(field_name):
    return Cast(F(field_name), DEC)


def _safe_url(pattern_name, *args):
    try:
        return reverse(pattern_name, args=args)
    except NoReverseMatch:
        return "#"


def _to_date(d):
    if d is None:
        return None
    if isinstance(d, str):
        try:
            from django.utils import timezone
            return timezone.datetime.fromisoformat(d).date()
        except Exception:
            return None
    try:
        return d.date()
    except Exception:
        return d


def _customer_ar_balance_as_of(customer_id: int, as_of_date=None, company=None) -> Decimal:
    """
    LIVE A/R balance for a customer from GL Customer Subledger (A/R),
    optionally as-of a date (inclusive).

    A/R is an Asset => debit - credit
    We match the customer subledger account by account_name == customer.customer_name
    """
    customer_qs = Newcustomer.objects.all()
    if company is not None:
        customer_qs = customer_qs.for_company(company)

    customer = customer_qs.only("id", "customer_name").get(id=customer_id)

    qs = JournalLine.objects.filter(
        account__detail_type__in=["Customer Subledger (A/R)", "Customer ledger (A/R)"],
        account__account_name=customer.customer_name,
    )

    if company is not None:
        qs = qs.filter(entry__company=company)

    if as_of_date:
        try:
            if isinstance(as_of_date, str):
                as_of_date = timezone.datetime.fromisoformat(as_of_date).date()
            elif hasattr(as_of_date, "date"):
                as_of_date = as_of_date.date()
        except Exception:
            as_of_date = None

    if as_of_date:
        qs = qs.filter(entry__date__lte=as_of_date)

    totals = qs.aggregate(
        deb=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
        cred=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
    )

    return (totals["deb"] or Decimal("0.00")) - (totals["cred"] or Decimal("0.00"))


# Customer detail
@login_required
@company_required
@module_required("home")
def customer_detail(request, pk):
    company = request.company

    customer = get_object_or_404(Newcustomer.objects.for_company(company), pk=pk)
    tab = request.GET.get("tab", "transactions")

    # ----------------------------
    # Invoices for this customer
    # ----------------------------
    inv_qs_base = (
        Newinvoice.objects
        .filter(customer=customer, company=company)
        .annotate(
            total_due_dec=Cast("total_due", DecimalField(max_digits=18, decimal_places=2)),
            total_paid=Coalesce(Sum("payments_applied__amount_paid"), Value(Decimal("0.00")))
        )
    )

    today = now().date()

    # ----------------------------
    # OPEN BALANCE = GL Customer Subledger (A/R)
    # ----------------------------
    gl_qs = JournalLine.objects.filter(
        entry__company=company,
        account__detail_type__in=["Customer Subledger (A/R)", "Customer ledger (A/R)"],
        account__account_name=customer.customer_name,
    )

    gl_totals = gl_qs.aggregate(
        deb=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
        cred=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
    )

    open_balance = (gl_totals["deb"] or Decimal("0.00")) - (gl_totals["cred"] or Decimal("0.00"))

    # ----------------------------
    # OVERDUE (invoice-based)
    # ----------------------------
    overdue_balance = sum(
        (row.total_due_dec or Decimal("0.00")) - (row.total_paid or Decimal("0.00"))
        for row in inv_qs_base.filter(due_date__lt=today)
    )
    count_invoices = inv_qs_base.count()

    # ----------------------------
    # Transactions
    # ----------------------------
    transactions_rows = []
    inv_qs = inv_qs_base.select_related("customer").order_by("-date_created", "-id")

    def _fallback_status(inv, total, paid, bal):
        if bal <= 0:
            return "Paid"
        if getattr(inv, "due_date", None) and inv.due_date < today:
            return "Overdue"
        return "Open"

    for inv in inv_qs:
        total = inv.total_due_dec or Decimal("0.00")
        paid = inv.total_paid or Decimal("0.00")
        bal = max(total - paid, Decimal("0.00"))
        try:
            status = status_for_invoice(inv, total, paid, bal)
        except NameError:
            status = _fallback_status(inv, total, paid, bal)

        transactions_rows.append({
            "id": inv.id,
            "date": getattr(inv, "date_created", None),
            "type": "Invoice",
            "no": f"INV-{inv.id:04d}",
            "customer": customer.customer_name,
            "memo": (getattr(inv, "memo", "") or "")[:140],
            "amount": total,
            "status": status,
            "edit_url": reverse("sales:edit-invoice", args=[inv.id]),
            "view_url": reverse("sales:invoice-detail", args=[inv.id]),
            "print_url": reverse("sales:invoice-print", args=[inv.id]),
        })

    # ----------------------------
    # Payments
    # ----------------------------
    pay_qs = (
        Payment.objects
        .filter(customer=customer, company=company)
        .annotate(applied_total=Coalesce(Sum("applied_invoices__amount_paid"), Value(Decimal("0.00"))))
        .order_by("-payment_date", "-id")
    )

    for p in pay_qs:
        transactions_rows.append({
            "id": p.id,
            "date": p.payment_date,
            "type": "Payment",
            "no": (getattr(p, "reference_no", None) or f"{p.id:04d}"),
            "customer": customer.customer_name,
            "memo": (getattr(p, "memo", "") or "")[:140],
            "amount": p.applied_total or Decimal("0.00"),
            "status": "Closed" if (p.applied_total or 0) > 0 else "Unapplied",
            "edit_url": reverse("sales:payment-edit", args=[p.id]),
            "view_url": reverse("sales:payment-detail", args=[p.id]),
            "print_url": reverse("sales:payment-print", args=[p.id]),
        })

    def _safe_url_local(name, obj_id):
        try:
            return reverse(name, args=[obj_id])
        except Exception:
            return "#"

    # ----------------------------
    # Sales Receipts
    # ----------------------------
    try:
        sr_qs = (
            SalesReceipt.objects
            .filter(customer=customer, company=company)
            .order_by("-receipt_date", "-id")
        )

        for r in sr_qs:
            amount = Decimal(str(getattr(r, "total_amount", "0") or "0"))
            transactions_rows.append({
                "id": r.id,
                "date": getattr(r, "receipt_date", None),
                "type": "Sales Receipt",
                "no": (getattr(r, "reference_no", None) or f"SR-{r.id:04d}"),
                "customer": customer.customer_name,
                "memo": (getattr(r, "memo", "") or "")[:140],
                "amount": amount,
                "status": "Closed",
                "edit_url": _safe_url_local("sales:receipt-edit", r.id),
                "view_url": _safe_url_local("sales:receipt-detail", r.id),
                "print_url": _safe_url_local("sales:receipt-print", r.id),
            })
    except Exception:
        pass

    transactions_rows.sort(key=lambda r: (r["date"] or today, r["type"], r["id"]), reverse=True)

    # ----------------------------
    # Statements
    # ----------------------------
    statements_rows = []
    try:
        from sales.models import Statement

        st_qs = (
            Statement.objects
            .filter(customer=customer, company=company)
            .order_by("-statement_date", "-id")
        )

        for st in st_qs:
            st_type = getattr(st, "get_statement_type_display", None)
            st_type = st_type() if callable(st_type) else getattr(st, "statement_type", "")

            ar_live_balance = _customer_ar_balance_as_of(customer.id, st.end_date, company=company)

            statements_rows.append({
                "id": st.id,
                "date": st.statement_date,
                "no": f"ST-{st.id:04d}",
                "type": st_type,
                "start": st.start_date,
                "end": st.end_date,
                "balance": ar_live_balance,
                "view_url": _safe_url_local("sales:statement-detail", st.id),
                "print_url": _safe_url_local("sales:statement-print", st.id),
                "send_url": _safe_url_local("sales:statement-send", st.id),
            })
    except Exception:
        statements_rows = []

    return render(request, "customer_detail.html", {
        "customer": customer,
        "tab": tab,
        "open_balance": open_balance,
        "overdue_balance": overdue_balance,
        "count_invoices": count_invoices,
        "transactions_rows": transactions_rows,
        "statements_rows": statements_rows,
    })


# making a customer active and inactive
@login_required
@company_required
@module_required("home")
@transaction.atomic
def make_inactive_customer(request, pk):
    company = request.company
    customer = get_object_or_404(Newcustomer.objects.for_company(company), pk=pk)
    customer.is_active = False
    customer.save(update_fields=["is_active"])
    return redirect("sowaf:customers")


# reactivating
@login_required
@company_required
@module_required("home")
@transaction.atomic
def make_active_customer(request, pk):
    company = request.company
    customer = get_object_or_404(Newcustomer.objects.for_company(company), pk=pk)
    customer.is_active = True
    customer.save(update_fields=["is_active"])
    return redirect("sowaf:customers")


# customer form view

# -------------------------
# ADD CUSTOMER
# -------------------------
@login_required
@company_required
@module_required("home")
@transaction.atomic
def add_customer(request):
    company = request.company

    if request.method == "POST":
        logo = request.FILES.get("logo")
        if logo:
            if not logo.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)

        opening_balance = _dec(request.POST.get("balance"), "0.00")
        registration_date = parse_date_or_none(request.POST.get("registration_date"))

        new_customer = Newcustomer(
            company=company,
            logo=logo,
            customer_name=request.POST.get("name"),
            company_name=request.POST.get("company"),
            email=request.POST.get("email"),
            phone_number=request.POST.get("phonenum"),
            mobile_number=request.POST.get("mobilenum"),
            website=request.POST.get("website"),
            tin_number=request.POST.get("tin"),
            opening_balance=opening_balance,
            registration_date=registration_date,
            street_one=request.POST.get("street1"),
            street_two=request.POST.get("street2"),
            city=request.POST.get("city"),
            province=request.POST.get("province"),
            postal_code=request.POST.get("postalcode"),
            country=request.POST.get("country"),
            notes=request.POST.get("notes"),
            attachments=request.FILES.get("attachments"),
        )
        new_customer.save()

        # Opening balance journal entry
        if opening_balance != 0:
            opening_equity = _get_or_create_opening_equity(company)
            customer_ar_sub = _get_or_create_customer_ar_subaccount(new_customer)

            _upsert_opening_balance_je(
                company=company,
                source_type="CUSTOMER_OPENING_BALANCE",
                source_id=new_customer.id,
                je_date=registration_date or timezone.localdate(),
                description=f"Opening balance for customer {new_customer}",
                dr_account=customer_ar_sub,
                cr_account=opening_equity,
                amount=abs(opening_balance),
            )

        if request.POST.get("save_action") == "save&new":
            return redirect("sowaf:add-customer")

        return redirect("sowaf:customers")

    return render(request, "customers_form.html")

# -------------------------
# EDIT CUSTOMER
# -------------------------
@transaction.atomic
@login_required
@company_required
@module_required("home")
def edit_customer(request, pk):
    company = request.company

    # Prevent cross-company editing
    customer = get_object_or_404(Newcustomer.objects.for_company(company), pk=pk)

    if request.method == "POST":
        customer.customer_name = request.POST.get("name", customer.customer_name)
        customer.company_name = request.POST.get("company", customer.company_name)
        customer.email = request.POST.get("email", customer.email)
        customer.phone_number = request.POST.get("phonenum", customer.phone_number)
        customer.mobile_number = request.POST.get("mobilenum", customer.mobile_number)
        customer.website = request.POST.get("website", customer.website)
        customer.tin_number = request.POST.get("tin", customer.tin_number)

        new_opening_balance = _dec(request.POST.get("balance"), "0.00")
        customer.opening_balance = new_opening_balance

        customer.registration_date = (
            parse_date_or_none(request.POST.get("registration_date"))
            or customer.registration_date
        )

        customer.street_one = request.POST.get("street1", customer.street_one)
        customer.street_two = request.POST.get("street2", customer.street_two)
        customer.city = request.POST.get("city", customer.city)
        customer.province = request.POST.get("province", customer.province)
        customer.postal_code = request.POST.get("postalcode", customer.postal_code)
        customer.country = request.POST.get("country", customer.country)
        customer.notes = request.POST.get("notes", customer.notes)

        logo = request.FILES.get("logo")
        if logo:
            if not logo.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)
            customer.logo = logo

        if "attachments" in request.FILES:
            customer.attachments = request.FILES["attachments"]

        customer.save()

        # Update opening balance JE (to CUSTOMER subaccount)
        opening_equity = _get_or_create_opening_equity(company)
        customer_ar_sub = _get_or_create_customer_ar_subaccount(customer)
        amt = abs(new_opening_balance)

        _upsert_opening_balance_je(
            company=company,
            source_type="CUSTOMER_OPENING_BALANCE",
            source_id=customer.id,
            je_date=_safe_date(customer.registration_date, timezone.localdate()),
            description=f"Opening balance for customer {customer.customer_name or customer.company_name or customer.id}",
            dr_account=customer_ar_sub,
            cr_account=opening_equity,
            amount=amt,
        )

        return redirect("sowaf:customers")

    return render(request, "customers_form.html", {"customer": customer, "is_edit": True})


# template for the download
@login_required
@company_required
@module_required("home")
def download_customers_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Template"

    headers = [
        "name", "company", "email", "phone", "mobile", "website", "tin", "balance",
        "date_str", "street1", "street2", "city", "province", "postal_code",
        "country", "notes", "logo"
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="customer_template.xlsx"'
        return response


@login_required
@company_required
@module_required("home")
@transaction.atomic
def import_customers(request):
    company = request.company

    if request.method != "POST" or "excel_file" not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect("sowaf:customers")

    excel_file = request.FILES["excel_file"]
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            decoded_file = excel_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)

            for row in reader:
                if not row:
                    continue

                row = list(row) + [""] * (17 - len(row))
                (
                    name, company_name, email, phone, mobile, website, tin, balance,
                    date_str, street1, street2, city, province, postal_code,
                    country, notes, logo
                ) = row[:17]

                customer = Newcustomer.objects.create(
                    company=company,
                    customer_name=name,
                    company_name=company_name,
                    email=email,
                    phone_number=phone,
                    mobile_number=mobile,
                    website=website,
                    tin_number=tin,
                    opening_balance=_dec(balance, "0.00"),
                    registration_date=parse_date_or_none(date_str),
                    street_one=street1,
                    street_two=street2,
                    city=city,
                    province=province,
                    postal_code=postal_code,
                    country=country,
                    notes=notes,
                )

                if logo:
                    image_path = os.path.join(settings.MEDIA_ROOT, "uploads", str(logo).strip())
                    if os.path.exists(image_path):
                        with open(image_path, "rb") as f:
                            customer.logo.save(str(logo).strip(), File(f), save=True)
                    else:
                        messages.warning(request, f"Image file '{logo}' not found.")

                opening_balance = _dec(balance, "0.00")
                if opening_balance != 0:
                    opening_equity = _get_or_create_opening_equity(company)
                    customer_ar_sub = _get_or_create_customer_ar_subaccount(customer)

                    _upsert_opening_balance_je(
                        company=company,
                        source_type="CUSTOMER_OPENING_BALANCE",
                        source_id=customer.id,
                        je_date=parse_date_or_none(date_str) or timezone.localdate(),
                        description=f"Opening balance for customer {customer.customer_name or customer.company_name or customer.id}",
                        dr_account=customer_ar_sub,
                        cr_account=opening_equity,
                        amount=abs(opening_balance),
                    )

        elif file_name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                row = list(row) + [None] * (17 - len(row))
                (
                    name, company_name, email, phone, mobile, website, tin, balance,
                    date_str, street1, street2, city, province, postal,
                    country, notes, logo
                ) = row[:17]

                customer = Newcustomer.objects.create(
                    company=company,
                    customer_name=name,
                    company_name=company_name,
                    email=email,
                    phone_number=phone,
                    mobile_number=mobile,
                    website=website,
                    tin_number=tin,
                    opening_balance=_dec(balance, "0.00"),
                    registration_date=parse_date_or_none(str(date_str) if date_str else ""),
                    street_one=street1,
                    street_two=street2,
                    city=city,
                    province=province,
                    postal_code=postal,
                    country=country,
                    notes=notes,
                )

                if logo:
                    logo_name = str(logo).strip()
                    image_path = os.path.join(settings.MEDIA_ROOT, "uploads", logo_name)
                    if os.path.exists(image_path):
                        with open(image_path, "rb") as f:
                            customer.logo.save(logo_name, File(f), save=True)
                    else:
                        messages.warning(request, f"Image file '{logo_name}' not found.")

                opening_balance = _dec(balance, "0.00")
                if opening_balance != 0:
                    opening_equity = _get_or_create_opening_equity(company)
                    customer_ar_sub = _get_or_create_customer_ar_subaccount(customer)

                    _upsert_opening_balance_je(
                        company=company,
                        source_type="CUSTOMER_OPENING_BALANCE",
                        source_id=customer.id,
                        je_date=parse_date_or_none(str(date_str) if date_str else "") or timezone.localdate(),
                        description=f"Opening balance for customer {customer.customer_name or customer.company_name or customer.id}",
                        dr_account=customer_ar_sub,
                        cr_account=opening_equity,
                        amount=abs(opening_balance),
                    )
        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect("sowaf:customers")

        messages.success(request, "Customer data imported successfully.")
        return redirect("sowaf:customers")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("sowaf:customers")


# adding employees
@login_required
@company_required
@module_required("home")
def employee(request):
    employees = Newemployee.objects.for_company(request.company)
    return render(request, "Employees.html", {"employees": employees})


# add employee form
@login_required
@company_required
@module_required("home")
@transaction.atomic
def add_employees(request):
    company = request.company

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        gender = request.POST.get("gender")
        dob = parse_date(request.POST.get("dob") or "")
        nationality = request.POST.get("nationality")
        nin_number = request.POST.get("nin_number")
        tin_number = request.POST.get("tin_number")

        profile_picture = request.FILES.get("profile_picture")
        if profile_picture:
            if not profile_picture.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the profile picture.")
                return redirect(request.path)
            if profile_picture.size > 1048576:
                messages.error(request, "Profile picture file size must not exceed 1MB.")
                return redirect(request.path)

        phone_number = request.POST.get("phone_number")
        email_address = request.POST.get("email_address")
        residential_address = request.POST.get("residential_address")
        emergency_person = request.POST.get("emergency_person")
        emergency_contact = request.POST.get("emergency_contact")
        relationship = request.POST.get("relationship")

        job_title = request.POST.get("job_title")
        department = request.POST.get("department")
        employment_type = request.POST.get("employment_type")
        status = request.POST.get("status")
        hire_date = parse_date(request.POST.get("hire_date") or "")
        supervisor = request.POST.get("supervisor")

        raw_salary = request.POST.get("salary")
        try:
            salary = Decimal(raw_salary) if raw_salary else Decimal("0.00")
        except Exception:
            salary = Decimal("0.00")

        raw_taxable = request.POST.get("taxable_allowances")
        try:
            taxable_allowances = Decimal(raw_taxable) if raw_taxable else Decimal("0.00")
        except Exception:
            taxable_allowances = Decimal("0.00")

        raw_intaxable = request.POST.get("intaxable_allowances")
        try:
            intaxable_allowances = Decimal(raw_intaxable) if raw_intaxable else Decimal("0.00")
        except Exception:
            intaxable_allowances = Decimal("0.00")

        payment_frequency = request.POST.get("payment_frequency")
        payment_method = request.POST.get("payment_method")
        bank_name = request.POST.get("bank_name")
        bank_account = request.POST.get("bank_account")
        bank_branch = request.POST.get("bank_branch")
        nssf_number = request.POST.get("nssf_number")
        insurance_provider = request.POST.get("insurance_provider")
        additional_notes = request.POST.get("additional_notes")
        doc_attachments = request.FILES.get("doc_attachments")

        employee = Newemployee(
            company=company,
            first_name=first_name,
            last_name=last_name,
            gender=gender,
            dob=dob,
            nationality=nationality,
            nin_number=nin_number,
            tin_number=tin_number,
            profile_picture=profile_picture,
            phone_number=phone_number,
            email_address=email_address,
            residential_address=residential_address,
            emergency_person=emergency_person,
            emergency_contact=emergency_contact,
            relationship=relationship,
            job_title=job_title,
            department=department,
            employment_type=employment_type,
            status=status,
            hire_date=hire_date,
            supervisor=supervisor,
            salary=salary,
            payment_frequency=payment_frequency,
            payment_method=payment_method,
            bank_name=bank_name,
            bank_account=bank_account,
            bank_branch=bank_branch,
            nssf_number=nssf_number,
            insurance_provider=insurance_provider,
            taxable_allowances=taxable_allowances,
            intaxable_allowances=intaxable_allowances,
            additional_notes=additional_notes,
            doc_attachments=doc_attachments,
        )
        employee.save()

        save_action = request.POST.get("save_action")
        if save_action == "save&new":
            return redirect("sowaf:add-employee")
        if save_action in ["save", "save&close"]:
            return redirect("sowaf:employees")

    return render(request, "employees_form.html")


@login_required
@company_required
@module_required("home")
@transaction.atomic
def edit_employee(request, pk):
    employee = get_object_or_404(Newemployee.objects.for_company(request.company), pk=pk)

    if request.method == "POST":
        employee.first_name = request.POST.get("first_name", employee.first_name)
        employee.last_name = request.POST.get("last_name", employee.last_name)
        employee.gender = request.POST.get("gender", employee.gender)
        employee.dob = parse_date(request.POST.get("dob") or "")
        employee.nationality = request.POST.get("nationality", employee.nationality)
        employee.nin_number = request.POST.get("nin_number", employee.nin_number)
        employee.tin_number = request.POST.get("tin_number", employee.tin_number)

        employee.phone_number = request.POST.get("phone_number", employee.phone_number)
        employee.email_address = request.POST.get("email_address", employee.email_address)
        employee.residential_address = request.POST.get("residential_address", employee.residential_address)
        employee.emergency_person = request.POST.get("emergency_person", employee.emergency_person)
        employee.emergency_contact = request.POST.get("emergency_contact", employee.emergency_contact)
        employee.relationship = request.POST.get("relationship", employee.relationship)

        employee.job_title = request.POST.get("job_title", employee.job_title)
        employee.department = request.POST.get("department", employee.department)
        employee.employment_type = request.POST.get("employment_type", employee.employment_type)
        employee.status = request.POST.get("status", employee.status)
        employee.hire_date = parse_date(request.POST.get("hire_date") or "")
        employee.supervisor = request.POST.get("supervisor", employee.supervisor)

        raw_salary = request.POST.get("salary")
        try:
            employee.salary = Decimal(raw_salary) if raw_salary else Decimal("0.00")
        except Exception:
            employee.salary = Decimal("0.00")

        raw_taxable = request.POST.get("taxable_allowances")
        try:
            employee.taxable_allowances = Decimal(raw_taxable) if raw_taxable else Decimal("0.00")
        except Exception:
            employee.taxable_allowances = Decimal("0.00")

        raw_intaxable = request.POST.get("intaxable_allowances")
        try:
            employee.intaxable_allowances = Decimal(raw_intaxable) if raw_intaxable else Decimal("0.00")
        except Exception:
            employee.intaxable_allowances = Decimal("0.00")

        employee.payment_frequency = request.POST.get("payment_frequency", employee.payment_frequency)
        employee.payment_method = request.POST.get("payment_method", employee.payment_method)
        employee.bank_name = request.POST.get("bank_name", employee.bank_name)
        employee.bank_account = request.POST.get("bank_account", employee.bank_account)
        employee.bank_branch = request.POST.get("bank_branch", employee.bank_branch)
        employee.nssf_number = request.POST.get("nssf_number", employee.nssf_number)
        employee.insurance_provider = request.POST.get("insurance_provider", employee.insurance_provider)
        employee.additional_notes = request.POST.get("additional_notes", employee.additional_notes)

        profile_picture = request.FILES.get("profile_picture")
        if profile_picture:
            if not profile_picture.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the profile picture.")
                return redirect(request.path)
            if profile_picture.size > 1048576:
                messages.error(request, "Profile picture file size must not exceed 1MB.")
                return redirect(request.path)
            employee.profile_picture = profile_picture

        if "doc_attachments" in request.FILES:
            employee.doc_attachments = request.FILES["doc_attachments"]

        employee.save()
        return redirect("sowaf:employees")

    return render(request, "employees_form.html", {"employee": employee})


# importing employees
@login_required
@company_required
@module_required("home")
def download_employees_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees Template"

    headers = [
        "first_name", "last_name", "gender", "dob", "nationality",
        "nin_number", "tin_number", "profile_picture", "phone_number", "email_address",
        "residential_address", "emergency_person", "emergency_contact", "relationship",
        "job_title", "department", "employment_type", "status", "hire_date", "supervisor",
        "salary", "payment_frequency", "payment_method", "bank_name", "bank_account",
        "bank_branch", "nssf_number", "insurance_provider", "taxable_allowances",
        "intaxable_allowances", "additional_notes"
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="employees_template.xlsx"'
        return response


def handle_profile_picture_upload(employee, profile_picture, request=None):
    if profile_picture:
        image_path = os.path.join(settings.MEDIA_ROOT, "uploads", profile_picture)
        if os.path.exists(image_path):
            with open(image_path, "rb") as f:
                employee.profile_picture.save(profile_picture, File(f), save=True)
        else:
            if request:
                messages.warning(request, f"Image file '{profile_picture}' not found.")


# Parse DOB (multiple formats)
def parse_dob_safe(dob):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(dob), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


# Parse Hire Date (multiple formats)
def parse_hire_date_safe(hire_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(hire_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


@login_required
@company_required
@module_required("home")
@transaction.atomic
def import_employees(request):
    company = request.company

    if request.method != "POST" or "excel_file" not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect("sowaf:employees")

    excel_file = request.FILES["excel_file"]
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            decoded_file = excel_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)

            for row in reader:
                if not row:
                    continue

                row = list(row) + [""] * (31 - len(row))
                (
                    first_name, last_name, gender, dob, nationality, nin_number, tin_number,
                    profile_picture, phone_number, email_address, residential_address,
                    emergency_person, emergency_contact, relationship, job_title, department,
                    employment_type, status, hire_date, supervisor, salary,
                    payment_frequency, payment_method, bank_name, bank_account,
                    bank_branch, nssf_number, insurance_provider, taxable_allowances,
                    intaxable_allowances, additional_notes
                ) = row[:31]

                dob = parse_dob_safe(dob)
                hire_date = parse_hire_date_safe(hire_date)
                profile_picture = profile_picture.strip() if profile_picture else ""

                employee = Newemployee.objects.create(
                    company=company,
                    first_name=first_name,
                    last_name=last_name,
                    gender=gender,
                    dob=dob,
                    nationality=nationality,
                    nin_number=nin_number,
                    tin_number=tin_number,
                    phone_number=str(phone_number).rstrip(".0") if phone_number else "",
                    email_address=email_address,
                    residential_address=residential_address,
                    emergency_person=emergency_person,
                    emergency_contact=emergency_contact,
                    relationship=relationship,
                    job_title=job_title,
                    department=department,
                    employment_type=employment_type,
                    status=status,
                    hire_date=hire_date,
                    supervisor=supervisor,
                    salary=_dec(salary, "0.00"),
                    payment_frequency=payment_frequency,
                    payment_method=payment_method,
                    bank_name=bank_name,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    nssf_number=nssf_number,
                    insurance_provider=insurance_provider,
                    taxable_allowances=_dec(taxable_allowances, "0.00"),
                    intaxable_allowances=_dec(intaxable_allowances, "0.00"),
                    additional_notes=additional_notes,
                )
                handle_profile_picture_upload(employee, profile_picture, request=request)

        elif file_name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                row = list(row) + [None] * (31 - len(row))
                (
                    first_name, last_name, gender, dob, nationality, nin_number, tin_number,
                    profile_picture, phone_number, email_address, residential_address,
                    emergency_person, emergency_contact, relationship, job_title, department,
                    employment_type, status, hire_date, supervisor, salary,
                    payment_frequency, payment_method, bank_name, bank_account,
                    bank_branch, nssf_number, insurance_provider, taxable_allowances,
                    intaxable_allowances, additional_notes
                ) = row[:31]

                dob = parse_dob_safe(dob)
                hire_date = parse_hire_date_safe(hire_date)
                profile_picture = str(profile_picture).strip() if profile_picture else ""

                employee = Newemployee.objects.create(
                    company=company,
                    first_name=first_name,
                    last_name=last_name,
                    gender=gender,
                    dob=dob,
                    nationality=nationality,
                    nin_number=nin_number,
                    tin_number=tin_number,
                    phone_number=str(phone_number).rstrip(".0") if phone_number else "",
                    email_address=email_address,
                    residential_address=residential_address,
                    emergency_person=emergency_person,
                    emergency_contact=emergency_contact,
                    relationship=relationship,
                    job_title=job_title,
                    department=department,
                    employment_type=employment_type,
                    status=status,
                    hire_date=hire_date,
                    supervisor=supervisor,
                    salary=_dec(salary, "0.00"),
                    payment_frequency=payment_frequency,
                    payment_method=payment_method,
                    bank_name=bank_name,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    nssf_number=nssf_number,
                    insurance_provider=insurance_provider,
                    taxable_allowances=_dec(taxable_allowances, "0.00"),
                    intaxable_allowances=_dec(intaxable_allowances, "0.00"),
                    additional_notes=additional_notes,
                )
                handle_profile_picture_upload(employee, profile_picture, request=request)

        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect("sowaf:employees")

        messages.success(request, "Employee data imported successfully.")
        return redirect("sowaf:employees")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("sowaf:employees")


# supplier view
# add new supplier form view
@transaction.atomic
@login_required
@company_required
@module_required("home")
def add_supplier(request):
    company = request.company

    if request.method == "POST":
        logo = request.FILES.get("logo")
        if logo:
            if not logo.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)

        company_name = request.POST.get("company_name")
        supplier_type = request.POST.get("supplier_type")
        contact_person = request.POST.get("contact_person")
        contact_position = request.POST.get("contact_position")
        contact = request.POST.get("contact")
        email = request.POST.get("email")

        open_balance = _dec(request.POST.get("open_balance"), "0.00")

        website = request.POST.get("website")
        address1 = request.POST.get("address1")
        address2 = request.POST.get("address2")
        city = request.POST.get("city")
        state = request.POST.get("state")
        zip_code = request.POST.get("zip_code")
        country = request.POST.get("country")
        bank = request.POST.get("bank")
        bank_account = request.POST.get("bank_account")
        bank_branch = request.POST.get("bank_branch")
        payment_terms = request.POST.get("payment_terms")
        currency = request.POST.get("currency")
        payment_method = request.POST.get("payment_method")
        tin = request.POST.get("tin")
        reg_number = request.POST.get("reg_number")
        attachments = request.FILES.get("attachments")

        new_supplier = Newsupplier(
            company=company,
            logo=logo,
            company_name=company_name,
            supplier_type=supplier_type,
            contact_person=contact_person,
            contact_position=contact_position,
            contact=contact,
            email=email,
            open_balance=open_balance,
            website=website,
            address1=address1,
            address2=address2,
            city=city,
            state=state,
            zip_code=zip_code,
            country=country,
            bank=bank,
            bank_account=bank_account,
            bank_branch=bank_branch,
            payment_terms=payment_terms,
            currency=currency,
            payment_method=payment_method,
            tin=tin,
            reg_number=reg_number,
            attachments=attachments,
        )
        new_supplier.save()

        # Post supplier opening balance to SUPPLIER subaccount under A/P
        if open_balance != 0:
            opening_equity = _get_or_create_opening_equity(company)
            supplier_ap_sub = _get_or_create_supplier_ap_subaccount(new_supplier)
            amt = abs(open_balance)

            _upsert_opening_balance_je(
                company=company,
                source_type="SUPP_OPEN_BALANCE",
                source_id=new_supplier.id,
                je_date=timezone.localdate(),
                description=f"Opening balance for supplier {new_supplier.company_name or new_supplier.id}",
                dr_account=opening_equity,
                cr_account=supplier_ap_sub,
                amount=amt,
            )

        save_action = request.POST.get("save_action")
        if save_action == "save&new":
            return redirect("sowaf:add-suppliers")
        return redirect("sowaf:suppliers")

    return render(request, "suppliers_entry_form.html", {})


@transaction.atomic
@login_required
@company_required
@module_required("home")
def edit_supplier(request, pk):
    company = request.company
    supplier = get_object_or_404(Newsupplier.objects.for_company(company), pk=pk)

    if request.method == "POST":
        supplier.company_name = request.POST.get("company_name", supplier.company_name)
        supplier.supplier_type = request.POST.get("supplier_type", supplier.supplier_type)
        supplier.contact_person = request.POST.get("contact_person", supplier.contact_person)
        supplier.contact_position = request.POST.get("contact_position", supplier.contact_position)
        supplier.contact = request.POST.get("contact", supplier.contact)
        supplier.email = request.POST.get("email", supplier.email)

        new_open_balance = _dec(request.POST.get("open_balance"), "0.00")
        supplier.open_balance = new_open_balance

        supplier.website = request.POST.get("website", supplier.website)
        supplier.address1 = request.POST.get("address1", supplier.address1)
        supplier.address2 = request.POST.get("address2", supplier.address2)
        supplier.city = request.POST.get("city", supplier.city)
        supplier.state = request.POST.get("state", supplier.state)
        supplier.zip_code = request.POST.get("zip_code", supplier.zip_code)
        supplier.country = request.POST.get("country", supplier.country)
        supplier.bank = request.POST.get("bank", supplier.bank)
        supplier.bank_account = request.POST.get("bank_account", supplier.bank_account)
        supplier.bank_branch = request.POST.get("bank_branch", supplier.bank_branch)
        supplier.payment_terms = request.POST.get("payment_terms", supplier.payment_terms)
        supplier.currency = request.POST.get("currency", supplier.currency)
        supplier.payment_method = request.POST.get("payment_method", supplier.payment_method)
        supplier.tin = request.POST.get("tin", supplier.tin)
        supplier.reg_number = request.POST.get("reg_number", supplier.reg_number)

        logo = request.FILES.get("logo")
        if logo:
            if not logo.name.lower().endswith(".png"):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)
            supplier.logo = logo

        if "attachments" in request.FILES:
            supplier.attachments = request.FILES["attachments"]

        supplier.save()

        opening_equity = _get_or_create_opening_equity(company)
        supplier_ap_sub = _get_or_create_supplier_ap_subaccount(supplier)
        amt = abs(new_open_balance)

        _upsert_opening_balance_je(
            company=company,
            source_type="SUPP_OPEN_BALANCE",
            source_id=supplier.id,
            je_date=timezone.localdate(),
            description=f"Opening balance for supplier {supplier.company_name or supplier.id}",
            dr_account=opening_equity,
            cr_account=supplier_ap_sub,
            amount=amt,
        )

        return redirect("sowaf:suppliers")

    return render(request, "suppliers_entry_form.html", {"supplier": supplier})


# LIVE SUPPLIER A/P OPEN BALANCE HELPER (GL-BASED)
# =========================================================
def _supplier_ap_balance_live(supplier_id: int, as_of_date=None, company=None) -> Decimal:
    """
    LIVE supplier open balance from GL (Supplier Subledger A/P).

    A/P is a LIABILITY -> normal CREDIT.
    So balance = credits - debits.

    Uses supplier.ap_account (OneToOne to Account).
    Optionally filter as-of date (<= as_of_date).
    """
    supplier_qs = Newsupplier.objects.select_related("ap_account")
    if company is not None:
        supplier_qs = supplier_qs.for_company(company)

    supplier = supplier_qs.filter(pk=supplier_id).first()
    if not supplier or not supplier.ap_account_id:
        return Decimal("0.00")

    qs = JournalLine.objects.filter(account_id=supplier.ap_account_id).select_related("entry")

    if company is not None:
        qs = qs.filter(entry__company=company)

    if as_of_date:
        qs = qs.filter(entry__date__lte=as_of_date)

    agg = qs.aggregate(
        deb=Coalesce(Sum("debit"), Value(Decimal("0.00")), output_field=DEC),
        cred=Coalesce(Sum("credit"), Value(Decimal("0.00")), output_field=DEC),
    )

    deb = Decimal(str(agg["deb"] or "0.00"))
    cred = Decimal(str(agg["cred"] or "0.00"))
    return cred - deb


# =========================================================
# SUPPLIER LIST VIEW (LIVE BALANCES + KPIs)
# =========================================================
@login_required
@company_required
@module_required("home")
def supplier(request):
    company = request.company
    suppliers = list(Newsupplier.objects.for_company(company))

    balances = _supplier_ap_balances_bulk([s.id for s in suppliers], company=company)

    for s in suppliers:
        s.open_balance = balances.get(s.id, Decimal("0.00"))

    open_total = Decimal("0.00")
    open_count = 0
    for s in suppliers:
        bal = s.open_balance or Decimal("0.00")
        if bal > 0:
            open_total += bal
            open_count += 1

    d30 = timezone.localdate() - timedelta(days=30)

    paid_30_rows = (
        JournalLine.objects
        .filter(
            entry__company=company,
            supplier_id__in=[s.id for s in suppliers],
            entry__date__gte=d30
        )
        .aggregate(
            paid_amount=Coalesce(Sum("debit"), Value(Decimal("0.00")))
        )
    )
    paid_30_amount = paid_30_rows["paid_amount"] or Decimal("0.00")

    supp_kpis = {
        "unbilled_amount": Decimal("0.00"),
        "unbilled_count": 0,

        "unpaid_amount": open_total,
        "unpaid_count": open_count,

        "open_bills_amount": open_total,
        "open_bills_count": open_count,

        "paid_30_amount": paid_30_amount,
        "paid_30_count": 0,
    }

    return render(request, "Supplier.html", {"suppliers": suppliers, "supp_kpis": supp_kpis})


# =========================================================
# SUPPLIER DETAIL VIEW (UPDATED TO USE HELPER)
# =========================================================
@login_required
@company_required
@module_required("home")
def supplier_detail(request, pk):
    company = request.company
    supplier = get_object_or_404(Newsupplier.objects.for_company(company), pk=pk)
    tab = request.GET.get("tab", "transactions")

    today = timezone.now().date()

    live_bal = _supplier_ap_balance_live(supplier.id, company=company)

    bills_qs = (
        Bill.objects
        .filter(company=company, supplier=supplier)
        .annotate(total_amount_dec=Cast("total_amount", DEC))
    )

    expenses_qs = (
        Expense.objects
        .filter(company=company, payee_supplier=supplier)
        .annotate(total_amount_dec=Cast("total_amount", DEC))
    )

    cheques_qs = (
        Cheque.objects
        .filter(company=company, payee_supplier=supplier)
        .annotate(total_amount_dec=Cast("total_amount", DEC))
    )

    bills_total = bills_qs.aggregate(
        t=Coalesce(Sum("total_amount_dec"), Value(Decimal("0.00")))
    )["t"] or Decimal("0.00")

    paid_total = (
        (expenses_qs.aggregate(t=Coalesce(Sum("total_amount_dec"), Value(Decimal("0.00"))))["t"] or Decimal("0.00"))
        + (cheques_qs.aggregate(t=Coalesce(Sum("total_amount_dec"), Value(Decimal("0.00"))))["t"] or Decimal("0.00"))
    )

    old_open_balance = bills_total - paid_total
    if old_open_balance < 0:
        old_open_balance = Decimal("0.00")

    overdue_bills_total = bills_qs.filter(due_date__lt=today).aggregate(
        t=Coalesce(Sum("total_amount_dec"), Value(Decimal("0.00")))
    )["t"] or Decimal("0.00")

    count_bills = bills_qs.count()

    rows = []

    for b in bills_qs.order_by("-bill_date", "-id"):
        rows.append({
            "id": b.id,
            "date": getattr(b, "bill_date", None),
            "type": "Bill",
            "no": b.bill_no or f"BILL-{b.id:04d}",
            "party": supplier.company_name or supplier.contact_person or "",
            "memo": (getattr(b, "memo", "") or "")[:140],
            "amount": b.total_amount_dec or Decimal("0.00"),
            "status": getattr(b, "status", "Open") or "Open",
            "edit_url": reverse("expenses:bill-edit", args=[b.id]),
            "view_url": reverse("expenses:bill-edit", args=[b.id]),
            "print_url": "#",
        })

    for e in expenses_qs.order_by("-payment_date", "-id"):
        rows.append({
            "id": e.id,
            "date": getattr(e, "payment_date", None),
            "type": "Expense",
            "no": getattr(e, "ref_no", "") or f"EXP-{e.id:04d}",
            "party": supplier.company_name or supplier.contact_person or "",
            "memo": (getattr(e, "memo", "") or "")[:140],
            "amount": e.total_amount_dec or Decimal("0.00"),
            "status": "Paid",
            "edit_url": reverse("expenses:expense-edit", args=[e.id]),
            "view_url": reverse("expenses:expense-edit", args=[e.id]),
            "print_url": "#",
        })

    for c in cheques_qs.order_by("-payment_date", "-id"):
        rows.append({
            "id": c.id,
            "date": getattr(c, "payment_date", None),
            "type": "Cheque",
            "no": getattr(c, "cheque_no", "") or f"CHQ-{c.id:04d}",
            "party": supplier.company_name or supplier.contact_person or "",
            "memo": (getattr(c, "memo", "") or "")[:140],
            "amount": c.total_amount_dec or Decimal("0.00"),
            "status": "Paid",
            "edit_url": reverse("expenses:cheque-edit", args=[c.id]),
            "view_url": reverse("expenses:cheque-edit", args=[c.id]),
            "print_url": "#",
        })

    rows.sort(key=lambda r: (r["date"] or today, r["type"], r["id"]), reverse=True)

    context = {
        "supplier": supplier,
        "tab": tab,
        "open_balance": live_bal,
        "old_open_balance": old_open_balance,
        "overdue_balance": overdue_bills_total,
        "count_bills": count_bills,
        "transactions_rows": rows,
    }
    return render(request, "supplier_detail.html", context)


# making the row active or inactive
@login_required
@company_required
@module_required("home")
@transaction.atomic
def make_inactive_supplier(request, pk):
    supplier = get_object_or_404(Newsupplier.objects.for_company(request.company), pk=pk)
    supplier.is_active = False
    supplier.save(update_fields=["is_active"])
    return redirect("sowaf:suppliers")


# reactivating
@login_required
@company_required
@module_required("home")
@transaction.atomic
def make_active_supplier(request, pk):
    supplier = get_object_or_404(Newsupplier.objects.for_company(request.company), pk=pk)
    supplier.is_active = True
    supplier.save(update_fields=["is_active"])
    return redirect("sowaf:suppliers")


# importing suppliers
@login_required
@company_required
@module_required("home")
def download_suppliers_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers Template"

    headers = [
        "logo", "company_name", "supplier_type", "contact_person", "contact_position",
        "contact", "email", "open_balance", "website", "address1", "address2", "city",
        "state", "zip_code", "country", "bank", "bank_account", "bank_branch",
        "payment_terms", "currency", "payment_method", "tin", "reg_number"
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="suppliers_template.xlsx"'
        return response


def handle_logo_upload(supplier, logo, request=None):
    if logo:
        image_path = os.path.join(settings.MEDIA_ROOT, "uploads", logo)
        if os.path.exists(image_path):
            with open(image_path, "rb") as f:
                supplier.logo.save(logo, File(f), save=True)
        else:
            if request:
                messages.warning(request, f"Image file '{logo}' not found.")


@login_required
@company_required
@module_required("home")
@transaction.atomic
def import_suppliers(request):
    company = request.company

    if request.method != "POST" or "excel_file" not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect("sowaf:suppliers")

    excel_file = request.FILES["excel_file"]
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            decoded_file = excel_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)

            for row in reader:
                if not row:
                    continue

                row = list(row) + [""] * (23 - len(row))
                (
                    logo, company_name, supplier_type, contact_person, contact_position,
                    contact, email, open_balance, website, address1, address2, city,
                    state, zip_code, country, bank, bank_account, bank_branch,
                    payment_terms, currency, payment_method, tin, reg_number
                ) = row[:23]

                logo = logo.strip() if logo else ""

                supplier = Newsupplier.objects.create(
                    company=company,
                    company_name=company_name,
                    supplier_type=supplier_type,
                    contact_person=contact_person,
                    contact_position=contact_position,
                    contact=contact,
                    email=email,
                    open_balance=_dec(open_balance, "0.00"),
                    website=website,
                    address1=address1,
                    address2=address2,
                    city=city,
                    state=state,
                    zip_code=zip_code,
                    country=country,
                    bank=bank,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    payment_terms=payment_terms,
                    currency=currency,
                    payment_method=payment_method,
                    tin=tin,
                    reg_number=reg_number,
                )
                handle_logo_upload(supplier, logo, request=request)

                supplier_open_balance = _dec(open_balance, "0.00")
                if supplier_open_balance != 0:
                    opening_equity = _get_or_create_opening_equity(company)
                    supplier_ap_sub = _get_or_create_supplier_ap_subaccount(supplier)

                    _upsert_opening_balance_je(
                        company=company,
                        source_type="SUPP_OPEN_BALANCE",
                        source_id=supplier.id,
                        je_date=timezone.localdate(),
                        description=f"Opening balance for supplier {supplier.company_name or supplier.id}",
                        dr_account=opening_equity,
                        cr_account=supplier_ap_sub,
                        amount=abs(supplier_open_balance),
                    )

        elif file_name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                row = list(row) + [None] * (23 - len(row))
                (
                    logo, company_name, supplier_type, contact_person, contact_position,
                    contact, email, open_balance, website, address1, address2, city,
                    state, zip_code, country, bank, bank_account, bank_branch,
                    payment_terms, currency, payment_method, tin, reg_number
                ) = row[:23]

                logo = str(logo).strip() if logo else ""

                supplier = Newsupplier.objects.create(
                    company=company,
                    company_name=company_name,
                    supplier_type=supplier_type,
                    contact_person=contact_person,
                    contact_position=contact_position,
                    contact=contact,
                    email=email,
                    open_balance=_dec(open_balance, "0.00"),
                    website=website,
                    address1=address1,
                    address2=address2,
                    city=city,
                    state=state,
                    zip_code=zip_code,
                    country=country,
                    bank=bank,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    payment_terms=payment_terms,
                    currency=currency,
                    payment_method=payment_method,
                    tin=tin,
                    reg_number=reg_number,
                )
                handle_logo_upload(supplier, logo, request=request)

                supplier_open_balance = _dec(open_balance, "0.00")
                if supplier_open_balance != 0:
                    opening_equity = _get_or_create_opening_equity(company)
                    supplier_ap_sub = _get_or_create_supplier_ap_subaccount(supplier)

                    _upsert_opening_balance_je(
                        company=company,
                        source_type="SUPP_OPEN_BALANCE",
                        source_id=supplier.id,
                        je_date=timezone.localdate(),
                        description=f"Opening balance for supplier {supplier.company_name or supplier.id}",
                        dr_account=opening_equity,
                        cr_account=supplier_ap_sub,
                        amount=abs(supplier_open_balance),
                    )
        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect("sowaf:suppliers")

        messages.success(request, "Supplier data imported successfully.")
        return redirect("sowaf:suppliers")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("sowaf:suppliers")


# tasks view
@login_required
@company_required
@module_required("home")
def tasks(request):
    return render(request, "tasks.html", {})


# taxes view
@login_required
@company_required
@module_required("home")
def taxes(request):
    return render(request, "Taxes.html", {})


# miscellaneous view
@login_required
@company_required
@module_required("home")
def miscellaneous(request):
    return render(request, "Miscellaneous.html", {})