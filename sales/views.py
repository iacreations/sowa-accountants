from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.http import HttpResponse, Http404
from django.template.loader import render_to_string
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from io import BytesIO
from tempfile import NamedTemporaryFile
from datetime import date, timedelta, datetime
from django.utils import timezone
from decimal import Decimal
from django.urls import reverse
from django.db import transaction
from django.templatetags.static import static
from django.db.models import DecimalField, Q, ExpressionWrapper
import openpyxl
import csv
import io
import os
from django.db.models.functions import Coalesce, Cast
from django.core.files import File
from django.conf import settings
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import (Newinvoice,InvoiceItem,Product,Payment,PaymentInvoice,PaymentOpenBalanceLine,SalesReceipt,SalesReceiptLine,CustomerRefund,RecurringInvoice, RecurringInvoiceLine)
from sowaf.models import Newcustomer
from .models import Statement, StatementLine
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, F, Value
from django.utils.dateparse import parse_date
from inventory.models import Product,Pclass
from accounts.models import Account, JournalEntry, JournalLine
from .services import (generate_unique_ref_no, parse_date_flexible, status_for_invoice, _payment_prefill_rows,_coerce_decimal,as_aware_datetime)
from accounts.utils import deposit_accounts_qs
from collections import defaultdict
from accounts.middleware import get_current_user, get_current_ip
from .recurring_service import generate_recurring_invoices_for_date
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from accounts.utils import (
    VAT_RATE,
    _get_inventory_asset_account,
    _get_cogs_account,
    _get_vat_payable_account,
)
from inventory.models import InventoryMovement, Product

def _as_date(d):
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return date.min  

def _customer_credit_balance(customer) -> Decimal:
    if not customer:
        return Decimal("0.00")

    adv = _get_customer_advance_account()
    agg = (
        JournalLine.objects
        .filter(account=adv, customer=customer)
        .aggregate(
            d=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
            c=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
        )
    )
    debit = Decimal(str(agg["d"] or "0.00"))
    credit = Decimal(str(agg["c"] or "0.00"))

    # Liability normal balance is CREDIT, so credit - debit
    bal = credit - debit
    return bal if bal > 0 else Decimal("0.00")


def _get_or_create_named_account(account_name: str, account_type: str, detail_type: str = "") -> Account:
    acc = Account.objects.filter(account_name=account_name, is_active=True).first()
    if acc:
        return acc
    return Account.objects.create(
        account_name=account_name,
        account_type=account_type,
        detail_type=detail_type or None,
        is_active=True,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )

def _get_customer_advance_account() -> Account:
    # Liability (you owe the customer)
    return _get_or_create_named_account(
        account_name="Customer Advances",
        account_type="CURRENT_LIABILITY",
        detail_type="Customer Credits",
    )

def _get_supplier_advance_account() -> Account:
    # Asset (supplier owes you / you prepaid)
    return _get_or_create_named_account(
        account_name="Supplier Advances",
        account_type="CURRENT_ASSET",
        detail_type="Supplier Prepayments",
    )


def _dec(val, default="0.00") -> Decimal:
    """
    Safe decimal parser.
    Accepts both: _dec(val) and _dec(val, "0.00")
    """
    try:
        s = str(val).strip() if val is not None else ""
        if s == "":
            return Decimal(str(default))
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(str(default))


def _account_debit_balance(account: "Account") -> Decimal:
    """
    Returns debit balance for an account:
      opening_balance + SUM(debit - credit)
    """
    opening = Decimal(str(getattr(account, "opening_balance", 0) or "0"))
    agg = (
        JournalLine.objects
        .filter(account=account)
        .aggregate(
            d=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
            c=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
        )
    )
    deb = Decimal(str(agg["d"] or "0"))
    cred = Decimal(str(agg["c"] or "0"))
    return opening + (deb - cred)


def _invoice_outstanding(inv: "Newinvoice") -> Decimal:
    total_due = Decimal(str(inv.total_due or "0"))
    paid = (
        PaymentInvoice.objects
        .filter(invoice=inv)
        .aggregate(s=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))["s"]
        or Decimal("0.00")
    )
    bal = total_due - paid
    return bal if bal > 0 else Decimal("0.00")


def _customer_open_balance_amount(customer: "Newcustomer") -> Decimal:
    """
    Open Balance = Customer A/R subaccount debit balance - total outstanding invoice balances
    (Anything in A/R not represented by invoice balances is considered "open balance".)
    """
    if not customer:
        return Decimal("0.00")

    customer_acc = _get_or_create_customer_ar_subaccount(customer)

    ar_debit_bal = _account_debit_balance(customer_acc)

    total_unpaid_invoices = Decimal("0.00")
    for inv in Newinvoice.objects.filter(customer=customer):
        total_unpaid_invoices += _invoice_outstanding(inv)

    open_bal = ar_debit_bal - total_unpaid_invoices
    return open_bal if open_bal > 0 else Decimal("0.00")

def _save_payment_open_balance(payment: "Payment", amount: Decimal):
    PaymentOpenBalanceLine.objects.filter(payment=payment).delete()
    if amount and amount > 0:
        PaymentOpenBalanceLine.objects.create(payment=payment, amount_applied=amount)

def _get_customer_advance_account() -> "Account":
    """
    Customer Advances = LIABILITY account (customer credit balance).
    """
    acc = Account.objects.filter(account_name__iexact="Customer Advances", is_active=True).first()
    if acc:
        return acc

    return Account.objects.create(
        account_name="Customer Advances",
        account_type="CURRENT_LIABILITY",   
        detail_type="Customer Advances",  
        is_active=True,
    )

def apply_audit_fields(obj):
    """
    Safely attach audit fields if the model supports them.
    Does NOT break models that don't have audit columns.
    """
    user = get_current_user()
    ip   = get_current_ip()

    if hasattr(obj, "created_by") and not obj.pk:
        obj.created_by = user
    if hasattr(obj, "updated_by"):
        obj.updated_by = user

    if hasattr(obj, "created_ip") and not obj.pk:
        obj.created_ip = ip
    if hasattr(obj, "updated_ip"):
        obj.updated_ip = ip


def _find_control_account(detail_type=None, name_contains=None):
    qs = Account.objects.filter(is_active=True)

    if detail_type:
        acc = qs.filter(detail_type__iexact=detail_type).first()
        if acc:
            return acc

    if name_contains:
        acc = qs.filter(account_name__icontains=name_contains).first()
        if acc:
            return acc

    return None


def _get_or_create_ar_control_account():
    """
    Finds or auto-creates the Accounts Receivable (A/R) control account.
    """
    ar = (
        _find_control_account(detail_type="Accounts Receivable (A/R)")
        or _find_control_account(name_contains="accounts receivable")
        or _find_control_account(name_contains="receivable")
    )
    if ar:
        return ar

    # Auto-create (adjust account_number if you already use a numbering scheme)
    ar = Account.objects.create(
        account_name="Accounts Receivable",
        account_number="1100",
        account_type="CURRENT_ASSET",          # <-- matches your Account model codes
        detail_type="Accounts Receivable (A/R)",
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
        is_active=True,
        description="System control account for customer balances (A/R).",
    )
    return ar


def _get_or_create_customer_ar_subaccount(customer: Newcustomer) -> Account:
    """
    Creates/gets a CUSTOMER subaccount under A/R control.
    This is your customer subledger.
    """
    ar_control = _get_or_create_ar_control_account()

    name = _safe_name(customer.customer_name) or _safe_name(customer.company_name) or f"Customer {customer.id}"
    sub_name = f"{name}"

    acc = Account.objects.filter(
        parent=ar_control,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()
    if acc:
        return acc

    # Create subaccount
    acc = Account.objects.create(
        account_name=sub_name,
        account_type=ar_control.account_type,   # still current asset
        detail_type="Customer Subledger (A/R)",
        is_active=True,
        is_subaccount=True,
        parent=ar_control,
        opening_balance=Decimal("0.00"),
    )
    return acc

def _safe_name(s: str) -> str:
    return (s or "").strip()
def _safe_date(val, default):
    try:
        if val:
            return val
    except Exception:
        pass
    return default

def _get_sales_income_account() -> "Account":
    """
    Returns a Sales/Revenue income account.
    Tries to find an existing one; otherwise creates a fallback using your Account codes.
    """
    acc = (
        Account.objects
        .filter(is_active=True)
        .filter(account_name__iexact="Sales")
        .first()
    )
    if acc:
        return acc

    acc = (
        Account.objects
        .filter(is_active=True)
        .filter(account_name__icontains="Sales")
        .first()
    )
    if acc:
        return acc

    acc = (
        Account.objects
        .filter(is_active=True)
        .filter(account_name__icontains="Revenue")
        .first()
    )
    if acc:
        return acc

    # fallback create (matches your ACCOUNT_TYPES codes)
    return Account.objects.create(
        account_name="Sales",
        account_type="OPERATING_INCOME",
        detail_type="Sales",
        is_active=True,
    )


def _post_customer_refund_to_ledger(refund: CustomerRefund):
    amt = Decimal(str(refund.amount or "0.00"))
    if amt <= 0:
        JournalEntry.objects.filter(source_type="CUSTOMER_REFUND", source_id=refund.id).delete()
        return

    JournalEntry.objects.filter(source_type="CUSTOMER_REFUND", source_id=refund.id).delete()

    adv = _get_customer_advance_account()
    bank = refund.paid_from

    entry = JournalEntry.objects.create(
        date=refund.refund_date or timezone.localdate(),
        description=f"Customer Refund {refund.id:04d} – {refund.customer.customer_name}",
        source_type="CUSTOMER_REFUND",
        source_id=refund.id,
    )

    # DR Customer Advances (reduce credit)
    JournalLine.objects.create(
        entry=entry, account=adv,
        debit=amt, credit=Decimal("0.00"),
        customer=refund.customer, supplier=None,
    )

    # CR Bank/Cash
    JournalLine.objects.create(
        entry=entry, account=bank,
        debit=Decimal("0.00"), credit=amt,
        customer=refund.customer, supplier=None,
    )

def _post_customer_refund_to_ledger(refund: CustomerRefund):
    amt = Decimal(str(refund.amount or "0.00"))
    if amt <= 0:
        JournalEntry.objects.filter(source_type="CUSTOMER_REFUND", source_id=refund.id).delete()
        return

    JournalEntry.objects.filter(source_type="CUSTOMER_REFUND", source_id=refund.id).delete()

    adv = _get_customer_advance_account()
    bank = refund.paid_from

    entry = JournalEntry.objects.create(
        date=refund.refund_date or timezone.localdate(),
        description=f"Customer Refund {refund.id:04d} – {refund.customer.customer_name}",
        source_type="CUSTOMER_REFUND",
        source_id=refund.id,
    )

    # DR Customer Advances (reduce credit)
    JournalLine.objects.create(
        entry=entry, account=adv,
        debit=amt, credit=Decimal("0.00"),
        customer=refund.customer, supplier=None,
    )

    # CR Bank/Cash
    JournalLine.objects.create(
        entry=entry, account=bank,
        debit=Decimal("0.00"), credit=amt,
        customer=refund.customer, supplier=None,
    )


def _post_invoice_to_ledger(invoice: Newinvoice):
    """
    INVOICE posting:

      DR Customer A/R Subaccount
      CR Revenue (by product.income_account)
      CR VAT Payable (if VAT exists)
    """

    total_due = Decimal(str(getattr(invoice, "total_due", None) or "0"))
    if total_due <= 0:
        JournalEntry.objects.filter(source_type="invoice", source_id=invoice.id).delete()
        return

    # delete & recreate style
    JournalEntry.objects.filter(source_type="invoice", source_id=invoice.id).delete()

    revenue_by_account = defaultdict(lambda: Decimal("0.00"))
    vat_total = Decimal("0.00")

    default_income_acc = (
        _find_control_account(name_contains="Sales")
        or _find_control_account(name_contains="Revenue")
    )

    items_qs = invoice.items.select_related("product").all()  #must match your related_name
    for line in items_qs:
        line_amount   = Decimal(str(getattr(line, "amount", None) or "0"))
        line_discount = Decimal(str(getattr(line, "discount_amount", None) or "0"))
        net_amount    = line_amount - line_discount
        if net_amount < 0:
            net_amount = Decimal("0.00")

        prod = getattr(line, "product", None)
        income_acc = getattr(prod, "income_account", None) if prod else None
        if not income_acc:
            income_acc = default_income_acc

        if income_acc and net_amount > 0:
            revenue_by_account[income_acc] += net_amount

        vat_total += Decimal(str(getattr(line, "vat", None) or "0"))

    shipping_fee = Decimal(str(getattr(invoice, "shipping_fee", None) or "0"))
    if shipping_fee > 0 and default_income_acc:
        revenue_by_account[default_income_acc] += shipping_fee

    if not getattr(invoice, "customer_id", None):
        raise ValueError("Invoice must have a customer.")
    customer_acc = _get_or_create_customer_ar_subaccount(invoice.customer)

    vat_account = _find_control_account(name_contains="VAT")

    entry_date = invoice.date_created.date() if getattr(invoice, "date_created", None) else timezone.localdate()
    cust_name = getattr(invoice.customer, "customer_name", None) or getattr(invoice.customer, "company_name", None) or ""
    desc = f"Invoice {invoice.id:04d} – {cust_name}".strip(" –")

    entry = JournalEntry.objects.create(
        date=entry_date,
        description=desc,
        source_type="invoice",
        source_id=invoice.id,
    )

    #DR Customer A/R (with customer link)
    JournalLine.objects.create(
        entry=entry,
        account=customer_acc,
        debit=total_due,
        credit=Decimal("0.00"),
        customer=invoice.customer,
        supplier=None,
    )

    # CR Revenue
    for acc, amt in revenue_by_account.items():
        if acc and amt > 0:
            JournalLine.objects.create(
                entry=entry,
                account=acc,
                debit=Decimal("0.00"),
                credit=amt,
                customer=None,
                supplier=None,
            )

    # CR VAT
    if vat_total > 0 and vat_account:
        JournalLine.objects.create(
            entry=entry,
            account=vat_account,
            debit=Decimal("0.00"),
            credit=vat_total,
            customer=None,
            supplier=None,
        )

def _post_payment_to_ledger(payment: Payment):
    """
    Customer payment posting:

      DR Bank/Cash (deposit_to)                = amount_received
      CR Customer A/R Subaccount               = invoices + open balance applied
      CR Customer Advances (liability)         = unapplied_amount (excess), if any
    """

    invoice_total = (
        PaymentInvoice.objects
        .filter(payment=payment)
        .aggregate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))["total"]
        or Decimal("0.00")
    )

    ob_total = (
        PaymentOpenBalanceLine.objects
        .filter(payment=payment)
        .aggregate(total=Coalesce(Sum("amount_applied"), Value(Decimal("0.00"))))["total"]
        or Decimal("0.00")
    )

    amount_received = Decimal(str(getattr(payment, "amount_received", Decimal("0.00")) or "0.00"))
    unapplied = Decimal(str(getattr(payment, "unapplied_amount", Decimal("0.00")) or "0.00"))

    # what should be applied to AR
    total_applied_to_ar = invoice_total + ob_total

    # ✅ SAFETY: normalize negative values
    if amount_received < 0:
        amount_received = Decimal("0.00")
    if unapplied < 0:
        unapplied = Decimal("0.00")
    if total_applied_to_ar < 0:
        total_applied_to_ar = Decimal("0.00")

    # ✅ SAFETY: enforce consistency:
    # amount_received should equal applied_to_ar + unapplied
    expected_total = total_applied_to_ar + unapplied

    # If view didn't compute unapplied correctly, auto-fix here
    # (Do not allow ledger to go out of sync.)
    if amount_received > 0 and expected_total != amount_received:
        # Recompute unapplied from amount_received - applied_to_ar
        recalculated_unapplied = amount_received - total_applied_to_ar

        # If recalculated_unapplied is negative, it means allocations exceed amount received
        # That should have been blocked in the view; but we protect ledger anyway.
        if recalculated_unapplied < 0:
            # clamp unapplied to 0 and clamp applied_to_ar down to amount_received
            unapplied = Decimal("0.00")
            total_applied_to_ar = amount_received
        else:
            unapplied = recalculated_unapplied

        # keep payment record in sync
        try:
            Payment.objects.filter(pk=payment.pk).update(unapplied_amount=unapplied)
            payment.unapplied_amount = unapplied
        except Exception:
            # if update fails, still continue posting correctly
            pass

    # if nothing meaningful, remove journal
    if amount_received <= 0:
        JournalEntry.objects.filter(source_type="payment", source_id=payment.id).delete()
        return

    # clear existing journal for this payment
    JournalEntry.objects.filter(source_type="payment", source_id=payment.id).delete()

    if not payment.customer_id:
        raise ValueError("Payment must have a customer.")

    customer_ar = _get_or_create_customer_ar_subaccount(payment.customer)
    deposit_acc = payment.deposit_to
    if not deposit_acc:
        return

    advance_acc = None
    if unapplied > 0:
        advance_acc = _get_customer_advance_account()

    entry_date = payment.payment_date or timezone.localdate()

    bits = [f"Sales Collection {payment.id:04d}"]
    cust_name = payment.customer.customer_name or getattr(payment.customer, "company_name", None)
    if cust_name:
        bits.append(f"– {cust_name}")
    if payment.reference_no:
        bits.append(f"(Ref {payment.reference_no})")
    description = " ".join(bits)

    entry = JournalEntry.objects.create(
        date=entry_date,
        description=description,
        source_type="payment",
        source_id=payment.id,
    )

    # DR Bank/Cash = full amount received
    JournalLine.objects.create(
        entry=entry,
        account=deposit_acc,
        debit=amount_received,
        credit=Decimal("0.00"),
        customer=payment.customer,
        supplier=None,
    )

    # CR A/R = what was applied to invoices + open balance
    if total_applied_to_ar > 0:
        JournalLine.objects.create(
            entry=entry,
            account=customer_ar,
            debit=Decimal("0.00"),
            credit=total_applied_to_ar,
            customer=payment.customer,
            supplier=None,
        )

    # CR Customer Advances = excess/credit
    if advance_acc and unapplied > 0:
        JournalLine.objects.create(
            entry=entry,
            account=advance_acc,
            debit=Decimal("0.00"),
            credit=unapplied,
            customer=payment.customer,
            supplier=None,
        )

def _post_sales_receipt_to_ledger(receipt: SalesReceipt):
    """
    SALES RECEIPT posting (GL-safe + supports overpayment):

      - If amount_paid <= total_amount:
            DR Bank/Cash (deposit_to)           = amount_paid
            DR Customer A/R (optional)          = (total_amount - amount_paid)
            CR Revenue (allocated by lines)     = sales portion
            CR VAT Payable (if any)             = vat_total

      - If amount_paid > total_amount:
            DR Bank/Cash (deposit_to)           = amount_paid
            CR Revenue (allocated by lines)     = sales portion
            CR VAT Payable (if any)             = vat_total
            CR Customer Advances (liability)    = (amount_paid - total_amount)
    """

    # ----------------------------
    # Read totals from receipt
    # ----------------------------
    total_amount = Decimal(str(getattr(receipt, "total_amount", None) or "0"))
    amount_paid  = Decimal(str(getattr(receipt, "amount_paid", None) or "0"))

    if total_amount < 0:
        total_amount = Decimal("0.00")
    if amount_paid < 0:
        amount_paid = Decimal("0.00")

    # If nothing meaningful, delete JE and return
    if total_amount <= 0 and amount_paid <= 0:
        JournalEntry.objects.filter(source_type="sales_receipt", source_id=receipt.id).delete()
        return

    # ----------------------------
    # Compute balance + excess (overpayment)
    # ----------------------------
    if amount_paid >= total_amount:
        balance = Decimal("0.00")
        excess  = amount_paid - total_amount
    else:
        balance = total_amount - amount_paid
        excess  = Decimal("0.00")

    # ----------------------------
    # Build revenue split (by product income accounts)
    # ----------------------------
    revenue_by_account = defaultdict(lambda: Decimal("0.00"))
    vat_total = Decimal("0.00")

    default_income_acc = (
        _find_control_account(name_contains="Sales")
        or _find_control_account(name_contains="Revenue")
    )

    # lines → revenue + VAT
    for line in receipt.lines.select_related("product").all():
        line_amount = Decimal(str(getattr(line, "amount", None) or "0"))
        if line_amount < 0:
            line_amount = Decimal("0.00")

        prod = getattr(line, "product", None)
        income_acc = getattr(prod, "income_account", None) if prod else None
        if not income_acc:
            income_acc = default_income_acc

        if income_acc and line_amount > 0:
            revenue_by_account[income_acc] += line_amount

        vat_total += Decimal(str(getattr(line, "vat_amt", None) or "0"))

    if vat_total < 0:
        vat_total = Decimal("0.00")

    # optional header adjustments you already do
    discount_amt = Decimal(str(getattr(receipt, "total_discount", None) or "0"))
    if discount_amt < 0:
        discount_amt = Decimal("0.00")

    shipping_fee = Decimal(str(getattr(receipt, "shipping_fee", None) or "0"))
    if shipping_fee < 0:
        shipping_fee = Decimal("0.00")

    # Apply discount/shipping to default income account (same as your logic)
    if default_income_acc:
        if discount_amt > 0:
            revenue_by_account[default_income_acc] -= discount_amt
        if shipping_fee > 0:
            revenue_by_account[default_income_acc] += shipping_fee

    # ----------------------------
    # Ensure SALES credits match expected sale portion
    # We assume: total_amount = sales_portion + vat_total
    # so sales_portion = total_amount - vat_total
    # ----------------------------
    sales_target = total_amount - vat_total
    if sales_target < 0:
        sales_target = Decimal("0.00")

    current_sales_credit = sum((amt for amt in revenue_by_account.values()), Decimal("0.00"))

    # Adjust rounding/differences into default income account to keep JE balanced
    diff = sales_target - current_sales_credit
    # allow tiny rounding differences
    if default_income_acc and diff != 0:
        revenue_by_account[default_income_acc] += diff

    # ----------------------------
    # Accounts
    # ----------------------------
    deposit_acc = getattr(receipt, "deposit_to", None)
    if not deposit_acc:
        return

    if not getattr(receipt, "customer_id", None):
        raise ValueError("Sales receipt must have a customer.")

    ar_posting_account = _get_or_create_customer_ar_subaccount(receipt.customer)

    # VAT Payable account (optional)
    vat_account = _find_control_account(name_contains="VAT")

    # Customer Advances account (needed only when excess > 0)
    advance_acc = None
    if excess > 0:
        advance_acc = _get_customer_advance_account()

    # ----------------------------
    # Recreate JE (edit-safe for your current approach)
    # ----------------------------
    JournalEntry.objects.filter(source_type="sales_receipt", source_id=receipt.id).delete()

    entry_date = _safe_date(getattr(receipt, "receipt_date", None), timezone.localdate())
    bits = [f"Receipt {receipt.id:04d}"]
    cust_name = getattr(receipt.customer, "customer_name", None) or getattr(receipt.customer, "company_name", None)
    if cust_name:
        bits.append(f"– {cust_name}")
    ref = getattr(receipt, "reference_no", None)
    if ref:
        bits.append(f"(Ref {ref})")
    description = " ".join(bits)

    entry = JournalEntry.objects.create(
        date=entry_date,
        description=description,
        source_type="sales_receipt",
        source_id=receipt.id,
    )

    # ----------------------------
    # DEBITS
    # ----------------------------
    if amount_paid > 0:
        JournalLine.objects.create(
            entry=entry,
            account=deposit_acc,
            debit=amount_paid,
            credit=Decimal("0.00"),
            customer=receipt.customer,
            supplier=None,
        )

    if balance > 0:
        JournalLine.objects.create(
            entry=entry,
            account=ar_posting_account,
            debit=balance,
            credit=Decimal("0.00"),
            customer=receipt.customer,
            supplier=None,
        )

    # ----------------------------
    # CREDITS (Revenue)
    # ----------------------------
    for acc, amt in revenue_by_account.items():
        # allow negative adjustments only if needed; but skip fully zero lines
        if not acc or amt == 0:
            continue

        if amt > 0:
            JournalLine.objects.create(
                entry=entry,
                account=acc,
                debit=Decimal("0.00"),
                credit=amt,
                customer=None,
                supplier=None,
            )
        else:
            # negative revenue adjustment → debit revenue account
            JournalLine.objects.create(
                entry=entry,
                account=acc,
                debit=abs(amt),
                credit=Decimal("0.00"),
                customer=None,
                supplier=None,
            )

    # VAT payable
    if vat_total > 0 and vat_account:
        JournalLine.objects.create(
            entry=entry,
            account=vat_account,
            debit=Decimal("0.00"),
            credit=vat_total,
            customer=None,
            supplier=None,
        )

    # Customer Advances (excess/overpayment)
    if advance_acc and excess > 0:
        JournalLine.objects.create(
            entry=entry,
            account=advance_acc,
            debit=Decimal("0.00"),
            credit=excess,
            customer=receipt.customer,
            supplier=None,
        )

def _get_vat_payable_account() -> "Account":
    """
    Returns VAT Payable (liability). If you’re not using VAT now, it will still be ready.
    """
    acc = (
        Account.objects
        .filter(is_active=True)
        .filter(account_name__iexact="VAT Payable")
        .first()
    )
    if acc:
        return acc

    acc = (
        Account.objects
        .filter(is_active=True)
        .filter(account_name__icontains="VAT")
        .filter(account_type__in=["CURRENT_LIABILITY", "NON_CURRENT_LIABILITY"])
        .first()
    )
    if acc:
        return acc
    
    
# sales analytics

def _invoice_analytics():
    today = timezone.localdate()

    invs = (
        Newinvoice.objects
        .prefetch_related("payments_applied")
        .only("id", "total_due", "due_date")
    )

    paid_amt = unpaid_amt = overdue_amt = Decimal("0.00")
    paid_cnt = unpaid_cnt = overdue_cnt = 0

    for inv in invs:
        total = _dec(inv.total_due)
        paid = sum((_dec(p.amount_paid) for p in inv.payments_applied.all()), Decimal("0"))
        bal = total - paid

        due = _as_date(getattr(inv, "due_date", None))  #normalize

        if bal <= Decimal("0.00001"):
            paid_cnt += 1
            paid_amt += total
        elif due and due < today:
            overdue_cnt += 1
            overdue_amt += bal
        else:
            unpaid_cnt += 1
            unpaid_amt += bal

    return {
        "paid_amount": paid_amt,
        "paid_count": paid_cnt,
        "unpaid_amount": unpaid_amt,
        "unpaid_count": unpaid_cnt,
        "over_amount": overdue_amt,
        "over_count": overdue_cnt,
    }
def sales(request):
    products = Product.objects.all()

    # You already use this:
    invoices = Newinvoice.objects.all().prefetch_related("invoiceitem_set")
    inv_analytics = _invoice_analytics()

    rows = []

    # ---- Invoices ----
    inv_qs = (
        Newinvoice.objects
        .select_related("customer")
        .prefetch_related("payments_applied")
        .order_by("-date_created", "-id")
    )
    for inv in inv_qs:
        total = _dec(inv.total_due)
        paid = sum((_dec(p.amount_paid) for p in inv.payments_applied.all()), Decimal("0"))
        bal = total - paid
        status = status_for_invoice(inv, total, paid, bal)

        rows.append({
            "date": inv.date_created,
            "type": "Invoice",
            "no": f"{inv.id:04d}",
            "customer": inv.customer.customer_name if inv.customer_id else "",
            "memo": (inv.memo or "")[:140],
            "amount": total,
            "status": status,
            "edit_url": reverse("sales:edit-invoice", args=[inv.id]),
            "view_url": reverse("sales:invoice-detail", args=[inv.id]),
            "print_url": reverse("sales:invoice-print", args=[inv.id]),
        })

    # ---- Payments ----
    pay_qs = (
        Payment.objects
        .select_related("customer", "deposit_to")
        .annotate(
            applied_total=Coalesce(Sum("applied_invoices__amount_paid"), Value(Decimal("0.00"))),
        )
        .order_by("-payment_date", "-id")
    )
    for p in pay_qs:
        rows.append({
            "date": p.payment_date,
            "type": "Payment",
            "no": (p.reference_no or f"{p.id:04d}"),
            "customer": p.customer.customer_name if p.customer_id else "",
            "memo": (p.memo or "")[:140],
            "amount": p.applied_total or Decimal("0"),
            "status": "Closed" if (p.applied_total or 0) > 0 else "Unapplied",
            "edit_url":  reverse("sales:payment-edit", args=[p.id]),
            "view_url":  reverse("sales:payment-detail", args=[p.id]),
            "print_url": reverse("sales:payment-print", args=[p.id]),
        })

    # ---- Sales Receipts ----
    sr_qs = (
        SalesReceipt.objects
        .select_related("customer", "deposit_to")
        .annotate(
            total_amount_dec=Cast("total_amount", DecimalField(max_digits=18, decimal_places=2)),
            amount_paid_dec=Cast(
                Coalesce(F("amount_paid"), Value(Decimal("0.00"))),
                DecimalField(max_digits=18, decimal_places=2),
            ),
        )
        .order_by("-receipt_date", "-id")
    )
    for r in sr_qs:
        total = _dec(r.total_amount)
        paid  = _dec(r.amount_paid)
        status = _receipt_status(r)  # you already have this

        rows.append({
            "date": r.receipt_date,
            "type": "Sales Receipt",
            "no": (r.reference_no or f"{r.id:04d}"),
            "customer": r.customer.customer_name if r.customer_id else "",
            "memo": (r.memo or "")[:140],
            "amount": total,
            "status": status,
            "edit_url":  reverse("sales:receipt-edit", args=[r.id]),
            "view_url":  reverse("sales:receipt-detail", args=[r.id]),
            "print_url": reverse("sales:receipt-print", args=[r.id]),
        })

    #sort newest first, safely even if a row has date=None
    def sort_key(x):
        d = _as_date(x.get("date"))
        return (d or date.min, x.get("type", ""))

    rows.sort(key=sort_key, reverse=True)

    return render(
        request,
        "Sales.html",
        {
            "products": products,
            "invoices": invoices,
            "inv_analytics": inv_analytics,
            "sales_rows": rows,
        },
    )
# invoice form view
def get_product_details(request, pk):
    """
    Returns key fields needed to auto-fill invoice row.
    """
    try:
        product = Product.objects.get(pk=pk)
        data = {
            "id": product.id,
            "name": product.name,
            "sales_description": product.sales_description or "",
            "sales_price": str(product.sales_price or 0),
            "taxable": bool(product.taxable),
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({"error": "Product not found"}, status=404)

TERMS_DAYS = {
    "due_on_receipt": 0, "one_day": 1, "two_days": 2, "net_7": 7,
    "net_15": 15, "net_30": 30, "net_60": 60,
    "credit_limit": 27, "credit_allowance": 29,
}
 
def parse_date_flexible(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None  # if nothing matched

@transaction.atomic
def add_invoice(request):
    if request.method == "POST":
        raw_date_created = (request.POST.get("date_created") or "").strip()
        raw_due_date     = (request.POST.get("due_date") or "").strip()

        customer_id = request.POST.get("customer")
        customer = None
        if customer_id:
            try:
                customer = Newcustomer.objects.get(pk=customer_id)
            except Newcustomer.DoesNotExist:
                customer = None

        email            = request.POST.get("email")
        billing_address  = request.POST.get("billing_address")
        shipping_address = request.POST.get("shipping_address")
        terms            = (request.POST.get("terms") or "").strip()
        sales_rep        = request.POST.get("sales_rep")

        class_field_id = request.POST.get("class_field")
        class_field = None
        if class_field_id:
            try:
                class_field = Pclass.objects.get(pk=class_field_id)
            except Pclass.DoesNotExist:
                class_field = None

        tags          = request.POST.get("tags")
        po_num        = request.POST.get("po_number")  #FIX: your form uses po_number
        memo          = request.POST.get("memo")
        customs_notes = request.POST.get("customs_notes")

        subtotal       = Decimal(request.POST.get("subtotal") or "0")
        total_discount = Decimal(request.POST.get("total_discount") or "0")
        shipping_fee   = Decimal(request.POST.get("shipping_fee") or "0")

        created_dt = parse_date_flexible(raw_date_created)
        due_dt     = parse_date_flexible(raw_due_date)

        if not due_dt and created_dt and terms in TERMS_DAYS:
            due_dt = created_dt + timedelta(days=TERMS_DAYS[terms])

        # Create invoice first
        invoice = Newinvoice.objects.create(
            customer=customer,
            email=email,
            date_created=as_aware_datetime(created_dt),
            due_date=as_aware_datetime(due_dt),
            billing_address=billing_address,
            shipping_address=shipping_address,
            class_field=class_field,
            terms=terms,
            sales_rep=sales_rep,
            tags=tags,
            po_num=po_num,
            memo=memo,
            customs_notes=customs_notes,
            subtotal=subtotal,
            total_discount=total_discount,
            total_vat=Decimal("0"),
            shipping_fee=shipping_fee,
            total_due=Decimal("0"),
        )

        # Line items arrays
        products          = request.POST.getlist("product[]")
        descriptions      = request.POST.getlist("description[]")
        qtys              = request.POST.getlist("qty[]")
        rates             = request.POST.getlist("unit_price[]")
        amounts           = request.POST.getlist("amount[]")
        vats              = request.POST.getlist("vat[]")
        discount_nums     = request.POST.getlist("discount_num[]")
        discount_amounts  = request.POST.getlist("discount_amount[]")

        total_vat = Decimal("0")

        for i in range(len(products)):
            if not products[i]:
                continue

            product = get_object_or_404(Product, pk=products[i])

            qty_val = Decimal(qtys[i] or "0") if i < len(qtys) else Decimal("0")
            rate_val = Decimal(rates[i] or "0") if i < len(rates) else Decimal("0")
            amt_val = Decimal(amounts[i] or "0") if i < len(amounts) else Decimal("0")
            vat_val = Decimal(vats[i] or "0") if i < len(vats) else Decimal("0")
            disc_num_val = Decimal(discount_nums[i] or "0") if i < len(discount_nums) else Decimal("0")
            disc_amt_val = Decimal(discount_amounts[i] or "0") if i < len(discount_amounts) else Decimal("0")
            desc_val = descriptions[i] if i < len(descriptions) else ""

            InvoiceItem.objects.create(
                invoice=invoice,
                product=product,
                description=desc_val,
                qty=qty_val,
                unit_price=rate_val,
                amount=amt_val,
                vat=vat_val,
                discount_num=disc_num_val,
                discount_amount=disc_amt_val,
            )

            total_vat += vat_val

        # Update totals
        total_due = (subtotal - total_discount) + shipping_fee + total_vat
        invoice.total_vat = total_vat
        invoice.total_due = total_due
        apply_audit_fields(invoice)
        invoice.save()

        _post_invoice_to_ledger(invoice)

        save_action = request.POST.get("save_action")
        if save_action == "save":
            return redirect("sales:invoices")
        if save_action == "save&new":
            return redirect("sales:add-invoice")
        if save_action == "save&close":
            return redirect("sales:sales")

        return redirect("sales:add-invoice")

    products  = Product.objects.all()
    customers = Newcustomer.objects.all()
    classes   = Pclass.objects.all()

    last_invoice = Newinvoice.objects.order_by("-id").first()
    next_id = 1 if not last_invoice else last_invoice.id + 1
    next_invoice_id = f"{next_id:03d}"

    return render(request, "invoice_form.html", {
        "customers": customers,
        "classes": classes,
        "products": products,
        "next_invoice_id": next_invoice_id,
    })
# edit invoice

def edit_invoice(request, pk: int):
    """
    Edit an invoice:
      - GET: prefill your existing invoice_form.html
      - POST: update header + replace line items, recompute totals on the server
    """
    inv = get_object_or_404(
        Newinvoice.objects.select_related("customer", "class_field"),
        pk=pk
    )

    if request.method == "POST":
        # ----- Header fields -----
        customer_id   = request.POST.get("customer")
        email         = request.POST.get("email")
        billing_addr  = request.POST.get("billing_address")
        shipping_addr = request.POST.get("shipping_address")
        terms         = (request.POST.get("terms") or "").strip()
        sales_rep     = request.POST.get("sales_rep")
        class_id      = request.POST.get("class_field")
        tags          = request.POST.get("tags")
        po_num        = request.POST.get("po_number") or request.POST.get("po_num")
        memo          = request.POST.get("memo")
        customs_notes = request.POST.get("customs_notes")

        customer    = Newcustomer.objects.filter(pk=customer_id).first() if customer_id else None
        class_field = Pclass.objects.filter(pk=class_id).first() if class_id else None

        created_dt = parse_date_flexible(request.POST.get("date_created"))
        due_dt     = parse_date_flexible(request.POST.get("due_date"))

        if not due_dt and created_dt and terms in TERMS_DAYS:
            due_dt = created_dt + timedelta(days=TERMS_DAYS[terms])

        # We will recompute totals from lines; only shipping is taken from POST
        shipping_fee = Decimal(request.POST.get("shipping_fee") or 0)

        # ----- Replace line items & recompute totals (authoritative) -----
        InvoiceItem.objects.filter(invoice=inv).delete()

        products       = request.POST.getlist("product[]")
        descriptions   = request.POST.getlist("description[]")
        qtys           = request.POST.getlist("qty[]")
        rates          = request.POST.getlist("unit_price[]")
        discount_percs = request.POST.getlist("discount_num[]")

        line_rows = []

        subtotal       = Decimal("0.00")
        total_discount = Decimal("0.00")
        total_vat      = Decimal("0.00")

        for i in range(len(products)):
            if not products[i]:
                continue

            product = get_object_or_404(Product, pk=products[i])

            desc = descriptions[i] if i < len(descriptions) else ""
            qty  = Decimal((qtys[i] or "0").strip() if i < len(qtys) else "0")
            rate = Decimal((rates[i] or "0").strip() if i < len(rates) else "0")
            dpc  = Decimal((discount_percs[i] or "0").strip() if i < len(discount_percs) else "0")

            line_amount = (qty * rate).quantize(Decimal("0.01"))
            line_discount_amt = (line_amount * dpc / Decimal("100")).quantize(Decimal("0.01"))

            # VAT on pre-discount amount (same as your comment)
            if getattr(product, "taxable", False):
                line_vat = (line_amount * Decimal("0.18")).quantize(Decimal("0.01"))
            else:
                line_vat = Decimal("0.00")

            subtotal       += line_amount
            total_discount += line_discount_amt
            total_vat      += line_vat

            line_rows.append(InvoiceItem(
                invoice=inv,
                product=product,
                description=desc,
                qty=qty,
                unit_price=rate,
                amount=line_amount,
                vat=line_vat,
                discount_num=dpc,
                discount_amount=line_discount_amt,
            ))

        if line_rows:
            InvoiceItem.objects.bulk_create(line_rows)

        #Save header fields (FIX: store aware datetime into DateTimeFields)
        inv.customer         = customer
        inv.email            = email
        inv.date_created     = as_aware_datetime(created_dt)  #FIX
        inv.due_date         = as_aware_datetime(due_dt)      #FIX
        inv.billing_address  = billing_addr
        inv.shipping_address = shipping_addr
        inv.class_field      = class_field
        inv.terms            = terms
        inv.sales_rep        = sales_rep
        inv.tags             = tags
        inv.po_num           = po_num
        inv.memo             = memo
        inv.customs_notes    = customs_notes

        inv.subtotal        = subtotal
        inv.total_discount  = total_discount
        inv.total_vat       = total_vat
        inv.shipping_fee    = shipping_fee
        inv.total_due       = (subtotal - total_discount + total_vat + shipping_fee).quantize(Decimal("0.01"))

        apply_audit_fields(inv)
        inv.save()

        # ⇨ Re-post to General Ledger
        _post_invoice_to_ledger(inv)

        return redirect("sales:invoice-detail", pk=inv.pk)

    # ----- GET: prefill form -----
    products  = Product.objects.all()
    customers = Newcustomer.objects.all()
    classes   = Pclass.objects.all()
    items     = InvoiceItem.objects.filter(invoice=inv).select_related("product").order_by("id")

    return render(request, "invoice_form.html", {
        "edit_mode": True,
        "inv": inv,
        "items": items,
        "products": products,
        "customers": customers,
        "classes": classes,
        "next_invoice_id": f"{inv.id:03d}",
    })
def add_class_ajax(request):
    if request.method == "POST":
        name = request.POST.get("name")
        if not name:
            return JsonResponse({"success": False, "error": "Class name required"})
        
        cls, created = Pclass.objects.get_or_create(class_name=name)
        return JsonResponse({
            "success": True,
            "id": cls.id,
            "name": cls.class_name,
        })
    
#  invoice list
def invoice_list(request):
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "all").lower()
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    invoices_qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), DecimalField(max_digits=18, decimal_places=2)),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid"),
                Value(Decimal("0.00"))
            )
        )
        .order_by("-date_created", "-id")
    )

    # ✅ SEARCH (FIXED)
    if q:
        search_q = (
            Q(customer__customer_name__icontains=q) |
            Q(email__icontains=q) |
            Q(billing_address__icontains=q)
        )
        if q.isdigit():
            search_q |= Q(id=int(q))
        invoices_qs = invoices_qs.filter(search_q)

    invoices = []
    counts = {"all": 0, "overdue": 0, "paid": 0, "partial": 0}

    for inv in invoices_qs:
        total_due = inv.total_due_dec or Decimal("0")
        total_paid = inv.total_paid or Decimal("0")
        balance = max(total_due - total_paid, Decimal("0"))

        inv.status = status_for_invoice(inv, total_due, total_paid, balance)
        s = inv.status.lower()

        counts["all"] += 1
        if "overdue" in s:
            counts["overdue"] += 1
        if "paid" in s or "deposited" in s:
            counts["paid"] += 1
        if "partially paid" in s:
            counts["partial"] += 1

        # status filter
        keep = (
            status_filter == "all" or
            (status_filter == "overdue" and "overdue" in s) or
            (status_filter == "paid" and ("paid" in s or "deposited" in s)) or
            (status_filter == "partial" and "partially paid" in s)
        )

        if keep:
            invoices.append(inv)

    # ✅ AJAX RESPONSE (for live search)
    if is_ajax:
        rows = []
        for inv in invoices:
            rows.append({
                "id": f"{inv.id:04d}",
                "customer": inv.customer.customer_name,
                "created": inv.date_created.strftime("%d/%m/%Y"),
                "due": inv.due_date.strftime("%d/%m/%Y") if inv.due_date else "—",
                "email": inv.email or "—",
                "billing": inv.billing_address or "—",
                "status": inv.status,
                "edit_url": f"/sales/invoices/{inv.id}/edit/",
                "view_url": f"/sales/invoices/{inv.id}/",
                "print_url": f"/sales/invoices/{inv.id}/print/",
            })
        return JsonResponse({"rows": rows})

    customers = Newcustomer.objects.all()
    return render(request, "invoice_lists.html", {
        "invoices": invoices,
        "customers": customers,
        "q": q,
        "status_filter": status_filter,
        "counts": counts,
    })

# ednd
def full_invoice_details(request):
    invoices=Newinvoice.objects.all()
    customers=Newcustomer.objects.all()
    return render(request, 'full_invoice_details.html',{
        'invoices':invoices,
        'customers':customers
    })
# edit and view  views 

def invoice_detail(request, pk: int):
    inv = get_object_or_404(
        Newinvoice.objects.select_related("customer", "class_field"),
        pk=pk
    )

    agg = (
        Newinvoice.objects.filter(pk=pk)
        .annotate(
            total_due_dec=Cast("total_due", DecimalField(max_digits=18, decimal_places=2)),
            total_paid=Coalesce(Sum("payments_applied__amount_paid"), Value(Decimal("0.00"))),
        )
        .values("total_due_dec", "total_paid")
        .first()
    ) or {"total_due_dec": Decimal("0"), "total_paid": Decimal("0")}

    total_due = agg["total_due_dec"] or Decimal("0")
    total_paid = agg["total_paid"] or Decimal("0")
    balance   = max(total_due - total_paid, Decimal("0"))

    # status (same rules you already use in the list)
    today = date.today()
    overdue_days = (today - inv.due_date).days if inv.due_date and balance > 0 and today > inv.due_date else None

    deposited = False
    if total_due > 0 and balance == 0:
        aps = inv.payments_applied.select_related("payment__deposit_to").all()
        if aps:
            def is_bankish(acc):
                if not acc: return False
                at = (acc.account_type or "").lower()
                dt = (acc.detail_type or "").lower()
                return at in ("bank", "cash and cash equivalents", "cash_equiv", "cash & cash equivalents") or "bank" in dt
            deposited = all(is_bankish(pi.payment.deposit_to) for pi in aps if pi.payment)

    if total_due == 0:
        status_text = "Cleared"
    elif balance == 0:
        status_text = "Deposited" if deposited else "Paid"
    else:
        if overdue_days:
            status_text = f"Overdue {overdue_days} days"
            if total_paid > 0:
                status_text += f" — Partially paid now {balance:,.0f} remaining"
            else:
                status_text += f" — {balance:,.0f} remaining"
        elif inv.due_date and inv.due_date == today and balance > 0:
            status_text = f"Due today — {balance:,.0f} remaining"
        else:
            status_text = f"Partially paid, {balance:,.0f} remaining" if total_paid > 0 else f"{balance:,.0f} remaining"

    items = InvoiceItem.objects.filter(invoice=inv).select_related("product").order_by("id")
    payments = (
        PaymentInvoice.objects
        .filter(invoice=inv)
        .select_related("payment", "payment__deposit_to")
        .order_by("-payment__payment_date", "-id")
    )

    payment_rows = [{
        "id":p.payment.id,
        "date": p.payment.payment_date,
        "ref": p.payment.reference_no,
        "method": (p.payment.payment_method or "").replace("_", " ").title(),
        "deposit_to": p.payment.deposit_to.account_name if p.payment.deposit_to else "",
        "amount": p.amount_paid,
    } for p in payments]

    return render(request, "invoice_detail.html", {
        "inv": inv,
        "items": items,
        "status_text": status_text,
        "total_due": total_due,
        "total_paid": total_paid,
        "balance": balance,
        "payment_rows": payment_rows,
    })
# invoice print view
def invoice_print(request, pk: int):
    inv = get_object_or_404(
        Newinvoice.objects.select_related("customer", "class_field"),
        pk=pk
    )

    agg = (
        Newinvoice.objects.filter(pk=pk)
        .annotate(
            total_due_dec=Cast("total_due", DecimalField(max_digits=18, decimal_places=2)),
            total_paid=Coalesce(Sum("payments_applied__amount_paid"), Value(Decimal("0.00"))),
        )
        .values("total_due_dec", "total_paid")
        .first()
    ) or {"total_due_dec": Decimal("0"), "total_paid": Decimal("0")}

    total_due = agg["total_due_dec"] or Decimal("0")
    total_paid = agg["total_paid"] or Decimal("0")
    balance   = max(total_due - total_paid, Decimal("0"))
    status_text = status_for_invoice(inv, total_due, total_paid, balance)

    items = InvoiceItem.objects.filter(invoice=inv).select_related("product").order_by("id")

    payments = (
        PaymentInvoice.objects
        .filter(invoice=inv)
        .select_related("payment", "payment__deposit_to")
        .order_by("-payment__payment_date", "-id")
    )
    payment_rows = [{
        "date": p.payment.payment_date,
        "ref": p.payment.reference_no,
        "method": (p.payment.payment_method or "").replace("_", " ").title(),
        "deposit_to": p.payment.deposit_to.account_name if p.payment.deposit_to else "",
        "amount": p.amount_paid,
    } for p in payments]

    # Optional: company/org details (replace these with your real ones or pull from a Company model)
    org = {
        "name": "Sowa Accountants Ltd",
        "address": "Plot 123, Kampala Road, Kampala",
        "phone": "+256 700 000 000",
        "email": "accounts@sowaf.co.ug",
        "website": "www.sowaf.co.ug",
        "logo_url": request.build_absolute_uri(static("sowaf/images/yo-logo.png")),
    }

    return render(request, "invoice_print.html", {
        "inv": inv,
        "items": items,
        "status_text": status_text,
        "total_due": total_due,
        "total_paid": total_paid,
        "balance": balance,
        "payment_rows": payment_rows,
        "org": org,
    })

# ------------------------------------------------------------
# RECEIVE PAYMENT (CREATE)
# ------------------------------------------------------------
@transaction.atomic
def receive_payment_view(request):
    customers = Newcustomer.objects.order_by("customer_name")
    accounts = deposit_accounts_qs()  # only Bank + Cash & Cash Equivalents

    if request.method == "POST":
        customer_id    = (request.POST.get("customer") or "").strip()
        payment_date   = parse_date(request.POST.get("payment_date") or "")
        payment_method = (request.POST.get("payment_method") or "cash").strip()
        deposit_to_id  = (request.POST.get("deposit_to") or "").strip()
        reference_no   = (request.POST.get("reference_no") or "").strip()
        tags           = (request.POST.get("tags") or "").strip()
        memo           = (request.POST.get("memo") or "").strip()

        # ✅ Amount received (from your HTML)
        amount_received = _dec(request.POST.get("amount_received"), "0.00")

        # resolve deposit account (only allowed set)
        deposit_account = None
        if deposit_to_id.isdigit():
            deposit_account = accounts.filter(id=int(deposit_to_id)).first()

        # ensure reference
        if not (len(reference_no) == 8 and reference_no.isdigit()):
            reference_no = generate_unique_ref_no()
        if Payment.objects.filter(reference_no=reference_no).exists():
            reference_no = generate_unique_ref_no()

        if not (customer_id.isdigit() and payment_date and deposit_account):
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "reference_no": reference_no,
                "form_error": "Please select a customer, a valid Bank/Cash account, and a date.",
            })

        if amount_received <= 0:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "reference_no": reference_no,
                "form_error": "Amount Received must be greater than 0.",
            })

        customer = get_object_or_404(Newcustomer, pk=int(customer_id))

        # --------------------------
        # invoice allocations: amount_paid_<invoice_id>
        # --------------------------
        allocations = []
        for key, val in request.POST.items():
            if key.startswith("amount_paid_"):
                inv_id = key.split("_")[-1]
                if inv_id.isdigit():
                    amt = _dec(val, "0.00")
                    if amt > 0:
                        allocations.append((int(inv_id), amt))

        # open balance typed (optional)
        raw_ob = request.POST.get("open_balance_amount")
        open_balance_manual = _dec(raw_ob, "0.00") if raw_ob is not None else Decimal("0.00")
        if open_balance_manual < 0:
            open_balance_manual = Decimal("0.00")

        # =====================================================================
        # ✅ FIX: validate invoice allocations not exceeding their balances
        #    (force Decimal arithmetic in annotations)
        # =====================================================================
        if allocations:
            dec_out = DecimalField(max_digits=18, decimal_places=2)

            invoice_ids = [i for i, _ in allocations]

            balances = (
                Newinvoice.objects
                .filter(id__in=invoice_ids)
                .annotate(
                    # If total_due is FloatField in DB, cast it to Decimal for safe math
                    total_due_dec=Cast(F("total_due"), output_field=dec_out),
                    total_paid=Coalesce(
                        Sum("payments_applied__amount_paid", output_field=dec_out),
                        Value(Decimal("0.00"), output_field=dec_out),
                        output_field=dec_out,
                    ),
                )
                .annotate(
                    outstanding_balance=ExpressionWrapper(
                        F("total_due_dec") - F("total_paid"),
                        output_field=dec_out,
                    )
                )
                .values_list("id", "outstanding_balance")
            )

            balance_map = {iid: Decimal(str(bal or "0")) for iid, bal in balances}

            for invoice_id, amount in allocations:
                max_allowed = balance_map.get(invoice_id)
                if max_allowed is None or amount > max_allowed:
                    return render(request, "receive_payment.html", {
                        "customers": customers,
                        "accounts": accounts,
                        "reference_no": reference_no,
                        "form_error": f"Allocation {amount} exceeds outstanding balance {max_allowed} on invoice {invoice_id}.",
                    })

        invoice_total = sum((amt for _, amt in allocations), Decimal("0.00"))

        # current open balance BEFORE this payment
        current_open_balance = _customer_open_balance_amount(customer)

        # remaining after invoices
        remaining_after_invoices = amount_received - invoice_total
        if remaining_after_invoices < 0:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "reference_no": reference_no,
                "form_error": "Amount Received is less than invoice allocations.",
            })

        # decide open balance apply:
        # if user typed open_balance_amount use that, else auto-apply remaining to OB (up to current_open_balance)
        if open_balance_manual > 0:
            open_balance_apply = open_balance_manual
        else:
            open_balance_apply = min(remaining_after_invoices, current_open_balance)

        if open_balance_apply < 0:
            open_balance_apply = Decimal("0.00")

        remaining_after_open_balance = remaining_after_invoices - open_balance_apply

        # ✅ any remaining is customer credit (unapplied)
        unapplied = remaining_after_open_balance if remaining_after_open_balance > 0 else Decimal("0.00")

        # must have something meaningful
        if invoice_total <= 0 and open_balance_apply <= 0 and unapplied <= 0:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "reference_no": reference_no,
                "form_error": "Nothing to apply. Enter invoice amounts or open balance.",
            })

        payment = Payment.objects.create(
            customer=customer,
            payment_date=payment_date,
            payment_method=payment_method,
            deposit_to=deposit_account,
            reference_no=reference_no,
            tags=tags,
            memo=memo,
            amount_received=amount_received,  # ✅ stored
            unapplied_amount=unapplied,       # ✅ stored
        )
        apply_audit_fields(payment)
        payment.save()

        if allocations:
            PaymentInvoice.objects.bulk_create([
                PaymentInvoice(payment=payment, invoice_id=inv_id, amount_paid=amt)
                for inv_id, amt in allocations
            ])

        # save open balance line
        _save_payment_open_balance(payment, open_balance_apply)

        # ✅ Post to GL (DR Bank, CR Customer AR) for full amount_received
        _post_payment_to_ledger(payment)

        return redirect(f"{request.path}?ok=1")

    reference_no = generate_unique_ref_no()
    return render(request, "receive_payment.html", {
        "customers": customers,
        "accounts": accounts,
        "reference_no": reference_no,
    })

# ------------------------------------------------------------
# PAYMENT EDIT (NO amount_received logic)
# ------------------------------------------------------------
@transaction.atomic
def payment_edit(request, pk: int):
    payment = get_object_or_404(
        Payment.objects.select_related("customer", "deposit_to"),
        pk=pk
    )
    customers = Newcustomer.objects.order_by("customer_name")
    accounts  = deposit_accounts_qs()

    if request.method == "POST":
        customer_id    = (request.POST.get("customer") or "").strip()
        payment_date   = parse_date(request.POST.get("payment_date") or "")
        payment_method = (request.POST.get("payment_method") or "cash").strip()
        deposit_to_id  = (request.POST.get("deposit_to") or "").strip()
        reference_no   = (request.POST.get("reference_no") or "").strip()
        tags           = (request.POST.get("tags") or "").strip()
        memo           = (request.POST.get("memo") or "").strip()

        amount_received = _dec(request.POST.get("amount_received"), "0.00")

        if not (customer_id.isdigit() and payment_date and deposit_to_id.isdigit()):
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "payment": payment,
                "reference_no": payment.reference_no or generate_unique_ref_no(),
                "prefill_rows": _payment_prefill_rows(payment),
                "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                "edit_mode": True,
                "form_error": "Please select a customer, a valid Bank/Cash account, and a date.",
            })

        if amount_received <= 0:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "payment": payment,
                "reference_no": payment.reference_no or generate_unique_ref_no(),
                "prefill_rows": _payment_prefill_rows(payment),
                "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                "edit_mode": True,
                "form_error": "Enter Amount Received (must be > 0).",
            })

        customer = get_object_or_404(Newcustomer, pk=int(customer_id))
        deposit_account = get_object_or_404(accounts, pk=int(deposit_to_id))

        # allocations
        allocations = []
        for key, val in request.POST.items():
            if key.startswith("amount_paid_"):
                inv_id = key.split("_")[-1]
                if inv_id.isdigit():
                    amt = _dec(val, "0.00")
                    if amt > 0:
                        allocations.append((int(inv_id), amt))

        # open balance
        raw_ob = request.POST.get("open_balance_amount")
        open_balance_amount = _dec(raw_ob, "0.00")
        if open_balance_amount < 0:
            open_balance_amount = Decimal("0.00")

        # validate allocations (edit-safe)
        prev_alloc_qs = PaymentInvoice.objects.filter(payment=payment).values_list("invoice_id", "amount_paid")
        prev_map = {}
        for iid, amt in prev_alloc_qs:
            prev_map[iid] = prev_map.get(iid, Decimal("0.00")) + Decimal(amt or 0)

        if allocations:
            invoice_ids = [i for i, _ in allocations]
            balances = (
                Newinvoice.objects.filter(id__in=invoice_ids)
                .annotate(
                    total_due_dec=F("total_due"),
                    total_paid=Coalesce(Sum("payments_applied__amount_paid"), Value(Decimal("0.00"))),
                )
                .annotate(outstanding_balance=F("total_due_dec") - F("total_paid"))
                .values_list("id", "outstanding_balance")
            )
            balance_map = {iid: Decimal(str(bal or "0")) for iid, bal in balances}

            for invoice_id, new_amt in allocations:
                allowed = balance_map.get(invoice_id, Decimal("0.00")) + prev_map.get(invoice_id, Decimal("0.00"))
                if new_amt > allowed:
                    return render(request, "receive_payment.html", {
                        "customers": customers,
                        "accounts": accounts,
                        "payment": payment,
                        "reference_no": payment.reference_no or generate_unique_ref_no(),
                        "prefill_rows": _payment_prefill_rows(payment),
                        "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                        "edit_mode": True,
                        "form_error": f"Allocation {new_amt} exceeds allowed {allowed} on invoice {invoice_id}.",
                    })

        invoice_total = sum((amt for _, amt in allocations), Decimal("0.00"))

        current_open_balance = _customer_open_balance_amount(customer)
        if open_balance_amount > current_open_balance:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "payment": payment,
                "reference_no": payment.reference_no or generate_unique_ref_no(),
                "prefill_rows": _payment_prefill_rows(payment),
                "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                "edit_mode": True,
                "form_error": f"Open Balance amount {open_balance_amount} exceeds current Open Balance {current_open_balance}.",
            })

        total_applied_to_ar = invoice_total + open_balance_amount
        if total_applied_to_ar <= 0:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "payment": payment,
                "reference_no": payment.reference_no or generate_unique_ref_no(),
                "prefill_rows": _payment_prefill_rows(payment),
                "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                "edit_mode": True,
                "form_error": "Enter at least one invoice amount or an Open Balance amount.",
            })

        if amount_received < total_applied_to_ar:
            return render(request, "receive_payment.html", {
                "customers": customers,
                "accounts": accounts,
                "payment": payment,
                "reference_no": payment.reference_no or generate_unique_ref_no(),
                "prefill_rows": _payment_prefill_rows(payment),
                "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
                "edit_mode": True,
                "form_error": "Amount Received is less than total applied (Invoices + Open Balance).",
            })

        unapplied = amount_received - total_applied_to_ar
        if unapplied < 0:
            unapplied = Decimal("0.00")

        # save header
        payment.customer = customer
        payment.payment_date = payment_date
        payment.payment_method = payment_method
        payment.deposit_to = deposit_account
        payment.reference_no = reference_no if reference_no else payment.reference_no
        payment.tags = tags
        payment.memo = memo
        payment.amount_received = amount_received
        payment.unapplied_amount = unapplied
        apply_audit_fields(payment)
        payment.save()

        # replace allocations
        PaymentInvoice.objects.filter(payment=payment).delete()
        if allocations:
            PaymentInvoice.objects.bulk_create([
                PaymentInvoice(payment=payment, invoice_id=inv_id, amount_paid=amt)
                for inv_id, amt in allocations
            ])

        _save_payment_open_balance(payment, open_balance_amount)

        _post_payment_to_ledger(payment)

        return redirect('sales:payment-detail', pk=payment.pk)

    context = {
        "customers": customers,
        "accounts": accounts,
        "payment": payment,
        "reference_no": payment.reference_no or generate_unique_ref_no(),
        "prefill_rows": _payment_prefill_rows(payment),
        "open_balance_prefill": getattr(getattr(payment, "open_balance_line", None), "amount_applied", Decimal("0.00")),
        "edit_mode": True,
    }
    return render(request, "receive_payment.html", context)

@require_GET
def outstanding_invoices_api(request):
    customer_id = request.GET.get("customer")
    if not customer_id:
        return JsonResponse({"invoices": [], "open_balance": "0.00"})

    customer = Newcustomer.objects.filter(pk=customer_id).first()
    if not customer:
        return JsonResponse({"invoices": [], "open_balance": "0.00"})

    invoices_payload = []

    # map invoice -> sum applied
    applied_map = dict(
        PaymentInvoice.objects
        .filter(invoice__customer_id=customer_id)
        .values("invoice_id")
        .annotate(s=Sum("amount_paid"))
        .values_list("invoice_id", "s")
    )

    for inv in Newinvoice.objects.filter(customer_id=customer_id).order_by("-date_created"):
        total = Decimal(str(inv.total_due or "0"))
        applied = Decimal(str(applied_map.get(inv.id) or "0"))
        balance = total - applied
        if balance <= 0:
            continue

        invoices_payload.append({
            "id": inv.id,
            "date_created": inv.date_created.strftime("%Y-%m-%d") if inv.date_created else None,
            "due_date": inv.due_date.strftime("%Y-%m-%d") if getattr(inv, "due_date", None) else None,
            "total_due": str(total),
            "balance": str(balance),
        })

    open_balance = _customer_open_balance_amount(customer)

    return JsonResponse({
        "invoices": invoices_payload,
        "open_balance": str(open_balance),
    })

# payment lists
def payments_list(request):
    payments = (
        Payment.objects
        .select_related("customer", "deposit_to")
        .prefetch_related("applied_invoices__invoice")
        .order_by("-payment_date", "-id")
    )

    # collect invoice ids appearing in payment lines
    invoice_ids = set()
    for p in payments:
        for pli in p.applied_invoices.all():
            invoice_ids.add(pli.invoice_id)

    # total paid to date per invoice
    totals = (
        PaymentInvoice.objects
        .filter(invoice_id__in=invoice_ids)
        .values("invoice_id")
        .annotate(total_paid=Sum("amount_paid"))
    )
    total_paid_map = {row["invoice_id"]: row["total_paid"] for row in totals}

    # fetch invoice objects
    invoices_by_id = Newinvoice.objects.in_bulk(invoice_ids)

    rows = []
    for p in payments:
        line_rows = []
        for pli in p.applied_invoices.all():
            inv = invoices_by_id.get(pli.invoice_id)
            if not inv:
                continue
            total_due = Decimal(str(inv.total_due or "0"))
            amount_applied = pli.amount_paid
            remaining_this_payment = total_due - amount_applied
            outstanding_now = total_due - (total_paid_map.get(pli.invoice_id) or Decimal("0"))

            line_rows.append({
                "invoice": inv,
                "amount_applied": amount_applied,
                "total_due": total_due,
                "remaining_this_payment": remaining_this_payment,
                "outstanding_now": outstanding_now,
            })

        # precompute section totals so the template stays simple
        applied_total = sum((lr["amount_applied"] for lr in line_rows), Decimal("0"))
        remaining_total_this_payment = sum((lr["remaining_this_payment"] for lr in line_rows), Decimal("0"))
        outstanding_total_now = sum((lr["outstanding_now"] for lr in line_rows), Decimal("0"))

        rows.append({
            "payment": p,
            "lines": line_rows,
            "applied_total": applied_total,
            "remaining_total_this_payment": remaining_total_this_payment,
            "outstanding_total_now": outstanding_total_now,
        })

    return render(request, "payments_list.html", {"rows": rows})
# individual payment
def payment_detail(request, pk: int):
    payment = get_object_or_404(
        Payment.objects.select_related("customer", "deposit_to"),
        pk=pk
    )
    group = _payment_prefill_rows(payment)
    return render(request, "payment_detail.html", {"group": group, "payment":payment})

# payment printout  
def _lines_for_payment(payment: Payment):
    """
    Build per-invoice rows for this payment:
      - invoice basic info
      - total_due (as Decimal)
      - amount_applied (this payment)
      - previously_paid (all payments with id < this payment.id)
      - remaining_this_payment
      - outstanding_now
    """
    # ids of invoices touched by this payment
    ids = list(
        PaymentInvoice.objects.filter(payment=payment).values_list("invoice_id", flat=True)
    )
    if not ids:
        return [], Decimal("0.00"), Decimal("0.00"), Decimal("0.00")

    # how much each of those invoices got from THIS payment
    applied_map = {
        row["invoice_id"]: row["applied"]
        for row in PaymentInvoice.objects.filter(payment=payment)
        .values("invoice_id")
        .annotate(applied=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))
    }

    # how much each invoice had before this payment (use id ordering as a stable proxy)
    prev_paid_map = {
        row["invoice_id"]: row["paid_before"]
        for row in PaymentInvoice.objects.filter(
            invoice_id__in=ids,
            payment__id__lt=payment.id
        )
        .values("invoice_id")
        .annotate(paid_before=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))
    }

    # pull invoices with their total_due as Decimal
    invoices = (
        Newinvoice.objects.filter(id__in=ids)
        .annotate(total_due_dec=F("total_due"))
        .select_related("customer")
        .order_by("id")
    )

    rows = []
    applied_total = Decimal("0.00")
    remaining_total = Decimal("0.00")
    outstanding_total = Decimal("0.00")

    for inv in invoices:
        total_due = Decimal(str(inv.total_due or "0"))
        applied = Decimal(str(applied_map.get(inv.id, Decimal("0.00"))))
        paid_before = Decimal(str(prev_paid_map.get(inv.id, Decimal("0.00"))))

        remaining_this_payment = max(total_due - paid_before - applied, Decimal("0.00"))
        outstanding_now = max(total_due - (paid_before + applied), Decimal("0.00"))

        rows.append({
            "invoice": inv,
            "date_created": inv.date_created,
            "total_due": total_due,
            "amount_applied": applied,
            "remaining_this_payment": remaining_this_payment,
            "outstanding_now": outstanding_now,
        })

        applied_total += applied
        remaining_total += remaining_this_payment
        outstanding_total += outstanding_now

    return rows, applied_total, remaining_total, outstanding_total


def payment_print(request, pk: int):
    """
    Printable Payment Receipt.
    """
    payment = get_object_or_404(
        Payment.objects.select_related("customer", "deposit_to"),
        pk=pk
    )
    lines, applied_total, remaining_total, outstanding_total = _lines_for_payment(payment)

    # company / branding (replace with your own source if you store company profile elsewhere)
    company = {
        "name": "YoAccountant",
        "address": "Kampala, Uganda",
        "phone": "+256 000 000 000",
        "email": "info@yoaccountant.com",
        "logo_url": request.build_absolute_uri(static("sowaf/images/yo-logo.png")),
    }

    ctx = {
        "payment": payment,
        "lines": lines,
        "applied_total": applied_total,
        "remaining_total": remaining_total,
        "outstanding_total": outstanding_total,
        "company": company,
    }
    return render(request, "payment_print.html", ctx)
# end

# working on the receipt
VAT_RATE = Decimal("0.18")


@transaction.atomic
def sales_receipt_new(request):
    customers = Newcustomer.objects.order_by("customer_name")
    accounts  = deposit_accounts_qs()
    products  = Product.objects.all()

    if request.method == "POST":
        customer_id    = (request.POST.get("customer") or "").strip()
        receipt_date   = parse_date(request.POST.get("receipt_date") or "")
        payment_method = (request.POST.get("payment_method") or "cash").strip()
        deposit_to_id  = (request.POST.get("deposit_to") or "").strip()
        reference_no   = (request.POST.get("reference_no") or "").strip() or generate_unique_ref_no()
        tags           = (request.POST.get("tags") or "").strip()
        memo           = (request.POST.get("memo") or "").strip()

        if not (customer_id.isdigit() and receipt_date and deposit_to_id.isdigit()):
            return render(request, "receipt_form.html", {
                "customers": customers, "accounts": accounts, "products": products,
                "reference_no": reference_no,
                "form_error": "Please select customer, date and a deposit account.",
            })

        customer   = get_object_or_404(Newcustomer, pk=int(customer_id))
        deposit_to = get_object_or_404(accounts, pk=int(deposit_to_id))

        subtotal        = _coerce_decimal(request.POST.get("subtotal"))
        discount_amount = _coerce_decimal(request.POST.get("discount_amount"))
        shipping_fee    = _coerce_decimal(request.POST.get("shipping_fee"))
        total_amount    = _coerce_decimal(request.POST.get("total_amount"))
        amount_paid     = _coerce_decimal(request.POST.get("amount_paid"))

        # ✅ balance: allow overpayment (balance becomes 0; excess handled in posting)
        balance = total_amount - amount_paid
        if balance < 0:
            balance = Decimal("0.00")

        # ensure 8-digit numeric ref
        if not (len(reference_no) == 8 and reference_no.isdigit()):
            reference_no = generate_unique_ref_no()

        if Payment.objects.filter(reference_no=reference_no).exists() or \
           SalesReceipt.objects.filter(reference_no=reference_no).exists():
            reference_no = generate_unique_ref_no()

        receipt = SalesReceipt.objects.create(
            customer=customer,
            receipt_date=receipt_date,
            payment_method=payment_method,
            deposit_to=deposit_to,
            reference_no=reference_no,
            tags=tags,
            memo=memo,
            subtotal=subtotal,
            total_discount=discount_amount,
            total_vat=Decimal("0.00"),   # set after lines
            shipping_fee=shipping_fee,
            total_amount=total_amount,
            amount_paid=amount_paid,
            balance=balance,
        )
        apply_audit_fields(receipt)
        receipt.save()

        products_ids = request.POST.getlist("product[]")
        descriptions = request.POST.getlist("description[]")
        qtys         = request.POST.getlist("qty[]")
        unit_prices  = request.POST.getlist("unit_price[]")
        line_totals  = request.POST.getlist("line_total[]")

        bulk = []
        total_vat = Decimal("0.00")

        row_count = max(len(descriptions), len(qtys), len(unit_prices), len(line_totals), len(products_ids))
        for i in range(row_count):
            prod_id = products_ids[i] if i < len(products_ids) else None
            product = Product.objects.filter(pk=prod_id).first() if (prod_id and str(prod_id).isdigit()) else None

            desc = descriptions[i] if i < len(descriptions) else ""
            qty  = _coerce_decimal(qtys[i] if i < len(qtys) else "0")
            rate = _coerce_decimal(unit_prices[i] if i < len(unit_prices) else "0")
            amt  = _coerce_decimal(line_totals[i] if i < len(line_totals) else "0")

            if not (product or desc or (qty > 0) or (rate > 0) or (amt > 0)):
                continue

            # ✅ VAT per line (18% if taxable)
            line_vat = Decimal("0.00")
            if product and getattr(product, "taxable", False) and amt > 0:
                line_vat = (amt * VAT_RATE).quantize(Decimal("0.01"))

            total_vat += line_vat

            bulk.append(SalesReceiptLine(
                receipt=receipt,
                product=product,
                description=desc,
                qty=qty,
                unit_price=rate,
                amount=amt,
                discount_pct=Decimal("0.00"),
                discount_amt=Decimal("0.00"),
                vat_amt=line_vat,
            ))

        if bulk:
            SalesReceiptLine.objects.bulk_create(bulk)

        # ✅ sync header VAT
        receipt.total_vat = total_vat
        receipt.save(update_fields=["total_vat"])

        # ✅ Post to GL (includes overpayment + inventory)
        _post_sales_receipt_to_ledger(receipt)

        action = request.POST.get("save_action")
        if action == "save":
            return redirect("sales:sales-receipt-list")
        if action == "save&new":
            return redirect("sales:sales-receipt-new")
        if action == "save&close":
            return redirect("sales:sales-receipt-list")
        return redirect("sales:receipt-detail", pk=receipt.pk)

    reference_no = generate_unique_ref_no()
    return render(request, "receipt_form.html", {
        "customers": customers,
        "accounts": accounts,
        "products": products,
        "reference_no": reference_no,
    })
# Edit view

VAT_RATE = Decimal("0.18")

@transaction.atomic
def sales_receipt_edit(request, pk: int):
    receipt   = get_object_or_404(SalesReceipt.objects.select_related("customer", "deposit_to"), pk=pk)
    customers = Newcustomer.objects.order_by("customer_name")
    accounts  = deposit_accounts_qs()
    products  = Product.objects.all()

    if request.method == "POST":
        customer_id    = (request.POST.get("customer") or "").strip()
        receipt_date   = parse_date(request.POST.get("receipt_date") or "")
        payment_method = (request.POST.get("payment_method") or "cash").strip()
        deposit_to_id  = (request.POST.get("deposit_to") or "").strip()
        reference_no   = (request.POST.get("reference_no") or "").strip() or receipt.reference_no
        tags           = (request.POST.get("tags") or "").strip()
        memo           = (request.POST.get("memo") or "").strip()

        errors = []
        if not (customer_id.isdigit()):
            errors.append("customer")
        if not receipt_date:
            errors.append("date")
        if not (deposit_to_id.isdigit()):
            errors.append("deposit account")

        if errors:
            return render(request, "receipt_form.html", {
                "customers": customers, "accounts": accounts, "products": products,
                "edit_mode": True, "receipt": receipt, "items": receipt.lines.all(),
                "reference_no": reference_no,
                "form_error": "Please select: " + ", ".join(errors) + ".",
            })

        receipt.customer       = get_object_or_404(Newcustomer, pk=int(customer_id))
        receipt.receipt_date   = receipt_date
        receipt.payment_method = payment_method
        receipt.deposit_to     = get_object_or_404(accounts, pk=int(deposit_to_id))
        receipt.reference_no   = reference_no
        receipt.tags           = tags
        receipt.memo           = memo

        receipt.amount_paid    = _coerce_decimal(request.POST.get("amount_paid"))
        receipt.subtotal       = _coerce_decimal(request.POST.get("subtotal"))
        receipt.total_discount = _coerce_decimal(request.POST.get("discount_amount"))
        receipt.shipping_fee   = _coerce_decimal(request.POST.get("shipping_fee"))
        receipt.total_amount   = _coerce_decimal(request.POST.get("total_amount"))

        # allow overpayment; balance cannot be negative
        receipt.balance = receipt.total_amount - receipt.amount_paid
        if receipt.balance < 0:
            receipt.balance = Decimal("0.00")

        receipt.total_vat = Decimal("0.00")  # set after lines
        apply_audit_fields(receipt)
        receipt.save()

        # replace lines
        SalesReceiptLine.objects.filter(receipt=receipt).delete()

        products_ids = request.POST.getlist("product[]")
        descriptions = request.POST.getlist("description[]")
        qtys         = request.POST.getlist("qty[]")
        unit_prices  = request.POST.getlist("unit_price[]")
        line_totals  = request.POST.getlist("line_total[]")

        bulk = []
        total_vat = Decimal("0.00")

        n = max(len(descriptions), len(products_ids), len(qtys), len(unit_prices), len(line_totals))
        for i in range(n):
            prod_id = products_ids[i] if i < len(products_ids) else None
            product = Product.objects.filter(pk=prod_id).first() if (prod_id and str(prod_id).isdigit()) else None

            desc  = descriptions[i] if i < len(descriptions) else ""
            qty   = _coerce_decimal(qtys[i] if i < len(qtys) else "0")
            price = _coerce_decimal(unit_prices[i] if i < len(unit_prices) else "0")
            amt   = _coerce_decimal(line_totals[i] if i < len(line_totals) else "0")

            if not product and not desc and qty == 0 and price == 0 and amt == 0:
                continue

            line_vat = Decimal("0.00")
            if product and getattr(product, "taxable", False) and amt > 0:
                line_vat = (amt * VAT_RATE).quantize(Decimal("0.01"))

            total_vat += line_vat

            bulk.append(SalesReceiptLine(
                receipt=receipt,
                product=product,
                description=desc,
                qty=qty,
                unit_price=price,
                amount=amt,
                discount_pct=Decimal("0.00"),
                discount_amt=Decimal("0.00"),
                vat_amt=line_vat,
            ))

        if bulk:
            SalesReceiptLine.objects.bulk_create(bulk)

        receipt.total_vat = total_vat
        receipt.save(update_fields=["total_vat"])

        # repost GL (and reflect stock movement again)
        _post_sales_receipt_to_ledger(receipt)

        return redirect("sales:receipt-detail", pk=receipt.pk)

    return render(request, "receipt_form.html", {
        "customers": customers,
        "accounts": accounts,
        "products": products,
        "edit_mode": True,
        "receipt": receipt,
        "items": receipt.lines.all(),
        "reference_no": receipt.reference_no or generate_unique_ref_no(),
    })


# sales receipt detail page
def sales_receipt_detail(request, pk: int):
    receipt = get_object_or_404(SalesReceipt.objects.select_related("customer", "deposit_to"), pk=pk)
    lines = receipt.lines.select_related("product").all()

    return render(request, "receipt_detail.html", {
        "receipt": receipt,
        "lines": lines,
    })

# receipt lists and printout

def _is_bankish(acc) -> bool:
    if not acc:
        return False
    at = (acc.account_type or "").lower()
    dt = (acc.detail_type or "").lower()
    return (
        at in ("bank", "cash and cash equivalents", "cash_equiv", "cash & cash equivalents")
        or "bank" in dt
    )


def _receipt_status(r: SalesReceipt) -> str:
    """
    Simple, readable status like we did for invoices/payments:
    - Deposited (if fully paid & deposited to a bankish account)
    - Paid (if balance 0 but account not bankish)
    - <balance> due (if balance > 0)
    """
    total = r.total_amount or Decimal("0")
    paid  = getattr(r, "amount_paid", Decimal("0"))
    bal   = getattr(r, "balance", (total - paid))
    if total == 0:
        return "No amount"
    if bal <= 0:
        return "Deposited" if _is_bankish(r.deposit_to) else "Paid"
    return f"{bal:,.0f} due"


def sales_receipt_list(request):
    """
    Receipts table with customer, date, deposit_to, method, totals,
    plus Actions (Edit | View | Print).
    """
    qs = (
        SalesReceipt.objects
        .select_related("customer", "deposit_to")
        .annotate(
            total_amount_dec=Cast("total_amount", DecimalField(max_digits=18, decimal_places=2)),
            amount_paid_dec=Cast(Coalesce(F("amount_paid"), Value(Decimal("0.00"))), DecimalField(max_digits=18, decimal_places=2)),
        )
        .order_by("-receipt_date", "-id")
    )

    rows = []
    for r in qs:
        total   = r.total_amount_dec or Decimal("0")
        paid    = r.amount_paid_dec or Decimal("0")
        balance = getattr(r, "balance", (total - paid))
        if balance is None:
            balance = total - paid
        if balance < 0:
            balance = Decimal("0")

        rows.append({
            "r": r,
            "total": total,
            "paid": paid,
            "balance": balance,
            "status": _receipt_status(r),
        })

    return render(request, "receipt_list.html", {"rows": rows})


def receipt_print(request, pk: int):
    receipt = get_object_or_404(
        SalesReceipt.objects.select_related("customer", "deposit_to"), pk=pk
    )
    lines = receipt.lines.select_related("product").all()

    context = {
        "receipt": receipt,
        "lines": lines,
        # header info (use your real settings if you have them)
        "logo_url": request.build_absolute_url(static("sowaf/images/yo-logo.png")),
        "company_name": "YoAccountant",
        "company_address": "Kampala, Uganda",
        "company_phone": "+256 700 000 000",
        "company_email": "support@yoaccountant.com",
    }
    return render(request, "receipt_print.html", context)
# end


# working on the statements

def _customer_opening_balance(customer_id, start_date):
    """
    Opening balance = (all invoice totals before start) - (all credits before start).
    Credits = payments applied to those invoices + sales receipts amounts.
    """
    inv_total = (
        Newinvoice.objects
        .filter(customer_id=customer_id, date_created__lt=start_date)
        .aggregate(total=Coalesce(Sum(Cast("total_due", DecimalField(max_digits=18, decimal_places=2))),
                                  Value(Decimal("0.00"))))["total"]
        or Decimal("0.00")
    )

    paid_total = (
        PaymentInvoice.objects
        .filter(invoice__customer_id=customer_id, payment__payment_date__lt=start_date)
        .aggregate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))["total"]
        or Decimal("0.00")
    )

    # Treat Sales Receipts as immediate credits to A/R
    receipts_total = (
        SalesReceipt.objects
        .filter(customer_id=customer_id, receipt_date__lt=start_date)
        .aggregate(total=Coalesce(Sum("amount_paid"), Value(Decimal("0.00"))))["total"]
        or Decimal("0.00")
    )

    return _dec(inv_total) - _dec(paid_total) - _dec(receipts_total)


def _period_rows(customer_id, start_date, end_date):
    """
    Build period activity rows across invoices, payments (applied), and sales receipts.
    Amount sign convention: +invoice total, -payment amount, -receipt amount.
    """
    rows = []

    # Invoices in range
    inv_qs = (
        Newinvoice.objects
        .filter(customer_id=customer_id, date_created__gte=start_date, date_created__lte=end_date)
        .annotate(total_due_dec=Cast("total_due", DecimalField(max_digits=18, decimal_places=2)))
        .order_by("date_created", "id")
    )
    for inv in inv_qs:
        rows.append({
            "date": inv.date_created,
            "kind": "invoice",
            "ref": f"INV-{inv.id:04d}",
            "memo": (inv.memo or "")[:180] if hasattr(inv, "memo") else "",
            "amount": _dec(inv.total_due_dec),
            "source_type": "invoice",
            "source_id": inv.id,
        })

    # Payments (use applied part only) in range
    pay_lines = (
        PaymentInvoice.objects
        .filter(invoice__customer_id=customer_id, payment__payment_date__gte=start_date, payment__payment_date__lte=end_date)
        .select_related("payment")
        .order_by("payment__payment_date", "id")
    )
    for pli in pay_lines:
        p = pli.payment
        rows.append({
            "date": p.payment_date,
            "kind": "payment",
            "ref": p.reference_no or f"PAY-{p.id:04d}",
            "memo": (p.memo or "")[:180],
            "amount": -_dec(pli.amount_paid),
            "source_type": "payment",
            "source_id": p.id,
        })

    # Sales Receipts in range (reduce A/R)
    rec_qs = (
        SalesReceipt.objects
        .filter(customer_id=customer_id, receipt_date__gte=start_date, receipt_date__lte=end_date)
        .order_by("receipt_date", "id")
    )
    for r in rec_qs:
        rows.append({
            "date": r.receipt_date,
            "kind": "sales_receipt",
            "ref": r.reference_no or f"RCPT-{r.id:04d}",
            "memo": (r.memo or "")[:180],
            "amount": -_dec(r.amount_paid),
            "source_type": "sales_receipt",
            "source_id": r.id,
        })

    rows.sort(key=lambda x: (x["date"], x["source_type"], x["source_id"]))
    return rows


def _filter_by_type(rows, statement_type, customer_id, start_date):
    """
    Adapts the period rows for the selected statement type.
    """
    if statement_type == Statement.StatementType.OPEN_ITEM:
        # Only invoices that still have balance as of today
        today = timezone.now().date()
        inv_open = (
            Newinvoice.objects
            .filter(customer_id=customer_id)
            .annotate(
                total_due_dec=Cast("total_due", DecimalField(max_digits=18, decimal_places=2)),
                total_paid=Coalesce(Sum("payments_applied__amount_paid"), Value(Decimal("0.00"))),
            )
            .annotate(outstanding=F("total_due_dec") - F("total_paid"))
            .filter(outstanding__gt=0)
        )
        ids = set(inv_open.values_list("id", flat=True))
        return [r for r in rows if r["kind"] == "invoice" and r["source_id"] in ids]

    return rows


def _to_date(d):
    """Safely convert string/date/datetime into date."""
    if d is None:
        return None
    if hasattr(d, "date") and not isinstance(d, str):
        # datetime -> date OR date stays date
        try:
            return d.date()
        except Exception:
            return d
    if isinstance(d, str):
        try:
            return timezone.datetime.fromisoformat(d).date()
        except Exception:
            return None
    return d


def _customer_ar_balance_as_of(customer_id: int, as_of_date):
    """
    LIVE A/R balance (Customer Subledger A/R) up to and including as_of_date.
    A/R is Asset => debit - credit
    """
    as_of_date = _to_date(as_of_date)
    cust = Newcustomer.objects.only("customer_name").get(pk=customer_id)

    qs = JournalLine.objects.filter(
        account__detail_type="Customer Subledger (A/R)",
        account__account_name=cust.customer_name,
    )

    # If your JournalEntry date field is "date" (as in your COA code: entry__date)
    if as_of_date:
        qs = qs.filter(entry__date__lte=as_of_date)

    totals = qs.aggregate(
        deb=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
        cred=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
    )

    return (totals["deb"] or Decimal("0.00")) - (totals["cred"] or Decimal("0.00"))


def _customer_opening_balance_live(customer_id: int, statement_start_date):
    """
    LIVE opening = A/R balance as of day BEFORE statement_start_date.
    """
    sd = _to_date(statement_start_date)
    if not sd:
        return Decimal("0.00")
    day_before = sd - timezone.timedelta(days=1)
    return _customer_ar_balance_as_of(customer_id, day_before)


@require_http_methods(["GET", "POST"])
def statement_new(request):
    customer_id = request.GET.get("customer_id") or request.POST.get("customer_id")
    customer = get_object_or_404(Newcustomer, pk=int(customer_id)) if customer_id else None

    today = timezone.now().date()
    default_start = today - timedelta(days=30)
    default_end = today

    statement_type = (
        request.GET.get("type")
        or request.POST.get("statement_type")
        or Statement.StatementType.TRANSACTION
    )
    statement_date = request.POST.get("statement_date") or today.isoformat()
    start_date = request.POST.get("start_date") or default_start.isoformat()
    end_date = request.POST.get("end_date") or default_end.isoformat()
    email_to = request.POST.get("email_to") or (customer.email if customer else "")

    rows = []
    opening_balance = Decimal("0.00")

    if customer:
        sd = timezone.datetime.fromisoformat(start_date).date()
        ed = timezone.datetime.fromisoformat(end_date).date()

        opening_balance = _customer_opening_balance_live(customer.id, sd)

        rows = _period_rows(customer.id, sd, ed)
        rows = _filter_by_type(rows, statement_type, customer.id, sd)

    preview_lines = []
    run = opening_balance

    # ✅ Transaction + Balance Forward behave as you had
    if statement_type in (Statement.StatementType.TRANSACTION, Statement.StatementType.BAL_FWD):
        preview_lines.append({
            "date": start_date,
            "kind": "opening_balance" if statement_type == Statement.StatementType.TRANSACTION else "balance_forward",
            "ref": "",
            "memo": "Opening Balance" if statement_type == Statement.StatementType.TRANSACTION else "Balance Forward",
            "amount": Decimal("0.00") if statement_type == Statement.StatementType.TRANSACTION else opening_balance,
        })
        if statement_type == Statement.StatementType.BAL_FWD:
            run += opening_balance

    # ✅ FIX: OPEN ITEM should not produce an empty statement when opening balance exists
    elif statement_type == Statement.StatementType.OPEN_ITEM:
        # Add Balance Forward line for context (QuickBooks-like)
        preview_lines.append({
            "date": start_date,
            "kind": "balance_forward",
            "ref": "",
            "memo": "Balance Forward",
            "amount": opening_balance,
            "running_balance": opening_balance,
            "source_type": "",
            "source_id": None,
        })
        run = opening_balance

    # Add period lines
    for r in rows:
        amt = r["amount"]
        if statement_type != Statement.StatementType.TRANSACTION:
            run += amt
        preview_lines.append({**r, "running_balance": run})

    closing_balance = (
        run if statement_type != Statement.StatementType.TRANSACTION
        else opening_balance + sum((r["amount"] for r in rows), Decimal("0"))
    )

    if request.method == "POST":
        if not customer:
            return redirect(request.path)

        st = Statement.objects.create(
            customer=customer,
            statement_date=statement_date,
            start_date=start_date,
            end_date=end_date,
            statement_type=statement_type,
            email_to=email_to or None,
            opening_balance=opening_balance,
            closing_balance=closing_balance,
            memo=(request.POST.get("memo") or "").strip(),
        )

        # ✅ Persist from preview_lines (so OPEN_ITEM gets a line too)
        run_save = opening_balance

        for pl in preview_lines:
            kind_map = {
                "opening_balance": StatementLine.LineKind.OPENING,
                "balance_forward": StatementLine.LineKind.BAL_FWD,
                "invoice": StatementLine.LineKind.INVOICE,
                "payment": StatementLine.LineKind.PAYMENT,
                "sales_receipt": StatementLine.LineKind.SALES_RECEIPT,
            }

            k = pl.get("kind")
            kind_val = kind_map.get(k, StatementLine.LineKind.OPENING)

            # For transaction statements you originally saved None running_balance for detail lines
            # Keep that behavior:
            if statement_type == Statement.StatementType.TRANSACTION and k in ("invoice", "payment", "sales_receipt"):
                rb = None
            else:
                rb = pl.get("running_balance", run_save)

            StatementLine.objects.create(
                statement=st,
                date=pl["date"],
                kind=kind_val,
                ref_no=pl.get("ref", "") or "",
                memo=pl.get("memo", "") or "",
                amount=pl.get("amount", Decimal("0.00")) or Decimal("0.00"),
                running_balance=rb,
                source_type=pl.get("source_type", "") or "",
                source_id=pl.get("source_id", None),
            )

            # keep run_save aligned for non-transaction formats
            if statement_type != Statement.StatementType.TRANSACTION:
                run_save = rb if rb is not None else run_save

        return redirect("sales:statement-detail", pk=st.pk)

    return render(request, "statement_form.html", {
        "customer": customer,
        "statement_type": statement_type,
        "statement_date": statement_date,
        "start_date": start_date,
        "end_date": end_date,
        "email_to": email_to,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "preview_lines": preview_lines,
    })

def _build_statement_rows(st):
    st_no = f"ST-000{st.id}"
    st_type = st.get_statement_type_display()

    sd = _to_date(st.start_date)
    ed = _to_date(st.end_date)

    live_opening = _customer_opening_balance_live(st.customer_id, sd)
    live_closing = _customer_ar_balance_as_of(st.customer_id, ed)

    rows = []
    for ln in st.lines.all().order_by("date", "id"):
        rows.append({
            "date": ln.date,
            "type": st_type,     
            "no": st_no,            
            "memo": ln.memo or "",
            "running": live_closing 
        })

    return rows, live_opening, live_closing


def statement_detail(request, pk):
    st = get_object_or_404(
        Statement.objects.select_related("customer").prefetch_related("lines"),
        pk=pk
    )

    sd = st.start_date
    ed = st.end_date

    # LIVE balances
    live_opening = _customer_opening_balance_live(st.customer_id, sd)
    live_closing = _customer_ar_balance_as_of(st.customer_id, ed)

    # Build rows with LIVE running (starting from opening)
    running = live_opening
    lines_live = []

    for ln in st.lines.all().order_by("date", "id"):
        # add/subtract based on your stored statement line amounts
        # (Invoice positive, Payment/Sales receipt negative if you saved them that way)
        running += (ln.amount or Decimal("0.00"))

        lines_live.append({
            "date": ln.date,
            "memo": ln.memo,
            "running": running,
        })

    return render(request, "statement_detail.html", {
        "st": st,
        "lines": lines_live,       # ✅ template will use this
        "live_opening": live_opening,
        "live_closing": live_closing,
    })

# ----- Excel export (openpyxl) -----
def statement_export_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    st = get_object_or_404(Statement.objects.select_related("customer"), pk=pk)

    rows, live_opening, live_closing = _build_statement_rows(st)

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    ws["A1"] = "Customer"
    ws["B1"] = st.customer.customer_name
    ws["A2"] = "Statement Date"
    ws["B2"] = str(st.statement_date)
    ws["A3"] = "Period"
    ws["B3"] = f"{st.start_date} — {st.end_date}"
    ws["A4"] = "Type"
    ws["B4"] = st.get_statement_type_display()
    ws["A5"] = "Opening Balance (LIVE)"
    ws["B5"] = float(live_opening)
    ws["A6"] = "Closing Balance (LIVE)"
    ws["B6"] = float(live_closing)

    start = 8
    headers = ["Date", "Type", "No.", "Memo", "Running Balance"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)

    r = start + 1
    for ln in rows:
        ws.cell(row=r, column=1, value=str(ln["date"]))
        ws.cell(row=r, column=2, value=ln["type"])
        ws.cell(row=r, column=3, value=ln["no"])
        ws.cell(row=r, column=4, value=ln["memo"])
        ws.cell(row=r, column=5, value=float(ln["running"]))
        r += 1

    widths = [14, 22, 12, 44, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="Statement_{st.id}.xlsx"'
    wb.save(resp)
    return resp

# ----- PDF export (WeasyPrint) -----
def statement_export_pdf(request, pk: int):
    st = get_object_or_404(Statement, pk=pk)

    lines = (
        StatementLine.objects
        .filter(statement=st)
        .order_by("date", "id")
    )

    sd = st.start_date if not isinstance(st.start_date, str) else timezone.datetime.fromisoformat(st.start_date).date()
    ed = st.end_date if not isinstance(st.end_date, str) else timezone.datetime.fromisoformat(st.end_date).date()

    live_opening = _customer_opening_balance_live(st.customer_id, sd)
    live_closing = _customer_ar_balance_as_of(st.customer_id, ed)

    context = {
        "statement": st,
        "customer": getattr(st, "customer", None),
        "lines": lines,
        "generated_at": timezone.now(),
        "BASE_URL": request.build_absolute_uri("/"),
        "live_opening": live_opening,
        "live_closing": live_closing,
    }

    html = render_to_string("statement_pdf.html", context)

    try:
        from weasyprint import HTML, CSS
        pdf_bytes = HTML(
            string=html, base_url=request.build_absolute_uri("/")
        ).write_pdf(stylesheets=[CSS(string="""
            @page { size: A4; margin: 18mm; }
            body { font-family: Arial, Segoe UI, Roboto, sans-serif; font-size: 12px; color: #0b1220; }
            h1 { font-size: 18px; margin: 0 0 6px; }
            .muted { color:#64748b; }
            table { width:100%; border-collapse: collapse; margin-top: 12px; }
            th, td { border:1px solid #e6e9ee; padding: 6px 8px; }
            thead th { background:#f8fdfa; text-align:left; }
            tfoot td { font-weight: 700; }
            .num { text-align: right; }
        """)])
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    except Exception:
        from xhtml2pdf import pisa
        resp = HttpResponse(content_type="application/pdf")
        pisa.CreatePDF(html, dest=resp, link_callback=lambda uri, rel: uri)

    fname = f"Statement_{st.customer_id}_{st.id}.pdf"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

@login_required
def customer_credits_list(request):
    customers = Newcustomer.objects.order_by("customer_name")
    rows = []
    for c in customers:
        bal = _customer_credit_balance(c)
        if bal > 0:
            rows.append({"customer": c, "credit": bal})

    return render(request, "customer_credits_list.html", {"rows": rows})


@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def customer_refund_new(request, customer_id: int):
    customer = get_object_or_404(Newcustomer, pk=customer_id)
    accounts = deposit_accounts_qs()

    max_refundable = _customer_credit_balance(customer)

    if request.method == "POST":
        refund_date = parse_date(request.POST.get("refund_date") or "") or timezone.localdate()
        paid_from_id = (request.POST.get("paid_from") or "").strip()
        amount = _dec(request.POST.get("amount") or "0.00")
        memo = (request.POST.get("memo") or "").strip()
        reference_no = (request.POST.get("reference_no") or "").strip()

        paid_from = None
        if paid_from_id.isdigit():
            paid_from = accounts.filter(id=int(paid_from_id)).first()

        if not paid_from:
            return render(request, "customer_refund_form.html", {
                "customer": customer, "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": "Select a valid Bank/Cash account."
            })

        if amount <= 0:
            return render(request, "customer_refund_form.html", {
                "customer": customer, "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": "Refund amount must be > 0."
            })

        if amount > max_refundable:
            return render(request, "customer_refund_form.html", {
                "customer": customer, "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": f"Refund exceeds available customer credit ({max_refundable})."
            })

        refund = CustomerRefund.objects.create(
            customer=customer,
            refund_date=refund_date,
            paid_from=paid_from,
            amount=amount,
            memo=memo,
            reference_no=reference_no,
        )

        _post_customer_refund_to_ledger(refund)
        return redirect("sales:customer-credits-list")

    return render(request, "customer_refund_form.html", {
        "customer": customer,
        "accounts": accounts,
        "max_refundable": max_refundable,
    })


# =========================================================
# 1) SALES RECEIPT LIST REPORT
# =========================================================
def sales_receipt_list_report(request):
    today = timezone.localdate()

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = SalesReceipt.objects.select_related("customer").all().order_by("-receipt_date", "-id")

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(receipt_date__gte=date_from)
    if date_to:
        qs = qs.filter(receipt_date__lte=date_to)

    rows = []
    total_amount = Decimal("0.00")

    for r in qs:
        total_amount += _dec(r.total_amount)
        rows.append({
            "id": r.id,
            "customer": r.customer,
            "receipt_date": r.receipt_date,
            "payment_method": r.payment_method,
            "reference_no": r.reference_no,
            "deposit_to": r.deposit_to,
            "subtotal": _dec(r.subtotal),
            "discount": _dec(r.total_discount),
            "vat": _dec(r.total_vat),
            "shipping": _dec(r.shipping_fee),
            "total": _dec(r.total_amount),
        })

    customers = Newcustomer.objects.order_by("customer_name")

    return render(request, "sales_receipt_list_report.html", {
        "today": today,
        "rows": rows,
        "customers": customers,
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "date_from": date_from,
        "date_to": date_to,
        "total_amount": total_amount,
    })


def sales_receipt_list_export(request, fmt: str):
    """
    fmt = excel | pdf
    """
    fmt = (fmt or "").lower()
    today = timezone.localdate()

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = SalesReceipt.objects.select_related("customer").all().order_by("-receipt_date", "-id")
    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(receipt_date__gte=date_from)
    if date_to:
        qs = qs.filter(receipt_date__lte=date_to)

    rows = []
    for r in qs:
        rows.append([
            str(r.customer.customer_name),
            str(r.receipt_date),
            str(r.payment_method),
            str(r.reference_no or ""),
            str(getattr(r.deposit_to, "account_name", getattr(r.deposit_to, "name", ""))),
            float(_dec(r.subtotal)),
            float(_dec(r.total_discount)),
            float(_dec(r.total_vat)),
            float(_dec(r.shipping_fee)),
            float(_dec(r.total_amount)),
        ])

    headers = [
        "Customer", "Receipt Date", "Method", "Reference", "Deposit To",
        "Subtotal", "Discount", "VAT", "Shipping", "Total"
    ]

    if fmt == "excel":
        # CSV works perfectly with Excel and needs no extra libs
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="sales_receipts_{today}.csv"'
        import csv
        w = csv.writer(response)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
        return response

    if fmt == "pdf":
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)

        p.setFont("Helvetica-Bold", 14)
        p.drawString(30, height - 30, f"Sales Receipt List (As of {today})")

        y = height - 60
        p.setFont("Helvetica-Bold", 9)
        x_positions = [30, 150, 240, 320, 420, 520, 590, 660, 720, 800]

        for i, h in enumerate(headers):
            p.drawString(x_positions[i], y, h)

        p.setFont("Helvetica", 9)
        y -= 18

        for row in rows:
            if y < 40:
                p.showPage()
                y = height - 40
                p.setFont("Helvetica", 9)

            for i, cell in enumerate(row):
                p.drawString(x_positions[i], y, str(cell)[:25])
            y -= 14

        p.showPage()
        p.save()

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="sales_receipts_{today}.pdf"'
        return response

    return HttpResponse("Invalid format", status=400)


# =========================================================
# 2) CUSTOMER STATEMENTS REPORT (list + export)
# =========================================================
def customer_statements_report(request):
    today = timezone.localdate()

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = Statement.objects.select_related("customer").all().order_by("-statement_date", "-id")

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(statement_date__gte=date_from)
    if date_to:
        qs = qs.filter(statement_date__lte=date_to)

    rows = []
    for s in qs:
        rows.append({
            "id": s.id,
            "customer": s.customer,
            "statement_date": s.statement_date,
            "start_date": s.start_date,
            "end_date": s.end_date,
            "statement_type": s.statement_type,
            "opening_balance": _dec(s.opening_balance),
            "closing_balance": _dec(s.closing_balance),
            "email_to": s.email_to,
        })

    customers = Newcustomer.objects.order_by("customer_name")

    return render(request, "customer_statements_report.html", {
        "today": today,
        "rows": rows,
        "customers": customers,
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "date_from": date_from,
        "date_to": date_to,
    })


def customer_statements_export(request, fmt: str):
    fmt = (fmt or "").lower()
    today = timezone.localdate()

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = Statement.objects.select_related("customer").all().order_by("-statement_date", "-id")
    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(statement_date__gte=date_from)
    if date_to:
        qs = qs.filter(statement_date__lte=date_to)

    headers = [
        "Customer", "Statement Date", "Start", "End", "Type",
        "Opening Balance", "Closing Balance", "Email"
    ]

    rows = []
    for s in qs:
        rows.append([
            str(s.customer.customer_name),
            str(s.statement_date),
            str(s.start_date),
            str(s.end_date),
            str(s.statement_type),
            float(_dec(s.opening_balance)),
            float(_dec(s.closing_balance)),
            str(s.email_to or ""),
        ])

    if fmt == "excel":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="customer_statements_{today}.csv"'
        import csv
        w = csv.writer(response)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
        return response

    if fmt == "pdf":
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)

        p.setFont("Helvetica-Bold", 14)
        p.drawString(30, height - 30, f"Customer Statements (As of {today})")

        y = height - 60
        p.setFont("Helvetica-Bold", 9)
        x_positions = [30, 180, 280, 360, 430, 520, 630, 740]

        for i, h in enumerate(headers):
            p.drawString(x_positions[i], y, h)

        p.setFont("Helvetica", 9)
        y -= 18

        for row in rows:
            if y < 40:
                p.showPage()
                y = height - 40
                p.setFont("Helvetica", 9)

            for i, cell in enumerate(row):
                p.drawString(x_positions[i], y, str(cell)[:28])
            y -= 14

        p.showPage()
        p.save()

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="customer_statements_{today}.pdf"'
        return response

    return HttpResponse("Invalid format", status=400)


# =========================================================
# PLACEHOLDERS
# =========================================================

def sales_by_customer_report(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(balance_db=F("total_due_dec") - F("total_paid"))
        .only("id", "customer_id", "date_created", "total_due")
        .order_by("customer__customer_name", "date_created", "id")
    )

    # filters
    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(date_created__date__gte=date_from)
    if date_to:
        qs = qs.filter(date_created__date__lte=date_to)

    # group manually to avoid changing your model logic
    rows_map = {}
    grand = {
        "invoiced": Decimal("0.00"),
        "paid": Decimal("0.00"),
        "balance": Decimal("0.00"),
    }

    for inv in qs:
        cust = inv.customer
        cid = cust.id

        invoiced = _dec(inv.total_due_dec)
        paid = _dec(inv.total_paid)
        bal = _dec(inv.balance_db)

        if cid not in rows_map:
            rows_map[cid] = {
                "customer": cust,
                "invoice_count": 0,
                "invoiced": Decimal("0.00"),
                "paid": Decimal("0.00"),
                "balance": Decimal("0.00"),
            }

        rows_map[cid]["invoice_count"] += 1
        rows_map[cid]["invoiced"] += invoiced
        rows_map[cid]["paid"] += paid
        rows_map[cid]["balance"] += bal

        grand["invoiced"] += invoiced
        grand["paid"] += paid
        grand["balance"] += bal

    rows = list(rows_map.values())
    rows.sort(key=lambda r: (getattr(r["customer"], "customer_name", "") or "").lower())

    customers = Newcustomer.objects.order_by("customer_name")

    return render(request, "sales_by_customer_report.html", {
        "today": today,
        "rows": rows,
        "grand": grand,
        "customers": customers,
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "date_from": date_from,
        "date_to": date_to,
    })


def sales_by_customer_export(request, fmt: str):
    fmt = (fmt or "").lower()
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(balance_db=F("total_due_dec") - F("total_paid"))
        .only("id", "customer_id", "date_created", "total_due")
        .order_by("customer__customer_name", "date_created", "id")
    )

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))
    if date_from:
        qs = qs.filter(date_created__date__gte=date_from)
    if date_to:
        qs = qs.filter(date_created__date__lte=date_to)

    rows_map = {}
    for inv in qs:
        cust = inv.customer
        cid = cust.id
        if cid not in rows_map:
            rows_map[cid] = {
                "customer": cust,
                "invoice_count": 0,
                "invoiced": Decimal("0.00"),
                "paid": Decimal("0.00"),
                "balance": Decimal("0.00"),
            }

        rows_map[cid]["invoice_count"] += 1
        rows_map[cid]["invoiced"] += _dec(inv.total_due_dec)
        rows_map[cid]["paid"] += _dec(inv.total_paid)
        rows_map[cid]["balance"] += _dec(inv.balance_db)

    rows = list(rows_map.values())
    rows.sort(key=lambda r: (getattr(r["customer"], "customer_name", "") or "").lower())

    headers = ["Customer", "Invoices", "Invoiced", "Paid", "Balance"]

    if fmt == "excel":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="sales_by_customer_{today}.csv"'
        w = csv.writer(response)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r["customer"].customer_name,
                r["invoice_count"],
                float(r["invoiced"]),
                float(r["paid"]),
                float(r["balance"]),
            ])
        return response

    if fmt == "pdf":
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)

        p.setFont("Helvetica-Bold", 14)
        p.drawString(30, height - 30, f"Sales by Customer (As of {today})")

        y = height - 60
        p.setFont("Helvetica-Bold", 10)
        x = [30, 280, 380, 500, 620]
        for i, h in enumerate(headers):
            p.drawString(x[i], y, h)

        p.setFont("Helvetica", 10)
        y -= 18

        for r in rows:
            if y < 40:
                p.showPage()
                y = height - 40
                p.setFont("Helvetica", 10)

            p.drawString(x[0], y, str(r["customer"].customer_name)[:35])
            p.drawString(x[1], y, str(r["invoice_count"]))
            p.drawString(x[2], y, f"{r['invoiced']}")
            p.drawString(x[3], y, f"{r['paid']}")
            p.drawString(x[4], y, f"{r['balance']}")
            y -= 14

        p.showPage()
        p.save()

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="sales_by_customer_{today}.pdf"'
        return response

    return HttpResponse("Invalid format", status=400)
def sales_by_product_report(request):
    return HttpResponse("Sales by Product/Service report not implemented yet.", status=501)

def sales_summary_report(request):
    return HttpResponse("Sales Summary report not implemented yet.", status=501)

def invoice_payments_report(request):
    return HttpResponse("Invoice & Payments report not implemented yet.", status=501)

# working on the recurring invoices
# @login_required
def recurring_invoice_list(request):
    recs = RecurringInvoice.objects.select_related("customer").order_by("-id")
    return render(request, "recurring_invoice_list.html", {"recs": recs})


# @login_required
@transaction.atomic
def recurring_invoice_new(request):
    """
    Create a recurring invoice template using the same style/layout.
    """
    from .models import Product, Newcustomer, Pclass  # keep same pattern as your invoice view

    if request.method == "POST":
        customer_id = request.POST.get("customer")
        customer = get_object_or_404(Newcustomer, pk=customer_id)

        class_field_id = request.POST.get("class_field")
        class_field = get_object_or_404(Pclass, pk=class_field_id)

        email = request.POST.get("email")
        billing_address = request.POST.get("billing_address")
        shipping_address = request.POST.get("shipping_address")
        terms = (request.POST.get("terms") or "").strip()
        sales_rep = request.POST.get("sales_rep")
        tags = request.POST.get("tags")
        po_num = request.POST.get("po_number")
        memo = request.POST.get("memo")
        customs_notes = request.POST.get("customs_notes")

        # schedule fields
        frequency = (request.POST.get("frequency") or "monthly").strip()
        interval = int(request.POST.get("interval") or 1)

        start_dt = request.POST.get("start_date")
        end_dt = request.POST.get("end_date")
        max_occ = request.POST.get("max_occurrences")

        start_date = timezone.localdate()
        if start_dt:
            parsed = parse_date_flexible(start_dt)
            if parsed:
                start_date = parsed

        end_date = None
        if end_dt:
            parsed = parse_date_flexible(end_dt)
            if parsed:
                end_date = parsed

        max_occurrences = int(max_occ) if (max_occ and str(max_occ).isdigit()) else None

        shipping_fee = Decimal(request.POST.get("shipping_fee") or "0")

        rec = RecurringInvoice.objects.create(
            customer=customer,
            email=email,
            billing_address=billing_address,
            shipping_address=shipping_address,
            terms=terms,
            sales_rep=sales_rep,
            class_field=class_field,
            tags=tags,
            po_num=po_num if (po_num and str(po_num).isdigit()) else None,
            memo=memo,
            customs_notes=customs_notes,
            shipping_fee=shipping_fee,
            frequency=frequency,
            interval=interval if interval > 0 else 1,
            start_date=start_date,
            next_run_date=start_date,
            end_date=end_date,
            max_occurrences=max_occurrences,
            is_active=True,
        )

        # lines
        products = request.POST.getlist("product[]")
        descriptions = request.POST.getlist("description[]")
        qtys = request.POST.getlist("qty[]")
        rates = request.POST.getlist("unit_price[]")
        discount_nums = request.POST.getlist("discount_num[]")

        rows = []
        for i in range(len(products)):
            if not products[i]:
                continue

            product = get_object_or_404(Product, pk=products[i])
            desc = descriptions[i] if i < len(descriptions) else ""
            qty = Decimal(qtys[i] or "0") if i < len(qtys) else Decimal("0")
            rate = Decimal(rates[i] or "0") if i < len(rates) else Decimal("0")
            dpc = Decimal(discount_nums[i] or "0") if i < len(discount_nums) else Decimal("0")

            rows.append(RecurringInvoiceLine(
                recurring=rec,
                product=product,
                description=desc,
                qty=qty,
                unit_price=rate,
                discount_num=dpc,
            ))

        if rows:
            RecurringInvoiceLine.objects.bulk_create(rows)

        return redirect("sales:recurring-invoices")

    products = Product.objects.all()
    customers = Newcustomer.objects.all()
    classes = Pclass.objects.all()

    return render(request, "recurring_invoice_form.html", {
        "customers": customers,
        "classes": classes,
        "products": products,
    })


# @login_required
def recurring_run_today(request):
    """
    Manual trigger (useful while you later set cron/celery).
    Generates all due recurring invoices for today.
    """
    result = generate_recurring_invoices_for_date(
        run_date=timezone.localdate(),
        apply_audit_fields=apply_audit_fields,
        _post_invoice_to_ledger=_post_invoice_to_ledger,
        as_aware_datetime=as_aware_datetime,
    )
    return JsonResponse(result)
