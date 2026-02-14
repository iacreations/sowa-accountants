import csv
from collections import OrderedDict, defaultdict
from collections import OrderedDict
from django.core.paginator import Paginator
from decimal import Decimal
import json
import io
from io import BytesIO
from openpyxl import Workbook
import csv
from reportlab.lib.pagesizes import A4
from openpyxl.styles import Font
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from datetime import datetime,date
from django.db.models import Q, Sum
from django.utils.dateparse import parse_date
from django.db.models import Sum, Value, DecimalField, F
from django.db.models.functions import Coalesce, Cast
from django.urls import reverse
from django.http import HttpResponse,Http404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from sales.models import Newinvoice,Payment
from django.db.models import ExpressionWrapper
from sowaf.models import Newcustomer
from .utils import income_accounts_qs, expense_accounts_qs, deposit_accounts_qs
from .models import (Account, ColumnPreference, JournalEntry, JournalLine, AuditTrail)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


# default visible columns (match the checkboxes in the template)
DEFAULT_ACCOUNTS_COL_PREFS = {
    "account_name": True,
    "opening_balance": True,
    "as_of": True,
    "account_number": True,
    "detail_type": True,
    "description": True,
    "actions": True,
}
# accounts/views.py
def accounts_dropdown_data(request):
    """
    Returns latest income & expense accounts for Product form dropdowns.
    """
    income = income_accounts_qs().values("id", "account_name")
    expense = expense_accounts_qs().values("id", "account_name")

    return JsonResponse({
        "income_accounts": list(income),
        "expense_accounts": list(expense),
    })

# fixed 5 Level-1 sections
LEVEL1_ORDER = [
    "Assets",
    "Equity",
    "Liabilities",
    "Income",
    "Expenses",
]

CASH_DETAIL_TYPES = [
    "Cash and Cash equivalents",
    # later you can add:
    # "Petty Cash",
    # "Bank current account",
    # "Mobile Money",
]
from django.http import JsonResponse
from django.views.decorators.http import require_GET

# import your Account model + the same queryset function you already use
# from .models import Account
# from .queries import deposit_accounts_qs

@require_GET
def deposit_accounts_api(request):
    """
    Returns deposit accounts for dropdown refresh.
    """
    qs = deposit_accounts_qs().order_by("account_name")

    data = []
    for a in qs:
        label = a.account_name or ""
        if getattr(a, "account_number", None):
            label += f" ({a.account_number})"
        # keep your extra text same as template display
        at = getattr(a, "account_type", "") or ""
        dt = getattr(a, "detail_type", "") or ""
        if at:
            label += f" — {at}"
        if dt:
            label += f" / {dt}"

        data.append({
            "id": a.id,
            "label": label,
        })

    return JsonResponse({"results": data})

# @login_required
def accounts(request):
    status = request.GET.get("status", "active")  # default is active

    base_qs = Account.objects.all()

    # status filter
    if status == "inactive":
        qs = base_qs.filter(is_active=False)
    elif status == "all":
        qs = base_qs
    else:  # "active"
        qs = base_qs.filter(is_active=True)

    qs = qs.select_related("parent").order_by("account_type", "account_name")

    # -----------------------------
    # Build CURRENT balances for COA (JOURNAL-BASED)
    # -----------------------------
    all_accounts = list(qs)
    acc_by_id = {a.id: a for a in all_accounts}

    # children map
    children_by_parent = defaultdict(list)
    for a in all_accounts:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a)

    def sort_key(a: Account):
        return (a.account_number or "", a.account_name or "")

    for pid in list(children_by_parent.keys()):
        children_by_parent[pid].sort(key=sort_key)

    def collect_subtree_ids(root_id: int):
        out = []
        stack = [root_id]
        while stack:
            nid = stack.pop()
            out.append(nid)
            for ch in children_by_parent.get(nid, []):
                stack.append(ch.id)
        return out

    def normal_side(acc: Account):
        if acc.level1_group in ["Assets", "Expenses"]:
            return "debit"
        return "credit"

    def apply_movement(acc: Account, running: Decimal, debit: Decimal, credit: Decimal):
        side = normal_side(acc)
        if side == "debit":
            return running + (debit - credit)
        return running + (credit - debit)

    # pull all journal lines for these accounts
    lines = (
        JournalLine.objects
        .filter(account_id__in=acc_by_id.keys())
        .select_related("account", "entry")
        .order_by("entry__date", "id")
    )

    lines_by_account = defaultdict(list)
    for ln in lines:
        lines_by_account[ln.account_id].append(ln)

    def rolled_up_closing(acc: Account):
        """
        Rollup = sum of each node's journal-based closing in subtree.
        """
        total = Decimal("0.00")

        for aid in collect_subtree_ids(acc.id):
            node = acc_by_id.get(aid)
            if not node:
                continue

            running = Decimal("0.00")
            for ln in lines_by_account.get(aid, []):
                debit = Decimal(str(ln.debit or "0.00"))
                credit = Decimal(str(ln.credit or "0.00"))
                running = apply_movement(node, running, debit, credit)

            total += running

        return total

    # -----------------------------
    # Attach computed balance to each account instance (NO OBE OVERRIDE)
    # -----------------------------
    for a in all_accounts:
        a.current_balance = rolled_up_closing(a)

    # group into 5 Level-1 buckets using model property
    grouped = {label: [] for label in LEVEL1_ORDER}
    for acc in all_accounts:
        level1 = acc.level1_group or "Assets"
        grouped.setdefault(level1, []).append(acc)

    level1_sections = [
        {"label": label, "accounts": grouped.get(label, [])}
        for label in LEVEL1_ORDER
    ]

    # counts for tabs
    active_count = base_qs.filter(is_active=True).count()
    inactive_count = base_qs.filter(is_active=False).count()
    all_count = base_qs.count()

    # Column preferences
    if request.user.is_authenticated:
        prefs, _ = ColumnPreference.objects.get_or_create(
            user=request.user,
            table_name="accounts",
            defaults={"preferences": DEFAULT_ACCOUNTS_COL_PREFS},
        )
        merged_prefs = {**DEFAULT_ACCOUNTS_COL_PREFS, **(prefs.preferences or {})}
    else:
        merged_prefs = DEFAULT_ACCOUNTS_COL_PREFS

    return render(
        request,
        "accounts.html",
        {
            "status": status,
            "column_prefs": merged_prefs,
            "active_count": active_count,
            "inactive_count": inactive_count,
            "all_count": all_count,
            "level1_sections": level1_sections,
            "coas": all_accounts,  #
        },
    )


# audit trail view
def audit_trail(request):
    logs = AuditTrail.objects.select_related("user").all()

    return render(request, "audit_trail.html", {
        "logs": logs
    })

# @login_required
def save_column_prefs(request):
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "detail": "POST required"}, status=400
        )

    try:
        data = json.loads(request.body or "{}")
        preferences = data.get("preferences", {})
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "detail": "Bad JSON"}, status=400)

    prefs, _ = ColumnPreference.objects.get_or_create(
        user=request.user,
        table_name="accounts",
        defaults={"preferences": DEFAULT_ACCOUNTS_COL_PREFS},
    )

    # ensure only known keys are saved
    cleaned = {
        k: bool(preferences.get(k, True)) for k in DEFAULT_ACCOUNTS_COL_PREFS.keys()
    }

    prefs.preferences = cleaned
    prefs.save()
    return JsonResponse({"status": "ok"})

# ading an account in the coa
# @login_required
@transaction.atomic
def add_account(request):
    if request.method == "POST":
        account_name   = request.POST.get("account_name")
        account_number = request.POST.get("account_number")
        account_type   = request.POST.get("account_type")  # code from select
        detail_type    = request.POST.get("detail_type")
        tax_category   = request.POST.get("tax_category")
        is_subaccount  = request.POST.get("is_subaccount") == "on"

        parent = None
        parent_id = request.POST.get("parent")
        if is_subaccount and parent_id:
            parent = Account.objects.filter(id=parent_id).first()

        opening_balance_str = request.POST.get("opening_balance") or "0"
        try:
            opening_balance = Decimal(opening_balance_str)
        except Exception:
            opening_balance = Decimal("0")

        as_of_str = request.POST.get("as_of")
        as_of = as_of_str or timezone.now().date()

        description = request.POST.get("description")

        # 1) Save the account
        new_account = Account(
            account_name=account_name,
            account_number=account_number,
            account_type=account_type,
            detail_type=detail_type,
            tax_category=tax_category,
            is_subaccount=is_subaccount,
            parent=parent,
            opening_balance=opening_balance,
            as_of=as_of,
            description=description,
        )
        new_account.save()

        # 2) Opening balance JE (same as you had)
        if opening_balance != 0:
            level1 = new_account.level1_group  # Assets / Liabilities / Equity / Income / Expenses

            if level1 in ["Assets", "Liabilities", "Equity"]:
                opening_equity_acct, _ = Account.objects.get_or_create(
                    account_name="Opening Balance Equity",
                    account_type="OWNER_EQUITY",
                    defaults={
                        "detail_type": "Opening balances",
                        "is_active": True,
                    },
                )

                je = JournalEntry.objects.create(
                    date=as_of,
                    description=f"Opening balance for {new_account.account_name}",
                    source_type="OPENING_BALANCE",
                    source_id=new_account.id,
                )

                amount = abs(opening_balance)

                if level1 == "Assets":
                    # Debit the asset account, Credit Opening Balance Equity
                    JournalLine.objects.create(
                        entry=je,
                        account=new_account,
                        debit=amount,
                        credit=Decimal("0"),
                    )
                    JournalLine.objects.create(
                        entry=je,
                        account=opening_equity_acct,
                        debit=Decimal("0"),
                        credit=amount,
                    )
                else:
                    # Liabilities & Equity: Credit account, Debit Opening Balance Equity
                    JournalLine.objects.create(
                        entry=je,
                        account=new_account,
                        debit=Decimal("0"),
                        credit=amount,
                    )
                    JournalLine.objects.create(
                        entry=je,
                        account=opening_equity_acct,
                        debit=amount,
                        credit=Decimal("0"),
                    )

        # 3) Redirect as before
        save_action = request.POST.get("save_action")
        if save_action == "save&new":
            return redirect("accounts:add-account")
        elif save_action == "save&close":
            return redirect("accounts:accounts")

        return redirect("accounts:accounts")

    # GET
    parents = Account.objects.all()
    return render(request, "coa_form.html", {
        "parents": parents,
        "account": None,      # important so template knows this is ADD mode
    })


@transaction.atomic
def edit_account(request, pk):
    """
    Edit an existing Account including its opening balance journal entry.
    """
    account = get_object_or_404(Account, pk=pk)

    if request.method == "POST":
        account_name   = request.POST.get("account_name")
        account_number = request.POST.get("account_number")
        account_type   = request.POST.get("account_type")
        detail_type    = request.POST.get("detail_type")
        tax_category   = request.POST.get("tax_category")
        is_subaccount  = request.POST.get("is_subaccount") == "on"

        parent = None
        parent_id = request.POST.get("parent")
        if is_subaccount and parent_id:
            parent = Account.objects.filter(id=parent_id).first()

        opening_balance_str = request.POST.get("opening_balance") or "0"
        try:
            new_opening_balance = Decimal(opening_balance_str)
        except Exception:
            new_opening_balance = Decimal("0")

        as_of_str = request.POST.get("as_of")
        as_of = as_of_str or timezone.now().date()

        description = request.POST.get("description")

        # ---- update account fields ----
        account.account_name   = account_name
        account.account_number = account_number
        account.account_type   = account_type
        account.detail_type    = detail_type
        account.tax_category   = tax_category
        account.is_subaccount  = is_subaccount
        account.parent         = parent
        account.opening_balance = new_opening_balance
        account.as_of          = as_of
        account.description    = description
        account.save()

        # ---- update opening balance journal entry ----
        # Only for balance-sheet accounts
        level1 = account.level1_group
        je = JournalEntry.objects.filter(
            source_type="OPENING_BALANCE",
            source_id=account.id
        ).first()

        if level1 not in ["Assets", "Liabilities", "Equity"]:
            # Not a balance-sheet account: remove any OB entry
            if je:
                je.delete()
        else:
            if new_opening_balance == 0:
                # zero balance: remove any OB entry
                if je:
                    je.delete()
            else:
                # Ensure Opening Balance Equity exists
                opening_equity_acct, _ = Account.objects.get_or_create(
                    account_name="Opening Balance Equity",
                    account_type="OWNER_EQUITY",
                    defaults={
                        "detail_type": "Opening balances",
                        "is_active": True,
                    },
                )

                amount = abs(new_opening_balance)

                if not je:
                    je = JournalEntry.objects.create(
                        date=as_of,
                        description=f"Opening balance for {account.account_name}",
                        source_type="OPENING_BALANCE",
                        source_id=account.id,
                    )
                else:
                    je.date = as_of
                    je.description = f"Opening balance for {account.account_name}"
                    je.save()
                    # wipe existing lines and rebuild
                    JournalLine.objects.filter(entry=je).delete()

                if level1 == "Assets":
                    # DR asset, CR Opening Balance Equity
                    JournalLine.objects.create(
                        entry=je,
                        account=account,
                        debit=amount,
                        credit=Decimal("0"),
                    )
                    JournalLine.objects.create(
                        entry=je,
                        account=opening_equity_acct,
                        debit=Decimal("0"),
                        credit=amount,
                    )
                else:
                    # Liabilities/Equity: CR account, DR Opening Balance Equity
                    JournalLine.objects.create(
                        entry=je,
                        account=account,
                        debit=Decimal("0"),
                        credit=amount,
                    )
                    JournalLine.objects.create(
                        entry=je,
                        account=opening_equity_acct,
                        debit=amount,
                        credit=Decimal("0"),
                    )

        # redirects
        save_action = request.POST.get("save_action")
        if save_action == "save&new":
            return redirect("accounts:add-account")
        elif save_action == "save&close":
            return redirect("accounts:accounts")

        return redirect("accounts:accounts")

    # GET: show form with existing values
    parents = Account.objects.exclude(pk=account.pk)
    return render(request, "coa_form.html", {
        "parents": parents,
        "account": account,
    })
# working on the activate and make inactive 
# @login_required
def deactivate_account(request, pk):
    coa = get_object_or_404(Account, pk=pk)
    coa.is_active = False
    coa.save()
    return redirect("accounts:accounts")


# @login_required
def activate_account(request, pk):
    coa = get_object_or_404(Account, pk=pk)
    coa.is_active = True
    coa.save()
    return redirect("accounts:accounts")

# the general ledger logic
# making the general ledger rows linkable

def get_entry_link(entry):
    """
    Map a JournalEntry to the EDIT page of the original transaction.
    Falls back to a journal entry edit/detail view if we don't know the source.
    """

    # normalise source_type to uppercase to handle both 'expense'/'EXPENSE'
    st = (entry.source_type or "").upper()
    sid = entry.source_id

    # Opening balances from Chart of Accounts -> edit that account
    if st == "OPENING_BALANCE":
        if sid:
            # sid is the Account PK we stored when creating the JE
            return reverse("accounts:edit-account", args=[sid])
        # fallback if somehow source_id is missing
        return reverse("accounts:accounts")
    
        # ---- CUSTOMER OPENING BALANCE ---------------------------------------
    if st == "CUSTOMER_OPENING_BALANCE" and sid:
        return reverse("sowaf:edit-customer", args=[sid])


    # ---- SALES / INVOICES -------------------------------------------------
    if st == "INVOICE" and sid:
        # if you have an invoice-edit view, prefer that:
        try:
            return reverse("sales:edit-invoice", args=[sid])
        except Exception:
            # fall back to the detail view you already had wired
            return reverse("sales:invoice-detail", args=[sid])
# ---- SALES / PAYMENTS -------------------------------------------------
    if st == "PAYMENT" and sid:
        # if you have an payment-edit view, prefer that:
        try:
            return reverse("sales:payment-edit", args=[sid])
        except Exception:
            # fall back to the detail view you already had wired
            return reverse("sales:payment-detail", args=[sid])
# ---- SALES / RECEIPTS -------------------------------------------------
    if st == "SALES_RECEIPT" and sid:
        # if you have an receipt-edit view, prefer that:
        try:
            return reverse("sales:receipt-edit", args=[sid])
        except Exception:
            # fall back to the detail view you already had wired
            return reverse("sales:receipt-detail", args=[sid])

    # ---- EXPENSES MODULE ---------------------------------------------------
    # Bill
    if st == "BILL" and sid:
        return reverse("expenses:bill-edit", args=[sid])

    # Expense
    if st == "EXPENSE" and sid:
        return reverse("expenses:expense-edit", args=[sid])

    # Cheque
    if st == "CHEQUE" and sid:
        return reverse("expenses:cheque-edit", args=[sid])

    # Purchase Order
    if st == "PURCHASE_ORDER" and sid:
        return reverse("expenses:purchase-order-edit", args=[sid])

    # Supplier Credit
    if st == "SUPPLIER_CREDIT" and sid:
        return reverse("expenses:supplier-credit-edit", args=[sid])

    # Pay Down Credit Card
    if st == "PAYDOWN_CREDIT" and sid:
        return reverse("expenses:paydown-credit-edit", args=[sid])

    # Credit Card Credit
    if st == "CREDIT_CARD_CREDIT" and sid:
        return reverse("expenses:credit-card-credit-edit", args=[sid])
    if st == "CUSTOMER_REFUND" and sid:
        return reverse("sales:customer-refund-edit", args=[sid])

    if st == "SUPPLIER_REFUND" and sid:
        return reverse("expenses:supplier-refund-edit", args=[sid])  # if you later add edit

    if st == "SALES_RECEIPT" and sid:
        return reverse("sales:receipt-edit", args=[sid])
    # ---- FALLBACK: MANUAL JOURNALS ---------------------------------------
    # If we get here, either source_type is empty/unknown,
    # or we couldn't reverse a URL above.
    try:
        # if you have an edit view for journal entries
        return reverse("accounts:journal-entry-edit", args=[entry.id])
    except Exception:
        try:
            # fall back to your old detail view if edit doesn't exist
            return reverse("accounts:journal-entry-detail", args=[entry.id])
        except Exception:
            # last resort – no link, but don't break the GL page
            return "#" 

# @login_required

def general_ledger(request):
    # -------- filters ----------
    account_id = request.GET.get("account_id")
    query      = request.GET.get("search", "")
    date_from  = request.GET.get("date_from")
    date_to    = request.GET.get("date_to")
    export     = request.GET.get("export")
    page_num   = request.GET.get("page", 1)

    supplier_id = request.GET.get("supplier_id")
    customer_id = request.GET.get("customer_id")

    # include_children can be auto-enabled later when a parent is selected
    include_children_raw = request.GET.get("include_children")

    # SANITIZE
    if not account_id or account_id in ("None", "null", ""):
        account_id = None
    if not date_from or date_from in ("None", "null", ""):
        date_from = None
    if not date_to or date_to in ("None", "null", ""):
        date_to = None
    if not supplier_id or supplier_id in ("None", "null", ""):
        supplier_id = None
    if not customer_id or customer_id in ("None", "null", ""):
        customer_id = None

    view_mode = (request.GET.get("view") or "detail").lower()
    if view_mode not in ["summary", "detail"]:
        view_mode = "detail"

    accounts = Account.objects.filter(is_active=True).order_by("account_name")

    account_label = None
    selected_account = None
    if account_id:
        selected_account = Account.objects.filter(pk=account_id).select_related("parent").first()
        if selected_account:
            account_label = selected_account.account_name

    def normal_side(acc: Account):
        if acc.level1_group in ["Assets", "Expenses"]:
            return "debit"
        return "credit"

    def apply_movement(acc: Account, running: Decimal, debit: Decimal, credit: Decimal):
        side = normal_side(acc)
        if side == "debit":
            return running + (debit - credit)
        return running + (credit - debit)

    def split_to_tb_columns(acc: Account, ending: Decimal):
        """
        Same logic as Trial Balance:
        - Debit-normal: + => Debit, - => Credit
        - Credit-normal: + => Credit, - => Debit
        """
        side = normal_side(acc)

        if side == "debit":
            if ending >= 0:
                return ending, Decimal("0.00")
            return Decimal("0.00"), -ending

        if ending >= 0:
            return Decimal("0.00"), ending
        return -ending, Decimal("0.00")

    # =====================================================================
    # BUILD ACCOUNT TREE (needed to include children transactions)
    # =====================================================================
    all_active_accounts = list(Account.objects.filter(is_active=True).select_related("parent"))
    acc_by_id = {a.id: a for a in all_active_accounts}

    children_by_parent = defaultdict(list)
    for a in all_active_accounts:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a)

    def sort_key(a: Account):
        return (a.account_number or "", a.account_name or "")

    for pid in list(children_by_parent.keys()):
        children_by_parent[pid].sort(key=sort_key)

    def collect_subtree_ids(root_id: int):
        out = []
        stack = [root_id]
        while stack:
            nid = stack.pop()
            out.append(nid)
            for ch in children_by_parent.get(nid, []):
                stack.append(ch.id)
        return out

    # ---------------------------------------------------------------------
    # Decide include_children:
    # - If user explicitly passed include_children => honor it
    # - Else if selected account has children => AUTO include children so GL matches TB rolled-up parents
    # ---------------------------------------------------------------------
    include_children = False
    if include_children_raw is not None:
        include_children = True if str(include_children_raw).lower() in ("1", "true", "yes", "on") else False
    else:
        if selected_account and children_by_parent.get(selected_account.id):
            include_children = True

    # Decide which account ids should be included in the report
    selected_account_ids = None
    if selected_account:
        selected_account_ids = [selected_account.id]

        # If user requested include_children OR if this is AR/AP parent header -> auto include subtree
        ARAP_PARENT_DETAIL_TYPES = {"Accounts Receivable (A/R)", "Accounts Payable (A/P)"}
        dt = (selected_account.detail_type or "").strip()
        nm = (selected_account.account_name or "").strip().lower()
        is_arap_parent = (dt in ARAP_PARENT_DETAIL_TYPES) or ("accounts receivable" in nm) or ("accounts payable" in nm)

        if include_children or is_arap_parent:
            selected_account_ids = collect_subtree_ids(selected_account.id)

    # =====================================================================
    # BASE QUERYSETS
    # =====================================================================
    base_qs = JournalLine.objects.select_related("entry", "account").order_by("entry__date", "id")

    if supplier_id:
        base_qs = base_qs.filter(supplier_id=supplier_id)
    if customer_id:
        base_qs = base_qs.filter(customer_id=customer_id)

    # filter by selected account OR subtree
    if selected_account_ids:
        base_qs = base_qs.filter(account_id__in=selected_account_ids)

    # balance_qs: for computing balances (ignores query)
    balance_qs = base_qs
    if date_to:
        balance_qs = balance_qs.filter(entry__date__lte=date_to)

    # display_qs: for showing rows (respects query + date range)
    display_qs = base_qs
    if date_from:
        display_qs = display_qs.filter(entry__date__gte=date_from)
    if date_to:
        display_qs = display_qs.filter(entry__date__lte=date_to)

    if query:
        display_qs = display_qs.filter(
            Q(entry__description__icontains=query) |
            Q(account__account_name__icontains=query)
        )

    all_lines = list(display_qs)

    # ---------------------------------------------------------------------
    # Build DISPLAY maps (based on the filtered result set)
    # ---------------------------------------------------------------------
    lines_by_account = defaultdict(list)
    for ln in all_lines:
        lines_by_account[ln.account_id].append(ln)

    def get_root(acc: Account):
        cur = acc
        while cur and cur.parent_id:
            cur = cur.parent
        return cur

    selected_root_id = None
    if selected_account:
        root = get_root(selected_account)
        selected_root_id = root.id if root else None

    active_account_ids = set(lines_by_account.keys())

    visible_ids = set(active_account_ids)
    for acc_id0 in list(active_account_ids):
        cur = acc_by_id.get(acc_id0)
        while cur and cur.parent_id:
            visible_ids.add(cur.parent_id)
            cur = acc_by_id.get(cur.parent_id)

    roots = [a for a in all_active_accounts if a.parent_id is None and a.id in visible_ids]
    roots.sort(key=sort_key)

    if selected_root_id:
        roots = [r for r in roots if r.id == selected_root_id]

    # ---------------------------------------------------------------------
    # TRUE closing balances (as-of date_to)
    # ---------------------------------------------------------------------
    closing_by_account_id = defaultdict(lambda: Decimal("0.00"))

    for ln in balance_qs:
        acc = ln.account
        debit = Decimal(str(ln.debit or "0.00"))
        credit = Decimal(str(ln.credit or "0.00"))
        closing_by_account_id[acc.id] = apply_movement(acc, closing_by_account_id[acc.id], debit, credit)

    def rolled_up_closing(acc: Account) -> Decimal:
        total = Decimal("0.00")
        for aid in collect_subtree_ids(acc.id):
            total += closing_by_account_id.get(aid, Decimal("0.00"))
        return total

    # =========================
    # SUMMARY VIEW
    # =========================
    if view_mode == "summary":
        summary_rows = []
        grand_total = Decimal("0.00")

        for r in roots:
            has_any = False
            for aid in collect_subtree_ids(r.id):
                if lines_by_account.get(aid):
                    has_any = True
                    break
            if not has_any:
                continue

            closing = rolled_up_closing(r)

            summary_rows.append({
                "account_name": r.account_name,
                "account_number": r.account_number or "",
                "account_type": r.get_account_type_display() if r.account_type else "",
                "closing_balance": closing,
            })
            grand_total += closing

        if export == "excel":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="general_ledger_summary.csv"'
            writer = csv.writer(response)
            writer.writerow(["Account Name", "Account Number", "Account Type", "Closing Balance"])
            for row in summary_rows:
                writer.writerow([
                    row["account_name"],
                    row["account_number"],
                    row["account_type"],
                    float(row["closing_balance"]),
                ])
            writer.writerow(["TOTAL", "", "", float(grand_total)])
            return response

        paginator = Paginator(summary_rows, 50)
        page_obj = paginator.get_page(page_num)

        return render(request, "general_ledger.html", {
            "mode": "summary",
            "view_mode": view_mode,
            "accounts": accounts,
            "account_id": account_id,
            "account_label": account_label,
            "date_from": date_from,
            "date_to": date_to,
            "query": query,
            "supplier_id": supplier_id,
            "customer_id": customer_id,
            "include_children": include_children,
            "page_obj": page_obj,
            "summary_rows": page_obj.object_list,
            "total_balance": grand_total,
        })

    # =========================
    # DETAIL VIEW
    # =========================
    paginator = Paginator(roots, 10)
    page_obj = paginator.get_page(page_num)
    paged_roots = list(page_obj.object_list)

    detail_rows = []

    # Movement totals (transactions shown)
    tx_total_debit = Decimal("0.00")
    tx_total_credit = Decimal("0.00")

    def clean_desc(description: str):
        d = description or ""
        if d.lower().startswith("payment"):
            d = d.replace("Payment", "Sales Collection", 1)
        return d

    def build_path(parent_path: str, node_id: int):
        return f"{parent_path}/{node_id}" if parent_path else str(node_id)

    ARAP_PARENT_DETAIL_TYPES = {
        "Accounts Receivable (A/R)",
        "Accounts Payable (A/P)",
    }

    def is_arap_parent(acc: Account) -> bool:
        if not acc:
            return False
        dt = (acc.detail_type or "").strip()
        nm = (acc.account_name or "").strip().lower()
        if dt in ARAP_PARENT_DETAIL_TYPES:
            return True
        if "accounts receivable" in nm or "(a/r)" in nm:
            return True
        if "accounts payable" in nm or "(a/p)" in nm:
            return True
        return False

    def is_under_arap(acc: Account) -> bool:
        if not acc or not acc.parent:
            return False
        parent_dt = (acc.parent.detail_type or "").strip()
        parent_nm = (acc.parent.account_name or "").strip().lower()
        if parent_dt in ARAP_PARENT_DETAIL_TYPES:
            return True
        if "accounts receivable" in parent_nm or "(a/r)" in parent_nm:
            return True
        if "accounts payable" in parent_nm or "(a/p)" in parent_nm:
            return True
        return False

    def add_account_header(acc: Account, depth: int, depth_display: int, path: str, has_children: bool):
        closing = rolled_up_closing(acc)

        detail_rows.append({
            "kind": "account",
            "depth": depth,
            "depth_display": depth_display,
            "path": path,
            "account_id": acc.id,
            "has_children": has_children,
            "account_name": acc.account_name,
            "account_number": acc.account_number or "",
            "account_type": acc.get_account_type_display() if acc.account_type else "",
            "detail_type": acc.detail_type or "",
            "closing_balance": closing,
        })

    def starting_balance_for_account(acc: Account) -> Decimal:
        if not date_from:
            return Decimal("0.00")

        prior = base_qs.filter(entry__date__lt=date_from, account_id=acc.id)
        running0 = Decimal("0.00")
        for ln0 in prior:
            d0 = Decimal(str(ln0.debit or "0.00"))
            c0 = Decimal(str(ln0.credit or "0.00"))
            running0 = apply_movement(acc, running0, d0, c0)
        return running0

    def add_leaf_transactions(acc: Account, depth: int, depth_display: int, path: str):
        nonlocal tx_total_debit, tx_total_credit

        lines = lines_by_account.get(acc.id, [])
        if not lines:
            return

        running = starting_balance_for_account(acc)

        for ln in lines:
            debit = Decimal(str(ln.debit or "0.00"))
            credit = Decimal(str(ln.credit or "0.00"))

            running = apply_movement(acc, running, debit, credit)

            detail_rows.append({
                "kind": "tx",
                "depth": depth,
                "depth_display": depth_display,
                "path": path,
                "account_name": acc.account_name,
                "account_number": acc.account_number or "",
                "account_type": acc.get_account_type_display() if acc.account_type else "",
                "detail_type": acc.detail_type or "",
                "date": ln.entry.date,
                "description": clean_desc(ln.entry.description),
                "reference": ln.entry.id,
                "debit": debit,
                "credit": credit,
                "balance": running,
                "url": get_entry_link(ln.entry),
                "supplier_id": ln.supplier_id,
                "customer_id": ln.customer_id,
            })

            # these are MOVEMENT totals (not closing)
            tx_total_debit += debit
            tx_total_credit += credit

    def walk_tree(node: Account, depth: int, parent_path: str):
        if node.id not in visible_ids:
            return

        kids = [c for c in children_by_parent.get(node.id, []) if c.id in visible_ids]
        has_children = len(kids) > 0

        if is_arap_parent(node):
            for ch in kids:
                walk_tree(ch, depth, parent_path)
            return

        path = build_path(parent_path, node.id)

        flatten = is_under_arap(node)
        depth_display = 0 if flatten else depth
        force_header = flatten

        is_root = (depth == 0)

        if is_root or has_children or force_header:
            add_account_header(node, depth, depth_display, path, has_children)

        add_leaf_transactions(node, depth, depth_display, path)

        for ch in kids:
            walk_tree(ch, depth + 1, path)

    for r in paged_roots:
        walk_tree(r, 0, "")

    # ---------------------------------------------------------------------
    # TB-CONSISTENT TOTALS WHEN AN ACCOUNT IS SELECTED
    # ---------------------------------------------------------------------
    closing_total = None
    closing_total_debit = Decimal("0.00")
    closing_total_credit = Decimal("0.00")

    if selected_account:
        closing_total = rolled_up_closing(selected_account)
        closing_total_debit, closing_total_credit = split_to_tb_columns(selected_account, closing_total)

    # This one is still movement net (useful sometimes)
    tx_total_balance = tx_total_debit - tx_total_credit

    if export == "excel":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="general_ledger_detail.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Level", "Account Name", "Account Number", "Account Type", "Account Sub-Class",
            "Date", "Description", "Journal Ref", "Debit", "Credit", "Balance"
        ])

        for r in detail_rows:
            level = r.get("depth_display", 0)
            if r.get("kind") == "account":
                writer.writerow([
                    level,
                    r["account_name"],
                    r["account_number"],
                    r["account_type"],
                    r["detail_type"],
                    "", "", "",
                    "", "",
                    float(r.get("closing_balance") or Decimal("0.00")),
                ])
            else:
                writer.writerow([
                    level,
                    r["account_name"],
                    r["account_number"],
                    r["account_type"],
                    r["detail_type"],
                    r["date"].strftime("%Y-%m-%d"),
                    r["description"],
                    r["reference"],
                    float(r["debit"]),
                    float(r["credit"]),
                    float(r["balance"]),
                ])
        return response

    return render(request, "general_ledger.html", {
        "mode": "detail",
        "view_mode": view_mode,
        "accounts": accounts,
        "account_id": account_id,
        "account_label": account_label,
        "date_from": date_from,
        "date_to": date_to,
        "query": query,
        "supplier_id": supplier_id,
        "customer_id": customer_id,
        "include_children": include_children,
        "page_obj": page_obj,
        "detail_rows": detail_rows,

        # Movement totals (transactions shown)
        "total_debit": tx_total_debit,
        "total_credit": tx_total_credit,
        "total_balance": tx_total_balance,

        # TB-consistent totals (closing) when an account is selected
        "closing_total": closing_total,
        "closing_total_debit": closing_total_debit,
        "closing_total_credit": closing_total_credit,
    })


# general ledger print view

def general_ledger_print(request):
    """
    Clean print-friendly General Ledger view.
    No pagination, no export, no UI – just the report.
    """

    account_id = request.GET.get("account_id")
    query      = request.GET.get("search", "")
    date_from  = request.GET.get("date_from")
    date_to    = request.GET.get("date_to")

    include_children = request.GET.get("include_children")
    include_children = True if str(include_children).lower() in ("1", "true", "yes", "on") else False

    # clean "None"/"null"/"" values
    if account_id in ["None", "null", ""]:
        account_id = None
    if date_from in ["None", "null", ""]:
        date_from = None
    if date_to in ["None", "null", ""]:
        date_to = None

    selected_account = None
    account_ids = None

    # build account subtree when needed
    if account_id:
        selected_account = Account.objects.filter(pk=account_id).select_related("parent").first()
        if selected_account:
            all_active_accounts = list(Account.objects.filter(is_active=True).select_related("parent"))
            children_by_parent = {}
            for a in all_active_accounts:
                if a.parent_id:
                    children_by_parent.setdefault(a.parent_id, []).append(a.id)

            def collect_subtree_ids(root_id: int):
                out = []
                stack = [root_id]
                while stack:
                    nid = stack.pop()
                    out.append(nid)
                    for cid in children_by_parent.get(nid, []):
                        stack.append(cid)
                return out

            ARAP_PARENT_DETAIL_TYPES = {"Accounts Receivable (A/R)", "Accounts Payable (A/P)"}
            dt = (selected_account.detail_type or "").strip()
            nm = (selected_account.account_name or "").strip().lower()
            is_arap_parent = (dt in ARAP_PARENT_DETAIL_TYPES) or ("accounts receivable" in nm) or ("accounts payable" in nm)

            if include_children or is_arap_parent:
                account_ids = collect_subtree_ids(selected_account.id)
            else:
                account_ids = [selected_account.id]

    lines_qs = JournalLine.objects.select_related("entry", "account").order_by("entry__date", "id")

    if account_ids:
        lines_qs = lines_qs.filter(account_id__in=account_ids)

    if date_from:
        lines_qs = lines_qs.filter(entry__date__gte=date_from)
    if date_to:
        lines_qs = lines_qs.filter(entry__date__lte=date_to)

    if query:
        lines_qs = lines_qs.filter(
            Q(entry__description__icontains=query) |
            Q(account__account_name__icontains=query)
        )

    lines = list(lines_qs)

    running_rows = []
    balance = Decimal("0")
    total_debit = Decimal("0")
    total_credit = Decimal("0")

    def normal_side(acc: Account):
        if acc.level1_group in ["Assets", "Expenses"]:
            return "debit"
        return "credit"

    for ln in lines:
        side = normal_side(ln.account)

        debit = Decimal(str(ln.debit or "0"))
        credit = Decimal(str(ln.credit or "0"))

        if side == "debit":
            balance += (debit - credit)
        else:
            balance += (credit - debit)

        total_debit += debit
        total_credit += credit

        acc = ln.account
        running_rows.append({
            "account_name": acc.account_name,
            "account_number": getattr(acc, "account_number", ""),
            "account_type": getattr(acc, "level1_group", ""),
            "account_sub_class": getattr(acc, "detail_type", ""),
            "date": ln.entry.date,
            "description": ln.entry.description,
            "debit": debit,
            "credit": credit,
            "running_balance": balance,
        })

    context = {
        "rows": running_rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "total_balance": balance,
        "selected_account": selected_account,
        "date_from": date_from,
        "date_to": date_to,
        "query": query,
        "generated_on": timezone.now(),
    }
    return render(request, "general_ledger_print.html", context)


# journel entry lists
# @login_required
def journal_entries(request):
    entries = (
        JournalEntry.objects
        .prefetch_related("lines__account")
        .order_by("-date", "-id")
    )

    totals = JournalLine.objects.filter(entry__in=entries).aggregate(
        grand_debit=Sum("debit"),
        grand_credit=Sum("credit"),
    )

    # ADD EDIT URLs JUST LIKE GL
    for entry in entries:
        entry.edit_url = get_entry_link(entry)

    context = {
        "entries": entries,
        "grand_debit": totals.get("grand_debit") or Decimal("0"),
        "grand_credit": totals.get("grand_credit") or Decimal("0"),
    }

    return render(request, "journal_entries.html", context)
# journal entry detail view
# @login_required
def journal_entry_detail(request, pk):
    entry = (
        JournalEntry.objects
        .prefetch_related("lines__account")
        .filter(pk=pk)
        .first()
    )
    if not entry:
        raise Http404("Journal entry not found")

    totals = JournalLine.objects.filter(entry=entry).aggregate(
        grand_debit=Sum("debit"),
        grand_credit=Sum("credit"),
    )
     
    context = {
        "entries": [entry],  # reuse the same table layout
        "grand_debit": totals.get("grand_debit") or Decimal("0"),
        "grand_credit": totals.get("grand_credit") or Decimal("0"),
    }
    return render(request, "journal_entries.html", context)

# working on the reports
def reports(request):
    return render(request, "Reports.html", {})


def reports_export_hub(request, fmt):
    """
    fmt: 'excel' or 'pdf'
    Shows a list of report links. User clicks the report and uses that report's export buttons.
    """
    fmt = (fmt or "").lower()
    if fmt not in ("excel", "pdf"):
        fmt = "excel"

    report_groups = [
        {
            "title": "Accounts Receivable",
            "items": [
                {"label": "A/R Ageing Summary", "url": reverse("sales:aging-report")},
                {"label": "A/R Ageing Detail", "url": reverse("sales:aging-report-detail")},
                {"label": "Open Invoices", "url": reverse("sales:open-invoices-report")},
                {"label": "Customer Balances", "url": reverse("sales:customer-balances-report")},
                {"label": "Invoice List", "url": reverse("sales:invoice-list-report")},
                {"label": "Collections Report", "url": reverse("sales:collections-report")},
            ],
        },
        {
            "title": "Accounts Payable",
            "items": [
                {"label": "A/P Ageing Summary", "url": reverse("expenses:ap-aging-summary")},
                {"label": "A/P Ageing Detail", "url": reverse("expenses:ap-aging-detail")},
                {"label": "Unpaid Bills", "url": reverse("expenses:unpaid-bills-report")},
                {"label": "Vendor Balances", "url": reverse("expenses:vendor-balances-report")},
                {"label": "Bills List", "url": reverse("expenses:bills-list-report")},
                {"label": "Payments to Vendors", "url": reverse("expenses:payments-to-vendors-report")},
            ],
        },
    ]

    return render(request, "reports_export_hub.html", {
        "fmt": fmt,
        "report_groups": report_groups,
    })
# working on the reports 

def _parse_range(request):
    """
    Helper to parse ?from=YYYY-MM-DD&to=YYYY-MM-DD from query string.
    Returns (dfrom, dto) as date or None.
    """
    from_str = request.GET.get("from") or request.GET.get("date_from")
    to_str   = request.GET.get("to") or request.GET.get("date_to")

    dfrom = parse_date(from_str) if from_str else None
    dto   = parse_date(to_str) if to_str else None
    return dfrom, dto

# trial balance
def report_trial_balance(request):
    """
    Trial Balance (Journal-only + Parent-only display)

    - CURRENT balance per account as at dto (or all-time if dto is None)
    - Balance is computed from JournalLine ONLY:
        Debit-normal (Assets, Expenses):  debits - credits
        Credit-normal (Liab, Equity, Income): credits - debits

    - If an account has sub-accounts (visible children), show ONLY the parent row
      with rolled-up balance, and DO NOT show the children rows.
    - Totals sum only displayed rows.
    """

    dfrom, dto = _parse_range(request)

    # ----- helper: normal side -----
    def normal_side(acc):
        if acc.level1_group in ["Assets", "Expenses"]:
            return "debit"
        return "credit"

    def ending_balance_from_journal(acc, debit_sum: Decimal, credit_sum: Decimal) -> Decimal:
        """
        Journal-only ending:
        Debit-normal:  debits - credits
        Credit-normal: credits - debits
        """
        side = normal_side(acc)
        if side == "debit":
            return debit_sum - credit_sum
        return credit_sum - debit_sum

    def split_to_tb_columns(acc, ending: Decimal):
        """
        Put ending into TB debit/credit columns depending on normal side.
        """
        side = normal_side(acc)

        # Debit-normal: + => Debit, - => Credit
        if side == "debit":
            if ending >= 0:
                return ending, Decimal("0.00")
            return Decimal("0.00"), -ending

        # Credit-normal: + => Credit, - => Debit
        if ending >= 0:
            return Decimal("0.00"), ending
        return -ending, Decimal("0.00")

    # ----- Journal sums up to dto (or all time if dto is None) -----
    lines = JournalLine.objects.select_related("entry", "account")
    if dto:
        lines = lines.filter(entry__date__lte=dto)

    sums = (
        lines.values("account_id")
        .annotate(
            debit=Coalesce(
                Sum("debit"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
            credit=Coalesce(
                Sum("credit"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
        )
    )

    sums_map = {
        r["account_id"]: (r["debit"] or Decimal("0.00"), r["credit"] or Decimal("0.00"))
        for r in sums
    }

    # ----- Build account tree -----
    all_accounts = list(Account.objects.filter(is_active=True).select_related("parent"))
    acc_by_id = {a.id: a for a in all_accounts}

    children_by_parent = defaultdict(list)
    for a in all_accounts:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a)

    def sort_key(a):
        return (a.account_number or "", a.account_name or "")

    for pid in list(children_by_parent.keys()):
        children_by_parent[pid].sort(key=sort_key)

    # ----- Compute base ending per account (self-only, JOURNAL-ONLY) -----
    base_ending = {}
    for acc in all_accounts:
        d_sum, c_sum = sums_map.get(acc.id, (Decimal("0.00"), Decimal("0.00")))
        base_ending[acc.id] = ending_balance_from_journal(acc, d_sum, c_sum)

    # ----- Visible accounts: any with non-zero ending OR any activity, plus ancestors -----
    visible_ids = set()

    # activity
    visible_ids |= set(sums_map.keys())

    # non-zero ending
    for acc in all_accounts:
        if base_ending.get(acc.id, Decimal("0.00")) != 0:
            visible_ids.add(acc.id)

    # add ancestors
    for aid in list(visible_ids):
        cur = acc_by_id.get(aid)
        while cur and cur.parent_id:
            visible_ids.add(cur.parent_id)
            cur = acc_by_id.get(cur.parent_id)

    roots = [a for a in all_accounts if a.parent_id is None and a.id in visible_ids]
    roots.sort(key=sort_key)

    # ----- helper: collect subtree ids -----
    def collect_subtree_ids(root_id: int):
        out = []
        stack = [root_id]
        while stack:
            nid = stack.pop()
            out.append(nid)
            for ch in children_by_parent.get(nid, []):
                if ch.id in visible_ids:
                    stack.append(ch.id)
        return out

    # ----- rolled-up ending for a node (sum of base endings in subtree) -----
    def rolled_up_ending(acc: Account) -> Decimal:
        total = Decimal("0.00")
        for sid in collect_subtree_ids(acc.id):
            total += base_ending.get(sid, Decimal("0.00"))
        return total

    # ----- Build rows: PARENT-ONLY (stop recursion when parent has children) -----
    rows = []
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    def walk_parent_only(acc: Account, depth: int):
        nonlocal total_debit, total_credit

        if acc.id not in visible_ids:
            return

        children = [c for c in children_by_parent.get(acc.id, []) if c.id in visible_ids]
        if children:
            # show ONLY parent (rolled-up), do NOT show children
            ending = rolled_up_ending(acc)
            if ending == 0:
                return

            debit_bal, credit_bal = split_to_tb_columns(acc, ending)

            rows.append({
                "account_id": acc.id,
                "account": acc.account_name,
                "debit": debit_bal,
                "credit": credit_bal,
                "depth": depth,
                "kind": "parent",
            })

            total_debit += debit_bal
            total_credit += credit_bal
            return

        # leaf account
        ending = base_ending.get(acc.id, Decimal("0.00"))
        if ending == 0:
            return

        debit_bal, credit_bal = split_to_tb_columns(acc, ending)

        rows.append({
            "account_id": acc.id,
            "account": acc.account_name,
            "debit": debit_bal,
            "credit": credit_bal,
            "depth": depth,
            "kind": "leaf",
        })

        total_debit += debit_bal
        total_credit += credit_bal

    for r in roots:
        walk_parent_only(r, 0)

    # =========================================================
    # EXPORTS: CSV / EXCEL / PDF (without changing your styles)
    # =========================================================
    export = (request.GET.get("export") or "").strip().lower()

    def _period_label():
        a = dfrom.strftime("%d/%m/%Y") if dfrom else "None"
        b = dto.strftime("%d/%m/%Y") if dto else "None"
        return f"{a} – {b}"

    filename_base = f"trial_balance_{timezone.localdate().strftime('%Y%m%d')}"

    if export == "csv":
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'

        writer = csv.writer(resp)
        writer.writerow([f"Trial Balance - { _period_label() }"])
        writer.writerow([f"Company: YoAccountant", f"Currency: UGX"])
        writer.writerow([])
        writer.writerow(["Account", "Debit", "Credit"])

        for r in rows:
            writer.writerow([
                r["account"],
                f"{r['debit']:.2f}",
                f"{r['credit']:.2f}",
            ])

        writer.writerow([])
        writer.writerow(["Total", f"{total_debit:.2f}", f"{total_credit:.2f}"])
        return resp

    if export == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        ws.append([f"Trial Balance - { _period_label() }"])
        ws.append([f"Company: YoAccountant", f"Currency: UGX"])
        ws.append([])
        ws.append(["Account", "Debit", "Credit"])

        # headers bold-ish (simple)
        for cell in ws[4]:
            cell.font = cell.font.copy(bold=True)

        for r in rows:
            ws.append([r["account"], float(r["debit"]), float(r["credit"])])

        ws.append([])
        ws.append(["Total", float(total_debit), float(total_credit)])

        # number format
        for row in ws.iter_rows(min_row=5, min_col=2, max_col=3):
            for c in row:
                c.number_format = "#,##0.00"

        # column widths
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    if export == "pdf":
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Trial Balance", styles["Title"]))
        story.append(Paragraph(f"Company: YoAccountant", styles["Normal"]))
        story.append(Paragraph(f"Period: {_period_label()}", styles["Normal"]))
        story.append(Paragraph(f"Reporting Currency: UGX", styles["Normal"]))
        story.append(Spacer(1, 12))

        data = [["Account", "Debit", "Credit"]]
        for r in rows:
            data.append([
                r["account"],
                f"{r['debit']:,.2f}",
                f"{r['credit']:,.2f}",
            ])
        data.append(["Total", f"{total_debit:,.2f}", f"{total_credit:,.2f}"])

        table = Table(data, colWidths=[280, 110, 110])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e6f7ea")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f5132")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),

            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e7e7e7")),
            ("FONTSIZE", (0, 1), (-1, -2), 9),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("BACKGROUND", (0, 1), (-1, -2), colors.white),

            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8fff9")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))

        story.append(table)
        doc.build(story)

        pdf = buf.getvalue()
        buf.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
        return resp

    # =========================================================
    # NORMAL PAGE RENDER
    # =========================================================
    context = {
        "company_name": "YoAccountant",
        "reporting_currency": "UGX",
        "rows": rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "dfrom": dfrom,
        "dto": dto,
    }
    return render(request, "trial_balance.html", context)

# profit and loss
# Treat these detail types as COGS (from your COA form options)
COGS_DETAIL_TYPES = {
    "cost of goods sold",
    "cost of sales",
}


def dec(v) -> Decimal:
    """Safe Decimal converter."""
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0.00")


def bankish_q():
    """
    IMPORTANT: Use DETAIL TYPE (your COA has 3 layers).
    If your bank accounts have detail_type like 'Bank', 'Cash and Cash Equivalents', etc.
    this will catch them.
    """
    return (
        Q(detail_type__icontains="bank") |
        Q(detail_type__icontains="cash") |
        Q(detail_type__icontains="cash and cash equivalents") |
        Q(detail_type__icontains="cash on hand")
    )


# Your COA codes (recommended) + fallback text matching
INCOME_TYPES = {"OPERATING_INCOME", "INVESTING_INCOME"}
EXPENSE_TYPES = {"OPERATING_EXPENSE", "INVESTING_EXPENSE", "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE"}

def dec(x) -> Decimal:
    return Decimal(str(x or "0.00"))

# Example placeholders — keep your real ones
INCOME_TYPES = {"Income", "Other Income"}
EXPENSE_TYPES = {"Expense", "Other Expense", "Cost of Goods Sold"}
COGS_DETAIL_TYPES = {"cost of goods sold", "cogs"}  # keep your real set

def report_pnl(request):
    # ---------- 1. Parse date filters ----------
    dfrom_raw = request.GET.get("from")
    dto_raw = request.GET.get("to")

    date_from = None
    date_to = None

    try:
        if dfrom_raw:
            date_from = datetime.strptime(dfrom_raw, "%Y-%m-%d").date()
        if dto_raw:
            date_to = datetime.strptime(dto_raw, "%Y-%m-%d").date()
    except ValueError:
        date_from = None
        date_to = None

    # If no selected date, treat it as "As of today" for display (and to avoid "None to None")
    today = timezone.localdate()
    if not date_from and not date_to:
        date_to = today
    elif date_from and not date_to:
        date_to = today

    # ---------- 2. Accounting basis toggle (cash/accrual) ----------
    basis = (request.GET.get("basis") or "accrual").lower()
    if basis not in {"cash", "accrual"}:
        basis = "accrual"

    # ---------- 3. Base queryset from journal lines ----------
    jl = JournalLine.objects.select_related("account", "entry")

    if date_from:
        jl = jl.filter(entry__date__gte=date_from)
    if date_to:
        jl = jl.filter(entry__date__lte=date_to)

    # ---------- 4. Cash basis filter (QB-like simplified) ----------
    if basis == "cash":
        cash_bank_accounts = Account.objects.filter(is_active=True).filter(bankish_q())

        cash_entry_ids = (
            jl.filter(account__in=cash_bank_accounts)
              .values_list("entry_id", flat=True)
              .distinct()
        )
        jl = jl.filter(entry_id__in=cash_entry_ids)

    posted_accounts = (
        Account.objects
        .filter(journalline__in=jl)
        .distinct()
        .select_related("parent")
    )

    # ---------- 5. Helper: classify accounts into P&L buckets ----------
    def classify_account(acc: Account) -> str | None:
        code = (acc.account_type or "").strip()
        detail = (getattr(acc, "detail_type", "") or "").strip().lower()
        name = (acc.account_name or "").strip().lower()

        if code in INCOME_TYPES:
            return "income"

        if code in EXPENSE_TYPES:
            if detail in COGS_DETAIL_TYPES or any(k in name for k in ["cogs", "cost of sales", "cost of goods"]):
                return "cogs"
            return "expense"

        text = f"{(acc.account_type or '').lower()} {detail} {name}"
        if "income" in text or "revenue" in text or "sales" in text:
            return "income"
        if "cogs" in text or "cost of goods" in text or "cost of sales" in text:
            return "cogs"
        if "expense" in text:
            return "expense"
        return None

    # ============================================================
    # PARENT-ONLY ROLLUP LIKE TRIAL BALANCE
    # ============================================================

    all_active = list(Account.objects.filter(is_active=True).select_related("parent"))
    children_by_parent = defaultdict(list)

    for a in all_active:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a)

    def get_display_parent(acc: Account) -> Account:
        cur = acc
        while cur and cur.parent_id and getattr(cur, "is_subaccount", False):
            cur = cur.parent
        while cur and getattr(cur, "is_subaccount", False) and cur.parent_id:
            cur = cur.parent
        return cur or acc

    def collect_subtree_ids(root_id: int):
        out = []
        stack = [root_id]
        while stack:
            nid = stack.pop()
            out.append(nid)
            for ch in children_by_parent.get(nid, []):
                stack.append(ch.id)
        return out

    agg_map = jl.values("account_id").annotate(debit=Sum("debit"), credit=Sum("credit"))
    debit_by_id = defaultdict(lambda: Decimal("0.00"))
    credit_by_id = defaultdict(lambda: Decimal("0.00"))

    for row in agg_map:
        aid = row["account_id"]
        debit_by_id[aid] = dec(row["debit"])
        credit_by_id[aid] = dec(row["credit"])

    display_parents = {}
    for acc in posted_accounts:
        parent = get_display_parent(acc)
        if parent:
            display_parents[parent.id] = parent

    buckets = {"income": [], "cogs": [], "expense": []}
    totals = {
        "income": Decimal("0.00"),
        "cogs": Decimal("0.00"),
        "expense": Decimal("0.00"),
    }

    parent_list = sorted(display_parents.values(), key=lambda x: (x.account_name or "").lower())

    for parent_acc in parent_list:
        bucket = classify_account(parent_acc)
        if not bucket:
            continue

        subtree_ids = collect_subtree_ids(parent_acc.id)
        debit = sum((debit_by_id[aid] for aid in subtree_ids), Decimal("0.00"))
        credit = sum((credit_by_id[aid] for aid in subtree_ids), Decimal("0.00"))

        if bucket == "income":
            amount = credit - debit
        else:
            amount = debit - credit

        if amount == 0:
            continue

        buckets[bucket].append({
            "account": parent_acc.account_name,
            "account_id": parent_acc.id,
            "amount": amount,
        })
        totals[bucket] += amount

    # ---------- 7. Totals ----------
    gross_profit = totals["income"] - totals["cogs"]
    operating_profit = gross_profit - totals["expense"]
    net_profit = operating_profit

    # ---------- 8. Period text (nice display) ----------
    def _period_text(dfrom, dto, basis):
        if dfrom and dto:
            return f"For the period {dfrom.strftime('%d/%m/%Y')} – {dto.strftime('%d/%m/%Y')} ({basis})"
        if (not dfrom) and dto:
            return f"As of {dto.strftime('%d/%m/%Y')} ({basis})"
        return f"As of {timezone.localdate().strftime('%d/%m/%Y')} ({basis})"

    period_text = _period_text(date_from, date_to, basis)

    # ---------- 9. Export (CSV / Excel / PDF) ----------
    export = (request.GET.get("export") or "").lower().strip()
    if export in {"csv", "xlsx", "pdf"}:

        # -------- CSV --------
        if export == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="profit_and_loss.csv"'
            writer = csv.writer(response)

            writer.writerow(["Statement of Profit or Loss"])
            writer.writerow(["Company", "YoAccountant"])
            writer.writerow(["Period", period_text])
            writer.writerow(["Reporting Currency", "UGX"])
            writer.writerow([])
            writer.writerow(["Section", "Account", "Amount"])

            writer.writerow(["Revenue", "", ""])
            for r in buckets["income"]:
                writer.writerow(["Revenue", r["account"], f"{r['amount']:.2f}"])
            writer.writerow(["", "Total revenue", f"{totals['income']:.2f}"])
            writer.writerow([])

            writer.writerow(["Cost of goods sold", "", ""])
            for r in buckets["cogs"]:
                writer.writerow(["COGS", r["account"], f"{r['amount']:.2f}"])
            writer.writerow(["", "Total COGS", f"{totals['cogs']:.2f}"])
            writer.writerow([])

            writer.writerow(["", "Gross profit", f"{gross_profit:.2f}"])
            writer.writerow([])

            writer.writerow(["Operating expenses", "", ""])
            for r in buckets["expense"]:
                writer.writerow(["Expense", r["account"], f"{r['amount']:.2f}"])
            writer.writerow(["", "Total operating expenses", f"{totals['expense']:.2f}"])
            writer.writerow([])

            writer.writerow(["", "Operating profit", f"{operating_profit:.2f}"])
            writer.writerow(["", "Net profit", f"{net_profit:.2f}"])
            return response

        # -------- Excel (XLSX) --------
        if export == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"

            ws.append(["Statement of Profit or Loss"])
            ws.append(["Company", "YoAccountant"])
            ws.append(["Period", period_text])
            ws.append(["Reporting Currency", "UGX"])
            ws.append([])
            ws.append(["Section", "Account", "Amount"])

            ws.append(["Revenue", "", ""])
            for r in buckets["income"]:
                ws.append(["Revenue", r["account"], float(r["amount"])])
            ws.append(["", "Total revenue", float(totals["income"])])
            ws.append([])

            ws.append(["Cost of goods sold", "", ""])
            for r in buckets["cogs"]:
                ws.append(["COGS", r["account"], float(r["amount"])])
            ws.append(["", "Total COGS", float(totals["cogs"])])
            ws.append([])

            ws.append(["", "Gross profit", float(gross_profit)])
            ws.append([])

            ws.append(["Operating expenses", "", ""])
            for r in buckets["expense"]:
                ws.append(["Expense", r["account"], float(r["amount"])])
            ws.append(["", "Total operating expenses", float(totals["expense"])])
            ws.append([])

            ws.append(["", "Operating profit", float(operating_profit)])
            ws.append(["", "Net profit", float(net_profit)])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            response = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="profit_and_loss.xlsx"'
            return response

        # -------- PDF (STRUCTURED + PROFESSIONAL + NO "Account/Amount" LABELS) --------
        if export == "pdf":
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                leftMargin=18 * mm,
                rightMargin=18 * mm,
                topMargin=16 * mm,
                bottomMargin=16 * mm,
                title="Statement of Profit or Loss",
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "title_style",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=16,
                leading=20,
                textColor=colors.HexColor("#39B54B"),
                alignment=1,  # center
                spaceAfter=6,
            )
            meta_style = ParagraphStyle(
                "meta_style",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=10,
                leading=12,
                textColor=colors.HexColor("#475569"),
                alignment=1,  # center
                spaceAfter=2,
            )
            section_style = ParagraphStyle(
                "section_style",
                parent=styles["Heading3"],
                fontName="Helvetica-Bold",
                fontSize=11,
                textColor=colors.HexColor("#0f5132"),
                spaceBefore=10,
                spaceAfter=6,
            )

            story = []
            story.append(Paragraph("Statement of Profit or Loss", title_style))
            story.append(Paragraph("YoAccountant", meta_style))
            story.append(Paragraph(period_text, meta_style))
            story.append(Paragraph("Reporting Currency: UGX", meta_style))
            story.append(Spacer(1, 10))

            def _fmt(x: Decimal) -> str:
                return f"{x:,.2f}"

            def add_section(title, rows, total_label, total_value, extra_totals=None):
                """
             NO COLUMN HEADERS (Account/Amount) — professional statement layout
                rows: list of dicts with keys: account, amount
                extra_totals: list of tuples (label, value) shown after total (bold)
                """
                story.append(Paragraph(title, section_style))

                data = []
                for r in rows:
                    data.append([r["account"], _fmt(r["amount"])])

                # total row
                data.append([total_label, _fmt(total_value)])

                t = Table(data, colWidths=[120 * mm, 50 * mm])
                t.setStyle(TableStyle([
                    ("FONTNAME", (0, 0), (-1, -2), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -2), 10),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e7e7e7")),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),

                    # Total row styling
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F8FFF9")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]))

                story.append(t)

                if extra_totals:
                    story.append(Spacer(1, 6))
                    data2 = []
                    for lbl, val in extra_totals:
                        data2.append([lbl, _fmt(val)])

                    t2 = Table(data2, colWidths=[120 * mm, 50 * mm])
                    t2.setStyle(TableStyle([
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e7e7e7")),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FAFDFB")),
                    ]))
                    story.append(t2)

                story.append(Spacer(1, 10))

            # Revenue
            add_section(
                "Revenue",
                buckets["income"],
                "Total revenue",
                totals["income"],
            )

            # COGS
            add_section(
                "Cost of goods sold",
                buckets["cogs"],
                "Total COGS",
                totals["cogs"],
                extra_totals=[("Gross profit", gross_profit)]
            )

            # Expenses
            add_section(
                "Operating expenses",
                buckets["expense"],
                "Total operating expenses",
                totals["expense"],
                extra_totals=[
                    ("Operating profit", operating_profit),
                    ("Net profit", net_profit),
                ]
            )

            doc.build(story)

            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="profit_and_loss.pdf"'
            return response

    # ---------- 10. Normal HTML render ----------
    context = {
        "company_name": "YoAccountant",
        "reporting_currency": "UGX",
        "basis": basis,

        "buckets": buckets,
        "totals": totals,
        "gross_profit": gross_profit,
        "operating_profit": operating_profit,
        "profit_before_financing_tax": operating_profit,
        "profit_before_income_tax": operating_profit,
        "net_profit": net_profit,

        "dfrom": date_from,
        "dto": date_to,
    }

    return render(request, "pnl.html", context)

# balance sheet
def report_balance_sheet(request):
    """
    Balance sheet 'as of' a single date, grouped into
    non-current/current assets & liabilities, with each
    account clickable to the General Ledger.

    Method toggle:
    - accrual: use all JournalLines up to 'asof'
    - cash:    use ONLY JournalLines that hit bank/cash accounts

    Retained Earnings / Current Year Profit:
    - computed dynamically (reports only)
    - does NOT create Retained Earnings in DB
    - keeps accounting correct and balances the sheet even without closing entries
    """

    # 1) Read filters
    to_str = request.GET.get("to")
    method = request.GET.get("method") or "accrual"

    if to_str:
        try:
            asof = datetime.strptime(to_str, "%Y-%m-%d").date()
        except ValueError:
            asof = timezone.localdate()
    else:
        asof = timezone.localdate()

    def bankish_q():
        return (
            Q(account__detail_type__icontains="bank") |
            Q(account__detail_type__icontains="cash") |
            Q(account__detail_type__icontains="cash and cash equivalents") |
            Q(account__detail_type__icontains="cash on hand")
        )

    # 2) All lines up to that date
    base_lines = JournalLine.objects.select_related("entry", "account").filter(entry__date__lte=asof)

    # 2b) CASH method: only keep lines that hit bank/cash accounts
    if method == "cash":
        base_lines = base_lines.filter(bankish_q())

    # =========================================================
    # Roll up subaccounts into their parent accounts
    # Show ONLY non-subaccount accounts (like your TB behavior)
    # =========================================================
    all_accs = list(Account.objects.filter(is_active=True).select_related("parent"))
    acc_by_id = {a.id: a for a in all_accs}

    children_by_parent = defaultdict(list)
    for a in all_accs:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a.id)

    def collect_subtree_ids(root_id: int) -> list[int]:
        out = []
        stack = [root_id]
        while stack:
            nid = stack.pop()
            out.append(nid)
            for ch_id in children_by_parent.get(nid, []):
                stack.append(ch_id)
        return out

    def infer_is_current(acc_type_code: str, detail: str) -> bool:
        detail_l = (detail or "").lower()
        is_current_local = True
        if "non current" in detail_l or "non-current" in detail_l:
            is_current_local = False
        if any(k in detail_l for k in ["fixed", "property", "equipment", "long term", "long-term", "ppe"]):
            is_current_local = False
        return is_current_local

    # Containers for rows
    asset_nc_rows = []
    asset_curr_rows = []
    eq_rows = []
    liab_nc_rows = []
    liab_curr_rows = []

    # =========================================================
    # Dynamic Profit/Loss helper (for Retained Earnings)
    # =========================================================
    def normal_side(acc: Account):
        if acc.level1_group in ["Assets", "Expenses"]:
            return "debit"
        return "credit"

    def apply_movement(acc: Account, running: Decimal, debit: Decimal, credit: Decimal) -> Decimal:
        """
        Same movement logic you use elsewhere:
        - Debit-normal: +debit -credit
        - Credit-normal: +credit -debit
        """
        side = normal_side(acc)
        if side == "debit":
            return running + (debit - credit)
        return running + (credit - debit)

    def profit_loss_for_range(dfrom: date | None, dto: date) -> Decimal:
        """
        Returns NET PROFIT (positive = profit, negative = loss)
        computed from Income and Expense accounts using your movement logic.
        """
        qs = JournalLine.objects.select_related("entry", "account").filter(entry__date__lte=dto)

        if dfrom:
            qs = qs.filter(entry__date__gte=dfrom)

        if method == "cash":
            qs = qs.filter(bankish_q())

        # only income + expenses lines
        qs = qs.filter(account__account_type__in=[
            "OPERATING_INCOME", "INVESTING_INCOME",
            "OPERATING_EXPENSE", "INVESTING_EXPENSE", "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE",
        ])

        running_by_acc = defaultdict(lambda: Decimal("0.00"))

        for ln in qs:
            acc = ln.account
            debit = Decimal(str(ln.debit or "0.00"))
            credit = Decimal(str(ln.credit or "0.00"))
            running_by_acc[acc.id] = apply_movement(acc, running_by_acc[acc.id], debit, credit)

        total_income = Decimal("0.00")
        total_exp = Decimal("0.00")

        for acc_id, bal in running_by_acc.items():
            acc = acc_by_id.get(acc_id) or Account.objects.filter(pk=acc_id).first()
            if not acc:
                continue

            if acc.level1_group == "Income":
                total_income += bal
            elif acc.level1_group == "Expenses":
                total_exp += bal

        return total_income - total_exp  # profit (+) or loss (-)

    # Fiscal year start assumption (change later if you want)
    FY_START_MONTH = 1
    FY_START_DAY = 1
    fy_start = date(asof.year, FY_START_MONTH, FY_START_DAY)

    # Current year earnings = FY start .. asof
    current_year_profit = profit_loss_for_range(fy_start, asof)

    # Retained earnings (prior years) = up to day before FY start
    retained_earnings = Decimal("0.00")
    if fy_start > date.min:
        day_before_fy = fy_start.fromordinal(fy_start.toordinal() - 1)
        retained_earnings = profit_loss_for_range(None, day_before_fy)

    # =========================================================
    # Build normal BS rows (Assets/Liab/Equity accounts only)
    # =========================================================
    for acc in all_accs:
        if getattr(acc, "is_subaccount", False):
            continue

        subtree_ids = collect_subtree_ids(acc.id)

        if not base_lines.filter(account_id__in=subtree_ids).exists():
            continue

        agg = base_lines.filter(account_id__in=subtree_ids).aggregate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit"),
        )
        debit = agg["total_debit"] or 0
        credit = agg["total_credit"] or 0

        net = Decimal(debit) - Decimal(credit)  # debit-positive convention

        code = acc.account_type or ""
        level1 = Account.ACCOUNT_LEVEL1_MAP.get(code, "")
        detail_type = getattr(acc, "detail_type", "") or ""

        if level1 == "Assets":
            amount = net
            is_current = infer_is_current(code, detail_type)
            target_list = asset_curr_rows if is_current else asset_nc_rows

        elif level1 == "Liabilities":
            amount = -net
            is_current = infer_is_current(code, detail_type)
            target_list = liab_curr_rows if is_current else liab_nc_rows

        elif level1 == "Equity":
            amount = -net
            target_list = eq_rows

        else:
            continue

        if not amount:
            continue

        target_list.append({
            "account": acc.account_name,
            "account_id": acc.id,
            "amount": amount,
        })

    # =========================================================
    # Inject computed Retained Earnings + Current Year Earnings
    # =========================================================
    if retained_earnings != 0:
        eq_rows.append({
            "account": "Retained Earnings",
            "account_id": None,
            "amount": retained_earnings
        })

    if current_year_profit != 0:
        eq_rows.append({
            "account": "Current Year Profit/Loss",
            "account_id": None,
            "amount": current_year_profit
        })

    # Sorting
    asset_nc_rows.sort(key=lambda x: (x["account"] or "").lower())
    asset_curr_rows.sort(key=lambda x: (x["account"] or "").lower())
    liab_nc_rows.sort(key=lambda x: (x["account"] or "").lower())
    liab_curr_rows.sort(key=lambda x: (x["account"] or "").lower())

    eq_db = [r for r in eq_rows if r.get("account_id")]
    eq_calc = [r for r in eq_rows if not r.get("account_id")]
    eq_db.sort(key=lambda x: (x["account"] or "").lower())
    eq_rows = eq_db + eq_calc

    # Totals
    asset_nc_total = sum(r["amount"] for r in asset_nc_rows) if asset_nc_rows else Decimal("0.00")
    asset_curr_total = sum(r["amount"] for r in asset_curr_rows) if asset_curr_rows else Decimal("0.00")
    asset_total = asset_nc_total + asset_curr_total

    liab_nc_total = sum(r["amount"] for r in liab_nc_rows) if liab_nc_rows else Decimal("0.00")
    liab_curr_total = sum(r["amount"] for r in liab_curr_rows) if liab_curr_rows else Decimal("0.00")
    liab_total = liab_nc_total + liab_curr_total

    eq_total = sum(r["amount"] for r in eq_rows) if eq_rows else Decimal("0.00")

    # Balance check
    check_ok = round(asset_total, 2) == round(liab_total + eq_total, 2)

    method_label = "Cash basis" if method == "cash" else "Accrual basis"

    # =========================================================
    # EXPORTS: CSV / EXCEL / PDF
    # =========================================================
    export = (request.GET.get("export") or "").strip().lower()
    reporting_currency = "UGX"
    company_name = "YoAccountant"
    filename_base = f"balance_sheet_{asof.strftime('%Y%m%d')}_{method}"

    def money(x: Decimal) -> str:
        try:
            return f"{Decimal(x):,.2f}"
        except Exception:
            return "0.00"

    # Build a simple flat export layout (section + amount)
    def build_export_rows():
        out = []
        out.append(("ASSETS", ""))

        out.append(("Non-current Assets", ""))
        for r in asset_nc_rows:
            out.append((r["account"], money(r["amount"])))
        out.append(("Subtotal Non-current Assets", money(asset_nc_total)))

        out.append(("Current Assets", ""))
        for r in asset_curr_rows:
            out.append((r["account"], money(r["amount"])))
        out.append(("Subtotal Current Assets", money(asset_curr_total)))

        out.append(("Total Assets", money(asset_total)))
        out.append(("", ""))

        out.append(("EQUITY & LIABILITIES", ""))

        out.append(("Equity", ""))
        for r in eq_rows:
            out.append((r["account"], money(r["amount"])))
        out.append(("Total Equity", money(eq_total)))

        out.append(("Liabilities", ""))

        out.append(("Non-current Liabilities", ""))
        for r in liab_nc_rows:
            out.append((r["account"], money(r["amount"])))
        out.append(("Subtotal Non-current Liabilities", money(liab_nc_total)))

        out.append(("Current Liabilities", ""))
        for r in liab_curr_rows:
            out.append((r["account"], money(r["amount"])))
        out.append(("Subtotal Current Liabilities", money(liab_curr_total)))

        out.append(("Total Liabilities", money(liab_total)))
        out.append(("Total Equity & Liabilities", money(liab_total + eq_total)))

        return out

    export_rows = build_export_rows()

    if export == "csv":
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        w = csv.writer(resp)

        w.writerow(["Statement of Financial Position (Balance Sheet)"])
        w.writerow(["Company", company_name])
        w.writerow(["As of", asof.strftime("%d/%m/%Y")])
        w.writerow(["Method", method_label])
        w.writerow(["Currency", reporting_currency])
        w.writerow([])

        w.writerow(["Section / Account", "Amount"])
        for a, b in export_rows:
            w.writerow([a, b])

        w.writerow([])
        w.writerow(["Check Balanced", "YES" if check_ok else "NO"])
        return resp

    if export == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        ws.append(["Statement of Financial Position (Balance Sheet)"])
        ws.append(["Company", company_name])
        ws.append(["As of", asof.strftime("%d/%m/%Y")])
        ws.append(["Method", method_label])
        ws.append(["Currency", reporting_currency])
        ws.append([])

        ws.append(["Section / Account", "Amount"])
        ws["A7"].font = ws["A7"].font.copy(bold=True)
        ws["B7"].font = ws["B7"].font.copy(bold=True)

        for a, b in export_rows:
            ws.append([a, b])

        ws.append([])
        ws.append(["Check Balanced", "YES" if check_ok else "NO"])

        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 20

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    # UPDATED: totals label + amount bold in PDF
    if export == "pdf":
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()

        story = []
        story.append(Paragraph("Statement of Financial Position (Balance Sheet)", styles["Title"]))
        story.append(Paragraph(f"Company: {company_name}", styles["Normal"]))
        story.append(Paragraph(f"As of: {asof.strftime('%d/%m/%Y')}", styles["Normal"]))
        story.append(Paragraph(f"Method: {method_label}", styles["Normal"]))
        story.append(Paragraph(f"Currency: {reporting_currency}", styles["Normal"]))
        story.append(Spacer(1, 12))

        data = [["Section / Account", "Amount"]]
        for a, b in export_rows:
            data.append([a, b])

        data.append(["", ""])
        data.append(["Check Balanced", "YES" if check_ok else "NO"])

        table = Table(data, colWidths=[360, 140])

        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9fbef")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f5132")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),

            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e7e7e7")),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),

            # (keeps the very last row bold like before)
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]

        # Make Subtotal/Total rows bold on BOTH columns
        # data row index: 0 is header, export_rows start at 1
        for i, (label, amt) in enumerate(export_rows, start=1):
            label_txt = (label or "").strip().lower()
            if label_txt.startswith("subtotal") or label_txt.startswith("total"):
                ts.append(("FONTNAME", (0, i), (1, i), "Helvetica-Bold"))
                ts.append(("BACKGROUND", (0, i), (1, i), colors.HexColor("#F8FFF9")))

        table.setStyle(TableStyle(ts))
        story.append(table)
        doc.build(story)

        pdf = buf.getvalue()
        buf.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
        return resp

    # =========================================================
    # NORMAL RENDER
    # =========================================================
    context = {
        "company_name": company_name,
        "reporting_currency": reporting_currency,
        "asof": asof,
        "method": method,
        "method_label": method_label,

        "asset_nc_rows": asset_nc_rows,
        "asset_nc_total": asset_nc_total,
        "asset_curr_rows": asset_curr_rows,
        "asset_curr_total": asset_curr_total,
        "asset_total": asset_total,

        "eq_rows": eq_rows,
        "eq_total": eq_total,

        "liab_nc_rows": liab_nc_rows,
        "liab_nc_total": liab_nc_total,
        "liab_curr_rows": liab_curr_rows,
        "liab_curr_total": liab_curr_total,
        "liab_total": liab_total,

        "check_ok": check_ok,
    }

    return render(request, "balance_sheet.html", context)


# cash flow

# =========================================================
# DETAIL TYPES 
# =========================================================
DEPOSIT_DETAIL_TYPES = [
    "Cash and Cash equivalents",
    "Bank",
]


# -----------------------------
# Helpers
# -----------------------------
def _to_decimal(x) -> Decimal:
    try:
        return Decimal(str(x or "0")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _parse_date_or_none(s):
    if not s:
        return None
    try:
        return timezone.datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _normal_side(acc) -> str:
    if getattr(acc, "level1_group", None) in ["Assets", "Expenses"]:
        return "debit"
    return "credit"


def _apply_movement(acc, running: Decimal, debit: Decimal, credit: Decimal) -> Decimal:
    side = _normal_side(acc)
    if side == "debit":
        return running + (debit - credit)
    return running + (credit - debit)


def _collect_subtree_ids(root_ids):
    """
    Expand a set of account IDs to include all descendant subaccounts.
    """
    from accounts.models import Account

    all_active = list(Account.objects.filter(is_active=True).only("id", "parent_id"))
    children_by_parent = defaultdict(list)
    for a in all_active:
        if a.parent_id:
            children_by_parent[a.parent_id].append(a.id)

    out = set()
    for rid in root_ids:
        stack = [rid]
        while stack:
            nid = stack.pop()
            if nid in out:
                continue
            out.add(nid)
            for ch in children_by_parent.get(nid, []):
                stack.append(ch)
    return list(out)


def _balance_for_account_ids(account_ids, as_of_date, strict_lt=False) -> Decimal:
    """
    Journal-only balance for a group of accounts, using normal-side rules.
    """
    from accounts.models import JournalLine

    if not account_ids:
        return Decimal("0.00")

    qs = JournalLine.objects.select_related("entry", "account").filter(account_id__in=account_ids)

    if strict_lt:
        qs = qs.filter(entry__date__lt=as_of_date)
    else:
        qs = qs.filter(entry__date__lte=as_of_date)

    running_by_acc = defaultdict(lambda: Decimal("0.00"))

    for ln in qs:
        acc = ln.account
        d = _to_decimal(getattr(ln, "debit", None))
        c = _to_decimal(getattr(ln, "credit", None))
        running_by_acc[acc.id] = _apply_movement(acc, running_by_acc[acc.id], d, c)

    total = Decimal("0.00")
    for aid in account_ids:
        total += running_by_acc.get(aid, Decimal("0.00"))

    return total.quantize(Decimal("0.01"))


def _period_net_profit(dfrom, dto) -> Decimal:
    """
    Net profit from journal:
    Income = credits - debits
    Expenses = debits - credits
    """
    qs = JournalLine.objects.select_related("entry", "account").all()
    if dfrom:
        qs = qs.filter(entry__date__gte=dfrom)
    qs = qs.filter(entry__date__lte=dto)

    income_lines = qs.filter(account__account_type__in=["OPERATING_INCOME", "INVESTING_INCOME"])
    expense_lines = qs.filter(
        account__account_type__in=[
            "OPERATING_EXPENSE", "INVESTING_EXPENSE",
            "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE",
        ]
    )

    income_agg = income_lines.aggregate(d=Sum("debit"), c=Sum("credit"))
    expense_agg = expense_lines.aggregate(d=Sum("debit"), c=Sum("credit"))

    income_total = _to_decimal(income_agg["c"]) - _to_decimal(income_agg["d"])
    expense_total = _to_decimal(expense_agg["d"]) - _to_decimal(expense_agg["c"])

    return (income_total - expense_total).quantize(Decimal("0.01"))


def _period_depreciation(dfrom, dto) -> Decimal:
    """
    Depreciation add-back:
    match depreciation expense accounts by name/detail_type.
    """
    from accounts.models import JournalLine

    qs = JournalLine.objects.select_related("entry", "account").filter(
        account__account_type__in=[
            "OPERATING_EXPENSE", "INVESTING_EXPENSE",
            "FINANCING_EXPENSE", "INCOME_TAX_EXPENSE",
        ]
    ).filter(
        Q(account__detail_type__icontains="depreciation") |
        Q(account__account_name__icontains="depreciation")
    )

    if dfrom:
        qs = qs.filter(entry__date__gte=dfrom)
    qs = qs.filter(entry__date__lte=dto)

    agg = qs.aggregate(d=Sum("debit"), c=Sum("credit"))
    dep = _to_decimal(agg["d"]) - _to_decimal(agg["c"])  # debit-normal expenses
    return dep.quantize(Decimal("0.01"))


def _first_id_or_none(qs):
    first = qs.first()
    return first.id if first else None


# -----------------------------
# MAIN CASHFLOW VIEW
# -----------------------------

def report_cashflow(request):
    from_str = request.GET.get("from")
    to_str = request.GET.get("to")
    export = (request.GET.get("export") or "").strip().lower()

    dfrom = _parse_date_or_none(from_str)
    dto = _parse_date_or_none(to_str)

    today = timezone.localdate()
    if not dto:
        dto = today

    start_date = dfrom

    # =========================================================
    # CASH & CASH EQUIVALENTS (your exact detail types)
    # include subaccounts where parent is cash/bank too
    # =========================================================
    cash_roots = Account.objects.filter(
        is_active=True,
        account_type="CURRENT_ASSET",
    ).filter(
        Q(detail_type__in=DEPOSIT_DETAIL_TYPES) |
        Q(parent__detail_type__in=DEPOSIT_DETAIL_TYPES)
    )

    cash_ids = _collect_subtree_ids(list(cash_roots.values_list("id", flat=True)))
    cash_account_id = _first_id_or_none(cash_roots)

    # =========================================================
    # BUCKETS (keep your HTML row names)
    # =========================================================

    # "Change in Accounts Receivable":
    # CURRENT ASSETS excluding cash/bank
    ar_roots = Account.objects.filter(
        is_active=True,
        account_type="CURRENT_ASSET",
    ).exclude(id__in=cash_ids)

    # "Change in Inventory"
    inv_roots = Account.objects.filter(
        is_active=True,
        account_type__in=["CURRENT_ASSET", "NON_CURRENT_ASSET"]
    ).filter(
        Q(detail_type__icontains="inventory") |
        Q(account_name__icontains="inventory")
    )

    # "Change in Accounts Payable":
    # CURRENT LIABILITIES excluding overdrafts
    ap_roots = Account.objects.filter(
        is_active=True,
        account_type="CURRENT_LIABILITY",
    ).exclude(
        Q(detail_type__icontains="overdraft") |
        Q(account_name__icontains="overdraft")
    )

    # "Change in Fixed/Other Assets":
    fa_roots = Account.objects.filter(
        is_active=True,
        account_type="NON_CURRENT_ASSET"
    )

    # "Change in Loans":
    loan_roots = Account.objects.filter(
        is_active=True,
        account_type="NON_CURRENT_LIABILITY"
    )

    # "Change in Equity":
    equity_roots = Account.objects.filter(
        is_active=True,
        account_type="OWNER_EQUITY"
    )

    # Expand subtrees (include children)
    ar_ids = _collect_subtree_ids(list(ar_roots.values_list("id", flat=True)))
    inv_ids = _collect_subtree_ids(list(inv_roots.values_list("id", flat=True)))
    ap_ids = _collect_subtree_ids(list(ap_roots.values_list("id", flat=True)))
    fa_ids = _collect_subtree_ids(list(fa_roots.values_list("id", flat=True)))
    loan_ids = _collect_subtree_ids(list(loan_roots.values_list("id", flat=True)))
    eq_ids = _collect_subtree_ids(list(equity_roots.values_list("id", flat=True)))

    # IDs for GL drill links (your HTML expects these)
    ar_account_id = _first_id_or_none(ar_roots)
    inv_account_id = _first_id_or_none(inv_roots)
    ap_account_id = _first_id_or_none(ap_roots)
    fa_account_id = _first_id_or_none(fa_roots)
    loans_account_id = _first_id_or_none(loan_roots)
    equity_account_id = _first_id_or_none(equity_roots)

    # =========================================================
    # PROFIT + DEPRECIATION
    # =========================================================
    net_profit = _period_net_profit(dfrom, dto)
    depreciation = _period_depreciation(dfrom, dto)

    # =========================================================
    # OPENING BALANCES (strictly before start_date)
    # =========================================================
    if start_date:
        cash_start = _balance_for_account_ids(cash_ids, start_date, strict_lt=True)

        ar_start = _balance_for_account_ids(ar_ids, start_date, strict_lt=True)
        inv_start = _balance_for_account_ids(inv_ids, start_date, strict_lt=True)
        ap_start = _balance_for_account_ids(ap_ids, start_date, strict_lt=True)
        fa_start = _balance_for_account_ids(fa_ids, start_date, strict_lt=True)
        loans_start = _balance_for_account_ids(loan_ids, start_date, strict_lt=True)
        equity_start = _balance_for_account_ids(eq_ids, start_date, strict_lt=True)
    else:
        cash_start = Decimal("0.00")
        ar_start = inv_start = ap_start = fa_start = loans_start = equity_start = Decimal("0.00")

    # =========================================================
    # CLOSING BALANCES (<= dto)
    # =========================================================
    cash_end = _balance_for_account_ids(cash_ids, dto)

    ar_end = _balance_for_account_ids(ar_ids, dto)
    inv_end = _balance_for_account_ids(inv_ids, dto)
    ap_end = _balance_for_account_ids(ap_ids, dto)
    fa_end = _balance_for_account_ids(fa_ids, dto)
    loans_end = _balance_for_account_ids(loan_ids, dto)
    equity_end = _balance_for_account_ids(eq_ids, dto)

    # =========================================================
    # RAW DELTAS (end - start)
    # =========================================================
    raw_delta_ar = (ar_end - ar_start).quantize(Decimal("0.01"))
    raw_delta_inv = (inv_end - inv_start).quantize(Decimal("0.01"))
    raw_delta_ap = (ap_end - ap_start).quantize(Decimal("0.01"))
    raw_delta_fa = (fa_end - fa_start).quantize(Decimal("0.01"))
    raw_delta_loans = (loans_end - loans_start).quantize(Decimal("0.01"))
    raw_delta_equity = (equity_end - equity_start).quantize(Decimal("0.01"))

    # =========================================================
    # CASHFLOW PRESENTATION SIGN RULES (client requirement)
    # Assets: increase negative, decrease positive
    # Liabilities/Equity: increase positive, decrease negative
    # =========================================================
    delta_ar = (-raw_delta_ar).quantize(Decimal("0.01"))        # asset
    delta_inv = (-raw_delta_inv).quantize(Decimal("0.01"))      # asset
    delta_fa = (-raw_delta_fa).quantize(Decimal("0.01"))        # asset

    delta_ap = (raw_delta_ap).quantize(Decimal("0.01"))         # liability
    delta_loans = (raw_delta_loans).quantize(Decimal("0.01"))   # liability
    delta_equity = (raw_delta_equity).quantize(Decimal("0.01")) # equity

    # =========================================================
    # CASHFLOW (INDIRECT METHOD)
    # =========================================================
    cash_from_ops = (
        net_profit
        + depreciation
        + delta_ar
        + delta_inv
        + delta_ap
    ).quantize(Decimal("0.01"))

    cash_from_investing = (delta_fa).quantize(Decimal("0.01"))
    cash_from_financing = (delta_loans + delta_equity).quantize(Decimal("0.01"))

    net_change = (cash_from_ops + cash_from_investing + cash_from_financing).quantize(Decimal("0.01"))
    recon_ok = (cash_start + net_change).quantize(Decimal("0.01")) == cash_end.quantize(Decimal("0.01"))

    company_name = "YoAccountant"
    reporting_currency = "UGX"

    # =========================================================
    # EXPORT ROWS + BOLD ROWS LIST
    # =========================================================
    period_label = f"{dfrom.strftime('%Y-%m-%d') if dfrom else 'As_of_date'}_to_{dto.strftime('%Y-%m-%d') if dto else 'As_of_date'}"

    # rows whose LABEL must be bold
    BOLD_LABELS = {
        "Net Cash from Operating Activities",
        "Net Cash from Investing Activities",
        "Net Cash from Financing Activities",
        "Net Change in Cash",
    }

    export_rows = [
        ("Cash Flows from Operating Activities", "", ""),
        ("Net Profit", "", net_profit),
        ("Depreciation", "", depreciation),
        ("Change in Accounts Receivable", "", delta_ar),
        ("Change in Inventory", "", delta_inv),
        ("Change in Accounts Payable", "", delta_ap),
        ("Net Cash from Operating Activities", "", cash_from_ops),
        ("", "", ""),

        ("Cash Flows from Investing Activities", "", ""),
        ("Change in Fixed/Other Assets", "", delta_fa),
        ("Net Cash from Investing Activities", "", cash_from_investing),
        ("", "", ""),

        ("Cash Flows from Financing Activities", "", ""),
        ("Change in Loans", "", delta_loans),
        ("Change in Equity", "", delta_equity),
        ("Net Cash from Financing Activities", "", cash_from_financing),
        ("", "", ""),

        ("Net Change in Cash", "", ""),
        ("Net Change in Cash", "", net_change),
        ("Cash at Start", "", cash_start),
        ("Cash at End", "", cash_end),
        ("Reconciliation OK", "", "YES" if recon_ok else "NO"),
    ]

    def _amt_to_str(x):
        if isinstance(x, Decimal):
            return f"{x:,.2f}"
        if x is None:
            return ""
        return str(x)

    # -----------------------------
    # Exports (CSV / Excel / PDF)
    # -----------------------------
    if export == "csv":
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="cashflow_{period_label}.csv"'
        w = csv.writer(resp)

        w.writerow(["Statement of Cash Flows", company_name])
        w.writerow(["Period", f"{dfrom.strftime('%d/%m/%Y') if dfrom else 'As of date'} – {dto.strftime('%d/%m/%Y') if dto else 'As of date'}"])
        w.writerow(["Reporting Currency", reporting_currency])
        w.writerow([])

        for label, _cur, amt in export_rows:
            if label == "" and _cur == "" and amt == "":
                w.writerow([])
            else:
                if _cur == "" and (amt == "" or isinstance(amt, str)):
                    w.writerow([label, ""])
                else:
                    # CSV can't bold, but it keeps the same labels and signs
                    w.writerow([label, _amt_to_str(amt)])
        return resp

    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Cashflow"

        ws.append(["Statement of Cash Flows", company_name])
        ws.append(["Period", f"{dfrom.strftime('%d/%m/%Y') if dfrom else 'As of date'} – {dto.strftime('%d/%m/%Y') if dto else 'As of date'}"])
        ws.append(["Reporting Currency", reporting_currency])
        ws.append([])

        bold_font = Font(bold=True)

        for label, _cur, amt in export_rows:
            if label == "" and _cur == "" and amt == "":
                ws.append([])
                continue

            # headings
            if _cur == "" and (amt == "" or isinstance(amt, str)):
                row = ws.max_row + 1
                ws.append([label, ""])
                ws.cell(row=row, column=1).font = bold_font
                continue

            # normal numeric rows
            row = ws.max_row + 1
            ws.append([label, float(amt) if isinstance(amt, Decimal) else _amt_to_str(amt)])

            # bold subtotal lines
            if label in BOLD_LABELS:
                ws.cell(row=row, column=1).font = bold_font
                ws.cell(row=row, column=2).font = bold_font

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="cashflow_{period_label}.xlsx"'
        return resp

    if export == "pdf":
        bio = BytesIO()
        c = canvas.Canvas(bio, pagesize=A4)
        width, height = A4

        y = height - 50
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, "Statement of Cash Flows")
        y -= 18

        c.setFont("Helvetica", 10)
        c.drawString(50, y, f"Company: {company_name}")
        y -= 14
        c.drawString(50, y, f"Period: {dfrom.strftime('%d/%m/%Y') if dfrom else 'As of date'} – {dto.strftime('%d/%m/%Y') if dto else 'As of date'}")
        y -= 14
        c.drawString(50, y, f"Reporting Currency: {reporting_currency}")
        y -= 16

        y -= 6
        c.line(50, y, width - 50, y)
        y -= 16

        for label, _cur, amt in export_rows:
            if label == "" and _cur == "" and amt == "":
                y -= 10
                continue

            if y < 60:
                c.showPage()
                y = height - 60
                c.setFont("Helvetica", 10)
                c.line(50, y, width - 50, y)
                y -= 16

            # SECTION HEADINGS
            if _cur == "" and (amt == "" or isinstance(amt, str) and label and amt == ""):
                c.setFont("Helvetica-Bold", 10)
                c.drawString(50, y, str(label))
                c.setFont("Helvetica", 10)
                y -= 14
                continue

            # YES/NO row
            if isinstance(amt, str) and label == "Reconciliation OK":
                c.setFont("Helvetica", 10)
                c.drawString(50, y, str(label))
                c.drawRightString(width - 50, y, _amt_to_str(amt))
                y -= 14
                continue

            # NORMAL OR SUBTOTAL ROW
            is_bold = label in BOLD_LABELS
            c.setFont("Helvetica-Bold" if is_bold else "Helvetica", 10)

            c.drawString(50, y, str(label))
            c.drawRightString(width - 50, y, _amt_to_str(amt))

            # reset
            if is_bold:
                c.setFont("Helvetica", 10)

            y -= 14

        c.save()
        bio.seek(0)

        resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="cashflow_{period_label}.pdf"'
        return resp

    # -----------------------------
    # Render (your cashflow.html unchanged)
    # -----------------------------
    context = {
        "company_name": company_name,
        "reporting_currency": reporting_currency,
        "dfrom": dfrom,
        "dto": dto,

        "net_profit": net_profit,
        "delta_ar": delta_ar,
        "delta_inv": delta_inv,
        "delta_ap": delta_ap,
        "delta_fa": delta_fa,
        "delta_loans": delta_loans,
        "delta_equity": delta_equity,

        "cash_from_ops": cash_from_ops,
        "cash_from_investing": cash_from_investing,
        "cash_from_financing": cash_from_financing,

        "cash_start": cash_start,
        "net_change": net_change,
        "cash_end": cash_end,
        "recon_ok": recon_ok,

        "cash_account_id": cash_account_id,
        "ar_account_id": ar_account_id,
        "inv_account_id": inv_account_id,
        "ap_account_id": ap_account_id,
        "fa_account_id": fa_account_id,
        "loans_account_id": loans_account_id,
        "equity_account_id": equity_account_id,
    }

    return render(request, "cashflow.html", context)

# -----------------------------
# Helpers 
# -----------------------------

def _dec(x) -> Decimal:
    try:
        return Decimal(str(x or "0"))
    except Exception:
        return Decimal("0")


def _as_date(dt):
    if not dt:
        return None
    try:
        return dt.date()
    except Exception:
        return dt


def _bucket(due_date, today):
    """
    Safe keys for templates:
      current, b1_30, b31_60, b61_90, b90_plus
    """
    if not due_date:
        return "current"
    days = (today - due_date).days
    if days <= 0:
        return "current"
    if 1 <= days <= 30:
        return "b1_30"
    if 31 <= days <= 60:
        return "b31_60"
    if 61 <= days <= 90:
        return "b61_90"
    return "b90_plus"
def _bucket_label(key: str) -> str:
    return {
        "current": "Current",
        "b1_30": "1–30",
        "b31_60": "31–60",
        "b61_90": "61–90",
        "b90_plus": "90+",
    }.get(key, key)

def aging_report(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    invoices = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(outstanding_db=F("total_due_dec") - F("total_paid"))
        .only("id", "customer_id", "due_date", "date_created", "total_due")
    )

    if customer_id.isdigit():
        invoices = invoices.filter(customer_id=int(customer_id))

    rows_map = {}
    grand = {
        "current": Decimal("0.00"),
        "b1_30": Decimal("0.00"),
        "b31_60": Decimal("0.00"),
        "b61_90": Decimal("0.00"),
        "b90_plus": Decimal("0.00"),
        "total": Decimal("0.00"),
    }

    for inv in invoices:
        bal = _dec(inv.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(inv.due_date) or _as_date(inv.date_created)
        key = _bucket(due, today)

        if bucket_filter and bucket_filter != key:
            continue

        cust = inv.customer
        cid = cust.id

        if cid not in rows_map:
            rows_map[cid] = {
                "customer": cust,
                "current": Decimal("0.00"),
                "b1_30": Decimal("0.00"),
                "b31_60": Decimal("0.00"),
                "b61_90": Decimal("0.00"),
                "b90_plus": Decimal("0.00"),
                "total": Decimal("0.00"),
            }

        rows_map[cid][key] += bal
        rows_map[cid]["total"] += bal

        grand[key] += bal
        grand["total"] += bal

    rows = list(rows_map.values())
    rows.sort(key=lambda r: (r["customer"].customer_name or "").lower())

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                r["current"], r["b1_30"], r["b31_60"], r["b61_90"], r["b90_plus"], r["total"]
            ])
        # grand total row
        data_rows.append([
            "GRAND TOTAL",
            grand["current"], grand["b1_30"], grand["b31_60"], grand["b61_90"], grand["b90_plus"], grand["total"]
        ])

        subtitle = f"As of {today}"
        if exp == "excel":
            return export_excel_simple(_export_filename("ar_aging_summary", "xlsx"), headers, data_rows, sheet_name="AR Aging Summary")
        return export_pdf_table(_export_filename("ar_aging_summary", "pdf"), "Accounts Receivable Aging Summary", subtitle, headers, data_rows)

    customers = Newcustomer.objects.order_by("customer_name")

    return render(request, "ar_aging_report.html", {
        "today": today,
        "rows": rows,
        "grand": grand,
        "customers": customers,
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
    })

def aging_report_detail(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_due_dec") - F("total_paid"),
                output_field=dec_out
            )
        )
        .only("id", "customer_id", "date_created", "due_date", "total_due")
        .order_by("customer__customer_name", "due_date", "date_created", "id")
    )

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))

    rows = []
    for inv in qs:
        bal = _dec(inv.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(inv.due_date) or _as_date(inv.date_created)
        key = _bucket(due, today)
        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        rows.append({
            "customer": inv.customer,
            "invoice_id": inv.id,
            "invoice_date": _as_date(inv.date_created),
            "due_date": due,
            "days_overdue": days_overdue,
            "total_due": _dec(inv.total_due),
            "amount_paid": _dec(inv.total_paid),
            "balance": bal,
            "bucket": key,
            "bucket_label": _bucket_label(key),
        })

    totals = {"total": Decimal("0.00")}
    for r in rows:
        totals["total"] += r["balance"]

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Invoice #", "Invoice Date", "Due Date", "Bucket", "Days", "Total", "Paid", "Balance"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                f"{r['invoice_id']:03d}",
                r["invoice_date"],
                r["due_date"],
                r["bucket_label"],
                max(r["days_overdue"], 0),
                r["total_due"],
                r["amount_paid"],
                r["balance"],
            ])
        data_rows.append(["TOTAL OUTSTANDING", "", "", "", "", "", "", "", totals["total"]])

        subtitle = f"As of {today}"
        if exp == "excel":
            return export_excel_simple(_export_filename("ar_aging_detail", "xlsx"), headers, data_rows, sheet_name="AR Aging Detail")
        return export_pdf_table(_export_filename("ar_aging_detail", "pdf"), "Accounts Receivable Aging Detail", subtitle, headers, data_rows)

    customers = Newcustomer.objects.order_by("customer_name")

    return render(request, "ar_aging_detail_report.html", {
        "today": today,
        "rows": rows,
        "totals": {"total": totals["total"]},
        "customers": customers,
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
    })


def _customer_model():
    return Newinvoice._meta.get_field("customer").remote_field.model


def _customers_qs():
    Customer = _customer_model()
    # assumes customer_name exists (in your models it does)
    try:
        return Customer.objects.order_by("customer_name")
    except Exception:
        return Customer.objects.all()

def _customer_model():
    # gets the customer model attached to Newinvoice.customer FK (safe even if customer lives in another app)
    return Newinvoice._meta.get_field("customer").remote_field.model

def aging_report_customer(request, customer_id: int):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    Customer = _customer_model()
    customer = get_object_or_404(Customer, pk=customer_id)

    bucket_filter = (request.GET.get("bucket") or "").strip()

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .filter(customer_id=customer_id)
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_due_dec") - F("total_paid"),
                output_field=dec_out
            )
        )
        .only("id", "customer_id", "date_created", "due_date", "total_due")
        .order_by("due_date", "date_created", "id")
    )

    summary = {
        "current": Decimal("0.00"),
        "b1_30": Decimal("0.00"),
        "b31_60": Decimal("0.00"),
        "b61_90": Decimal("0.00"),
        "b90_plus": Decimal("0.00"),
        "total": Decimal("0.00"),
    }

    rows = []
    for inv in qs:
        bal = _dec(inv.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(inv.due_date) or _as_date(inv.date_created)
        key = _bucket(due, today)

        summary[key] += bal
        summary["total"] += bal

        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        rows.append({
            "invoice_id": inv.id,
            "invoice_date": _as_date(inv.date_created),
            "due_date": due,
            "bucket": key,
            "bucket_label": _bucket_label(key),
            "days_overdue": days_overdue,
            "total_due": _dec(inv.total_due),
            "amount_paid": _dec(inv.total_paid),
            "balance": bal,
        })

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        # 1) Summary block
        sum_headers = ["Customer", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        sum_rows = [[
            getattr(customer, "customer_name", str(customer)),
            summary["current"], summary["b1_30"], summary["b31_60"], summary["b61_90"], summary["b90_plus"], summary["total"]
        ]]

        # 2) Detail block (invoice rows)
        det_headers = ["Invoice #", "Invoice Date", "Due Date", "Bucket", "Days", "Total", "Paid", "Balance"]
        det_rows = []
        for r in rows:
            det_rows.append([
                f"{r['invoice_id']:03d}",
                r["invoice_date"],
                r["due_date"],
                r["bucket_label"],
                max(r["days_overdue"], 0),
                r["total_due"],
                r["amount_paid"],
                r["balance"],
            ])

        # Combine into one export table: Summary row, blank row, then details
        headers = sum_headers
        data_rows = sum_rows + [["", "", "", "", "", "", ""]]  # spacer row for excel
        data_rows += [det_headers]  # put detail header as a row
        data_rows += det_rows

        subtitle = f"As of {today}"
        if bucket_filter:
            subtitle += f" | Bucket: {_bucket_label(bucket_filter)}"

        if exp == "excel":
            return export_excel_simple(_export_filename("ar_aging_customer", "xlsx"), headers, data_rows, sheet_name="AR Aging Customer")
        return export_pdf_table(_export_filename("ar_aging_customer", "pdf"), f"A/R Aging — {getattr(customer,'customer_name','Customer')}", subtitle, headers, data_rows)

    return render(request, "ar_aging_customer_report.html", {
        "today": today,
        "customer": customer,
        "summary": summary,
        "rows": rows,
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
    })

# ---------------------------------------------------------
# 1) OPEN INVOICES REPORT
# ---------------------------------------------------------
def open_invoices_report(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_due_dec") - F("total_paid"),
                output_field=dec_out
            )
        )
        .only("id", "customer_id", "date_created", "due_date", "total_due")
        .order_by("customer__customer_name", "due_date", "date_created", "id")
    )

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))

    rows = []
    totals = {"total": Decimal("0.00")}

    for inv in qs:
        bal = _dec(inv.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(inv.due_date) or _as_date(inv.date_created)
        key = _bucket(due, today)

        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        rows.append({
            "customer": inv.customer,
            "invoice_id": inv.id,
            "invoice_date": _as_date(inv.date_created),
            "due_date": due,
            "bucket": key,
            "bucket_label": _bucket_label(key),
            "days_overdue": days_overdue,
            "total_due": _dec(inv.total_due),
            "amount_paid": _dec(inv.total_paid),
            "balance": bal,
        })
        totals["total"] += bal

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Invoice #", "Invoice Date", "Due Date", "Bucket", "Days", "Total", "Paid", "Balance"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                f"{r['invoice_id']:03d}",
                r["invoice_date"],
                r["due_date"],
                r["bucket_label"],
                max(r["days_overdue"], 0),
                r["total_due"],
                r["amount_paid"],
                r["balance"],
            ])
        data_rows.append(["TOTAL OUTSTANDING", "", "", "", "", "", "", "", totals["total"]])

        subtitle = f"As of {today}"
        if exp == "excel":
            return export_excel_simple(_export_filename("open_invoices", "xlsx"), headers, data_rows, sheet_name="Open Invoices")
        return export_pdf_table(_export_filename("open_invoices", "pdf"), "Open Invoices", subtitle, headers, data_rows)

    return render(request, "open_invoices_report.html", {
        "today": today,
        "rows": rows,
        "totals": totals,
        "customers": _customers_qs(),
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
    })


# ---------------------------------------------------------
# 2) CUSTOMER BALANCES REPORT
# ---------------------------------------------------------
def customer_balances_report(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_due_dec") - F("total_paid"),
                output_field=dec_out
            )
        )
        .only("id", "customer_id", "date_created", "due_date", "total_due")
    )

    cust_map = {}
    grand = {"current": Decimal("0.00"), "overdue": Decimal("0.00"), "total": Decimal("0.00")}

    for inv in qs:
        bal = _dec(inv.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(inv.due_date) or _as_date(inv.date_created)
        overdue = bool(due and due < today)

        cid = inv.customer.id
        if cid not in cust_map:
            cust_map[cid] = {
                "customer": inv.customer,
                "current": Decimal("0.00"),
                "overdue": Decimal("0.00"),
                "total": Decimal("0.00"),
            }

        if overdue:
            cust_map[cid]["overdue"] += bal
            grand["overdue"] += bal
        else:
            cust_map[cid]["current"] += bal
            grand["current"] += bal

        cust_map[cid]["total"] += bal
        grand["total"] += bal

    rows = list(cust_map.values())
    rows.sort(key=lambda r: (getattr(r["customer"], "customer_name", "") or "").lower())

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Current", "Overdue", "Total"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                r["current"],
                r["overdue"],
                r["total"],
            ])
        data_rows.append(["GRAND TOTAL", grand["current"], grand["overdue"], grand["total"]])

        subtitle = f"As of {today}"
        if exp == "excel":
            return export_excel_simple(_export_filename("customer_balances", "xlsx"), headers, data_rows, sheet_name="Customer Balances")
        return export_pdf_table(_export_filename("customer_balances", "pdf"), "Customer Balances", subtitle, headers, data_rows)

    return render(request, "customer_balances_report.html", {
        "today": today,
        "rows": rows,
        "grand": grand,
    })


# ---------------------------------------------------------
# 3) INVOICE LIST REPORT (all invoices)
# ---------------------------------------------------------
def invoice_list_report(request):
    today = timezone.localdate()
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    customer_id = (request.GET.get("customer") or "").strip()
    status = (request.GET.get("status") or "").strip()  # all|paid|unpaid|overdue

    qs = (
        Newinvoice.objects
        .select_related("customer")
        .annotate(
            total_due_dec=Cast(F("total_due"), output_field=dec_out),
            total_paid=Coalesce(
                Sum("payments_applied__amount_paid", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_due_dec") - F("total_paid"),
                output_field=dec_out
            )
        )
        .only("id", "customer_id", "date_created", "due_date", "total_due")
        .order_by("-date_created", "-id")
    )

    if customer_id.isdigit():
        qs = qs.filter(customer_id=int(customer_id))

    rows = []
    for inv in qs:
        bal = _dec(inv.outstanding_db)
        due = _as_date(inv.due_date) or _as_date(inv.date_created)

        is_paid = bal <= Decimal("0.00001")
        is_overdue = bool((not is_paid) and due and due < today)

        if status == "paid" and not is_paid:
            continue
        if status == "unpaid" and is_paid:
            continue
        if status == "overdue" and not is_overdue:
            continue

        rows.append({
            "customer": inv.customer,
            "invoice_id": inv.id,
            "invoice_date": _as_date(inv.date_created),
            "due_date": due,
            "total_due": _dec(inv.total_due),
            "amount_paid": _dec(inv.total_paid),
            "balance": bal if bal > 0 else Decimal("0.00"),
            "status": "PAID" if is_paid else ("OVERDUE" if is_overdue else "OPEN"),
        })

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Invoice #", "Invoice Date", "Due Date", "Status", "Total", "Paid", "Balance"]
        data_rows = []
        total_balance = Decimal("0.00")
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                f"{r['invoice_id']:03d}",
                r["invoice_date"],
                r["due_date"],
                r["status"],
                r["total_due"],
                r["amount_paid"],
                r["balance"],
            ])
            total_balance += _dec(r["balance"])
        data_rows.append(["TOTAL OUTSTANDING", "", "", "", "", "", "", total_balance])

        subtitle = f"As of {today}"
        if status:
            subtitle += f" | Status: {status}"
        if exp == "excel":
            return export_excel_simple(_export_filename("invoice_list", "xlsx"), headers, data_rows, sheet_name="Invoice List")
        return export_pdf_table(_export_filename("invoice_list", "pdf"), "Invoice List", subtitle, headers, data_rows)

    return render(request, "invoice_list_report.html", {
        "today": today,
        "rows": rows,
        "customers": _customers_qs(),
        "selected_customer": int(customer_id) if customer_id.isdigit() else "",
        "selected_status": status,
        "status_choices": [
            ("", "All"),
            ("paid", "Paid"),
            ("unpaid", "Unpaid"),
            ("overdue", "Overdue"),
        ],
    })


# ---------------------------------------------------------
# 4) COLLECTIONS REPORT (payments + applied + unapplied)
# ---------------------------------------------------------
def collections_report(request):
    today = timezone.localdate()

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    # default current month
    if not date_from and not date_to:
        first = today.replace(day=1)
        date_from = str(first)
        date_to = str(today)

    qs = Payment.objects.select_related("customer").all().order_by("-payment_date", "-id")

    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    dec_out = DecimalField(max_digits=18, decimal_places=2)
    qs = qs.annotate(
        total_applied=Coalesce(
            Sum("applied_invoices__amount_paid", output_field=dec_out),
            Value(Decimal("0.00"), output_field=dec_out),
            output_field=dec_out
        )
    )

    rows = []
    totals = {"received": Decimal("0.00"), "applied": Decimal("0.00"), "unapplied": Decimal("0.00")}

    for p in qs:
        received = _dec(p.amount_received)
        applied = _dec(p.total_applied)
        unapplied = _dec(getattr(p, "unapplied_amount", Decimal("0.00")))

        rows.append({
            "customer": p.customer,
            "date": p.payment_date,
            "ref": p.reference_no,
            "method": p.payment_method,
            "received": received,
            "applied": applied,
            "unapplied": unapplied,
        })

        totals["received"] += received
        totals["applied"] += applied
        totals["unapplied"] += unapplied

    # EXPORT
    exp = _export_wants(request)
    if exp in ("excel", "pdf"):
        headers = ["Customer", "Payment Date", "Reference", "Method", "Received", "Applied", "Unapplied"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r["customer"].customer_name,
                r["date"],
                r["ref"],
                r["method"],
                r["received"],
                r["applied"],
                r["unapplied"],
            ])
        data_rows.append(["TOTALS", "", "", "", totals["received"], totals["applied"], totals["unapplied"]])

        subtitle = f"From {date_from} to {date_to}"
        if exp == "excel":
            return export_excel_simple(_export_filename("collections_report", "xlsx"), headers, data_rows, sheet_name="Collections")
        return export_pdf_table(_export_filename("collections_report", "pdf"), "Collections Report", subtitle, headers, data_rows)

    return render(request, "collections_report.html", {
        "today": today,
        "rows": rows,
        "totals": totals,
        "date_from": date_from,
        "date_to": date_to,
    })
def _export_wants(request) -> str:
    """
    Returns 'excel', 'pdf', or '' (no export).
    """
    return (request.GET.get("export") or "").strip().lower()


def _export_filename(prefix: str, ext: str) -> str:
    return f"{prefix}.{ext}"


def export_excel_simple(filename: str, headers: list[str], rows: list[list], sheet_name="Report") -> HttpResponse:
    """
    Excel export using openpyxl.
    pip install openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # header row
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for r in rows:
        ws.append([_safe_excel_value(v) for v in r])

    # autosize columns (simple)
    for col in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _safe_excel_value(v):
    # openpyxl doesn't like Decimal sometimes; convert safely
    if isinstance(v, Decimal):
        return float(v)
    return v


def export_pdf_table(filename: str, title: str, subtitle: str, headers: list[str], rows: list[list]) -> HttpResponse:
    """
    PDF export using reportlab.
    pip install reportlab
    """
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)

    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(title, styles["Title"]))
    if subtitle:
        elems.append(Paragraph(subtitle, styles["Normal"]))
    elems.append(Spacer(1, 12))

    data = [headers] + [[_safe_pdf_value(x) for x in r] for r in rows]
    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8fbef")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))

    elems.append(table)
    doc.build(elems)

    pdf = bio.getvalue()
    bio.close()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _safe_pdf_value(v):
    if v is None:
        return ""
    if isinstance(v, Decimal):
        # show clean numbers in pdf
        return f"{v:.2f}"
    return str(v)
