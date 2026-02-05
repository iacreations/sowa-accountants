# Create your views here.
from decimal import Decimal
from urllib.parse import urlencode
from django.db.models import Sum, Value, F, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.db.models.functions import Coalesce
from django.db.models import Sum, Value, DecimalField, Prefetch
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_GET
from django.urls import reverse
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from .models import (Expense, ExpenseCategoryLine, ExpenseItemLine,ColumnPreference,Bill, BillCategoryLine, BillItemLine,PurchaseOrder, PurchaseOrderLine,Cheque, ChequeCategoryLine, ChequeItemLine,SupplierCredit,SupplierCreditLine,PayDownCredit,CreditCardCredit,CreditCardCreditCategoryLine,CreditCardCreditItemLine,ChequeBillLine,ChequeOpenBalanceLine,SupplierRefund)
from sowaf.models import Newcustomer, Newsupplier
from accounts.models import Account, JournalEntry, JournalLine
from accounts.utils import deposit_accounts_qs, expense_accounts_qs
from collections import defaultdict
from django.db import transaction
from django.views.decorators.http import require_http_methods
from inventory.models import Product,Pclass
from .utils import (generate_unique_ref_no,_save_cheque_bill_allocations,bankish_q)
from inventory.services import rebuild_movements_for_bill, rebuild_movements_for_expense
# Expenses view

DEFAULT_ACCOUNTS_COL_PREFS = {
    "payment_date": True,
    "payee_name": True,
    "payee_supplier": True,
    "payment_account": True,
    "payment_method": True,
    "ref_no": True,
    "memo": True,
    "attachments": True,  # keep actions togglable too
}
def _cat_label_from_lines(cat_lines, item_lines):
    """
    Build a single Category label for the row:
      - '--Split--' if more than one line
      - category account name for a single category line
      - product name for a single item line
      - '' if no lines
    Assumes related lines were prefetched.
    """
    cats = list(cat_lines) if hasattr(cat_lines, "__iter__") else list(cat_lines.all())
    items = list(item_lines) if hasattr(item_lines, "__iter__") else list(item_lines.all())
    n = len(cats) + len(items)
    if n == 0:
        return ""
    if n > 1:
        return "--Split--"
    if len(cats) == 1:
        return getattr(cats[0].category, "account_name", "")
    if len(items) == 1:
        return getattr(items[0].product, "name", "")
    return ""

# generate a po number
def generate_unique_po_no(prefix="PO"):
    """
    Generate a unique Purchase Order number, e.g. PO00000001.
    """
    last = PurchaseOrder.objects.order_by("-id").first()
    last_num = 0
    if last and last.po_number and last.po_number.startswith(prefix):
        tail = last.po_number[len(prefix):]
        if tail.isdigit():
            last_num = int(tail)
    next_num = last_num + 1
    return f"{prefix}{next_num:08d}"



# ----------------------------
# HELPERS: BILL BALANCE + SUPPLIER OPEN BALANCE
# ----------------------------

def _bill_balance(bill: "Bill", exclude_cheque_id: int | None = None) -> Decimal:
    """
    Bill balance = total_amount - sum(applied via cheques).
    Optionally excludes one cheque (useful during edit).
    """
    total = _dec(bill.total_amount)

    qs = ChequeBillLine.objects.filter(bill=bill)
    if exclude_cheque_id:
        qs = qs.exclude(cheque_id=exclude_cheque_id)

    applied = qs.aggregate(s=Sum("amount_applied"))["s"] or Decimal("0.00")

    bal = total - _dec(applied)
    return bal if bal > 0 else Decimal("0.00")


# def _supplier_open_balance(supplier_id: int | None, exclude_cheque_id: int | None = None) -> Decimal:
#     """
#     Total open balance across ALL supplier bills.
#     Optionally excludes allocations from one cheque (edit-safe).
#     """
#     if not supplier_id:
#         return Decimal("0.00")

#     bills = Bill.objects.filter(supplier_id=supplier_id)
#     total_open = Decimal("0.00")

#     for b in bills:
#         total_open += _bill_balance(b, exclude_cheque_id=exclude_cheque_id)

#     return total_open


def _account_credit_balance(account: "Account") -> Decimal:
    """
    Returns credit balance for an account:
      opening_balance + SUM(credit - debit)
    If negative, return 0 for our purpose here.
    """
    opening = Decimal(str(getattr(account, "opening_balance", 0) or "0"))
    agg = (
        JournalLine.objects
        .filter(account=account)
        .aggregate(
            d=Sum("debit"),
            c=Sum("credit"),
        )
    )
    deb = Decimal(str(agg["d"] or "0"))
    cred = Decimal(str(agg["c"] or "0"))

    bal = opening + (cred - deb)
    return bal if bal > 0 else Decimal("0.00")


def _supplier_open_balance_amount(supplier: "Newsupplier") -> Decimal:
    """
    Open balance = Supplier A/P subaccount credit balance - total unpaid bill balances
    (Anything in supplier A/P not represented by open bill balances is considered "open balance".)
    """
    if not supplier:
        return Decimal("0.00")

    supplier_acc = _get_or_create_supplier_ap_subaccount(supplier)

    supplier_ap_bal = _account_credit_balance(supplier_acc)  # credit balance in A/P

    # total unpaid bills for this supplier
    total_unpaid_bills = Decimal("0.00")
    for b in Bill.objects.filter(supplier=supplier):
        total_unpaid_bills += _bill_balance(b)

    open_bal = supplier_ap_bal - total_unpaid_bills
    return open_bal if open_bal > 0 else Decimal("0.00")




def _get_or_create_named_account(account_name: str, account_type: str, detail_type: str = "") -> Account:
    acc = Account.objects.filter(account_name=account_name, is_active=True).first()
    if acc:
        return acc

    return Account.objects.create(
        account_name=account_name,
        account_type=account_type,              # uses your codes e.g CURRENT_ASSET, CURRENT_LIABILITY
        detail_type=detail_type or None,
        is_active=True,
        opening_balance=Decimal("0.00"),
        as_of=timezone.localdate(),
    )

def _get_supplier_advance_account() -> "Account":
    acc = Account.objects.filter(account_name__iexact="Supplier Advances", is_active=True).first()
    if acc:
        return acc

    return Account.objects.create(
        account_name="Supplier Advances",
        account_type="CURRENT_LIABILITY",
        detail_type="Supplier Advances",
        is_active=True,
    )


def _supplier_prepayment_balance(supplier) -> Decimal:
    if not supplier:
        return Decimal("0.00")

    adv = _get_supplier_advance_account()
    agg = (
        JournalLine.objects
        .filter(account=adv, supplier=supplier)
        .aggregate(
            d=Coalesce(Sum("debit"), Value(Decimal("0.00"))),
            c=Coalesce(Sum("credit"), Value(Decimal("0.00"))),
        )
    )
    debit = Decimal(str(agg["d"] or "0.00"))
    credit = Decimal(str(agg["c"] or "0.00"))

    # Asset normal balance: DEBIT
    bal = debit - credit
    return bal if bal > 0 else Decimal("0.00")




def _save_cheque_open_balance(request, cheque: "Cheque"):
    """
    Reads posted field: open_balance_amount
    Saves it as ChequeOpenBalanceLine (edit-safe).
    """
    # delete old line if any (edit-safe)
    ChequeOpenBalanceLine.objects.filter(cheque=cheque).delete()

    raw = request.POST.get("open_balance_amount")
    if raw is None or raw == "":
        return

    amt = _dec(raw, "0")
    if amt <= 0:
        return

    # NOTE: do NOT clamp. If they enter more than current open balance, that becomes supplier credit.
    ChequeOpenBalanceLine.objects.create(
        cheque=cheque,
        amount_applied=amt
    )


# posting to the chart of accounts

def _post_supplier_refund_to_ledger(refund: SupplierRefund):
    """
    Supplier refund posting (supplier pays you back):

      DR Bank/Cash (received_to)
      CR Supplier Advances (Asset)     (reduces your prepaid balance)
    """
    amt = Decimal(str(refund.amount or "0.00"))
    if amt <= 0:
        JournalEntry.objects.filter(source_type="SUPPLIER_REFUND", source_id=refund.id).delete()
        return

    JournalEntry.objects.filter(source_type="SUPPLIER_REFUND", source_id=refund.id).delete()

    adv = _get_supplier_advance_account()
    bank = refund.received_to

    entry = JournalEntry.objects.create(
        date=refund.refund_date or timezone.localdate(),
        description=f"Supplier Refund {refund.id:04d} – {getattr(refund.supplier, 'supplier_name', str(refund.supplier))}",
        source_type="SUPPLIER_REFUND",
        source_id=refund.id,
    )

    # DR Bank/Cash
    JournalLine.objects.create(
        entry=entry,
        account=bank,
        debit=amt,
        credit=Decimal("0.00"),
        supplier=refund.supplier,
        customer=None,
    )

    # CR Supplier Advances
    JournalLine.objects.create(
        entry=entry,
        account=adv,
        debit=Decimal("0.00"),
        credit=amt,
        supplier=refund.supplier,
        customer=None,
    )


def _post_expense_to_ledger(expense: Expense):
    """
    Post an Expense document into the General Ledger.

    Pattern (cash expense):

        DR individual expense / cost accounts (category lines + item lines)
        CR payment_account  (cash / bank)

    Replaces any existing journal entry for this expense.
    """

    total = Decimal(str(expense.total_amount or "0.00"))

    # If total is zero, remove any existing journal and stop.
    if total == 0:
        JournalEntry.objects.filter(
            source_type="expense",
            source_id=expense.id,
        ).delete()
        return

    # Remove previous journal for this expense (for edits)
    JournalEntry.objects.filter(
        source_type="expense",
        source_id=expense.id,
    ).delete()

    # Build description
    bits = [f"Expense {expense.ref_no or expense.id}"]
    if expense.payee_name:
        bits.append(f"– {expense.payee_name}")
    elif getattr(expense, "payee_supplier", None) and getattr(expense.payee_supplier, "company_name", None):
        bits.append(f"– {expense.payee_supplier.company_name}")
    description = " ".join(bits)

    entry_date = expense.payment_date or timezone.localdate()

    entry = JournalEntry.objects.create(
        date=entry_date,
        description=description,
        source_type="expense",
        source_id=expense.id,
    )

    # ----- Collect debits per account -----
    debits_by_account = defaultdict(lambda: Decimal("0.00"))

    # 1) Category lines → use the chosen category account directly
    for cl in ExpenseCategoryLine.objects.filter(expense=expense).select_related("category"):
        acc = cl.category
        amt = Decimal(str(cl.amount or "0.00"))
        if acc and amt > 0:
            debits_by_account[acc] += amt

    # 2) Item lines → use product's expense / COGS account if configured,
    #    otherwise fall back to some generic expense account.
    for il in ExpenseItemLine.objects.filter(expense=expense).select_related("product"):
        line_amt = Decimal(str(il.amount or "0.00"))
        if line_amt <= 0:
            continue

        acc = None
        prod = il.product

        if prod is not None:
            # Try common attribute names; only use if present
            acc = getattr(prod, "expense_account", None) or getattr(prod, "cogs_account", None)

        if not acc:
            # Fallback: some broad expense / COGS account
            acc = (
                _find_control_account(name_contains="Cost of Goods") or
                _find_control_account(name_contains="Cost of Sales") or
                _find_control_account(name_contains="Expense")
            )

        if acc:
            debits_by_account[acc] += line_amt

    # 3) Create debit lines
    for acc, amt in debits_by_account.items():
        if not acc or amt <= 0:
            continue
        JournalLine.objects.create(
            entry=entry,
            account=acc,
            debit=amt,
            credit=Decimal("0.00"),
        )

    # 4) Credit payment account (cash / bank)
    if expense.payment_account and total > 0:
        JournalLine.objects.create(
            entry=entry,
            account=expense.payment_account,
            debit=Decimal("0.00"),
            credit=total,
        )
# post bill to ledger
def _get_or_create_ap_control_account() -> "Account":
    """
    Returns the A/P control account.
    Creates it if missing.
    """
    acc = (
        Account.objects.filter(is_active=True, detail_type__iexact="Accounts Payable (A/P)")
        .first()
    )
    if acc:
        return acc

    # Auto-create A/P control account
    # Adjust account_number if you have a numbering rule
    return Account.objects.create(
        account_name="Accounts Payable",
        account_number="2000",
        account_type="CURRENT_LIABILITY",
        detail_type="Accounts Payable (A/P)",
        is_subaccount=False,
        parent=None,
        opening_balance=Decimal("0.00"),
        is_active=True,
    )


def _get_or_create_supplier_ap_subaccount(supplier: Newsupplier) -> Account:
    """
    Creates/gets a SUPPLIER subaccount under A/P control.
    This is your supplier subledger.
    Also links it to supplier.ap_account (IMPORTANT for live balance).
    """
    ap_control = _get_or_create_ap_control_account()

    name = _safe_name(supplier.company_name) or f"Supplier {supplier.id}"
    sub_name = f"{name}"

    acc = Account.objects.filter(
        parent=ap_control,
        account_name__iexact=sub_name,
        is_active=True,
    ).first()

    if not acc:
        acc = Account.objects.create(
            account_name=sub_name,
            account_type=ap_control.account_type,   # current liability
            detail_type="Supplier Subledger (A/P)",
            is_active=True,
            is_subaccount=True,
            parent=ap_control,
            opening_balance=Decimal("0.00"),
        )

    # ✅ LINK THIS ACCOUNT TO THE SUPPLIER (missing before)
    if getattr(supplier, "ap_account_id", None) != acc.id:
        supplier.ap_account = acc
        supplier.save(update_fields=["ap_account"])

    return acc

def _safe_name(s: str) -> str:
    return (s or "").strip()

def _find_control_account(detail_type=None, name_contains=None):
    qs = Account.objects.filter(is_active=True)

    if detail_type:
        acc = qs.filter(detail_type__iexact=detail_type).first()
        if acc:
            return acc

    if name_contains:
        acc = qs.filter(account_name__icontains=name_contains).first()
        if acc:
            return acc

    return None


# posting bill to gl
def _post_bill_to_ledger(bill: "Bill"):
    """
    BILL posting (Correct Accounting + Supplier Subledger):

        DR  Expense/COGS accounts (from bill lines)
        CR  Supplier A/P Subaccount (child under Accounts Payable control)

    Uses bill.journal_entry so edits UPDATE the same JournalEntry (no duplicates).
    """

    total = Decimal(str(bill.total_amount or "0"))
    if total <= 0:
        if bill.journal_entry_id:
            bill.journal_entry.delete()
            bill.journal_entry = None
            bill.save(update_fields=["journal_entry"])
        return

    if not bill.supplier_id:
        # You can decide to allow bills without suppliers, but subledger requires supplier
        raise ValueError("Bill must have a Supplier selected to post to Accounts Payable subledger.")

    # 1) Collect debits (expense accounts)
    expense_by_account = defaultdict(lambda: Decimal("0.00"))

    for cl in bill.category_lines.select_related("category"):
        acc = cl.category
        amt = Decimal(str(cl.amount or "0"))
        if acc and amt > 0:
            expense_by_account[acc] += amt

    default_exp_acc = (
        _find_control_account(name_contains="Cost of Sales")
        or _find_control_account(name_contains="Expense")
    )

    for il in bill.item_lines.select_related("product"):
        amt = Decimal(str(il.amount or "0"))
        if amt <= 0:
            continue

        acc = None
        if il.product:
            acc = getattr(il.product, "expense_account", None) or getattr(il.product, "cogs_account", None)

        if not acc:
            acc = default_exp_acc

        if acc:
            expense_by_account[acc] += amt

    expense_total = sum(expense_by_account.values())
    if expense_total <= 0:
        return

    # 2) Supplier A/P subaccount (creates A/P control if missing)
    supplier_acc = _get_or_create_supplier_ap_subaccount(bill.supplier)

    # 3) Create/update JournalEntry
    entry_date = bill.bill_date or timezone.localdate()
    vendor = bill.supplier.company_name if bill.supplier else (bill.supplier_name or "")
    description = f"Bill {bill.bill_no}" + (f" – {vendor}" if vendor else "")

    entry = bill.journal_entry
    if not entry:
        entry = JournalEntry.objects.create(
            date=entry_date,
            description=description,
            source_type="bill",
            source_id=bill.id,
        )
        bill.journal_entry = entry
        bill.save(update_fields=["journal_entry"])
    else:
        entry.date = entry_date
        entry.description = description
        entry.source_type = "bill"
        entry.source_id = bill.id
        entry.save(update_fields=["date", "description", "source_type", "source_id"])

    # 4) Replace JE lines
    JournalLine.objects.filter(entry=entry).delete()

    # 5) DR Expenses
    for acc, amt in expense_by_account.items():
        if amt > 0:
            JournalLine.objects.create(
                entry=entry,
                account=acc,
                debit=amt,
                credit=Decimal("0.00"),
            )

    # 6) CR Supplier subledger account
    JournalLine.objects.create(
        entry=entry,
        account=supplier_acc,
        debit=Decimal("0.00"),
        credit=expense_total,
    )

# posting cheque to ledger 
# ============================
def _post_cheque_to_ledger(cheq: "Cheque"):
    """
    CHEQUE posting (supports):
      A) Paying Bills (allocations)     -> DR Supplier A/P subledger, CR Bank
      B) Paying Open Balance (new)      -> DR Supplier A/P subledger, CR Bank
      C) Direct expenses (cat/item)     -> DR Expense/COGS, CR Bank

    Total CR Bank = (alloc_total + open_total + direct_total)
    """

    # -------- totals --------
    alloc_total = (
        ChequeBillLine.objects.filter(cheque=cheq)
        .aggregate(s=Sum("amount_applied"))["s"]
        or Decimal("0.00")
    )

    open_total = (
        ChequeOpenBalanceLine.objects.filter(cheque=cheq)
        .aggregate(s=Sum("amount_applied"))["s"]
        or Decimal("0.00")
    )

    # collect direct expenses
    expense_by_account = defaultdict(lambda: Decimal("0.00"))

    # Category lines -> expense accounts directly
    for cl in cheq.category_lines.select_related("category"):
        acc = cl.category
        amt = Decimal(str(cl.amount or "0"))
        if acc and amt > 0:
            expense_by_account[acc] += amt

    # Item lines -> product expense/cogs else fallback
    default_exp_acc = (
        _find_control_account(name_contains="Cost of Sales")
        or _find_control_account(name_contains="Expense")
    )

    for il in cheq.item_lines.select_related("product"):
        amt = Decimal(str(il.amount or "0"))
        if amt <= 0:
            continue

        acc = None
        if il.product:
            acc = getattr(il.product, "expense_account", None) or getattr(il.product, "cogs_account", None)

        if not acc:
            acc = default_exp_acc

        if acc:
            expense_by_account[acc] += amt

    direct_total = sum(expense_by_account.values()) if expense_by_account else Decimal("0.00")

    total_bank_credit = (alloc_total + open_total + direct_total)

    # Nothing to post?
    if total_bank_credit <= 0:
        # if you have a journal entry link on Cheque, delete/clear here if you want
        return

    # -------- accounts --------
    if not cheq.bank_account_id:
        raise ValueError("Cheque must have a bank account.")

    bank_acc = cheq.bank_account

    # Supplier AP subledger account (only if supplier exists AND we have alloc/open totals)
    supplier_acc = None
    supplier_ap_debit = (alloc_total + open_total)
    if cheq.payee_supplier_id and supplier_ap_debit > 0:
        supplier_acc = _get_or_create_supplier_ap_subaccount(cheq.payee_supplier)

    # -------- journal entry header --------
    entry_date = cheq.payment_date or timezone.localdate()
    vendor = cheq.payee_supplier.company_name if cheq.payee_supplier else (cheq.payee_name or "")
    desc_bits = [f"Cheque {cheq.cheque_no}"]
    if vendor:
        desc_bits.append(vendor)
    description = " – ".join(desc_bits)

    # If you already have cheq.journal_entry like Bill, keep it edit-safe
    entry = getattr(cheq, "journal_entry", None)
    if entry is None:
        entry = JournalEntry.objects.create(
            date=entry_date,
            description=description,
            source_type="cheque",
            source_id=cheq.id,
        )
        # only if your Cheque has journal_entry field:
        if hasattr(cheq, "journal_entry_id"):
            cheq.journal_entry = entry
            cheq.save(update_fields=["journal_entry"])
    else:
        entry.date = entry_date
        entry.description = description
        entry.source_type = "cheque"
        entry.source_id = cheq.id
        entry.save(update_fields=["date", "description", "source_type", "source_id"])

    # Replace JE lines (edit-safe)
    JournalLine.objects.filter(entry=entry).delete()

    # Supplier reference (used on all lines for reporting/filtering)
    supplier_ref = cheq.payee_supplier if cheq.payee_supplier_id else None

    # -------- DR: expenses (direct) --------
    for acc, amt in expense_by_account.items():
        if amt > 0:
            JournalLine.objects.create(
                entry=entry,
                account=acc,
                debit=amt,
                credit=Decimal("0.00"),
                supplier=supplier_ref,  # ✅ tag supplier so supplier reports/lists can see it
            )

    # -------- DR: Supplier A/P (bills + open balance) --------
    if supplier_acc and supplier_ap_debit > 0:
        JournalLine.objects.create(
            entry=entry,
            account=supplier_acc,
            debit=supplier_ap_debit,
            credit=Decimal("0.00"),
            supplier=supplier_ref,  # ✅ CRITICAL
        )

    # -------- CR: Bank (total cheque) --------
    JournalLine.objects.create(
        entry=entry,
        account=bank_acc,
        debit=Decimal("0.00"),
        credit=total_bank_credit,
        supplier=supplier_ref,  
    )

# ageing reports 

# ==========================================================
# Helpers (same style as A/R, safe keys)
# ==========================================================
def _dec(x) -> Decimal:
    try:
        return Decimal(str(x or "0"))
    except Exception:
        return Decimal("0")


def _as_date(dt):
    if not dt:
        return None
    try:
        return dt.date()
    except Exception:
        return dt


def _ap_bucket(due_date, today):
    """
    Safe keys:
      current, b1_30, b31_60, b61_90, b90_plus
    """
    if not due_date:
        return "current"

    days = (today - due_date).days
    if days <= 0:
        return "current"
    if 1 <= days <= 30:
        return "b1_30"
    if 31 <= days <= 60:
        return "b31_60"
    if 61 <= days <= 90:
        return "b61_90"
    return "b90_plus"


def _bucket_label(key: str) -> str:
    return {
        "current": "Current",
        "b1_30": "1–30",
        "b31_60": "31–60",
        "b61_90": "61–90",
        "b90_plus": "90+",
    }.get(key, key)


def _vendor_model():
    return Bill._meta.get_field("supplier").remote_field.model


def _vendors_qs():
    Vendor = _vendor_model()
    # assumes company_name exists on Newsupplier
    try:
        return Vendor.objects.order_by("company_name")
    except Exception:
        return Vendor.objects.all()


def _export_urls(request):
    """
    Keep filters and just toggle export=excel/pdf.
    """
    base = request.GET.copy()
    base.pop("export", None)

    excel_qs = base.copy()
    excel_qs["export"] = "excel"

    pdf_qs = base.copy()
    pdf_qs["export"] = "pdf"

    excel_url = f"?{excel_qs.urlencode()}" if excel_qs else "?export=excel"
    pdf_url = f"?{pdf_qs.urlencode()}" if pdf_qs else "?export=pdf"
    return excel_url, pdf_url


# ==========================================================
# Export (Excel)
# ==========================================================
def _excel_response(filename: str, sheet_name: str, headers: list, data_rows: list):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(headers)
    for row in data_rows:
        ws.append(row)

    # autosize columns
    for col_idx in range(1, len(headers) + 1):
        max_len = 10
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            v = r[0].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# ==========================================================
# Export (PDF) - simple clean reportlab table-like output
# ==========================================================
def _pdf_response(filename: str, title: str, subtitle: str, headers: list, data_rows: list):
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(resp, pagesize=landscape(A4))
    width, height = landscape(A4)

    x = 24
    y = height - 30

    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, title)
    y -= 16

    c.setFont("Helvetica", 10)
    c.drawString(x, y, subtitle)
    y -= 18

    # headers
    c.setFont("Helvetica-Bold", 9)
    col_w = max((width - 2 * x) / max(len(headers), 1), 70)
    col_w = min(col_w, 140)

    for i, h in enumerate(headers):
        c.drawString(x + i * col_w, y, str(h)[:25])
    y -= 14

    c.setFont("Helvetica", 9)
    for row in data_rows:
        if y < 30:
            c.showPage()
            y = height - 30
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x, y, title)
            y -= 16
            c.setFont("Helvetica", 10)
            c.drawString(x, y, subtitle)
            y -= 18
            c.setFont("Helvetica-Bold", 9)
            for i, h in enumerate(headers):
                c.drawString(x + i * col_w, y, str(h)[:25])
            y -= 14
            c.setFont("Helvetica", 9)

        for i, val in enumerate(row):
            c.drawString(x + i * col_w, y, str(val)[:28])
        y -= 12

    c.showPage()
    c.save()
    return resp


# ==========================================================
# Shared Bill queryset (Outstanding = total_amount - sum(applied))
# ==========================================================
def _bills_with_outstanding_qs():
    dec_out = DecimalField(max_digits=18, decimal_places=2)

    return (
        Bill.objects
        .select_related("supplier")
        .annotate(
            total_amt=Coalesce(F("total_amount"), Value(Decimal("0.00"), output_field=dec_out)),
            total_paid=Coalesce(
                Sum("cheque_bill_lines__amount_applied", output_field=dec_out),
                Value(Decimal("0.00"), output_field=dec_out),
                output_field=dec_out
            ),
        )
        .annotate(
            outstanding_db=ExpressionWrapper(
                F("total_amt") - F("total_paid"),
                output_field=dec_out
            )
        )
    )


# ==========================================================
# 1) A/P Ageing Summary (per vendor)
# ==========================================================
def ap_aging_summary(request):
    today = timezone.localdate()

    vendor_id = (request.GET.get("vendor") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    bills = _bills_with_outstanding_qs().only(
        "id", "supplier_id", "supplier_name", "bill_no", "bill_date", "due_date", "total_amount"
    )

    if vendor_id.isdigit():
        bills = bills.filter(supplier_id=int(vendor_id))

    rows_map = {}
    grand = {
        "current": Decimal("0.00"),
        "b1_30": Decimal("0.00"),
        "b31_60": Decimal("0.00"),
        "b61_90": Decimal("0.00"),
        "b90_plus": Decimal("0.00"),
        "total": Decimal("0.00"),
    }

    for b in bills:
        bal = _dec(b.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(b.due_date) or _as_date(b.bill_date)
        key = _ap_bucket(due, today)

        if bucket_filter and bucket_filter != key:
            continue

        vendor = b.supplier
        if not vendor:
            # skip orphan bills without supplier FK (supplier_name only)
            # still could be supported later, but cleanest now
            continue

        vid = vendor.id
        if vid not in rows_map:
            rows_map[vid] = {
                "vendor": vendor,
                "current": Decimal("0.00"),
                "b1_30": Decimal("0.00"),
                "b31_60": Decimal("0.00"),
                "b61_90": Decimal("0.00"),
                "b90_plus": Decimal("0.00"),
                "total": Decimal("0.00"),
            }

        rows_map[vid][key] += bal
        rows_map[vid]["total"] += bal

        grand[key] += bal
        grand["total"] += bal

    rows = list(rows_map.values())
    rows.sort(key=lambda r: (getattr(r["vendor"], "company_name", "") or "").lower())

    vendors = _vendors_qs()
    excel_url, pdf_url = _export_urls(request)

    # EXPORT
    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        data = []
        for r in rows:
            data.append([
                getattr(r["vendor"], "company_name", "—"),
                float(r["current"]),
                float(r["b1_30"]),
                float(r["b31_60"]),
                float(r["b61_90"]),
                float(r["b90_plus"]),
                float(r["total"]),
            ])
        data.append(["GRAND TOTAL", float(grand["current"]), float(grand["b1_30"]), float(grand["b31_60"]),
                     float(grand["b61_90"]), float(grand["b90_plus"]), float(grand["total"])])
        return _excel_response("ap_aging_summary.xlsx", "AP Aging Summary", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        data = []
        for r in rows:
            data.append([
                getattr(r["vendor"], "company_name", "—"),
                r["current"], r["b1_30"], r["b31_60"], r["b61_90"], r["b90_plus"], r["total"]
            ])
        data.append(["GRAND TOTAL", grand["current"], grand["b1_30"], grand["b31_60"],
                     grand["b61_90"], grand["b90_plus"], grand["total"]])
        return _pdf_response("ap_aging_summary.pdf", "A/P Ageing Summary", f"As of {today}", headers, data)

    return render(request, "ap_aging_summary.html", {
        "today": today,
        "rows": rows,
        "grand": grand,
        "vendors": vendors,
        "selected_vendor": int(vendor_id) if vendor_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 2) A/P Ageing Detail (bill-level)
# ==========================================================
def ap_aging_detail(request):
    today = timezone.localdate()

    vendor_id = (request.GET.get("vendor") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    bills = (
        _bills_with_outstanding_qs()
        .only("id", "supplier_id", "bill_no", "bill_date", "due_date", "total_amount")
        .order_by("supplier__company_name", "due_date", "bill_date", "id")
    )

    if vendor_id.isdigit():
        bills = bills.filter(supplier_id=int(vendor_id))

    rows = []
    totals = {
        "current": Decimal("0.00"),
        "b1_30": Decimal("0.00"),
        "b31_60": Decimal("0.00"),
        "b61_90": Decimal("0.00"),
        "b90_plus": Decimal("0.00"),
        "total": Decimal("0.00"),
    }

    for b in bills:
        bal = _dec(b.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(b.due_date) or _as_date(b.bill_date)
        key = _ap_bucket(due, today)
        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        vendor = b.supplier
        vendor_name = getattr(vendor, "company_name", "") if vendor else (b.supplier_name or "—")

        rows.append({
            "bill": b,
            "vendor": vendor,
            "vendor_name": vendor_name,
            "bill_id": b.id,
            "bill_no": b.bill_no,
            "bill_date": _as_date(b.bill_date),
            "due_date": due,
            "days_overdue": days_overdue,
            "total_amount": _dec(b.total_amount),
            "amount_paid": _dec(b.total_paid),
            "balance": bal,
            "bucket": key,
            "bucket_label": _bucket_label(key),
        })

        totals[key] += bal
        totals["total"] += bal

    vendors = _vendors_qs()
    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Days", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"],
                r["bill_no"],
                str(r["bill_date"]),
                str(r["due_date"]),
                r["bucket_label"],
                int(r["days_overdue"]) if r["days_overdue"] > 0 else 0,
                float(r["total_amount"]),
                float(r["amount_paid"]),
                float(r["balance"]),
            ])
        data.append(["TOTAL", "", "", "", "", "", "", "", float(totals["total"])])
        return _excel_response("ap_aging_detail.xlsx", "AP Aging Detail", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Days", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], r["bill_no"], r["bill_date"], r["due_date"],
                r["bucket_label"], int(r["days_overdue"]) if r["days_overdue"] > 0 else 0,
                r["total_amount"], r["amount_paid"], r["balance"],
            ])
        data.append(["TOTAL", "", "", "", "", "", "", "", totals["total"]])
        return _pdf_response("ap_aging_detail.pdf", "A/P Ageing Detail", f"As of {today}", headers, data)

    return render(request, "ap_aging_detail.html", {
        "today": today,
        "rows": rows,
        "totals": totals,
        "vendors": vendors,
        "selected_vendor": int(vendor_id) if vendor_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 3) Vendor-specific A/P Ageing (summary+detail)
# ==========================================================
def ap_aging_vendor(request, vendor_id: int):
    today = timezone.localdate()
    Vendor = _vendor_model()
    vendor = get_object_or_404(Vendor, pk=vendor_id)

    bucket_filter = (request.GET.get("bucket") or "").strip()

    bills = (
        _bills_with_outstanding_qs()
        .filter(supplier_id=vendor_id)
        .only("id", "bill_no", "bill_date", "due_date", "total_amount")
        .order_by("due_date", "bill_date", "id")
    )

    summary = {
        "current": Decimal("0.00"),
        "b1_30": Decimal("0.00"),
        "b31_60": Decimal("0.00"),
        "b61_90": Decimal("0.00"),
        "b90_plus": Decimal("0.00"),
        "total": Decimal("0.00"),
    }

    rows = []
    for b in bills:
        bal = _dec(b.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(b.due_date) or _as_date(b.bill_date)
        key = _ap_bucket(due, today)

        summary[key] += bal
        summary["total"] += bal

        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        rows.append({
            "bill_no": b.bill_no,
            "bill_date": _as_date(b.bill_date),
            "due_date": due,
            "bucket": key,
            "bucket_label": _bucket_label(key),
            "days_overdue": days_overdue,
            "total_amount": _dec(b.total_amount),
            "amount_paid": _dec(b.total_paid),
            "balance": bal,
        })

    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        data = [[
            getattr(vendor, "company_name", "—"),
            float(summary["current"]),
            float(summary["b1_30"]),
            float(summary["b31_60"]),
            float(summary["b61_90"]),
            float(summary["b90_plus"]),
            float(summary["total"]),
        ]]
        return _excel_response("ap_aging_vendor.xlsx", "AP Aging Vendor", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Current", "1–30", "31–60", "61–90", "90+", "Total"]
        data = [[
            getattr(vendor, "company_name", "—"),
            summary["current"], summary["b1_30"], summary["b31_60"],
            summary["b61_90"], summary["b90_plus"], summary["total"]
        ]]
        return _pdf_response("ap_aging_vendor.pdf", "Vendor A/P Ageing", f"As of {today}", headers, data)

    return render(request, "ap_aging_vendor.html", {
        "today": today,
        "vendor": vendor,
        "summary": summary,
        "rows": rows,
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 4) Unpaid Bills (OPEN bills)
# ==========================================================
def unpaid_bills_report(request):
    today = timezone.localdate()

    vendor_id = (request.GET.get("vendor") or "").strip()
    bucket_filter = (request.GET.get("bucket") or "").strip()

    bills = (
        _bills_with_outstanding_qs()
        .only("id", "supplier_id", "bill_no", "bill_date", "due_date", "total_amount")
        .order_by("supplier__company_name", "due_date", "bill_date", "id")
    )

    if vendor_id.isdigit():
        bills = bills.filter(supplier_id=int(vendor_id))

    rows = []
    totals = {"total": Decimal("0.00")}

    for b in bills:
        bal = _dec(b.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(b.due_date) or _as_date(b.bill_date)
        key = _ap_bucket(due, today)

        if bucket_filter and bucket_filter != key:
            continue

        days_overdue = (today - due).days if due else 0

        vendor = b.supplier
        vendor_name = getattr(vendor, "company_name", "") if vendor else (b.supplier_name or "—")

        rows.append({
            "vendor": vendor,
            "vendor_name": vendor_name,
            "bill_no": b.bill_no,
            "bill_date": _as_date(b.bill_date),
            "due_date": due,
            "bucket": key,
            "bucket_label": _bucket_label(key),
            "days_overdue": days_overdue,
            "total_amount": _dec(b.total_amount),
            "amount_paid": _dec(b.total_paid),
            "balance": bal,
        })
        totals["total"] += bal

    vendors = _vendors_qs()
    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Days", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], r["bill_no"], str(r["bill_date"]), str(r["due_date"]),
                r["bucket_label"], int(r["days_overdue"]) if r["days_overdue"] > 0 else 0,
                float(r["total_amount"]), float(r["amount_paid"]), float(r["balance"])
            ])
        data.append(["TOTAL", "", "", "", "", "", "", "", float(totals["total"])])
        return _excel_response("unpaid_bills.xlsx", "Unpaid Bills", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Days", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], r["bill_no"], r["bill_date"], r["due_date"],
                r["bucket_label"], int(r["days_overdue"]) if r["days_overdue"] > 0 else 0,
                r["total_amount"], r["amount_paid"], r["balance"]
            ])
        data.append(["TOTAL", "", "", "", "", "", "", "", totals["total"]])
        return _pdf_response("unpaid_bills.pdf", "Unpaid Bills", f"As of {today}", headers, data)

    return render(request, "ap_unpaid_bills.html", {
        "today": today,
        "rows": rows,
        "totals": totals,
        "vendors": vendors,
        "selected_vendor": int(vendor_id) if vendor_id.isdigit() else "",
        "selected_bucket": bucket_filter,
        "bucket_choices": [
            ("", "All"),
            ("current", "Current"),
            ("b1_30", "1–30"),
            ("b31_60", "31–60"),
            ("b61_90", "61–90"),
            ("b90_plus", "90+"),
        ],
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 5) Vendor Balances (current vs overdue)
# ==========================================================
def vendor_balances_report(request):
    today = timezone.localdate()

    bills = _bills_with_outstanding_qs().only("id", "supplier_id", "supplier_name", "bill_date", "due_date", "total_amount")

    vendor_map = {}
    grand = {"current": Decimal("0.00"), "overdue": Decimal("0.00"), "total": Decimal("0.00")}

    for b in bills:
        bal = _dec(b.outstanding_db)
        if bal <= Decimal("0.00001"):
            continue

        due = _as_date(b.due_date) or _as_date(b.bill_date)
        overdue = bool(due and due < today)

        vendor = b.supplier
        if not vendor:
            continue

        vid = vendor.id
        if vid not in vendor_map:
            vendor_map[vid] = {
                "vendor": vendor,
                "current": Decimal("0.00"),
                "overdue": Decimal("0.00"),
                "total": Decimal("0.00"),
            }

        if overdue:
            vendor_map[vid]["overdue"] += bal
            grand["overdue"] += bal
        else:
            vendor_map[vid]["current"] += bal
            grand["current"] += bal

        vendor_map[vid]["total"] += bal
        grand["total"] += bal

    rows = list(vendor_map.values())
    rows.sort(key=lambda r: (getattr(r["vendor"], "company_name", "") or "").lower())

    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Current", "Overdue", "Total"]
        data = []
        for r in rows:
            data.append([
                getattr(r["vendor"], "company_name", "—"),
                float(r["current"]),
                float(r["overdue"]),
                float(r["total"]),
            ])
        data.append(["GRAND TOTAL", float(grand["current"]), float(grand["overdue"]), float(grand["total"])])
        return _excel_response("vendor_balances.xlsx", "Vendor Balances", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Current", "Overdue", "Total"]
        data = []
        for r in rows:
            data.append([getattr(r["vendor"], "company_name", "—"), r["current"], r["overdue"], r["total"]])
        data.append(["GRAND TOTAL", grand["current"], grand["overdue"], grand["total"]])
        return _pdf_response("vendor_balances.pdf", "Vendor Balances", f"As of {today}", headers, data)

    return render(request, "ap_vendor_balances.html", {
        "today": today,
        "rows": rows,
        "grand": grand,
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 6) Bills List (all bills, paid/open/overdue)
# ==========================================================
def bills_list_report(request):
    today = timezone.localdate()

    vendor_id = (request.GET.get("vendor") or "").strip()
    status = (request.GET.get("status") or "").strip()  # all|paid|unpaid|overdue

    bills = (
        _bills_with_outstanding_qs()
        .only("id", "supplier_id", "bill_no", "bill_date", "due_date", "total_amount")
        .order_by("-bill_date", "-id")
    )

    if vendor_id.isdigit():
        bills = bills.filter(supplier_id=int(vendor_id))

    rows = []
    for b in bills:
        bal = _dec(b.outstanding_db)
        due = _as_date(b.due_date) or _as_date(b.bill_date)

        is_paid = bal <= Decimal("0.00001")
        is_overdue = bool((not is_paid) and due and due < today)

        if status == "paid" and not is_paid:
            continue
        if status == "unpaid" and is_paid:
            continue
        if status == "overdue" and not is_overdue:
            continue

        vendor = b.supplier
        vendor_name = getattr(vendor, "company_name", "") if vendor else (b.supplier_name or "—")

        rows.append({
            "vendor": vendor,
            "vendor_name": vendor_name,
            "bill_no": b.bill_no,
            "bill_date": _as_date(b.bill_date),
            "due_date": due,
            "total_amount": _dec(b.total_amount),
            "amount_paid": _dec(b.total_paid),
            "balance": bal if bal > 0 else Decimal("0.00"),
            "status": "PAID" if is_paid else ("OVERDUE" if is_overdue else "OPEN"),
        })

    vendors = _vendors_qs()
    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], r["bill_no"], str(r["bill_date"]), str(r["due_date"]),
                r["status"], float(r["total_amount"]), float(r["amount_paid"]), float(r["balance"])
            ])
        return _excel_response("bills_list.xlsx", "Bills List", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Status", "Total", "Paid", "Balance"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], r["bill_no"], r["bill_date"], r["due_date"],
                r["status"], r["total_amount"], r["amount_paid"], r["balance"]
            ])
        return _pdf_response("bills_list.pdf", "Bills List", f"As of {today}", headers, data)

    return render(request, "ap_bills_list.html", {
        "today": today,
        "rows": rows,
        "vendors": vendors,
        "selected_vendor": int(vendor_id) if vendor_id.isdigit() else "",
        "selected_status": status,
        "status_choices": [
            ("", "All"),
            ("paid", "Paid"),
            ("unpaid", "Unpaid"),
            ("overdue", "Overdue"),
        ],
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })


# ==========================================================
# 7) Payments to Vendors (Cheques)
# ==========================================================
def payments_to_vendors_report(request):
    today = timezone.localdate()

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    if not date_from and not date_to:
        first = today.replace(day=1)
        date_from = str(first)
        date_to = str(today)

    dec_out = DecimalField(max_digits=18, decimal_places=2)

    qs = Cheque.objects.select_related("payee_supplier").all().order_by("-payment_date", "-id")

    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    qs = qs.annotate(
        total_applied=Coalesce(
            Sum("bill_lines__amount_applied", output_field=dec_out),
            Value(Decimal("0.00"), output_field=dec_out),
            output_field=dec_out
        ),
        open_balance=Coalesce(
            F("open_balance_line__amount_applied"),
            Value(Decimal("0.00"), output_field=dec_out),
            output_field=dec_out
        )
    )

    rows = []
    totals = {"paid": Decimal("0.00"), "applied": Decimal("0.00"), "unapplied": Decimal("0.00")}

    for ch in qs:
        paid = _dec(ch.total_amount)
        applied = _dec(ch.total_applied)
        open_bal = _dec(ch.open_balance)
        unapplied = paid - applied - open_bal
        if unapplied < 0:
            unapplied = Decimal("0.00")

        vendor_name = (ch.payee_supplier.company_name if ch.payee_supplier else (ch.payee_name or "—"))

        rows.append({
            "vendor_name": vendor_name,
            "date": ch.payment_date,
            "ref": ch.cheque_no,
            "method": "Cheque",
            "paid": paid,
            "applied": applied,
            "unapplied": unapplied,
        })

        totals["paid"] += paid
        totals["applied"] += applied
        totals["unapplied"] += unapplied

    excel_url, pdf_url = _export_urls(request)

    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        headers = ["Vendor", "Date", "Cheque #", "Method", "Paid", "Applied", "Unapplied"]
        data = []
        for r in rows:
            data.append([
                r["vendor_name"], str(r["date"]), r["ref"], r["method"],
                float(r["paid"]), float(r["applied"]), float(r["unapplied"])
            ])
        data.append(["TOTALS", "", "", "", float(totals["paid"]), float(totals["applied"]), float(totals["unapplied"])])
        return _excel_response("payments_to_vendors.xlsx", "Payments to Vendors", headers, data)

    if export == "pdf":
        headers = ["Vendor", "Date", "Cheque #", "Method", "Paid", "Applied", "Unapplied"]
        data = []
        for r in rows:
            data.append([r["vendor_name"], r["date"], r["ref"], r["method"], r["paid"], r["applied"], r["unapplied"]])
        data.append(["TOTALS", "", "", "", totals["paid"], totals["applied"], totals["unapplied"]])
        return _pdf_response("payments_to_vendors.pdf", "Payments to Vendors", f"From {date_from} to {date_to}", headers, data)

    return render(request, "ap_payments_to_vendors.html", {
        "today": today,
        "rows": rows,
        "totals": totals,
        "date_from": date_from,
        "date_to": date_to,
        "export_excel_url": excel_url,
        "export_pdf_url": pdf_url,
    })

  





























































# posting supplier credit to ledger

# def _post_supplier_credit_to_ledger(credit: SupplierCredit):
#     """
#     Post a Supplier Credit to the GL.

#     Pattern (reverse of Bill):
#       DR Accounts Payable
#       CR Expense / Cost accounts

#     This effectively reduces liability and reverses expense.
#     """

#     total = Decimal(str(credit.total_amount or "0"))
#     if total == 0:
#         JournalEntry.objects.filter(
#             source_type="supplier_credit",
#             source_id=credit.id,
#         ).delete()
#         return

#     # Remove previous journal if editing
#     JournalEntry.objects.filter(
#         source_type="supplier_credit",
#         source_id=credit.id,
#     ).delete()

#     # Collect credits per expense account
#     expense_by_account: dict[Account, Decimal] = defaultdict(lambda: Decimal("0.00"))

#     for line in credit.lines.select_related("category"):
#         if not line.category:
#             continue
#         amt = Decimal(str(line.amount or "0"))
#         if amt <= 0:
#             continue
#         expense_by_account[line.category] += amt

#     expense_total = sum(expense_by_account.values())
#     if expense_total == 0:
#         return

#     # Accounts Payable control account
#     ap_account = (
#         _find_control_account(detail_type="Accounts Payable (A/P)")
#         or _find_control_account(name_contains="payable")
#     )
#     if not ap_account:
#         # No A/P configured; skip posting
#         return

#     entry_date = credit.credit_date or timezone.localdate()
#     bits = [f"Supplier Credit {credit.ref_no}"]
#     if credit.supplier and getattr(credit.supplier, "company_name", None):
#         bits.append(f"– {credit.supplier.company_name}")
#     elif credit.supplier_name:
#         bits.append(f"– {credit.supplier_name}")
#     description = " ".join(bits)

#     entry = JournalEntry.objects.create(
#         date=entry_date,
#         description=description,
#         source_type="supplier_credit",
#         source_id=credit.id,
#     )

#     # CR Expense accounts
#     for acc, amt in expense_by_account.items():
#         if not acc or amt <= 0:
#             continue
#         JournalLine.objects.create(
#             entry=entry,
#             account=acc,
#             debit=Decimal("0.00"),
#             credit=amt,
#         )

#     # DR Accounts Payable
#     JournalLine.objects.create(
#         entry=entry,
#         account=ap_account,
#         debit=expense_total,
#         credit=Decimal("0.00"),
#     )
# posting pay down credit to gl
# def _post_paydown_credit_to_ledger(pdc: PayDownCredit):
#     """
#     Post a PayDownCredit into the GL.

#       DR Credit Card liability
#       CR Bank / Cash account
#     """
#     amt = Decimal(str(pdc.amount or "0"))

#     # If zero → delete any existing journal & stop
#     if amt == 0:
#         JournalEntry.objects.filter(
#             source_type="paydown_credit",
#             source_id=pdc.id,
#         ).delete()
#         return

#     # Remove previous entry (for edits)
#     JournalEntry.objects.filter(
#         source_type="paydown_credit",
#         source_id=pdc.id,
#     ).delete()

#     bits = [f"Pay down credit card {pdc.credit_card.account_name}"]
#     if pdc.ref_no:
#         bits.append(f"(Ref {pdc.ref_no})")
#     if pdc.payee_supplier and getattr(pdc.payee_supplier, "company_name", None):
#         bits.append(f"– {pdc.payee_supplier.company_name}")
#     elif pdc.payee_name:
#         bits.append(f"– {pdc.payee_name}")

#     description = " ".join(bits)
#     entry_date = pdc.payment_date or timezone.localdate()

#     entry = JournalEntry.objects.create(
#         date=entry_date,
#         description=description,
#         source_type="paydown_credit",
#         source_id=pdc.id,
#     )

#     # DR Credit Card
#     JournalLine.objects.create(
#         entry=entry,
#         account=pdc.credit_card,
#         debit=amt,
#         credit=Decimal("0.00"),
#     )

#     # CR Bank
#     JournalLine.objects.create(
#         entry=entry,
#         account=pdc.bank_account,
#         debit=Decimal("0.00"),
#         credit=amt,
#     )
# post credit card 
# def _post_credit_card_credit_to_ledger(cc: CreditCardCredit):
#     """
#     Post a CreditCardCredit into the GL.

#     Concept (QuickBooks-style):

#       For each category line:
#         DR Credit card liability
#         CR Expense/Other account

#       (Item logic can be extended later once inventory mapping is ready.)
#     """
#     # Remove any previous entries (for edits or zeroing)
#     JournalEntry.objects.filter(
#         source_type="credit_card_credit",
#         source_id=cc.id,
#     ).delete()

#     # Sum category amounts (only these affect GL for now)
#     cat_amount = cc.category_lines.aggregate(
#         total=Coalesce(
#             Sum("amount"),
#             Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))
#         )
#     )["total"] or Decimal("0.00")

#     total = Decimal(cat_amount)

#     # Update stored total
#     cc.total_amount = total
#     cc.save(update_fields=["total_amount"])

#     if total == 0:
#         # nothing to post
#         return

#     bits = [f"Credit card credit – {cc.credit_card.account_name}"]
#     if cc.ref_no:
#         bits.append(f"(Ref {cc.ref_no})")
#     if cc.payee_supplier and getattr(cc.payee_supplier, "company_name", None):
#         bits.append(f"– {cc.payee_supplier.company_name}")
#     elif cc.payee_name:
#         bits.append(f"– {cc.payee_name}")
#     description = " ".join(bits)

#     entry_date = cc.credit_date or timezone.localdate()

#     entry = JournalEntry.objects.create(
#         date=entry_date,
#         description=description,
#         source_type="credit_card_credit",
#         source_id=cc.id,
#     )

#     # DR credit card (liability decreases)
#     JournalLine.objects.create(
#         entry=entry,
#         account=cc.credit_card,
#         debit=total,
#         credit=Decimal("0.00"),
#     )

#     # Credit per category line
#     for line in cc.category_lines.select_related("category"):
#         amt = Decimal(str(line.amount or "0"))
#         if amt == 0:
#             continue
#         JournalLine.objects.create(
#             entry=entry,
#             account=line.category,
#             debit=Decimal("0.00"),
#             credit=amt,
#         )

# all expenses

def expenses(request):
    # ---------------- Expenses ----------------
    exp_qs = (
        Expense.objects
        .select_related("payee_supplier")
        .prefetch_related("cat_lines__category", "item_lines__product")
        .order_by("-payment_date", "-id")
    )

    for e in exp_qs:
        e._total_lines = len(list(e.cat_lines.all())) + len(list(e.item_lines.all()))

    # ---------------- Bills ----------------
    bill_qs = (
        Bill.objects
        .select_related("supplier")
        .prefetch_related("category_lines__category", "item_lines__product")
        .order_by("-bill_date", "-id")
    )

    # ---------------- Cheques ----------------
    cheque_qs = (
        Cheque.objects
        .select_related("payee_supplier", "bank_account")
        .prefetch_related("category_lines__category", "item_lines__product")
        .order_by("-payment_date", "-id")
    )

    # ---------------- Purchase Orders ----------------
    po_qs = (
        PurchaseOrder.objects
        .select_related("vendor")
        .prefetch_related("lines__product")
        .order_by("-po_date", "-id")
    )

    # ---------------- Supplier Credits ----------------
    supplier_credit_qs = (
        SupplierCredit.objects
        .select_related("supplier")
        .prefetch_related("lines__category")
        .order_by("-credit_date", "-id")
    )

    # ---------------- Pay Down Credit ----------------
    paydown_qs = (
        PayDownCredit.objects
        .select_related("credit_card", "bank_account", "payee_supplier")
        .order_by("-payment_date", "-id")
    )

    # ---------------- Credit Card Credits ----------------
    cc_credit_qs = (
        CreditCardCredit.objects
        .select_related("credit_card", "payee_supplier")
        .order_by("-credit_date", "-id")
    )

    # ---------------- Normalize into one list ----------------
    rows = []

    # Expenses
    for e in exp_qs:
        rows.append({
            "id": e.id,
            "kind": "Expense",
            "date": e.payment_date,
            "number": getattr(e, "ref_no", "") or "",
            "payee": (
                e.payee_supplier.company_name
                if e.payee_supplier else (e.payee_name or "")
            ),
            "category": _cat_label_from_lines(e.cat_lines.all(), e.item_lines.all()),
            "total_before_tax": e.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": e.total_amount,
            "edit_url": reverse("expenses:expense-edit", args=[e.id]),
        })

    # Bills
    for b in bill_qs:
        rows.append({
            "id": b.id,
            "kind": "Bill",
            "date": b.bill_date,
            "number": b.bill_no or "",
            "payee": (
                b.supplier.company_name
                if b.supplier else (b.supplier_name or "")
            ),
            "category": _cat_label_from_lines(b.category_lines.all(), b.item_lines.all()),
            "total_before_tax": b.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": b.total_amount,
            "edit_url": reverse("expenses:bill-edit", args=[b.id]),
        })

    # Cheques
    for c in cheque_qs:
        rows.append({
            "id": c.id,
            "kind": "Cheque",
            "date": c.payment_date,
            "number": c.cheque_no or "",
            "payee": (
                c.payee_supplier.company_name
                if c.payee_supplier else (c.payee_name or "")
            ),
            "category": _cat_label_from_lines(c.category_lines.all(), c.item_lines.all()),
            "total_before_tax": c.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": c.total_amount,
            "edit_url": reverse("expenses:cheque-edit", args=[c.id]),
        })

    # Purchase Orders
    for po in po_qs:
        rows.append({
            "id": po.id,
            "kind": "Purchase Order",
            "date": po.po_date,
            "number": po.po_number or "",
            "payee": (
                po.vendor.company_name
                if po.vendor else (po.vendor_name or "")
            ),
            "category": _cat_label_from_lines([], po.lines.all()),
            "total_before_tax": po.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": po.total_amount,
            "edit_url": reverse("expenses:purchase-order-edit", args=[po.id]),
        })

    # Supplier Credits
    for sc in supplier_credit_qs:
        rows.append({
            "id": sc.id,
            "kind": "Supplier Credit",
            "date": sc.credit_date,
            "number": sc.ref_no or "",
            "payee": (
                sc.supplier.company_name
                if sc.supplier else (sc.supplier_name or "")
            ),
            "category": _cat_label_from_lines(sc.lines.all(), []),
            "total_before_tax": sc.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": sc.total_amount,
            "edit_url": reverse("expenses:supplier-credit-edit", args=[sc.id]),
        })

    # Paydown credit
    for pdc in paydown_qs:
        rows.append({
            "id": pdc.id,
            "kind": "Credit Card Payment",
            "date": pdc.payment_date,
            "number": pdc.ref_no or "",
            "payee": (
                pdc.payee_supplier.company_name
                if pdc.payee_supplier else (pdc.payee_name or "")
            ),
            "category": getattr(pdc.credit_card, "account_name", ""),
            "total_before_tax": pdc.amount,
            "sales_tax": Decimal("0.00"),
            "total": pdc.amount,
            "edit_url": reverse("expenses:paydown-credit-edit", args=[pdc.id]),
        })

    # Credit Card Credits
    for c in cc_credit_qs:
        rows.append({
            "id": c.id,
            "kind": "Credit Card Credit",
            "date": c.credit_date,
            "number": c.ref_no or "",
            "payee": (
                c.payee_supplier.company_name
                if c.payee_supplier else (c.payee_name or "")
            ),
            "category": _cat_label_from_lines(
                c.category_lines.all(),
                c.item_lines.all()
            ),
            "total_before_tax": c.total_amount,
            "sales_tax": Decimal("0.00"),
            "total": c.total_amount,
            "edit_url": reverse("expenses:credit-card-credit-edit", args=[c.id]),
        })

    rows = sorted(rows, key=lambda r: (r["date"], r["id"]), reverse=True)

    # Column prefs as before
    if getattr(request.user, "is_authenticated", False):
        prefs, _ = ColumnPreference.objects.get_or_create(
            user=request.user,
            table_name="accounts",
            defaults={"preferences": DEFAULT_ACCOUNTS_COL_PREFS},
        )
        merged_prefs = {**DEFAULT_ACCOUNTS_COL_PREFS, **(prefs.preferences or {})}
    else:
        merged_prefs = DEFAULT_ACCOUNTS_COL_PREFS

    return render(request, "expenses.html", {
        "expenses": exp_qs,
        "transactions": rows,
        "column_prefs": merged_prefs,
    })

@csrf_exempt
def save_column_prefs(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "detail": "POST required"}, status=400)

    try:
        data = json.loads(request.body or "{}")
        preferences = data.get("preferences", {})
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "detail": "Bad JSON"}, status=400)

    prefs, _ = ColumnPreference.objects.get_or_create(
        user=request.user,
        table_name="accounts",
    )
    # also ensure unknown keys don’t sneak in (optional)
    cleaned = {k: bool(preferences.get(k, True)) for k in DEFAULT_ACCOUNTS_COL_PREFS.keys()}
    prefs.preferences = cleaned
    prefs.save()
    return JsonResponse({"status": "ok"})

# adding an expense
def _dec(v, default="0.00"):
    try:
        return Decimal(str(v or default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

@transaction.atomic
def add_expense(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                # --- Allowed account sets ---
                payment_accounts_qs = deposit_accounts_qs()   # only cash/bank
                expense_accounts_qs_ = expense_accounts_qs()  # only expense-type accounts

                # ----- Header -----
                payee_name         = request.POST.get("payee_name") or ""
                supplier_id        = request.POST.get("payee_supplier") or ""
                payment_account_id = request.POST.get("payment_account") or ""
                payment_date       = request.POST.get("payment_date") or timezone.localdate()
                payment_method     = request.POST.get("payment_method") or "cash"
                ref_no             = request.POST.get("ref_no") or ""
                location           = request.POST.get("location") or ""
                memo               = request.POST.get("memo") or ""
                attachment         = request.FILES.get("attachments")

                supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None

                # Resolve payment account strictly from cash/bank accounts
                payment_account = get_object_or_404(payment_accounts_qs, pk=payment_account_id)

                # ensure we have a valid 8-digit numeric ref; if not, generate one
                if not (len(ref_no) == 8 and ref_no.isdigit()):
                    ref_no = generate_unique_ref_no()
                if Expense.objects.filter(ref_no=ref_no).exists():
                    ref_no = generate_unique_ref_no()

                exp = Expense.objects.create(
                    payee_name=payee_name,
                    payee_supplier=supplier,
                    payment_account=payment_account,
                    payment_date=payment_date,
                    payment_method=payment_method,
                    ref_no=ref_no,
                    location=location,
                    memo=memo,
                    attachments=attachment,
                )

                total = Decimal("0.00")

                # -------- Category lines --------
                cat_category_ids = request.POST.getlist("cat_category[]")
                cat_descs        = request.POST.getlist("cat_desc[]")
                cat_amounts      = request.POST.getlist("cat_amount[]")
                cat_billable     = set(request.POST.getlist("cat_billable[]"))  # contains row idx strings
                cat_customer_ids = request.POST.getlist("cat_customer[]")
                cat_class_ids    = request.POST.getlist("cat_class[]")

                for idx, cat_id in enumerate(cat_category_ids):
                    if not cat_id:
                        continue

                    category = expense_accounts_qs_.filter(pk=cat_id).first()
                    if not category:
                        continue

                    amt = _dec(cat_amounts[idx])
                    if amt <= 0:
                        continue

                    is_bill  = str(idx) in cat_billable
                    customer = Newcustomer.objects.filter(pk=cat_customer_ids[idx] or None).first()
                    klass    = Pclass.objects.filter(pk=cat_class_ids[idx] or None).first()

                    ExpenseCategoryLine.objects.create(
                        expense=exp, category=category,
                        description=(cat_descs[idx] or ""),
                        amount=amt, is_billable=is_bill,
                        customer=customer, class_field=klass
                    )
                    total += amt

                # -------- Item lines --------
                item_product_ids  = request.POST.getlist("item_product[]")
                item_descs        = request.POST.getlist("item_desc[]")
                item_qtys         = request.POST.getlist("item_qty[]")
                item_rates        = request.POST.getlist("item_rate[]")
                item_amounts      = request.POST.getlist("item_amount[]")
                item_billable     = set(request.POST.getlist("item_billable[]"))
                item_customer_ids = request.POST.getlist("item_customer[]")
                item_class_ids    = request.POST.getlist("item_class[]")

                for idx, prod_id in enumerate(item_product_ids):
                    if not prod_id:
                        continue
                    product = Product.objects.filter(pk=prod_id).first()
                    if not product:
                        continue

                    qty  = _dec(item_qtys[idx], "0")
                    rate = _dec(item_rates[idx], "0")
                    amt  = _dec(item_amounts[idx]) if (idx < len(item_amounts) and item_amounts[idx]) else (qty * rate)

                    if amt <= 0:
                        continue

                    is_bill  = str(idx) in item_billable
                    customer = Newcustomer.objects.filter(pk=item_customer_ids[idx] or None).first()
                    klass    = Pclass.objects.filter(pk=item_class_ids[idx] or None).first()

                    ExpenseItemLine.objects.create(
                        expense=exp, product=product,
                        description=(item_descs[idx] or ""),
                        qty=qty, rate=rate, amount=amt,
                        is_billable=is_bill, customer=customer, class_field=klass
                    )
                    total += amt

                if total <= 0:
                    exp.delete()
                    messages.error(request, "Expense total must be greater than 0.")
                    return redirect("expenses:add-expenses")

                exp.total_amount = total
                exp.save(update_fields=["total_amount"])

                # ✅ Post GL + ✅ Rebuild inventory movements (stock-in for inventory products)
                _post_expense_to_ledger(exp)
                rebuild_movements_for_expense(exp)

                # redirect behaviour
                action = request.POST.get("save_action") or "save"
                if action == "save":
                    return redirect("expenses:expenses")
                if action == "save&new":
                    return redirect("expenses:add-expenses")
                return redirect("expenses:expenses")

        except Exception as e:
            messages.error(request, f"Could not save expense: {e}")

    # GET: load form lists
    ref_no = generate_unique_ref_no()
    context = {
        "accounts": deposit_accounts_qs(),
        "expense_accounts": expense_accounts_qs(),

        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "ref_no": ref_no,
        "payment_methods": Expense.PAYMENT_METHODS,
    }
    return render(request, "expenses_form.html", context)

# edit expenses
@transaction.atomic
def expense_edit(request, pk: int):
    exp = get_object_or_404(
        Expense.objects
        .select_related("payee_supplier", "payment_account")
        .prefetch_related("cat_lines__category", "item_lines__product"),
        pk=pk
    )

    if request.method == "POST":
        try:
            with transaction.atomic():
                # --- Allowed account sets ---
                payment_accounts_qs = deposit_accounts_qs()   # only cash/bank
                expense_accounts_qs_ = expense_accounts_qs()  # only expense-type accounts

                # ---- Header
                exp.payee_name = request.POST.get("payee_name") or ""

                supplier_id = request.POST.get("payee_supplier") or ""
                exp.payee_supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None

                payment_account_id = request.POST.get("payment_account") or ""
                exp.payment_account = get_object_or_404(payment_accounts_qs, pk=payment_account_id)

                exp.payment_date   = request.POST.get("payment_date") or timezone.localdate()
                exp.payment_method = request.POST.get("payment_method") or "cash"
                exp.ref_no         = request.POST.get("ref_no") or exp.ref_no or ""
                exp.location       = request.POST.get("location") or ""
                exp.memo           = request.POST.get("memo") or ""

                if request.FILES.get("attachments"):
                    exp.attachments = request.FILES["attachments"]

                exp.save()

                # ---- Replace lines
                ExpenseCategoryLine.objects.filter(expense=exp).delete()
                ExpenseItemLine.objects.filter(expense=exp).delete()

                total = Decimal("0.00")

                # -------- Category lines --------
                cat_category_ids = request.POST.getlist("cat_category[]")
                cat_descs        = request.POST.getlist("cat_desc[]")
                cat_amounts      = request.POST.getlist("cat_amount[]")
                cat_billable     = set(request.POST.getlist("cat_billable[]"))
                cat_customer_ids = request.POST.getlist("cat_customer[]")
                cat_class_ids    = request.POST.getlist("cat_class[]")

                for idx, cat_id in enumerate(cat_category_ids):
                    if not cat_id:
                        continue

                    category = expense_accounts_qs_.filter(pk=cat_id).first()
                    if not category:
                        continue

                    amt = _dec(cat_amounts[idx])
                    if amt <= 0:
                        continue

                    is_bill  = str(idx) in cat_billable
                    customer = Newcustomer.objects.filter(pk=cat_customer_ids[idx] or None).first()
                    klass    = Pclass.objects.filter(pk=cat_class_ids[idx] or None).first()

                    ExpenseCategoryLine.objects.create(
                        expense=exp,
                        category=category,
                        description=(cat_descs[idx] or ""),
                        amount=amt,
                        is_billable=is_bill,
                        customer=customer,
                        class_field=klass,
                    )
                    total += amt

                # -------- Item lines --------
                item_product_ids  = request.POST.getlist("item_product[]")
                item_descs        = request.POST.getlist("item_desc[]")
                item_qtys         = request.POST.getlist("item_qty[]")
                item_rates        = request.POST.getlist("item_rate[]")
                item_amounts      = request.POST.getlist("item_amount[]")
                item_billable     = set(request.POST.getlist("item_billable[]"))
                item_customer_ids = request.POST.getlist("item_customer[]")
                item_class_ids    = request.POST.getlist("item_class[]")

                for idx, prod_id in enumerate(item_product_ids):
                    if not prod_id:
                        continue
                    product = Product.objects.filter(pk=prod_id).first()
                    if not product:
                        continue

                    qty  = _dec(item_qtys[idx], "0")
                    rate = _dec(item_rates[idx], "0")
                    amt  = _dec(item_amounts[idx]) if (idx < len(item_amounts) and item_amounts[idx]) else (qty * rate)

                    if amt <= 0:
                        continue

                    is_bill  = str(idx) in item_billable
                    customer = Newcustomer.objects.filter(pk=item_customer_ids[idx] or None).first()
                    klass    = Pclass.objects.filter(pk=item_class_ids[idx] or None).first()

                    ExpenseItemLine.objects.create(
                        expense=exp,
                        product=product,
                        description=(item_descs[idx] or ""),
                        qty=qty,
                        rate=rate,
                        amount=amt,
                        is_billable=is_bill,
                        customer=customer,
                        class_field=klass,
                    )
                    total += amt

                if total <= 0:
                    messages.error(request, "Expense total must be greater than 0.")
                    return redirect("expenses:expense-edit", pk=exp.pk)

                # ---- Update total + post to GL + inventory ----
                exp.total_amount = total
                exp.save(update_fields=["total_amount"])

                _post_expense_to_ledger(exp)
                rebuild_movements_for_expense(exp)

                action = request.POST.get("save_action") or "save"
                if action == "save&new":
                    return redirect("expenses:add-expenses")
                if action == "save":
                    return redirect("expenses:expenses")
                return redirect("expenses:expense-detail", pk=exp.pk)

        except Exception as e:
            messages.error(request, f"Could not update expense: {e}")
            return redirect("expenses:expense-edit", pk=exp.pk)

    context = {
        "expense": exp,
        "accounts": deposit_accounts_qs(),
        "expense_accounts": expense_accounts_qs(),

        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "classes": Pclass.objects.all().order_by("class_name"),

        "cat_lines": exp.cat_lines.select_related("category", "customer", "class_field").all(),
        "item_lines": exp.item_lines.select_related("product", "customer", "class_field").all(),

        "payment_methods": Expense.PAYMENT_METHODS,
    }
    return render(request, "expenses_form.html", context)

# expense list 
def expense_list(request):
    qs = (
        Expense.objects
        .select_related("payee_supplier", "payment_account")
        .prefetch_related("cat_lines__category", "item_lines__product")
        .order_by("-payment_date", "-id")
    )
    return render(request, "expenses_list.html", {"expenses": qs})
# expense detail
def expense_detail(request, pk: int):
    exp = get_object_or_404(
        Expense.objects
        .select_related("payee_supplier", "payment_account")
        .prefetch_related("cat_lines__category", "item_lines__product"),
        pk=pk
    )
    return render(request, "expense_detail.html", {"e": exp,})

# bill views
def _dec(v, default="0.00"):
    try:
        return Decimal(str(v if v not in (None, "") else default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def generate_unique_bill_no(prefix="BILL"):
    """
    8-digit numeric suffix (like 00001234) with a prefix for readability. Ensures uniqueness.
    """
    base_date = timezone.now().strftime("%y%m")  # e.g., '2510'
    seed = f"{base_date}0001"
    suffix = int(seed)
    while True:
        candidate = f"{prefix}{suffix:08d}"
        if not Bill.objects.filter(bill_no=candidate).exists():
            return candidate
        suffix += 1

def _dec(v, default="0.00"):
    try:
        return Decimal(str(v if v not in (None, "",) else default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

def _parse_ymd(s, fallback=None):
    """Parse 'YYYY-MM-DD' safely; returns fallback if empty/invalid."""
    if not s:
        return fallback
    try:
        return timezone.datetime.fromisoformat(s).date()
    except Exception:
        return fallback

def generate_unique_bill_no():
    # Simple example. Replace with your existing generator if you already have one.
    last = Bill.objects.order_by("-id").first()
    base = 10000000 if not last else (int(str(last.bill_no or 0).strip()[-8:]) if str(last.bill_no or "").isdigit() else last.id) + 1
    return f"{base:08d}"

# adding a bill
def _dec(x, default="0.00") -> Decimal:
    try:
        return Decimal(str(x if x not in (None, "") else default))
    except Exception:
        return Decimal(default)

@transaction.atomic
def add_bill(request):
    if request.method == "POST":
        expense_qs = expense_accounts_qs()

        supplier_id      = request.POST.get("supplier_id") or ""
        supplier_name    = request.POST.get("supplier") or ""
        mailing_address  = request.POST.get("mailing_address") or ""
        terms            = request.POST.get("terms") or ""
        bill_date        = _parse_ymd(request.POST.get("bill_date"), timezone.localdate())
        due_date         = _parse_ymd(request.POST.get("due_date"))
        bill_no          = (request.POST.get("bill_no") or "").strip()
        location         = request.POST.get("location") or ""
        memo             = request.POST.get("memo") or ""
        attachment       = request.FILES.get("attachments")

        supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None

        if (not bill_no) or Bill.objects.filter(bill_no=bill_no).exists():
            bill_no = generate_unique_bill_no()

        bill = Bill.objects.create(
            supplier=supplier,
            supplier_name=None if supplier else supplier_name,
            mailing_address=mailing_address,
            terms=terms,
            bill_date=bill_date,
            due_date=due_date,
            bill_no=bill_no,
            location=location,
            memo=memo,
            attachments=attachment,
        )

        total = Decimal("0.00")

        # ---------- Category lines ----------
        cat_category_ids = request.POST.getlist("cat_category[]")
        cat_descs        = request.POST.getlist("cat_desc[]")
        cat_amounts      = request.POST.getlist("cat_amount[]")
        cat_billable     = set(request.POST.getlist("cat_billable[]"))
        cat_customer_ids = request.POST.getlist("cat_customer[]")
        cat_class_ids    = request.POST.getlist("cat_class[]")

        for idx, acc_id in enumerate(cat_category_ids):
            if not acc_id:
                continue
            account = expense_qs.filter(pk=acc_id).first()
            if not account:
                continue

            amt = _dec(cat_amounts[idx])
            if amt <= 0:
                continue

            is_bill  = str(idx) in cat_billable
            customer = Newcustomer.objects.filter(pk=(cat_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(cat_class_ids[idx] or None)).first()

            BillCategoryLine.objects.create(
                bill=bill, category=account,
                description=(cat_descs[idx] or ""),
                amount=amt, is_billable=is_bill,
                customer=customer, class_field=klass
            )
            total += amt

        # ---------- Item lines ----------
        item_product_ids  = request.POST.getlist("item_product[]")
        item_descs        = request.POST.getlist("item_desc[]")
        item_qtys         = request.POST.getlist("item_qty[]")
        item_rates        = request.POST.getlist("item_rate[]")
        item_amounts      = request.POST.getlist("item_amount[]")
        item_billable     = set(request.POST.getlist("item_billable[]"))
        item_customer_ids = request.POST.getlist("item_customer[]")
        item_class_ids    = request.POST.getlist("item_class[]")

        for idx, prod_id in enumerate(item_product_ids):
            if not prod_id:
                continue
            product = Product.objects.filter(pk=prod_id).first()
            if not product:
                continue

            qty  = _dec(item_qtys[idx], "0")
            rate = _dec(item_rates[idx], "0")
            amt  = _dec(item_amounts[idx]) if (idx < len(item_amounts) and item_amounts[idx]) else (qty * rate)
            if amt <= 0:
                continue

            is_bill  = str(idx) in item_billable
            customer = Newcustomer.objects.filter(pk=(item_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(item_class_ids[idx] or None)).first()

            BillItemLine.objects.create(
                bill=bill, product=product,
                description=(item_descs[idx] or ""),
                qty=qty, rate=rate, amount=amt,
                is_billable=is_bill, customer=customer, class_field=klass
            )
            total += amt

        if total <= 0:
            bill.delete()
            return redirect("expenses:add-bill")

        bill.total_amount = total
        bill.save(update_fields=["total_amount"])

        # Post to ledger: Dr Expenses/Items, Cr Accounts Payable
        _post_bill_to_ledger(bill)
        rebuild_movements_for_bill(bill)
        action = request.POST.get("save_action") or "save"
        if action == "save":
            return redirect("expenses:bills-list")
        if action == "save&new":
            return redirect("expenses:add-bill")
        return redirect("expenses:bills-list")

    context = {
        "expense_accounts": expense_accounts_qs(),
        "all_accounts": Account.objects.all().order_by("account_name"),
        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "generated_bill_no": generate_unique_bill_no(),
    }
    return render(request, "bill_form.html", context)


# edit bill
@transaction.atomic
def edit_bill(request, pk: int):
    bill = get_object_or_404(
        Bill.objects.select_related("supplier")
            .prefetch_related("category_lines__category", "item_lines__product"),
        pk=pk
    )

    if request.method == "POST":
        expense_qs = expense_accounts_qs()

        supplier_id      = request.POST.get("supplier_id") or ""
        supplier_manual  = request.POST.get("supplier") or ""

        bill.mailing_address = request.POST.get("mailing_address") or ""
        bill.terms           = request.POST.get("terms") or ""
        bill.bill_date       = _parse_ymd(request.POST.get("bill_date"), bill.bill_date or timezone.localdate())
        bill.due_date        = _parse_ymd(request.POST.get("due_date"))
        new_bill_no          = (request.POST.get("bill_no") or "").strip()
        bill.location        = request.POST.get("location") or ""
        bill.memo            = request.POST.get("memo") or ""

        if new_bill_no and new_bill_no != (bill.bill_no or ""):
            if Bill.objects.exclude(pk=bill.pk).filter(bill_no=new_bill_no).exists():
                messages.error(request, "Bill No. already exists. Please use another number.")
                return redirect("expenses:bill-edit", pk=bill.pk)
            bill.bill_no = new_bill_no

        supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None
        bill.supplier = supplier
        bill.supplier_name = None if supplier else supplier_manual

        if request.FILES.get("attachments"):
            bill.attachments = request.FILES["attachments"]

        bill.save()

        # replace lines
        BillCategoryLine.objects.filter(bill=bill).delete()
        BillItemLine.objects.filter(bill=bill).delete()

        total = Decimal("0.00")

        # ---- Category lines ----
        cat_category_ids = request.POST.getlist("cat_category[]")
        cat_descs        = request.POST.getlist("cat_desc[]")
        cat_amounts      = request.POST.getlist("cat_amount[]")
        cat_billable     = set(request.POST.getlist("cat_billable[]"))
        cat_customer_ids = request.POST.getlist("cat_customer[]")
        cat_class_ids    = request.POST.getlist("cat_class[]")

        for idx, acc_id in enumerate(cat_category_ids):
            if not acc_id:
                continue
            account = expense_qs.filter(pk=acc_id).first()
            if not account:
                continue

            amt = _dec(cat_amounts[idx])
            if amt <= 0:
                continue

            is_bill  = str(idx) in cat_billable
            customer = Newcustomer.objects.filter(pk=(cat_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(cat_class_ids[idx] or None)).first()

            BillCategoryLine.objects.create(
                bill=bill, category=account,
                description=(cat_descs[idx] or ""),
                amount=amt, is_billable=is_bill,
                customer=customer, class_field=klass
            )
            total += amt

        # ---- Item lines ----
        item_product_ids  = request.POST.getlist("item_product[]")
        item_descs        = request.POST.getlist("item_desc[]")
        item_qtys         = request.POST.getlist("item_qty[]")
        item_rates        = request.POST.getlist("item_rate[]")
        item_amounts      = request.POST.getlist("item_amount[]")
        item_billable     = set(request.POST.getlist("item_billable[]"))
        item_customer_ids = request.POST.getlist("item_customer[]")
        item_class_ids    = request.POST.getlist("item_class[]")

        for idx, prod_id in enumerate(item_product_ids):
            if not prod_id:
                continue
            product = Product.objects.filter(pk=prod_id).first()
            if not product:
                continue

            qty  = _dec(item_qtys[idx], "0")
            rate = _dec(item_rates[idx], "0")
            amt  = _dec(item_amounts[idx]) if (idx < len(item_amounts) and item_amounts[idx]) else (qty * rate)
            if amt <= 0:
                continue

            is_bill  = str(idx) in item_billable
            customer = Newcustomer.objects.filter(pk=(item_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(item_class_ids[idx] or None)).first()

            BillItemLine.objects.create(
                bill=bill, product=product,
                description=(item_descs[idx] or ""),
                qty=qty, rate=rate, amount=amt,
                is_billable=is_bill, customer=customer, class_field=klass
            )
            total += amt

        if total <= 0:
            return redirect("expenses:bill-edit", pk=bill.pk)

        bill.total_amount = total
        bill.save(update_fields=["total_amount"])

        # Update ledger entry (rewrite lines)
        _post_bill_to_ledger(bill)
        rebuild_movements_for_bill(bill)
        action = request.POST.get("save_action") or "save"
        if action == "save&new":
            return redirect("expenses:add-bill")
        return redirect("expenses:bills-list")

    context = {
        "bill": bill,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "expense_accounts": expense_accounts_qs(),
        "all_accounts": Account.objects.all().order_by("account_name"),
        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "cat_lines": BillCategoryLine.objects.filter(bill=bill).select_related("category","customer","class_field"),
        "item_lines": BillItemLine.objects.filter(bill=bill).select_related("product","customer","class_field"),
    }
    return render(request, "bill_form.html", context)

# bill list
def bills_list(request):
    """
    Bills list with search, date filter and pagination.
    """
    today = timezone.localdate()
    qs = (
        Bill.objects
        .select_related("supplier")
        .order_by("-bill_date", "-id")
    )

    # ----- Filters (GET) -----
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(bill_no__icontains=q) |
            Q(location__icontains=q) |
            Q(supplier__company_name__icontains=q)
        )

    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")
    if date_from:
        qs = qs.filter(bill_date__gte=date_from)
    if date_to:
        qs = qs.filter(bill_date__lte=date_to)

    # Simple status chip (Open/Overdue/Closed) computed on the fly:
    # If you have a stored status field, you can display that instead.
    rows = []
    for b in qs:
        status = "Open"
        if b.due_date and b.due_date < today:
            status = "Overdue"
        # if you later add payments + balance logic, set "Closed" when fully paid
        rows.append((b, status))

    # Totals (for the current filtered set)
    totals = qs.aggregate(
        grand=Coalesce(Sum("total_amount"), Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))),
    )
    # Keep keys your template expects; if you don't track them on the model, set 0
    totals["subtotal"] = Decimal("0.00")
    totals["tax"] = Decimal("0.00")
    # ----- Pagination -----
    paginator = Paginator(rows, 25)  # 25 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
        "from": date_from,
        "to": date_to,
        "totals": totals,
        "count_all": qs.count(),
    }
    return render(request, "bill_list.html", context)
# bill detail

def bill_detail(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related("supplier").prefetch_related(
            Prefetch(
                "category_lines",
                queryset=BillCategoryLine.objects.select_related("category", "customer", "class_field"),
            ),
            Prefetch(
                "item_lines",
                queryset=BillItemLine.objects.select_related("product", "customer", "class_field"),
            ),
        ),
        pk=pk,
    )

    cat_total  = bill.category_lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    item_total = bill.item_lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    subtotal   = cat_total + item_total

    context = {
        "bill": bill,
        "cat_total": cat_total,
        "item_total": item_total,
        "subtotal": subtotal,
    }
    return render(request, "bill_detail.html", context)

# end

# cheque view

def _dec(v, default="0.00"):
    try:
        return Decimal(str(v or default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

def generate_unique_cheque_no():
    # Simple sequential fallback: last cheque id + 1, padded
    last = Cheque.objects.order_by("-id").first()
    nxt = (last.id + 1) if last else 1
    return f"{nxt:06d}"
def _account_credit_balance(account: "Account") -> Decimal:
    """
    Returns credit balance for an account:
      opening_balance + SUM(credit - debit)
    If negative, return 0 for our purpose here.
    """
    opening = Decimal(str(getattr(account, "opening_balance", 0) or "0"))
    agg = (
        JournalLine.objects
        .filter(account=account)
        .aggregate(
            d=Sum("debit"),
            c=Sum("credit"),
        )
    )
    deb = Decimal(str(agg["d"] or "0"))
    cred = Decimal(str(agg["c"] or "0"))

    bal = opening + (cred - deb)
    return bal if bal > 0 else Decimal("0.00")


def _supplier_open_balance_amount(supplier: "Newsupplier") -> Decimal:
    """
    Open balance = Supplier A/P subaccount credit balance - total unpaid bill balances
    (Anything in supplier A/P not represented by open bill balances is considered "open balance".)
    """
    if not supplier:
        return Decimal("0.00")

    supplier_acc = _get_or_create_supplier_ap_subaccount(supplier)

    supplier_ap_bal = _account_credit_balance(supplier_acc)  # credit balance in A/P

    # total unpaid bills for this supplier
    total_unpaid_bills = Decimal("0.00")
    for b in Bill.objects.filter(supplier=supplier):
        total_unpaid_bills += _bill_balance(b)

    open_bal = supplier_ap_bal - total_unpaid_bills
    return open_bal if open_bal > 0 else Decimal("0.00")


def _save_cheque_open_balance(request, cheque: "Cheque"):
    """
    Reads posted field: open_balance_amount
    Saves it as ChequeOpenBalanceLine (edit-safe).
    """
    # delete old line if any (edit-safe)
    ChequeOpenBalanceLine.objects.filter(cheque=cheque).delete()

    raw = request.POST.get("open_balance_amount")
    if raw is None or raw == "":
        return

    amt = _dec(raw, "0")
    if amt <= 0:
        return

    # NOTE: do NOT clamp. If they enter more than current open balance, that becomes supplier credit.
    ChequeOpenBalanceLine.objects.create(
        cheque=cheque,
        amount_applied=amt
    )

@require_GET
def outstanding_bills_api(request):
    supplier_id = request.GET.get("supplier")
    if not supplier_id:
        return JsonResponse({"bills": [], "open_balance": "0.00"})

    supplier = Newsupplier.objects.filter(pk=supplier_id).first()
    if not supplier:
        return JsonResponse({"bills": [], "open_balance": "0.00"})

    bills_payload = []

    # compute balances map
    applied_map = dict(
        ChequeBillLine.objects
        .filter(bill__supplier_id=supplier_id)
        .values("bill_id").annotate(s=Sum("amount_applied"))
        .values_list("bill_id", "s")
    )

    for b in Bill.objects.filter(supplier_id=supplier_id).order_by("-bill_date"):
        total = Decimal(str(b.total_amount or "0"))
        applied = Decimal(str(applied_map.get(b.id) or "0"))
        balance = total - applied
        if balance <= 0:
            continue

        bills_payload.append({
            "id": b.id,
            "bill_no": b.bill_no,
            "bill_date": b.bill_date.strftime("%Y-%m-%d") if b.bill_date else None,
            "due_date": b.due_date.strftime("%Y-%m-%d") if b.due_date else None,
            "total": str(total),
            "balance": str(balance),
        })

    open_balance = _supplier_open_balance_amount(supplier)

    return JsonResponse({
        "bills": bills_payload,
        "open_balance": str(open_balance),
    })

# add cheque
@transaction.atomic
def add_cheque(request):
    if request.method == "POST":
        expense_qs = expense_accounts_qs()

        supplier_id = request.POST.get("payee_supplier") or ""
        payee_name  = request.POST.get("payee_name") or ""
        bank_id     = request.POST.get("bank_account") or ""
        mailing     = request.POST.get("mailing_address") or ""
        payment_date= _parse_ymd(request.POST.get("payment_date"), timezone.localdate())
        cheque_no   = (request.POST.get("cheque_no") or "").strip()
        location    = request.POST.get("location") or ""
        memo        = request.POST.get("memo") or ""
        attachment  = request.FILES.get("attachments")

        supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None
        bank_acc = Account.objects.filter(pk=bank_id).first() if bank_id else None

        with transaction.atomic():
            cheque = Cheque.objects.create(
                payee_supplier=supplier,
                payee_name="" if supplier else (payee_name or ""),
                bank_account=bank_acc,
                mailing_address=mailing,
                payment_date=payment_date,
                cheque_no=cheque_no,
                location=location,
                memo=memo,
                attachments=attachment,
            )

            total_direct = Decimal("0.00")

            # ---------- Category lines ----------
            cat_ids   = request.POST.getlist("cat_category[]")
            cat_descs = request.POST.getlist("cat_desc[]")
            cat_amts  = request.POST.getlist("cat_amount[]")
            cat_billable = set(request.POST.getlist("cat_billable[]"))
            cat_cust  = request.POST.getlist("cat_customer[]")
            cat_cls   = request.POST.getlist("cat_class[]")

            for idx, acc_id in enumerate(cat_ids):
                if not acc_id:
                    continue
                acc = expense_qs.filter(pk=acc_id).first() or Account.objects.filter(pk=acc_id).first()
                if not acc:
                    continue
                amt = _dec(cat_amts[idx])
                if amt <= 0:
                    continue

                is_bill = str(idx) in cat_billable
                customer = Newcustomer.objects.filter(pk=(cat_cust[idx] or None)).first()
                klass    = Pclass.objects.filter(pk=(cat_cls[idx] or None)).first()

                ChequeCategoryLine.objects.create(
                    cheque=cheque,
                    category=acc,
                    description=(cat_descs[idx] or ""),
                    amount=amt,
                    is_billable=is_bill,
                    customer=customer,
                    class_field=klass,
                )
                total_direct += amt

            # ---------- Item lines ----------
            item_prod = request.POST.getlist("item_product[]")
            item_desc = request.POST.getlist("item_desc[]")
            item_qty  = request.POST.getlist("item_qty[]")
            item_rate = request.POST.getlist("item_rate[]")
            item_amt  = request.POST.getlist("item_amount[]")
            item_billable = set(request.POST.getlist("item_billable[]"))
            item_cust = request.POST.getlist("item_customer[]")
            item_cls  = request.POST.getlist("item_class[]")

            for idx, pid in enumerate(item_prod):
                if not pid:
                    continue
                product = Product.objects.filter(pk=pid).first()
                if not product:
                    continue

                qty  = _dec(item_qty[idx], "0")
                rate = _dec(item_rate[idx], "0")
                amt  = _dec(item_amt[idx]) if (idx < len(item_amt) and item_amt[idx]) else (qty * rate)
                if amt <= 0:
                    continue

                is_bill = str(idx) in item_billable
                customer = Newcustomer.objects.filter(pk=(item_cust[idx] or None)).first()
                klass    = Pclass.objects.filter(pk=(item_cls[idx] or None)).first()

                ChequeItemLine.objects.create(
                    cheque=cheque,
                    product=product,
                    description=(item_desc[idx] or ""),
                    qty=qty,
                    rate=rate,
                    amount=amt,
                    is_billable=is_bill,
                    customer=customer,
                    class_field=klass,
                )
                total_direct += amt

            # ✅ Save allocations (bills)
            _save_cheque_bill_allocations(request, cheque)

            alloc_total = (
                ChequeBillLine.objects.filter(cheque=cheque)
                .aggregate(s=Sum("amount_applied"))["s"]
                or Decimal("0.00")
            )

            # ✅ Save open balance amount
            _save_cheque_open_balance(request, cheque)

            open_total = (
                ChequeOpenBalanceLine.objects.filter(cheque=cheque)
                .aggregate(s=Sum("amount_applied"))["s"]
                or Decimal("0.00")
            )

            # ✅ Total cheque = bills + open balance + direct expenses
            cheque.total_amount = alloc_total + open_total + total_direct
            cheque.save(update_fields=["total_amount"])

            # ✅ Post to GL (supports bills + open balance + direct)
            _post_cheque_to_ledger(cheque)

        action = request.POST.get("save_action") or "save"
        if action == "save&new":
            return redirect("expenses:add-cheque")
        if action == "save&close":
            return redirect("expenses:expenses")
        return redirect("expenses:expenses")

    context = {
        "cheque": None,
        "today": timezone.localdate(),
        "generated_cheque_no": "CHQ-" + timezone.now().strftime("%H%M%S"),
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "products": Product.objects.all().order_by("name"),
        "expense_accounts": expense_accounts_qs(),
        "bank_accounts": Account.objects.filter(is_active=True).filter(bankish_q()).order_by("account_name"),
        "cat_lines": [],
        "item_lines": [],
        "bill_prefill": [],
        "open_balance_prefill": None,   # for edit only (kept)
    }
    return render(request, "cheque_form.html", context)



# ----------------------------
# CHEQUE EDIT
# ----------------------------
def cheque_edit(request, pk: int):
    cheque = get_object_or_404(Cheque, pk=pk)

    if request.method == "POST":
        expense_qs = expense_accounts_qs()

        supplier_id = request.POST.get("payee_supplier") or ""
        payee_name  = request.POST.get("payee_name") or ""
        bank_id     = request.POST.get("bank_account") or ""
        mailing     = request.POST.get("mailing_address") or ""
        payment_date= _parse_ymd(request.POST.get("payment_date"), cheque.payment_date)
        location    = request.POST.get("location") or ""
        memo        = request.POST.get("memo") or ""
        attachment  = request.FILES.get("attachments")

        supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None
        bank_acc = Account.objects.filter(pk=bank_id).first() if bank_id else None

        with transaction.atomic():
            cheque.payee_supplier  = supplier
            cheque.payee_name      = "" if supplier else (payee_name or "")
            cheque.bank_account    = bank_acc
            cheque.mailing_address = mailing
            cheque.payment_date    = payment_date
            cheque.location        = location
            cheque.memo            = memo
            if attachment:
                cheque.attachments = attachment
            cheque.save()

            ChequeCategoryLine.objects.filter(cheque=cheque).delete()
            ChequeItemLine.objects.filter(cheque=cheque).delete()

            total_direct = Decimal("0.00")

            # ---------- Category lines ----------
            cat_ids   = request.POST.getlist("cat_category[]")
            cat_descs = request.POST.getlist("cat_desc[]")
            cat_amts  = request.POST.getlist("cat_amount[]")
            cat_billable = set(request.POST.getlist("cat_billable[]"))
            cat_cust  = request.POST.getlist("cat_customer[]")
            cat_cls   = request.POST.getlist("cat_class[]")

            for idx, acc_id in enumerate(cat_ids):
                if not acc_id:
                    continue
                acc = expense_qs.filter(pk=acc_id).first() or Account.objects.filter(pk=acc_id).first()
                if not acc:
                    continue
                amt = _dec(cat_amts[idx])
                if amt <= 0:
                    continue

                is_bill = str(idx) in cat_billable
                customer = Newcustomer.objects.filter(pk=(cat_cust[idx] or None)).first()
                klass    = Pclass.objects.filter(pk=(cat_cls[idx] or None)).first()

                ChequeCategoryLine.objects.create(
                    cheque=cheque,
                    category=acc,
                    description=(cat_descs[idx] or ""),
                    amount=amt,
                    is_billable=is_bill,
                    customer=customer,
                    class_field=klass,
                )
                total_direct += amt

            # ---------- Item lines ----------
            item_prod = request.POST.getlist("item_product[]")
            item_desc = request.POST.getlist("item_desc[]")
            item_qty  = request.POST.getlist("item_qty[]")
            item_rate = request.POST.getlist("item_rate[]")
            item_amt  = request.POST.getlist("item_amount[]")
            item_billable = set(request.POST.getlist("item_billable[]"))
            item_cust = request.POST.getlist("item_customer[]")
            item_cls  = request.POST.getlist("item_class[]")

            for idx, pid in enumerate(item_prod):
                if not pid:
                    continue
                product = Product.objects.filter(pk=pid).first()
                if not product:
                    continue

                qty  = _dec(item_qty[idx], "0")
                rate = _dec(item_rate[idx], "0")
                amt  = _dec(item_amt[idx]) if (idx < len(item_amt) and item_amt[idx]) else (qty * rate)
                if amt <= 0:
                    continue

                is_bill = str(idx) in item_billable
                customer = Newcustomer.objects.filter(pk=(item_cust[idx] or None)).first()
                klass    = Pclass.objects.filter(pk=(item_cls[idx] or None)).first()

                ChequeItemLine.objects.create(
                    cheque=cheque,
                    product=product,
                    description=(item_desc[idx] or ""),
                    qty=qty,
                    rate=rate,
                    amount=amt,
                    is_billable=is_bill,
                    customer=customer,
                    class_field=klass,
                )
                total_direct += amt

            # ✅ Save allocations (bills)
            _save_cheque_bill_allocations(request, cheque)
            alloc_total = (
                ChequeBillLine.objects.filter(cheque=cheque)
                .aggregate(s=Sum("amount_applied"))["s"]
                or Decimal("0.00")
            )

            # ✅ Save open balance
            _save_cheque_open_balance(request, cheque)
            open_total = (
                ChequeOpenBalanceLine.objects.filter(cheque=cheque)
                .aggregate(s=Sum("amount_applied"))["s"]
                or Decimal("0.00")
            )

            cheque.total_amount = alloc_total + open_total + total_direct
            cheque.save(update_fields=["total_amount"])

            _post_cheque_to_ledger(cheque)

        action = request.POST.get("save_action") or "save"
        if action == "save&new":
            return redirect("expenses:add-cheque")
        if action == "save&close":
            return redirect("expenses:expenses")
        return redirect("expenses:expenses")

    # GET prefill
    cat_lines = list(ChequeCategoryLine.objects.filter(cheque=cheque).values(
        "category_id", "description", "amount", "is_billable", "customer_id", "class_field_id"
    ))
    item_lines = list(ChequeItemLine.objects.filter(cheque=cheque).values(
        "product_id", "description", "qty", "rate", "amount", "is_billable", "customer_id", "class_field_id"
    ))

    bill_prefill = []
    open_balance_prefill = None
    open_balance_value = Decimal("0.00")

    if cheque.payee_supplier_id:
        allocs = {x.bill_id: x.amount_applied for x in ChequeBillLine.objects.filter(cheque=cheque)}

        applied_map = dict(
            ChequeBillLine.objects
            .filter(bill__supplier_id=cheque.payee_supplier_id)
            .values("bill_id").annotate(s=Sum("amount_applied"))
            .values_list("bill_id", "s")
        )

        for b in Bill.objects.filter(supplier_id=cheque.payee_supplier_id).order_by("-bill_date"):
            total = Decimal(str(b.total_amount or "0"))
            applied_all = Decimal(str(applied_map.get(b.id) or "0"))
            balance = total - applied_all
            if balance <= 0:
                continue

            bill_prefill.append({
                "id": b.id,
                "bill_no": b.bill_no,
                "bill_date": b.bill_date,
                "due_date": b.due_date,
                "total": total,
                "balance": balance,
                "applied": allocs.get(b.id, Decimal("0.00")),
            })

        # current computed open balance
        open_balance_value = _supplier_open_balance_amount(cheque.payee_supplier)

        # prefill what user applied on this cheque
        ob_line = ChequeOpenBalanceLine.objects.filter(cheque=cheque).first()
        open_balance_prefill = Decimal(str(ob_line.amount_applied)) if ob_line else Decimal("0.00")

    context = {
        "cheque": cheque,
        "today": timezone.localdate(),
        "generated_cheque_no": cheque.cheque_no,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "products": Product.objects.all().order_by("name"),
        "expense_accounts": expense_accounts_qs(),
        "bank_accounts": Account.objects.filter(is_active=True).filter(bankish_q()).order_by("account_name"),
        "cat_lines": cat_lines,
        "item_lines": item_lines,
        "bill_prefill": bill_prefill,
        "open_balance_value": open_balance_value,
        "open_balance_prefill": open_balance_prefill,
    }
    return render(request, "cheque_form.html", context)

# cheque lists 
def cheque_list(request):
    """
    List of cheques with search, date filter and pagination.
    """
    today = timezone.localdate()
    qs = (
        Cheque.objects
        .select_related("payee_supplier", "bank_account")
        .order_by("-payment_date", "-id")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(cheque_no__icontains=q) |
            Q(payee_name__icontains=q) |
            Q(payee_supplier__company_name__icontains=q) |
            Q(bank_account__account_name__icontains=q)
        )

    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")
    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    totals = qs.aggregate(
        grand=Coalesce(
            Sum("total_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
        "from": date_from,
        "to": date_to,
        "totals": totals,
        "count_all": qs.count(),
    }
    return render(request, "cheque_list.html", context) 
# cheque detail 
def cheque_detail(request, pk: int):
    chq = get_object_or_404(
        Cheque.objects
        .select_related("payee_supplier", "bank_account")
        .prefetch_related(
            Prefetch(
                "category_lines",
                queryset=ChequeCategoryLine.objects.select_related("category", "customer", "class_field"),
            ),
            Prefetch(
                "item_lines",
                queryset=ChequeItemLine.objects.select_related("product", "customer", "class_field"),
            ),
        ),
        pk=pk,
    )

    cat_total  = chq.category_lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    item_total = chq.item_lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    subtotal   = cat_total + item_total

    context = {
        "cheque": chq,
        "cat_total": cat_total,
        "item_total": item_total,
        "subtotal": subtotal,
    }
    return render(request, "cheque_detail.html", context)
# adding a purchase order
@transaction.atomic
def purchase_order(request):
    """
    Create a new Purchase Order.
    Non-posting (no GL) – just records the order and shows in All Expenses.
    """
    if request.method == "POST":
        vendor_id       = request.POST.get("vendor_id") or ""
        vendor_name     = request.POST.get("vendor_name") or ""
        mailing_address = request.POST.get("mailing_address") or ""
        po_date         = _parse_ymd(request.POST.get("po_date"), timezone.localdate())
        deliver_by      = _parse_ymd(request.POST.get("deliver_by"))
        ship_to         = request.POST.get("ship_to") or ""
        location        = request.POST.get("location") or ""
        memo            = request.POST.get("memo") or ""
        attachment      = request.FILES.get("attachments")
        po_number       = (request.POST.get("po_number") or "").strip()

        vendor = Newsupplier.objects.filter(pk=vendor_id).first() if vendor_id else None

        if (not po_number) or PurchaseOrder.objects.filter(po_number=po_number).exists():
            po_number = generate_unique_po_no()

        po = PurchaseOrder.objects.create(
            vendor=vendor,
            vendor_name=None if vendor else vendor_name,
            mailing_address=mailing_address,
            po_date=po_date,
            deliver_by=deliver_by,
            ship_to=ship_to,
            location=location,
            po_number=po_number,
            memo=memo,
            attachments=attachment,
        )

        total = Decimal("0.00")

        # ----- Item lines -----
        item_product_ids  = request.POST.getlist("item_product[]")
        item_descs        = request.POST.getlist("item_desc[]")
        item_qtys         = request.POST.getlist("item_qty[]")
        item_rates        = request.POST.getlist("item_rate[]")
        item_amounts      = request.POST.getlist("item_amount[]")
        item_customer_ids = request.POST.getlist("item_customer[]")
        item_class_ids    = request.POST.getlist("item_class[]")

        for idx, prod_id in enumerate(item_product_ids):
            if not prod_id:
                continue
            product = Product.objects.filter(pk=prod_id).first()
            if not product:
                continue

            qty  = _dec(item_qtys[idx], "0")
            rate = _dec(item_rates[idx], "0")
            amt  = _dec(item_amounts[idx]) if item_amounts[idx] else (qty * rate)
            if amt <= 0:
                continue

            customer = Newcustomer.objects.filter(pk=(item_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(item_class_ids[idx] or None)).first()

            PurchaseOrderLine.objects.create(
                purchase_order=po,
                product=product,
                description=(item_descs[idx] or ""),
                qty=qty,
                rate=rate,
                amount=amt,
                customer=customer,
                class_field=klass,
            )
            total += amt

        if total <= 0:
            po.delete()
            messages.error(request, "You must add at least one line with an amount.")
            return redirect("expenses:purchase_order")

        po.total_amount = total
        po.save(update_fields=["total_amount"])

        # No GL posting yet (PO is a non-posting document)

        action = request.POST.get("save_action") or "save"
        if action == "save&new":
            return redirect("expenses:purchase_order")
        # save / save&close – send to All Expenses
        return redirect("expenses:expenses")

    # GET – render blank form
    context = {
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "generated_po_number": generate_unique_po_no(),
        "today": timezone.localdate(),
        "po": None,
        "item_lines": [],
    }
    return render(request, "purchase_order_form.html", context)

# purchase order edit
@transaction.atomic
def purchase_order_edit(request, pk: int):
    po = get_object_or_404(
        PurchaseOrder.objects.prefetch_related("lines__product", "lines__customer", "lines__class_field"),
        pk=pk
    )

    if request.method == "POST":
        vendor_id       = request.POST.get("vendor_id") or ""
        vendor_name     = request.POST.get("vendor_name") or ""
        po.mailing_address = request.POST.get("mailing_address") or ""
        po.po_date         = _parse_ymd(request.POST.get("po_date"), po.po_date or timezone.localdate())
        po.deliver_by      = _parse_ymd(request.POST.get("deliver_by"))
        po.ship_to         = request.POST.get("ship_to") or ""
        po.location        = request.POST.get("location") or ""
        po.memo            = request.POST.get("memo") or ""

        new_po_number = (request.POST.get("po_number") or "").strip()
        if new_po_number and new_po_number != (po.po_number or ""):
            if PurchaseOrder.objects.exclude(pk=po.pk).filter(po_number=new_po_number).exists():
                messages.error(request, "PO No. already exists. Please use another number.")
                return redirect("expenses:purchase-order-edit", pk=po.pk)
            po.po_number = new_po_number

        vendor = Newsupplier.objects.filter(pk=vendor_id).first() if vendor_id else None
        po.vendor = vendor
        po.vendor_name = None if vendor else vendor_name

        if request.FILES.get("attachments"):
            po.attachments = request.FILES["attachments"]

        po.save()

        # Replace lines
        PurchaseOrderLine.objects.filter(purchase_order=po).delete()

        total = Decimal("0.00")

        item_product_ids  = request.POST.getlist("item_product[]")
        item_descs        = request.POST.getlist("item_desc[]")
        item_qtys         = request.POST.getlist("item_qty[]")
        item_rates        = request.POST.getlist("item_rate[]")
        item_amounts      = request.POST.getlist("item_amount[]")
        item_customer_ids = request.POST.getlist("item_customer[]")
        item_class_ids    = request.POST.getlist("item_class[]")

        for idx, prod_id in enumerate(item_product_ids):
            if not prod_id:
                continue
            product = Product.objects.filter(pk=prod_id).first()
            if not product:
                continue

            qty  = _dec(item_qtys[idx], "0")
            rate = _dec(item_rates[idx], "0")
            amt  = _dec(item_amounts[idx]) if item_amounts[idx] else (qty * rate)
            if amt <= 0:
                continue

            customer = Newcustomer.objects.filter(pk=(item_customer_ids[idx] or None)).first()
            klass    = Pclass.objects.filter(pk=(item_class_ids[idx] or None)).first()

            PurchaseOrderLine.objects.create(
                purchase_order=po,
                product=product,
                description=(item_descs[idx] or ""),
                qty=qty,
                rate=rate,
                amount=amt,
                customer=customer,
                class_field=klass,
            )
            total += amt

        if total <= 0:
            messages.error(request, "You must add at least one line with an amount.")
            return redirect("expenses:purchase-order-edit", pk=po.pk)

        po.total_amount = total
        po.save(update_fields=["total_amount"])

        action = request.POST.get("save_action") or "save"
        if action == "save&new":
            return redirect("expenses:purchase_order")
        return redirect("expenses:purchase-order-detail", pk=po.pk)

    context = {
        "po": po,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "products": Product.objects.all().order_by("name"),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "item_lines": po.lines.select_related("product", "customer", "class_field").all(),
    }
    return render(request, "purchase_order_form.html", context)

# purchase order detail page
def purchase_order_detail(request, pk: int):
    po = get_object_or_404(
        PurchaseOrder.objects.prefetch_related(
            "lines__product", "lines__customer", "lines__class_field"
        ),
        pk=pk,
    )
    subtotal = po.lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    context = {
        "po": po,
        "subtotal": subtotal,
    }
    return render(request, "purchase_order_detail.html", context)

# purchase order lists
def purchase_order_list(request):
    qs = (
        PurchaseOrder.objects
        .select_related("vendor")
        .order_by("-po_date", "-id")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(po_number__icontains=q) |
            Q(location__icontains=q) |
            Q(vendor__company_name__icontains=q)
        )

    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")
    if date_from:
        qs = qs.filter(po_date__gte=date_from)
    if date_to:
        qs = qs.filter(po_date__lte=date_to)

    totals = qs.aggregate(
        grand=Coalesce(
            Sum("total_amount"),
            Value(Decimal("0.00"),
                  output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
        "from": date_from,
        "to": date_to,
        "totals": totals,
    }
    return render(request, "purchase_order_list.html", context)
# supplier credit
@transaction.atomic
def add_supplier_credit(request):
    """
    Create a Supplier Credit (Supplier Credit Note).
    Behaves similarly to a 'negative bill'.
    """
    if request.method == "POST":
        try:
            with transaction.atomic():
                supplier_id     = request.POST.get("supplier_id") or ""
                supplier_name   = request.POST.get("supplier_name") or ""
                mailing_address = request.POST.get("mailing_address") or ""
                credit_date     = request.POST.get("credit_date") or timezone.localdate()
                ref_no          = (request.POST.get("ref_no") or "").strip()
                location        = request.POST.get("location") or ""
                memo            = request.POST.get("memo") or ""
                attachment      = request.FILES.get("attachments")

                supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None

                # ensure unique ref_no (reusing your generate_unique_ref_no helper)
                if not ref_no:
                    ref_no = generate_unique_ref_no()
                if SupplierCredit.objects.filter(ref_no=ref_no).exists():
                    ref_no = generate_unique_ref_no()

                credit = SupplierCredit.objects.create(
                    supplier=supplier,
                    supplier_name=None if supplier else supplier_name,
                    mailing_address=mailing_address,
                    credit_date=credit_date,
                    ref_no=ref_no,
                    location=location,
                    memo=memo,
                    attachments=attachment,
                )

                total = Decimal("0.00")
                expense_accounts_qs_ = expense_accounts_qs()

                # ----- Category lines -----
                cat_dates        = request.POST.getlist("cat_date[]")
                cat_category_ids = request.POST.getlist("cat_category[]")
                cat_descs        = request.POST.getlist("cat_desc[]")
                cat_amounts      = request.POST.getlist("cat_amount[]")
                cat_billable     = set(request.POST.getlist("cat_billable[]"))
                cat_customer_ids = request.POST.getlist("cat_customer[]")
                cat_class_ids    = request.POST.getlist("cat_class[]")

                for idx, acc_id in enumerate(cat_category_ids):
                    if not acc_id:
                        continue

                    account = expense_accounts_qs_.filter(pk=acc_id).first()
                    if not account:
                        continue

                    amt = _dec(cat_amounts[idx])
                    if amt <= 0:
                        continue

                    line_date = cat_dates[idx] or None
                    is_bill   = str(idx) in cat_billable
                    customer  = Newcustomer.objects.filter(pk=(cat_customer_ids[idx] or None)).first()
                    klass     = Pclass.objects.filter(pk=(cat_class_ids[idx] or None)).first()

                    SupplierCreditLine.objects.create(
                        supplier_credit=credit,
                        line_date=line_date or None,
                        category=account,
                        description=(cat_descs[idx] or ""),
                        amount=amt,
                        is_billable=is_bill,
                        customer=customer,
                        class_field=klass,
                    )
                    total += amt

                if total <= 0:
                    credit.delete()
                    messages.error(request, "Supplier credit must have a positive total.")
                    return redirect("expenses:supplier-credit")

                credit.total_amount = total
                credit.save(update_fields=["total_amount"])

                # Post to GL
                # _post_supplier_credit_to_ledger(credit)

                action = request.POST.get("save_action") or "save"
                if action == "save":
                    return redirect("expenses:supplier-credit-list")
                if action == "save&new":
                    return redirect("expenses:supplier-credit")
                if action == "save&close":
                    return redirect("expenses:supplier-credit-list")

        except Exception as e:
            messages.error(request, f"Could not save supplier credit: {e}")
            return redirect("expenses:supplier-credit")

    # GET → show empty form
    context = {
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "expense_accounts": expense_accounts_qs(),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "today": timezone.localdate(),
        "ref_no": generate_unique_ref_no(),
    }
    return render(request, "supplier_credit_form.html", context)

# edit supplier credit

@transaction.atomic
def supplier_credit_edit(request, pk: int):
    credit = get_object_or_404(
        SupplierCredit.objects.select_related("supplier")
        .prefetch_related("lines__category", "lines__customer", "lines__class_field"),
        pk=pk,
    )

    if request.method == "POST":
        try:
            with transaction.atomic():
                supplier_id     = request.POST.get("supplier_id") or ""
                supplier_name   = request.POST.get("supplier_name") or ""
                credit.mailing_address = request.POST.get("mailing_address") or ""
                credit.credit_date     = request.POST.get("credit_date") or credit.credit_date or timezone.localdate()
                new_ref_no             = (request.POST.get("ref_no") or "").strip()
                credit.location        = request.POST.get("location") or ""
                credit.memo            = request.POST.get("memo") or ""

                if new_ref_no and new_ref_no != (credit.ref_no or ""):
                    if SupplierCredit.objects.exclude(pk=credit.pk).filter(ref_no=new_ref_no).exists():
                        messages.error(request, "Ref No. already exists. Use another.")
                        return redirect("expenses:supplier-credit-edit", pk=credit.pk)
                    credit.ref_no = new_ref_no

                supplier = Newsupplier.objects.filter(pk=supplier_id).first() if supplier_id else None
                credit.supplier = supplier
                credit.supplier_name = None if supplier else supplier_name

                if request.FILES.get("attachments"):
                    credit.attachments = request.FILES["attachments"]

                credit.save()

                # Replace lines
                SupplierCreditLine.objects.filter(supplier_credit=credit).delete()

                total = Decimal("0.00")
                expense_accounts_qs_ = expense_accounts_qs()

                cat_dates        = request.POST.getlist("cat_date[]")
                cat_category_ids = request.POST.getlist("cat_category[]")
                cat_descs        = request.POST.getlist("cat_desc[]")
                cat_amounts      = request.POST.getlist("cat_amount[]")
                cat_billable     = set(request.POST.getlist("cat_billable[]"))
                cat_customer_ids = request.POST.getlist("cat_customer[]")
                cat_class_ids    = request.POST.getlist("cat_class[]")

                for idx, acc_id in enumerate(cat_category_ids):
                    if not acc_id:
                        continue
                    account = expense_accounts_qs_.filter(pk=acc_id).first()
                    if not account:
                        continue

                    amt = _dec(cat_amounts[idx])
                    if amt <= 0:
                        continue

                    line_date = cat_dates[idx] or None
                    is_bill   = str(idx) in cat_billable
                    customer  = Newcustomer.objects.filter(pk=(cat_customer_ids[idx] or None)).first()
                    klass     = Pclass.objects.filter(pk=(cat_class_ids[idx] or None)).first()

                    SupplierCreditLine.objects.create(
                        supplier_credit=credit,
                        line_date=line_date or None,
                        category=account,
                        description=(cat_descs[idx] or ""),
                        amount=amt,
                        is_billable=is_bill,
                        customer=customer,
                        class_field=klass,
                    )
                    total += amt

                if total <= 0:
                    messages.error(request, "Supplier credit must have a positive total.")
                    return redirect("expenses:supplier-credit-edit", pk=credit.pk)

                credit.total_amount = total
                credit.save(update_fields=["total_amount"])

                # _post_supplier_credit_to_ledger(credit)

                action = request.POST.get("save_action") or "save"
                if action == "save&new":
                    return redirect("expenses:supplier-credit")
                return redirect("expenses:supplier-credit-list")

        except Exception as e:
            messages.error(request, f"Could not update supplier credit: {e}")
            return redirect("expenses:supplier-credit-edit", pk=credit.pk)

    # GET: prefill
    context = {
        "credit": credit,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "expense_accounts": expense_accounts_qs(),
        "customers": Newcustomer.objects.all().order_by("customer_name"),
        "classes": Pclass.objects.all().order_by("class_name"),
        "lines": credit.lines.select_related("category", "customer", "class_field"),
        "today": timezone.localdate(),
        "ref_no": generate_unique_ref_no(),
    }
    return render(request, "supplier_credit_form.html", context)

# supplier list.

def supplier_credit_list(request):
    qs = (
        SupplierCredit.objects
        .select_related("supplier")
        .order_by("-credit_date", "-id")
    )

    today = timezone.localdate()

    # ----- Search text -----
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(ref_no__icontains=q) |
            Q(location__icontains=q) |
            Q(supplier__company_name__icontains=q) |
            Q(supplier_name__icontains=q)
        )

    # ----- Date filters -----
    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")

    if date_from:
        qs = qs.filter(credit_date__gte=date_from)
    if date_to:
        qs = qs.filter(credit_date__lte=date_to)

    # ----- Build rows (you can just pass qs, but keeping your pattern) -----
    rows = list(qs)

    # ----- Totals -----
    totals = qs.aggregate(
        grand=Coalesce(
            Sum("total_amount"),
            Value(Decimal("0.00"),
                  output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )
    totals["subtotal"] = Decimal("0.00")
    totals["tax"] = Decimal("0.00")

    return render(request, "supplier_credit_list.html", {
        "credits": rows,
        "totals": totals,
        "count_all": qs.count(),
        "q": q,
        "from": date_from,
        "to": date_to,
    })

# supplier detail

def supplier_credit_detail(request, pk: int):
    credit = get_object_or_404(
        SupplierCredit.objects.select_related("supplier").prefetch_related(
            Prefetch(
                "lines",
                queryset=SupplierCreditLine.objects.select_related(
                    "category", "customer", "class_field"
                ),
            )
        ),
        pk=pk,
    )

    total = credit.lines.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")

    context = {
        "credit": credit,
        "total_lines": total,
    }
    return render(request, "supplier_credit_detail.html", context)

# pay down credit
@transaction.atomic
def add_paydown_credit(request):
    # credit card accounts: simple filter
    credit_cards_qs = Account.objects.filter(
        Q(account_type__icontains="credit") |
        Q(detail_type__icontains="credit card")
    ).order_by("account_name")

    bank_accounts_qs = deposit_accounts_qs()  # your existing helper

    if request.method == "POST":
        try:
            with transaction.atomic():
                credit_card_id   = request.POST.get("credit_card") or ""
                bank_account_id  = request.POST.get("bank_account") or ""
                supplier_id      = request.POST.get("payee_supplier") or ""
                payee_name       = request.POST.get("payee_name") or ""
                amount_raw       = request.POST.get("amount") or "0"
                payment_date     = request.POST.get("payment_date") or timezone.localdate()
                ref_no           = request.POST.get("ref_no") or ""
                location         = request.POST.get("location") or ""
                memo             = request.POST.get("memo") or ""
                attachment       = request.FILES.get("attachments")

                credit_card = get_object_or_404(credit_cards_qs, pk=credit_card_id)
                bank_account = get_object_or_404(bank_accounts_qs, pk=bank_account_id)
                supplier = (
                    Newsupplier.objects.filter(pk=supplier_id).first()
                    if supplier_id else None
                )

                amount = _dec(amount_raw)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    raise ValueError("Zero amount")

                if not ref_no:
                    ref_no = generate_unique_ref_no()

                pdc = PayDownCredit.objects.create(
                    credit_card=credit_card,
                    bank_account=bank_account,
                    payee_supplier=supplier,
                    payee_name=payee_name,
                    payment_date=payment_date,
                    amount=amount,
                    ref_no=ref_no,
                    location=location,
                    memo=memo,
                    attachments=attachment,
                )

                # _post_paydown_credit_to_ledger(pdc)

                action = request.POST.get("save_action") or "save"
                if action == "save":
                    return redirect("expenses:paydown-credit-list")
                if action == "save&new":
                    return redirect("expenses:pay-down-credit")
                # save&close
                return redirect("expenses:expenses")

        except Exception as e:
            messages.error(request, f"Could not save payment: {e}")

    # GET → render empty form
    context = {
        "credit_cards": credit_cards_qs,
        "bank_accounts": bank_accounts_qs,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
        "today": timezone.localdate(),
        "ref_no": generate_unique_ref_no(),
    }
    return render(request, "paydown_credit_form.html", context)

# padown edit view 
@transaction.atomic
def paydown_credit_edit(request, pk: int):
    pdc = get_object_or_404(
        PayDownCredit.objects.select_related(
            "credit_card", "bank_account", "payee_supplier"
        ),
        pk=pk,
    )

    credit_cards_qs = Account.objects.filter(
        Q(account_type__icontains="credit") |
        Q(detail_type__icontains="credit card")
    ).order_by("account_name")
    bank_accounts_qs = deposit_accounts_qs()

    if request.method == "POST":
        try:
            with transaction.atomic():
                credit_card_id   = request.POST.get("credit_card") or ""
                bank_account_id  = request.POST.get("bank_account") or ""
                supplier_id      = request.POST.get("payee_supplier") or ""
                pdc.payee_name   = request.POST.get("payee_name") or ""
                pdc.payment_date = request.POST.get("payment_date") or timezone.localdate()
                pdc.ref_no       = request.POST.get("ref_no") or pdc.ref_no or ""
                pdc.location     = request.POST.get("location") or ""
                pdc.memo         = request.POST.get("memo") or ""

                amount_raw       = request.POST.get("amount") or "0"
                amount           = _dec(amount_raw)

                pdc.credit_card = get_object_or_404(credit_cards_qs, pk=credit_card_id)
                pdc.bank_account = get_object_or_404(bank_accounts_qs, pk=bank_account_id)
                pdc.payee_supplier = (
                    Newsupplier.objects.filter(pk=supplier_id).first()
                    if supplier_id else None
                )

                if request.FILES.get("attachments"):
                    pdc.attachments = request.FILES["attachments"]

                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    raise ValueError("Zero amount")
                pdc.amount = amount

                pdc.save()

                # _post_paydown_credit_to_ledger(pdc)

                action = request.POST.get("save_action") or "save"
                if action == "save&new":
                    return redirect("expenses:pay-down-credit")
                return redirect("expenses:paydown-credit-detail", pk=pdc.pk)

        except Exception as e:
            messages.error(request, f"Could not update payment: {e}")
            return redirect("expenses:paydown-credit-edit", pk=pdc.pk)

    context = {
        "payment": pdc,
        "credit_cards": credit_cards_qs,
        "bank_accounts": bank_accounts_qs,
        "suppliers": Newsupplier.objects.all().order_by("company_name"),
    }
    return render(request, "paydown_credit_form.html", context)

# pay down detail
def paydown_credit_detail(request, pk: int):
    pdc = get_object_or_404(
        PayDownCredit.objects.select_related(
            "credit_card", "bank_account", "payee_supplier"
        ),
        pk=pk,
    )
    return render(request, "paydown_credit_detail.html", {"p": pdc})
#
# pay down lists
def paydown_credit_list(request):
    qs = (
        PayDownCredit.objects
        .select_related("credit_card", "bank_account", "payee_supplier")
        .order_by("-payment_date", "-id")
    )

    # search
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(ref_no__icontains=q) |
            Q(location__icontains=q) |
            Q(payee_name__icontains=q) |
            Q(credit_card__account_name__icontains=q) |
            Q(bank_account__account_name__icontains=q) |
            Q(payee_supplier__company_name__icontains=q)
        )

    # date filters
    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")
    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    rows = list(qs)

    totals = qs.aggregate(
        grand=Coalesce(
            Sum("amount"),
            Value(Decimal("0.00"),
                  output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )
    totals["subtotal"] = Decimal("0.00")
    totals["tax"] = Decimal("0.00")

    return render(request, "paydown_credit_list.html", {
        "payments": rows,
        "totals": totals,
        "count_all": qs.count(),
        "q": q,
        "from": date_from,
        "to": date_to,
    })


# credit card credit 


@transaction.atomic
def add_credit_card_credit(request):
    # Use the same accounts as "Deposit to"
    deposit_accounts = deposit_accounts_qs()

    # expense/other accounts for category lines
    expense_accounts = Account.objects.filter(
        Q(account_type__icontains="expense")
        | Q(account_type__icontains="cost of goods sold")
        | Q(account_type__icontains="other expense")
    ).order_by("account_name")

    products = Product.objects.all().order_by("id")
    suppliers = Newsupplier.objects.all().order_by("company_name")
    customers = Newcustomer.objects.all().order_by("customer_name")
    classes = Pclass.objects.all().order_by("class_name") if "Pclass" in globals() else []

    if request.method == "POST":
        try:
            with transaction.atomic():
                credit_card_id = request.POST.get("credit_card") or ""
                supplier_id = request.POST.get("payee_supplier") or ""
                payee_name = request.POST.get("payee_name") or ""
                credit_date = request.POST.get("credit_date") or timezone.localdate()
                ref_no = request.POST.get("ref_no") or ""
                location = request.POST.get("location") or ""
                tags = request.POST.get("tags") or ""
                memo = request.POST.get("memo") or ""
                attachment = request.FILES.get("attachments")

                credit_card = get_object_or_404(deposit_accounts, pk=credit_card_id)
                supplier = (
                    Newsupplier.objects.filter(pk=supplier_id).first()
                    if supplier_id
                    else None
                )

                if not ref_no:
                    ref_no = generate_unique_ref_no()

                cc = CreditCardCredit.objects.create(
                    credit_card=credit_card,
                    payee_supplier=supplier,
                    payee_name=payee_name,
                    credit_date=credit_date,
                    ref_no=ref_no,
                    location=location,
                    tags=tags,
                    memo=memo,
                    attachments=attachment,
                )

                # ---- Category lines ----
                cat_categories = request.POST.getlist("cat_category[]")
                cat_descs = request.POST.getlist("cat_description[]")
                cat_amounts = request.POST.getlist("cat_amount[]")
                cat_billables = request.POST.getlist("cat_billable[]")
                cat_customers = request.POST.getlist("cat_customer[]")
                cat_classes = request.POST.getlist("cat_class[]")

                total_amount = 0

                for idx, cat_id in enumerate(cat_categories):
                    if not (cat_id or "").strip():
                        continue

                    category = get_object_or_404(expense_accounts, pk=cat_id)
                    desc = cat_descs[idx] if idx < len(cat_descs) else ""
                    amt_raw = cat_amounts[idx] if idx < len(cat_amounts) else "0"
                    amt = _dec(amt_raw)
                    if amt == 0:
                        continue

                    total_amount += amt

                    billable_flag = False
                    if idx < len(cat_billables):
                        billable_flag = cat_billables[idx] == "on"

                    cust_id = cat_customers[idx] if idx < len(cat_customers) else ""
                    cls_id = cat_classes[idx] if idx < len(cat_classes) else ""
                    customer = (
                        Newcustomer.objects.filter(pk=cust_id).first()
                        if cust_id
                        else None
                    )
                    pclass = (
                        Pclass.objects.filter(pk=cls_id).first()
                        if cls_id
                        else None
                    )

                    CreditCardCreditCategoryLine.objects.create(
                        credit=cc,
                        category=category,
                        description=desc,
                        amount=amt,
                        billable=billable_flag,
                        customer=customer,
                        pclass=pclass,
                    )

                # (If you later wire item lines, add their amounts into total_amount too)
                cc.total_amount = total_amount
                cc.save()

                # _post_credit_card_credit_to_ledger(cc)

                action = request.POST.get("save_action") or "save"
                if action == "save":
                    return redirect("expenses:credit-card-credit-list")
                if action == "save&new":
                    return redirect("expenses:credit-card")
                if action == "save&close":
                    return redirect("expenses:expenses")
                return redirect("expenses:credit-card-credit-list")

        except Exception as e:
            messages.error(request, f"Could not save credit card credit: {e}")

    context = {
        "credit": None,
        "deposit_accounts": deposit_accounts_qs(),
        "suppliers": suppliers,
        "expense_accounts": expense_accounts,
        "customers": customers,
        "classes": classes,
        "products": products,
        "today": timezone.localdate(),
        "ref_no": generate_unique_ref_no(),
    }
    return render(request, "credit_card_credit_form.html", context)


@transaction.atomic
def credit_card_credit_edit(request, pk: int):
    cc = get_object_or_404(
        CreditCardCredit.objects.select_related("credit_card", "payee_supplier"),
        pk=pk,
    )

    deposit_accounts = deposit_accounts_qs()

    expense_accounts = Account.objects.filter(
        Q(account_type__icontains="expense")
        | Q(account_type__icontains="cost of goods sold")
        | Q(account_type__icontains="other expense")
    ).order_by("account_name")

    products = Product.objects.all().order_by("id")
    suppliers = Newsupplier.objects.all().order_by("company_name")
    customers = Newcustomer.objects.all().order_by("customer_name")
    classes = Pclass.objects.all().order_by("class_name") if "Pclass" in globals() else []

    if request.method == "POST":
        try:
            with transaction.atomic():
                credit_card_id = request.POST.get("credit_card") or ""
                supplier_id = request.POST.get("payee_supplier") or ""
                cc.payee_name = request.POST.get("payee_name") or ""
                cc.credit_date = request.POST.get("credit_date") or timezone.localdate()
                cc.ref_no = request.POST.get("ref_no") or cc.ref_no or ""
                cc.location = request.POST.get("location") or ""
                cc.tags = request.POST.get("tags") or ""
                cc.memo = request.POST.get("memo") or ""

                cc.credit_card = get_object_or_404(deposit_accounts, pk=credit_card_id)
                cc.payee_supplier = (
                    Newsupplier.objects.filter(pk=supplier_id).first()
                    if supplier_id
                    else None
                )

                if request.FILES.get("attachments"):
                    cc.attachments = request.FILES["attachments"]

                cc.save()

                # wipe old lines, rebuild
                cc.category_lines.all().delete()
                cc.item_lines.all().delete()

                cat_categories = request.POST.getlist("cat_category[]")
                cat_descs = request.POST.getlist("cat_description[]")
                cat_amounts = request.POST.getlist("cat_amount[]")
                cat_billables = request.POST.getlist("cat_billable[]")
                cat_customers = request.POST.getlist("cat_customer[]")
                cat_classes = request.POST.getlist("cat_class[]")

                total_amount = 0

                for idx, cat_id in enumerate(cat_categories):
                    if not (cat_id or "").strip():
                        continue
                    category = get_object_or_404(expense_accounts, pk=cat_id)
                    desc = cat_descs[idx] if idx < len(cat_descs) else ""
                    amt_raw = cat_amounts[idx] if idx < len(cat_amounts) else "0"
                    amt = _dec(amt_raw)
                    if amt == 0:
                        continue

                    total_amount += amt

                    billable_flag = False
                    if idx < len(cat_billables):
                        billable_flag = cat_billables[idx] == "on"

                    cust_id = cat_customers[idx] if idx < len(cat_customers) else ""
                    cls_id = cat_classes[idx] if idx < len(cat_classes) else ""
                    customer = (
                        Newcustomer.objects.filter(pk=cust_id).first()
                        if cust_id
                        else None
                    )
                    pclass = (
                        Pclass.objects.filter(pk=cls_id).first()
                        if cls_id
                        else None
                    )

                    CreditCardCreditCategoryLine.objects.create(
                        credit=cc,
                        category=category,
                        description=desc,
                        amount=amt,
                        billable=billable_flag,
                        customer=customer,
                        pclass=pclass,
                    )

                cc.total_amount = total_amount
                cc.save()

                # _post_credit_card_credit_to_ledger(cc)

                action = request.POST.get("save_action") or "save"
                if action == "save&new":
                    return redirect("expenses:credit-card")
                if action == "save&close":
                    return redirect("expenses:expenses")
                return redirect("expenses:credit-card-credit-detail", pk=cc.pk)

        except Exception as e:
            messages.error(request, f"Could not update credit card credit: {e}")
            return redirect("expenses:credit-card-credit-edit", pk=cc.pk)

    context = {
        "credit": cc,
        "deposit_accounts": deposit_accounts,
        "suppliers": suppliers,
        "expense_accounts": expense_accounts,
        "customers": customers,
        "classes": classes,
        "products": products,
        "today": timezone.localdate(),
        "ref_no": generate_unique_ref_no(),
    }
    return render(request, "credit_card_credit_form.html", context)
# detail view 
def credit_card_credit_detail(request, pk: int):
    cc = get_object_or_404(
        CreditCardCredit.objects
        .select_related("credit_card", "payee_supplier")
        .prefetch_related("category_lines__category", "item_lines__product"),
        pk=pk,
    )
    return render(request, "credit_card_credit_detail.html", {"credit": cc})
# credit lists 

def credit_card_credit_list(request):
    qs = (
        CreditCardCredit.objects
        .select_related("credit_card", "payee_supplier")
        .order_by("-credit_date", "-id")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(ref_no__icontains=q) |
            Q(location__icontains=q) |
            Q(tags__icontains=q) |
            Q(payee_name__icontains=q) |
            Q(credit_card__account_name__icontains=q) |
            Q(payee_supplier__company_name__icontains=q)
        )

    date_from = request.GET.get("from", "")
    date_to   = request.GET.get("to", "")
    if date_from:
        qs = qs.filter(credit_date__gte=date_from)
    if date_to:
        qs = qs.filter(credit_date__lte=date_to)

    totals = qs.aggregate(
        grand=Coalesce(
            Sum("total_amount"),
            Value(Decimal("0.00"),
                  output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )
    totals["subtotal"] = totals["grand"] or Decimal("0.00")
    totals["tax"]      = Decimal("0.00")

    return render(request, "credit_card_credit_list.html", {
        "credits": qs,
        "totals": totals,
        "count_all": qs.count(),
        "q": q,
        "from": date_from,
        "to": date_to,
    })

# @login_required
def supplier_prepayments_list(request):
    suppliers = Newsupplier.objects.order_by("company_name")

    rows = []
    for s in suppliers:
        bal = _supplier_prepayment_balance(s)
        if bal and bal > Decimal("0.00"):
            rows.append({"supplier": s, "prepayment": bal})

    return render(request, "supplier_prepayments_list.html", {"rows": rows})


# @login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def supplier_refund_new(request, supplier_id: int):
    supplier = get_object_or_404(Newsupplier, pk=supplier_id)
    accounts = deposit_accounts_qs()  # ✅ Bank/Cash list

    max_refundable = _supplier_prepayment_balance(supplier)

    if request.method == "POST":
        refund_date = request.POST.get("refund_date") or ""
        received_to_id = (request.POST.get("received_to") or "").strip()
        amount = _dec(request.POST.get("amount") or "0.00")
        memo = (request.POST.get("memo") or "").strip()
        reference_no = (request.POST.get("reference_no") or "").strip()

        received_to = None
        if received_to_id.isdigit():
            received_to = accounts.filter(id=int(received_to_id)).first()

        if not received_to:
            return render(request, "supplier_refund_form.html", {
                "supplier": supplier,
                "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": "Select a valid Bank/Cash account.",
            })

        if amount <= 0:
            return render(request, "supplier_refund_form.html", {
                "supplier": supplier,
                "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": "Refund amount must be > 0.",
            })

        if amount > max_refundable:
            return render(request, "supplier_refund_form.html", {
                "supplier": supplier,
                "accounts": accounts,
                "max_refundable": max_refundable,
                "form_error": f"Refund exceeds available supplier prepayment ({max_refundable}).",
            })

        # parse date safely
        try:
            from django.utils.dateparse import parse_date
            d = parse_date(refund_date) or None
        except Exception:
            d = None

        refund = SupplierRefund.objects.create(
            supplier=supplier,
            refund_date=d,
            received_to=received_to,
            amount=amount,
            memo=memo,
            reference_no=reference_no,
        )

        _post_supplier_refund_to_ledger(refund)
        return redirect("expenses:supplier-prepayments-list")

    return render(request, "supplier_refund_form.html", {
        "supplier": supplier,
        "accounts": accounts,
        "max_refundable": max_refundable,
    })






def add_time_activity(request):
   
    return render(request, 'time_activity_form.html', {})

def supplier_credit(request):
   
    return render(request, 'supplier_credit_form.html', {})

def pay_down_credit(request):
   
    return render(request, 'pay_down_credit_form.html', {})

def import_bills(request):
   
    return render(request, 'import_bills_form.html', {})
def credit_card(request):
   
    return render(request, 'credit_card_credit_form.html', {})
